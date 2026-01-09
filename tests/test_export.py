import os
import sys
import django
from datetime import datetime

# Set up Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import RequestFactory
from django.contrib.auth import get_user_model
from main.views.core_views import export_sheet

print("=" * 100)
print("TESTING EXCEL EXPORT")
print("=" * 100)

# Create a mock request with authenticated user
factory = RequestFactory()
today = datetime.now().strftime('%Y-%m-%d')
request = factory.get('/export_sheet/', {
    'date_from': '2025-12-01',
    'date_to': today
})

# Add authenticated user
User = get_user_model()
try:
    user = User.objects.get(username='developer')
except:
    user = User.objects.filter(is_superuser=True).first()
    if not user:
        user = User.objects.first()

request.user = user

print(f"\nGenerating export for December 1 - {today}...")
print("-" * 100)

try:
    response = export_sheet(request)

    if response.status_code == 200:
        # Save the Excel file
        filename = 'test_export_output.xlsx'
        with open(filename, 'wb') as f:
            f.write(response.content)

        print(f"\n[OK] Export successful!")
        print(f"[OK] File saved as: {filename}")
        print(f"[OK] File size: {len(response.content)} bytes")

        # Read the Excel file and show first few rows
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        print(f"\n[OK] Total rows in export: {ws.max_row}")
        print(f"\n[OK] First 20 rows:")
        print("-" * 100)

        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx <= 20:
                # Show ContactName, InvoiceNumber, Description
                contact = row[0] if len(row) > 0 else ''
                inv_num = row[10] if len(row) > 10 else ''
                desc = row[12] if len(row) > 12 else ''
                print(f"{idx:3}. {contact:40} | {inv_num:10} | {desc[:60] if desc else ''}")

        print("\n" + "=" * 100)

    else:
        print(f"\n[ERROR] Export failed with status code: {response.status_code}")

except Exception as e:
    print(f"\n[ERROR] Error during export: {e}")
    import traceback
    traceback.print_exc()
