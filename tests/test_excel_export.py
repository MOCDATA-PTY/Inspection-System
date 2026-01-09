"""Test Excel export with proper headers using openpyxl"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

print("=" * 100)
print("TESTING EXCEL EXPORT WITH PROPER HEADERS")
print("=" * 100)

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoice Data"

# Define headers
headers = [
    '*ContactName', 'EmailAddress', 'POAddressLine1', 'POAddressLine2',
    'POAddressLine3', 'POAddressLine4', 'POCity', 'PORegion',
    'POPostalCode', 'POCountry', '*InvoiceNumber', 'Reference',
    '*InvoiceDate', '*DueDate', 'Total', 'InventoryItemCode',
    '*Description', '*Quantity', '*UnitAmount', 'Discount',
    '*AccountCode', '*TaxType', 'Tax Amount', 'TrackingName1',
    'TrackingOption1', 'TrackingName2', 'TrackingOption2',
    'Currency', 'BrandingTheme'
]

# Write headers
ws.append(headers)

# Format header row (bold)
for cell in ws[1]:
    cell.font = Font(bold=True)

# Add sample data rows
sample_data = [
    [
        'Test Client 1', 'test1@example.com', '', '', '', '',
        'Johannesburg', 'Gauteng', '2000', 'South Africa',
        'INV-001', 'REF-001', '2025-01-15', '2025-02-14', '',
        '', 'PMP 120 - Inspector Travel (Johannesburg)', '1', '150.00', '',
        '200', 'Tax Exempt (0%)', '', '', '', '', '', 'ZAR', ''
    ],
    [
        'Test Client 1', 'test1@example.com', '', '', '', '',
        'Johannesburg', 'Gauteng', '2000', 'South Africa',
        'INV-001', 'REF-001', '2025-01-15', '2025-02-14', '',
        '', 'PMP 062 - Fat Test (Lancet Labs)', '1', '450.00', '',
        '200', 'Tax Exempt (0%)', '', '', '', '', '', 'ZAR', ''
    ],
    [
        'Test Client 2', 'test2@example.com', '', '', '', '',
        'Pretoria', 'Gauteng', '0001', 'South Africa',
        'INV-002', 'REF-002', '2025-01-15', '2025-02-14', '',
        '', 'RAW 052 - Fat Test (Lancet Labs)', '1', '450.00', '',
        '200', 'Tax Exempt (0%)', '', '', '', '', '', 'ZAR', ''
    ],
]

for row_data in sample_data:
    ws.append(row_data)

# Set column widths
column_widths = [
    25, 20, 20, 20, 20, 20, 15, 15, 12, 15,
    20, 20, 12, 12, 12, 15, 50, 10, 12, 10,
    15, 25, 12, 15, 15, 15, 15, 10, 15
]

for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Create a table (this makes Excel recognize headers properly)
tab = Table(displayName="InvoiceTable", ref=f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(sample_data) + 1}")

# Add a default style with striped rows
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

# Add the table to the worksheet
ws.add_table(tab)

# Save the file
filename = f'test_excel_export_{date.today()}.xlsx'
wb.save(filename)

print(f"\n[OK] Excel file created: {filename}")
print("\nNow test it:")
print("1. Open the file in Excel")
print("2. Go to Data -> Sort")
print("3. Check if the sort dialog shows column names like:")
print("   - *ContactName")
print("   - EmailAddress")
print("   - *InvoiceNumber")
print("   - *Description")
print("   etc.")
print("\nIf it still shows 'Column A, Column B, Column C', then Excel has an issue.")
print("If it shows the actual names, then the SheetJS library is the problem.")
print("=" * 100)
