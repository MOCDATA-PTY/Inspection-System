#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Commodity Data to Excel
Pulls data from SQL Server for Poultry, RAW, Eggs, and PMP inspections
Creates two Excel files: one for approved records, one for not approved
"""

import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging
import sys
import os

# Fix encoding issues on Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class CommodityDataExporter:
    """Export commodity inspection data to formatted Excel files"""

    def __init__(self):
        # SQL Server connection details
        self.server = '102.67.140.12'
        self.port = '1053'
        self.database = 'AFS'
        self.username = 'FSAUser2'
        self.password = 'password'
        self.driver = 'ODBC Driver 17 for SQL Server'
        self.connection = None

        # Define queries for each commodity type
        self.queries = {
            'Poultry': """
                SELECT [Id]
                      ,[CreatedOn]
                      ,[ModifiedOn]
                      ,[IsActive]
                      ,[ClientId]
                      ,[NewClientName]
                      ,[ProductName]
                      ,[DateOfInspection]
                      ,[InspectionLocationTypeId]
                      ,[InspectorId]
                      ,[ProductTypeId]
                      ,[InspectedGradeId]
                      ,[PoultryTypeId]
                      ,[ClassDesignationId]
                      ,[AlternativeClassDesignationId]
                      ,[GradeId]
                      ,[RegistrationNumber]
                      ,[ClientManagerName]
                      ,[ClientManagerEmail]
                      ,[EmailAddress1]
                      ,[EmailAddress2]
                      ,[IsExported]
                      ,[UniqueReferenceNumber]
                      ,[OriginalDirectionGuidReference]
                      ,[GuidId]
                      ,[ExpireDate]
                      ,[DispensationExpireDate]
                      ,[DispensationNumber]
                      ,[IsApproved]
                      ,[StatusId]
                      ,[ApprovedById]
                      ,[CorrectByDate]
                      ,[OwnerManagerRepresentativeEmail]
                      ,[InspectionReasonTypeId]
                      ,[DirectionTypeId]
                      ,[Comments]
                      ,[ApproveOnDate]
                FROM [AFS].[dbo].[PoultryDirectionRecordTypes]
                WHERE [DateOfInspection] > '2025-09-30 00:00:00.000'
                ORDER BY [IsApproved] DESC, [DateOfInspection] DESC
            """,

            'RAW': """
                SELECT [Id]
                      ,[CreatedOn]
                      ,[ModifiedOn]
                      ,[IsActive]
                      ,[InspectionRecordLinkId]
                      ,[DateOfInspection]
                      ,[CorrectByDate]
                      ,[OwnerManagerRepresentativeName]
                      ,[OwnerManagerRepresentativeEmail]
                      ,[EmailAddress1]
                      ,[EmailAddress2]
                      ,[InspectorId]
                      ,[ClientId]
                      ,[NewClientName]
                      ,[ManufactureBBDate]
                      ,[BatchNumber]
                      ,[IsSent]
                      ,[GuidId]
                      ,[InspectionReasonTypeId]
                      ,[UniqueReferenceNumber]
                      ,[ProducerId]
                      ,[NewProducerName]
                      ,[ProductItemId]
                      ,[NewProductItemDetails]
                      ,[OriginalDirectionGuidReference]
                      ,[PrimarySampleSize]
                      ,[DispensationExpireDate]
                      ,[DispensationNumber]
                      ,[IsApproved]
                      ,[StatusId]
                      ,[ApprovedById]
                      ,[IsDirectionType]
                      ,[Comments]
                      ,[ApproveOnDate]
                      ,[ManufactureBBDateStamp]
                FROM [AFS].[dbo].[RawRMPDirectionRecordTypes]
                WHERE [DateOfInspection] > '2025-09-30 00:00:00.000'
                ORDER BY [IsApproved] DESC, [DateOfInspection] DESC
            """,

            'Eggs': """
                SELECT [Id]
                      ,[CreatedOn]
                      ,[ModifiedOn]
                      ,[IsActive]
                      ,[InspectionRecordLinkId]
                      ,[DateOfInspection]
                      ,[OwnerManagerRepresentativeName]
                      ,[OwnerManagerRepresentativeEmail]
                      ,[InspectorId]
                      ,[ClientId]
                      ,[ClientBranch]
                      ,[EggProducer]
                      ,[ProductRemovedQuantity]
                      ,[ProductRemovedQuantityText]
                      ,[BBDate]
                      ,[BatchNumber]
                      ,[IsExported]
                      ,[GuidId]
                      ,[InspectionReasonTypeId]
                      ,[UniqueReferenceNumber]
                      ,[IsMarkPackDirectivePartPresent]
                      ,[IsGradeSizeDirectivePartPresent]
                      ,[LabelCorrectByDate]
                      ,[QualityCorrectByDate]
                      ,[EmailAddress1]
                      ,[EmailAddress2]
                      ,[EggProducerId]
                      ,[DispensationExpireDate]
                      ,[DispensationNumber]
                      ,[SizeId]
                      ,[GradeId]
                      ,[IsApproved]
                      ,[StatusId]
                      ,[OriginalDirectionReference]
                      ,[ApprovedById]
                      ,[TraySizeId]
                      ,[Comments]
                      ,[ApproveOnDate]
                FROM [AFS].[dbo].[PoultryEggDirections]
                WHERE [DateOfInspection] > '2025-09-30 00:00:00.000'
                ORDER BY [IsApproved] DESC, [DateOfInspection] DESC
            """,

            'PMP': """
                SELECT [Id]
                      ,[CreatedOn]
                      ,[ModifiedOn]
                      ,[IsActive]
                      ,[InspectionRecordLinkId]
                      ,[DateOfInspection]
                      ,[CorrectByDate]
                      ,[OwnerManagerRepresentativeName]
                      ,[OwnerManagerRepresentativeEmail]
                      ,[EmailAddress1]
                      ,[EmailAddress2]
                      ,[InspectorId]
                      ,[ClientId]
                      ,[NewClientName]
                      ,[ManufactureBBDate]
                      ,[BatchNumber]
                      ,[IsSent]
                      ,[GuidId]
                      ,[InspectionReasonTypeId]
                      ,[UniqueReferenceNumber]
                      ,[ProducerId]
                      ,[NewProducerName]
                      ,[PMPItemId]
                      ,[NewPMPItemDetails]
                      ,[OriginalDirectionGuidReference]
                      ,[PrimarySampleSize]
                      ,[DispensationExpireDate]
                      ,[DispensationNumber]
                      ,[IsApproved]
                      ,[StatusId]
                      ,[ApprovedById]
                      ,[IsDirectionType]
                      ,[Comments]
                      ,[ApproveOnDate]
                FROM [AFS].[dbo].[PMPDirectionRecordTypes]
                WHERE [DateOfInspection] > '2025-09-30 00:00:00.000'
                ORDER BY [IsApproved] DESC, [DateOfInspection] DESC
            """
        }

    def get_connection_string(self):
        """Get the connection string for SQL Server"""
        return (
            f"DRIVER={{{self.driver}}};"
            f"SERVER={self.server},{self.port};"
            f"DATABASE={self.database};"
            f"UID={self.username};"
            f"PWD={self.password};"
            f"TrustServerCertificate=yes;"
            f"Encrypt=yes;"
        )

    def connect(self):
        """Establish connection to SQL Server"""
        try:
            connection_string = self.get_connection_string()
            self.connection = pyodbc.connect(connection_string, timeout=30)
            logger.info("✅ Successfully connected to SQL Server")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to connect to SQL Server: {e}")
            return False

    def disconnect(self):
        """Close SQL Server connection"""
        if self.connection:
            self.connection.close()
            self.connection = None
            logger.info("🔌 Disconnected from SQL Server")

    def fetch_commodity_data(self, commodity_name):
        """Fetch data for a specific commodity"""
        try:
            query = self.queries[commodity_name]
            logger.info(f"📥 Fetching {commodity_name} data...")

            # Use pandas to execute query and return DataFrame
            df = pd.read_sql(query, self.connection)

            logger.info(f"✅ Retrieved {len(df)} records for {commodity_name}")
            return df

        except Exception as e:
            logger.error(f"❌ Error fetching {commodity_name} data: {e}")
            return pd.DataFrame()

    def format_excel_sheet(self, writer, sheet_name, df, is_approved):
        """Apply nice formatting to the Excel sheet"""
        # Write the dataframe to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style for data cells
        cell_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border

                # Alternate row colors for better readability
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter

            # Calculate max length
            max_length = len(str(column))
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=col_num, max_col=col_num):
                try:
                    cell_value = str(row[0].value) if row[0].value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass

            # Set column width (with max limit)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

        logger.info(f"✅ Formatted {sheet_name} sheet")

    def export_to_excel(self):
        """Main export function"""
        try:
            # Connect to database
            if not self.connect():
                return False

            # Fetch all data
            all_data = {}
            for commodity in ['Poultry', 'RAW', 'Eggs', 'PMP']:
                df = self.fetch_commodity_data(commodity)
                if not df.empty:
                    all_data[commodity] = df

            if not all_data:
                logger.error("❌ No data retrieved from any commodity type")
                return False

            # Generate timestamp for filenames
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")

            # Create Excel file for APPROVED records (IsApproved = 1)
            approved_filename = f"Commodity_Approved_{timestamp}.xlsx"
            logger.info(f"📝 Creating approved records file: {approved_filename}")

            with pd.ExcelWriter(approved_filename, engine='openpyxl') as writer:
                for commodity, df in all_data.items():
                    # Filter for approved records
                    approved_df = df[df['IsApproved'] == 1].copy()

                    if not approved_df.empty:
                        # Keep original 1/0 values in IsApproved column
                        # Format and write to Excel
                        self.format_excel_sheet(writer, commodity, approved_df, is_approved=True)
                        logger.info(f"  ✓ {commodity}: {len(approved_df)} approved records")
                    else:
                        logger.info(f"  ⚠ {commodity}: No approved records found")

            logger.info(f"✅ Created: {approved_filename}")

            # Create Excel file for NOT APPROVED records (IsApproved = 0)
            not_approved_filename = f"Commodity_Not_Approved_{timestamp}.xlsx"
            logger.info(f"📝 Creating not approved records file: {not_approved_filename}")

            with pd.ExcelWriter(not_approved_filename, engine='openpyxl') as writer:
                for commodity, df in all_data.items():
                    # Filter for not approved records
                    not_approved_df = df[df['IsApproved'] == 0].copy()

                    if not not_approved_df.empty:
                        # Keep original 1/0 values in IsApproved column
                        # Format and write to Excel
                        self.format_excel_sheet(writer, commodity, not_approved_df, is_approved=False)
                        logger.info(f"  ✓ {commodity}: {len(not_approved_df)} not approved records")
                    else:
                        logger.info(f"  ⚠ {commodity}: No not approved records found")

            logger.info(f"✅ Created: {not_approved_filename}")

            # Print summary
            print("\n" + "="*60)
            print("📊 EXPORT SUMMARY")
            print("="*60)
            for commodity, df in all_data.items():
                approved_count = len(df[df['IsApproved'] == 1])
                not_approved_count = len(df[df['IsApproved'] == 0])
                total = len(df)
                print(f"{commodity:12} - Total: {total:4} | Approved: {approved_count:4} | Not Approved: {not_approved_count:4}")
            print("="*60)
            print(f"✅ Approved records exported to: {approved_filename}")
            print(f"✅ Not approved records exported to: {not_approved_filename}")
            print("="*60 + "\n")

            return True

        except Exception as e:
            logger.error(f"❌ Error during export: {e}")
            import traceback
            traceback.print_exc()
            return False

        finally:
            self.disconnect()


def main():
    """Main function"""
    print("\n" + "="*60)
    print("🚀 COMMODITY DATA EXPORT UTILITY")
    print("="*60)
    print("This script exports inspection data for:")
    print("  • Poultry")
    print("  • RAW (Raw RMP)")
    print("  • Eggs")
    print("  • PMP (Processed Meat Products)")
    print("\nData will be split into two Excel files:")
    print("  1. Approved records (IsApproved = 1)")
    print("  2. Not approved records (IsApproved = 0)")
    print("="*60 + "\n")

    exporter = CommodityDataExporter()
    success = exporter.export_to_excel()

    if success:
        print("✅ Export completed successfully!")
        return 0
    else:
        print("❌ Export failed. Please check the logs above.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
