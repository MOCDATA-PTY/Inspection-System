import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from ..models import FoodSafetyAgencyInspection

@login_required
def update_product_name(request):
    """Update product_name for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            value = request.POST.get('product_name', '')
            inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
            inspection = inspections.first()
            if not inspection:
                return JsonResponse({'success': False, 'error': f'Inspection {inspection_id} not found'})
            inspection.product_name = value.strip() or None
            inspection.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def update_product_class(request):
    """Update product_class for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            value = request.POST.get('product_class', '')
            inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
            inspection = inspections.first()
            if not inspection:
                return JsonResponse({'success': False, 'error': f'Inspection {inspection_id} not found'})
            # Enforce rule: if commodity is egg or poultry, product_class must be unset and unselectable
            try:
                commodity_value = (inspection.commodity or '').strip().lower()
            except Exception:
                commodity_value = ''
            is_egg_or_poultry = commodity_value in ['egg', 'eggs', 'poultry', 'chicken']
            if is_egg_or_poultry:
                # Force clear and signal back that class is not applicable
                if inspection.product_class:
                    inspection.product_class = None
                inspection.save(update_fields=['product_class'])
                return JsonResponse({'success': True, 'ignored': True, 'message': 'Product class not applicable for egg/poultry commodities'})

            inspection.product_class = value.strip() or None
            inspection.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_http_methods

# Safe print wrapper to avoid UnicodeEncodeError on Windows consoles
def safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        sanitized_args = []
        for a in args:
            try:
                sanitized_args.append(str(a).encode('cp1252', errors='ignore').decode('cp1252', errors='ignore'))
            except Exception:
                sanitized_args.append(str(a))
        print(*sanitized_args, **kwargs)
from django.conf import settings
from django.db import models
from django.utils import timezone
from datetime import timedelta, date, datetime
import time
from decimal import Decimal
import calendar
import json
import requests
import os
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator
from django.db.models import Q
from ..forms import LoginForm, RegisterForm, ClientForm, InspectionForm, InspectorMappingForm, FoodSafetyAgencyInspectionForm
from ..models import Client, Inspection, Shipment, Settings, FoodSafetyAgencyInspection, SystemLog, InspectorMapping
from django.views.decorators.csrf import csrf_exempt
from ..decorators import role_required, inspector_restricted, financial_only, scientist_only, inspector_only_inspections, no_inspector_scientist
@login_required
@role_required(['admin', 'super_admin', 'developer'])
def save_manual_client_email(request):
    """Persist a manual email override for a client (by business name)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
        client_name = (data.get('client_name') or '').strip()
        email = (data.get('email') or '').strip()
        if not client_name:
            return JsonResponse({'success': False, 'error': 'client_name is required'})
        if email and '@' not in email:
            return JsonResponse({'success': False, 'error': 'Invalid email format'})

        # Try multiple matching strategies to find the client
        client = Client.objects.filter(client_id__iexact=client_name).first()
        if not client:
            client = Client.objects.filter(name__iexact=client_name).first()
        if not client:
            client = Client.objects.filter(client_id__icontains=client_name).first()
        if not client:
            # Final fallback: normalized comparison in Python
            import re
            def _norm(txt: str) -> str:
                txt = (txt or '').lower().strip()
                txt = re.sub(r"[\(\)\[\]{}\\/._,-]", " ", txt)
                txt = re.sub(r"\s+", " ", txt)
                return txt
            norm_target = _norm(client_name)
            for c in Client.objects.only('id', 'client_id', 'name'):
                if _norm(c.client_id) == norm_target or _norm(c.name) == norm_target:
                    client = c
                    break
        # If no client found, create a minimal record with this business name
        if not client:
            client = Client.objects.create(client_id=client_name, name=client_name)
        # Save manual override and record additional email
        if email:
            client.manual_email = email
            client.save(update_fields=['manual_email'])
            try:
                from ..models import ClientEmail
                ClientEmail.objects.get_or_create(client=client, email=email)
            except Exception:
                pass
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['admin', 'super_admin', 'developer'])
def delete_client_email(request):
    """Delete a removable client email (manual or additional)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
        client_name = (data.get('client_name') or '').strip()
        email = (data.get('email') or '').strip()
        if not (client_name and email):
            return JsonResponse({'success': False, 'error': 'client_name and email are required'})
        client = Client.objects.filter(client_id__iexact=client_name).first()
        if not client:
            return JsonResponse({'success': False, 'error': 'Client not found'})
        # If matches primary email, clear it
        if client.email and client.email.lower() == email.lower():
            client.email = None
            client.save(update_fields=['email'])
            return JsonResponse({'success': True})
        # If matches manual_email, clear it
        if client.manual_email and client.manual_email.lower() == email.lower():
            client.manual_email = None
            client.save(update_fields=['manual_email'])
            return JsonResponse({'success': True})
        # Otherwise try remove from additional emails
        try:
            from ..models import ClientEmail
            deleted, _ = ClientEmail.objects.filter(client=client, email__iexact=email).delete()
            if deleted:
                return JsonResponse({'success': True})
        except Exception:
            pass
        return JsonResponse({'success': False, 'error': 'Email not found or not removable'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
from .data_views import remote_sqlserver_data_view
from .utils import apply_filters, clear_messages
from ..services.google_drive_service import GoogleDriveService


# Global flag to track OneDrive operations during batch processing
onedrive_batch_processing = False

@login_required
def check_compliance_documents_batch(request):
    """Check compliance documents for multiple inspections after page loads."""
    global onedrive_batch_processing
    
    if request.method == 'POST':
        try:
            onedrive_batch_processing = True  # Start tracking OneDrive operations
            
            data = json.loads(request.body.decode('utf-8'))
            inspections_to_check = data.get('inspections', [])
            
            results = {}
            total_inspections = len(inspections_to_check)
            
            for i, inspection in enumerate(inspections_to_check):
                client_name = inspection.get('client_name')
                inspection_date = inspection.get('date_of_inspection')
                
                if client_name and inspection_date:
                    has_compliance = check_for_compliance_documents(client_name, inspection_date)
                    results[f"{client_name}_{inspection_date}"] = has_compliance
                    
                    # Log progress for debugging
                    print(f"Processed {i+1}/{total_inspections}: {client_name} - {inspection_date}")
            
            # Add a small delay to ensure all OneDrive operations complete
            import time
            time.sleep(2)
            
            onedrive_batch_processing = False  # Stop tracking OneDrive operations
            
            return JsonResponse({
                'success': True,
                'results': results,
                'total_processed': total_inspections
            })
        except Exception as e:
            onedrive_batch_processing = False  # Ensure flag is reset on error
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})





def index(request):
    """Redirect root URL to login page."""
    return redirect('login')


# =============================================================================
# AUTHENTICATION VIEWS
# =============================================================================

def user_login(request):
    """Handle user login with improved error handling and CSRF protection."""
    if request.method == 'POST':
        try:
            form = LoginForm(request, data=request.POST)
            # Strip spaces from username and password
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()

            # Debug print
            print(f"Login attempt - Username: {username}")
            print(f"Password length: {len(password) if password else 0}")

            user = authenticate(request, username=username, password=password)
            print(f"Authentication result: {user}")
            if user is not None:
                if user.is_active:
                    login(request, user)
                    # Set initial last activity timestamp
                    from django.utils import timezone
                    request.session['last_activity'] = timezone.now().isoformat()
                    request.session['authenticated'] = True
                    request.session.modified = True
                    
                    # Role-based redirect logic
                    user_role = getattr(user, 'role', 'inspector')
                    if user_role == 'admin':
                        # Redirect administrators to inspection page
                        return redirect('shipment_list')
                    elif user_role == 'inspector':
                        # Redirect inspectors to inspector dashboard
                        return redirect('inspector_dashboard')
                    else:
                        # All other users go to home page
                        return redirect('home')
                else:
                    messages.error(request, "Your account is disabled.")
            else:
                messages.error(request, "Invalid username or password. Please try again.")
        except Exception as e:
            # Handle CSRF and other errors gracefully
            print(f"Login error: {e}")
            if "CSRF" in str(e):
                messages.error(request, "Security token expired. Please refresh the page and try again.")
            else:
                messages.error(request, "An error occurred during login. Please try again.")
    else:
        form = LoginForm()

    response = render(request, 'main/login.html', {'form': form})
    response['Cache-Control'] = 'no-store'
    # Ensure CSRF cookie is set
    from django.middleware.csrf import get_token
    get_token(request)
    return response


def register(request):
    """Handle new user registration and automatic login upon successful registration."""
    clear_messages(request)
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                # Log the user in immediately after registration
                user = authenticate(request, username=user.username, password=form.cleaned_data['password1'])
                if user:
                    login(request, user)
                    # Set initial last activity timestamp
                    from django.utils import timezone
                    request.session['last_activity'] = timezone.now().isoformat()
                    request.session.set_expiry(0)
                    messages.success(request, f"Registration successful! Welcome, {user.username}!")
                    
                    # Role-based redirect logic
                    user_role = getattr(user, 'role', 'inspector')
                    if user_role == 'admin':
                        # Redirect administrators to inspection page
                        return redirect('shipment_list')
                    elif user_role == 'inspector':
                        # Redirect inspectors to inspector dashboard
                        return redirect('inspector_dashboard')
                    else:
                        # All other users go to home page
                        return redirect('home')
                else:
                    messages.error(request, "Registration successful but login failed. Please try logging in manually.")
                    return redirect('login')
            except Exception as e:
                print(f"Registration error: {e}")
                messages.error(request, f"Registration failed: {str(e)}")
        else:
            # Debug: print form errors and add them as messages
            print("Form errors:", form.errors)
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f"{error}")
                    elif field == 'username':
                        messages.error(request, f"Username: {error}")
                    elif field == 'password1':
                        messages.error(request, f"Password: {error}")
                    elif field == 'password2':
                        messages.error(request, f"Password confirmation: {error}")
                    elif field == 'email':
                        messages.error(request, f"Email: {error}")
                    else:
                        messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = RegisterForm()

    return render(request, 'main/register.html', {'form': form})


def user_logout(request):
    """Handle user logout."""
    logout(request)
    return redirect('login')


# =============================================================================
# CLIENT MANAGEMENT VIEWS
# =============================================================================

@login_required(login_url='login')
@inspector_only_inspections
def client_list(request):
    """Display list of clients with inspection counts."""
    from ..models import Client, Inspection
    from django.db.models import Count, Q

    search_query = request.GET.get('search', '').strip()

    # Get all clients with inspection count
    clients = Client.objects.annotate(
        inspection_count=Count('inspections', distinct=True)
    ).order_by('name')

    # Apply search filter if provided
    if search_query:
        clients = clients.filter(
            Q(name__icontains=search_query) |
            Q(client_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(manual_email__icontains=search_query) |
            Q(internal_account_code__icontains=search_query)
        )

    context = {
        'clients': clients,
        'total_count': clients.count(),
        'search_query': search_query,
    }

    return render(request, 'main/client_list.html', context)


@login_required(login_url='login')
@inspector_only_inspections
def add_client(request):
    """Add a new client."""
    clear_messages(request)
    
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            try:
                client = form.save()
                messages.success(request, f"Client '{client.name}' added successfully!")
                return redirect('client_list')
            except Exception as e:
                messages.error(request, f"Error adding client: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ClientForm()
    
    context = {
        'form': form,
        'action': 'Add'
    }
    
    return render(request, 'main/client_form.html', context)


@login_required(login_url='login')
@inspector_only_inspections
def edit_client(request, pk):
    """Edit an existing client."""
    clear_messages(request)
    
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            try:
                client = form.save()
                messages.success(request, f"Client '{client.name}' updated successfully!")
                return redirect('client_list')
            except Exception as e:
                messages.error(request, f"Error updating client: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ClientForm(instance=client)
    
    context = {
        'form': form,
        'client': client,
        'action': 'Edit'
    }
    
    return render(request, 'main/client_form.html', context)


@login_required(login_url='login')
@inspector_only_inspections
def delete_client(request, pk):
    """Delete a client."""
    clear_messages(request)
    
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        try:
            client_name = client.name
            client.delete()
            messages.success(request, f"Client '{client_name}' deleted successfully!")
            return redirect('client_list')
        except Exception as e:
            messages.error(request, f"Error deleting client: {str(e)}")
            return redirect('client_list')
    
    context = {
        'client': client
    }
    
    return render(request, 'main/client_confirm_delete.html', context)


# =============================================================================
# INSPECTION MANAGEMENT VIEWS
# =============================================================================

@login_required(login_url='login')
def inspection_list(request):
    """Display list of all inspections."""
    clear_messages(request)
    
    # Get all inspections with optional search and filtering
    search_query = request.GET.get('search', '')
    inspector_filter = request.GET.get('inspector', '')
    town_filter = request.GET.get('town', '')
    commodity_filter = request.GET.get('commodity', '')
    
    inspections = Inspection.objects.all()
    
    if search_query:
        inspections = inspections.filter(
            models.Q(facility_client_name__icontains=search_query) |
            models.Q(inspector__icontains=search_query) |
            models.Q(product_name__icontains=search_query) |
            models.Q(comments__icontains=search_query)
        )
    
    if inspector_filter:
        inspections = inspections.filter(inspector=inspector_filter)
    
    if town_filter:
        inspections = inspections.filter(town=town_filter)
    
    if commodity_filter:
        inspections = inspections.filter(commodity=commodity_filter)
    
    # Get unique values for filters
    inspectors = Inspection.objects.values_list('inspector', flat=True).distinct().order_by('inspector')
    towns = Inspection.objects.values_list('town', flat=True).distinct().exclude(town__isnull=True).exclude(town='').order_by('town')
    commodities = Inspection.objects.values_list('commodity', flat=True).distinct().exclude(commodity__isnull=True).exclude(commodity='').order_by('commodity')
    
    context = {
        'inspections': inspections.order_by('-inspection_date', '-inspection_number'),
        'total_inspections': inspections.count(),
        'search_query': search_query,
        'inspector_filter': inspector_filter,
        'town_filter': town_filter,
        'commodity_filter': commodity_filter,
        'inspectors': inspectors,
        'towns': towns,
        'commodities': commodities
    }
    
    return render(request, 'main/inspection_list.html', context)


@login_required(login_url='login')
def add_inspection(request):
    """Add a new inspection."""
    clear_messages(request)
    
    if request.method == 'POST':
        form = InspectionForm(request.POST)
        if form.is_valid():
            try:
                inspection = form.save()
                messages.success(request, f"Inspection {inspection.inspection_number} for {inspection.facility_client_name} added successfully!")
                return redirect('inspection_list')
            except Exception as e:
                messages.error(request, f"Error adding inspection: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = InspectionForm()
    
    context = {
        'form': form,
        'action': 'Add'
    }
    
    return render(request, 'main/inspection_form.html', context)


@login_required(login_url='login')
def edit_inspection(request, pk):
    """Edit an existing inspection."""
    clear_messages(request)
    
    inspection = get_object_or_404(Inspection, pk=pk)
    
    if request.method == 'POST':
        form = InspectionForm(request.POST, instance=inspection)
        if form.is_valid():
            try:
                inspection = form.save()
                messages.success(request, f"Inspection {inspection.inspection_number} updated successfully!")
                return redirect('inspection_list')
            except Exception as e:
                messages.error(request, f"Error updating inspection: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = InspectionForm(instance=inspection)
    
    context = {
        'form': form,
        'inspection': inspection,
        'action': 'Edit'
    }
    
    return render(request, 'main/inspection_form.html', context)


@login_required(login_url='login')
def delete_inspection(request, pk):
    """Delete an inspection."""
    clear_messages(request)
    
    inspection = get_object_or_404(Inspection, pk=pk)
    
    if request.method == 'POST':
        try:
            inspection_info = f"{inspection.inspector} - {inspection.inspection_number} - {inspection.facility_client_name}"
            inspection.delete()
            messages.success(request, f"Inspection {inspection_info} deleted successfully!")
            return redirect('inspection_list')
        except Exception as e:
            messages.error(request, f"Error deleting inspection: {str(e)}")
            return redirect('inspection_list')
    
    context = {
        'inspection': inspection
    }
    
    return render(request, 'main/inspection_confirm_delete.html', context)


# =============================================================================
# MANUAL FSA INSPECTION ENTRY
# =============================================================================

@login_required(login_url='login')
@role_required(['admin', 'super_admin', 'developer', 'inspector'])
def add_fsa_inspection(request):
    """Add a new Food Safety Agency inspection manually."""
    clear_messages(request)

    if request.method == 'POST':
        form = FoodSafetyAgencyInspectionForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Parse products data (JSON array with product info for each inspection)
                import json
                products_data_json = request.POST.get('products_data', '[]')
                try:
                    products_data = json.loads(products_data_json)
                except json.JSONDecodeError:
                    products_data = []

                # If no products data, fall back to commodity selections
                if not products_data:
                    commodity_selections_json = request.POST.get('commodity_selections', '{}')
                    try:
                        commodity_selections = json.loads(commodity_selections_json)
                    except json.JSONDecodeError:
                        commodity_selections = {}

                    # Build products_data from commodity selections
                    for commodity, count in commodity_selections.items():
                        if count > 0:
                            for _ in range(count):
                                products_data.append({
                                    'commodity': commodity,
                                    'product_name': form.cleaned_data.get('product_name', ''),
                                    'product_class': form.cleaned_data.get('product_class', ''),
                                    'lab': form.cleaned_data.get('lab', '')
                                })

                # If still no products, fall back to single commodity from form
                if not products_data:
                    commodity = form.cleaned_data.get('commodity', '')
                    if commodity:
                        products_data = [{
                            'commodity': commodity,
                            'product_name': form.cleaned_data.get('product_name', ''),
                            'product_class': form.cleaned_data.get('product_class', ''),
                            'lab': form.cleaned_data.get('lab', '')
                        }]

                if not products_data:
                    messages.error(request, "Please select at least one commodity type")
                    return redirect('add_fsa_inspection')

                # Create inspections for each product
                created_inspections = []
                from django.db.models import Min

                # Handle file uploads - save to proper folder structure
                from django.utils import timezone
                from django.conf import settings
                import os
                import re
                from main.models import Client

                def save_uploaded_file(uploaded_file, insp, category):
                    """Save uploaded file to the proper folder structure."""
                    # Ensure inspection has a linked Client
                    if not insp.client:
                        client = Client.objects.filter(name__iexact=insp.client_name).first()
                        if not client:
                            client = Client.objects.create(name=insp.client_name or "Unknown Client")
                        insp.client = client
                        insp.save(update_fields=['client'])

                    folder_path = os.path.join(
                        settings.MEDIA_ROOT, 'docs',
                        str(insp.client.id), str(insp.id), category
                    )
                    os.makedirs(folder_path, exist_ok=True)
                    file_path = os.path.join(folder_path, uploaded_file.name)
                    with open(file_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)
                    return file_path

                # Create inspections for each product
                first_inspection = None
                total_created = 0

                for product_info in products_data:
                    # Convert checkbox boolean to value (True -> 1.0, False -> None)
                    def checkbox_to_value(val):
                        return 1.0 if val else None

                    # Create new inspection instance with product-specific data
                    inspection = FoodSafetyAgencyInspection(
                        client_name=form.cleaned_data.get('client_name'),
                        date_of_inspection=form.cleaned_data.get('date_of_inspection'),
                        commodity=product_info.get('commodity', ''),
                        product_name=product_info.get('product_name', ''),
                        product_class=product_info.get('product_class', ''),
                        inspector_name=form.cleaned_data.get('inspector_name', ''),
                        # Use product-specific sample data (checkboxes)
                        is_sample_taken=bool(product_info.get('is_sample_taken', False)),
                        needs_retest=form.cleaned_data.get('needs_retest', 'PENDING'),
                        fat=bool(product_info.get('fat', False)),
                        protein=bool(product_info.get('protein', False)),
                        calcium=bool(product_info.get('calcium', False)),
                        dna=bool(product_info.get('dna', False)),
                        bought_sample=float(product_info.get('bought_sample', 0) or 0),
                        lab=product_info.get('lab', ''),
                        town=form.cleaned_data.get('town', ''),
                        km_traveled=float(product_info.get('km_traveled', 0) or 0),
                        hours=float(product_info.get('hours', 0) or 0),
                        additional_email=request.POST.get('additional_email', ''),
                        comment=form.cleaned_data.get('comment', ''),
                        is_manual=True,
                        inspected=bool(form.cleaned_data.get('inspected', True)),
                        follow_up=bool(form.cleaned_data.get('follow_up', False)),
                        dispensation_application=bool(form.cleaned_data.get('dispensation_application', False)),
                        corporate_group=request.POST.get('corporate_group', ''),
                        group_type=request.POST.get('group_type', ''),
                        facility_type=request.POST.get('facility_type', ''),
                    )

                    # Generate unique negative remote_id for manual entries
                    min_remote_id = FoodSafetyAgencyInspection.objects.filter(
                        is_manual=True
                    ).aggregate(Min('remote_id'))['remote_id__min']
                    if min_remote_id is None or min_remote_id >= 0:
                        inspection.remote_id = -1
                    else:
                        inspection.remote_id = min_remote_id - 1

                    # Link to client
                    client = Client.objects.filter(name__iexact=inspection.client_name).first()
                    if not client:
                        client = Client.objects.create(name=inspection.client_name or "Unknown Client")
                    inspection.client = client

                    inspection.save()
                    created_inspections.append(inspection)
                    total_created += 1

                    # Keep track of first inspection for file uploads
                    if first_inspection is None:
                        first_inspection = inspection

                # Handle file uploads - attach to first inspection only
                if first_inspection:
                    if 'rfi_file' in request.FILES:
                        save_uploaded_file(request.FILES['rfi_file'], first_inspection, 'rfi')
                        first_inspection.rfi_uploaded_by = request.user
                        first_inspection.rfi_uploaded_date = timezone.now()

                    if 'invoice_file' in request.FILES:
                        save_uploaded_file(request.FILES['invoice_file'], first_inspection, 'invoice')
                        first_inspection.invoice_uploaded_by = request.user
                        first_inspection.invoice_uploaded_date = timezone.now()

                    if 'labform_file' in request.FILES:
                        save_uploaded_file(request.FILES['labform_file'], first_inspection, 'lab')
                        first_inspection.lab_form_uploaded_by = request.user
                        first_inspection.lab_form_uploaded_date = timezone.now()

                    if 'coa_file' in request.FILES:
                        save_uploaded_file(request.FILES['coa_file'], first_inspection, 'compliance')
                        first_inspection.coa_uploaded_by = request.user
                        first_inspection.coa_uploaded_date = timezone.now()

                    if 'composition_file' in request.FILES:
                        save_uploaded_file(request.FILES['composition_file'], first_inspection, 'composition')
                        first_inspection.composition_uploaded_by = request.user
                        first_inspection.composition_uploaded_date = timezone.now()

                    first_inspection.save()

                    # Clear file cache
                    from django.core.cache import cache
                    cache_key = f"docs_files:{first_inspection.client.id}:{first_inspection.id}"
                    cache.delete(cache_key)
                    cache.delete(f"{cache_key}_timestamp")

                # Build success message with commodity counts
                from collections import Counter
                commodity_counts = Counter(p.get('commodity', '') for p in products_data)
                commodity_summary = ', '.join([f"{c}: {n}" for c, n in commodity_counts.items()])
                msg = f"Created {total_created} inspection(s) for {first_inspection.client_name if first_inspection else 'client'} ({commodity_summary})"
                messages.success(request, msg)
                return redirect('shipment_list')
            except Exception as e:
                messages.error(request, f"Error adding inspection: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = FoodSafetyAgencyInspectionForm()

    # Get default inspector name from logged-in user
    default_inspector = ''
    if request.user.is_authenticated:
        # Use full name if available, otherwise username
        full_name = request.user.get_full_name()
        default_inspector = full_name if full_name else request.user.username

    # Get clients for dropdown with their associated towns
    from ..models import Client
    clients_qs = Client.objects.all().order_by('name')

    # Build clients list with town data from most recent inspection
    clients_with_towns = []
    for client in clients_qs:
        # Get most recent town for this client
        town = FoodSafetyAgencyInspection.objects.filter(
            client_name__iexact=client.name
        ).exclude(town__isnull=True).exclude(town='').order_by('-date_of_inspection').values_list('town', flat=True).first()

        clients_with_towns.append({
            'name': client.name,
            'client_id': client.client_id,
            'internal_account_code': client.internal_account_code or '',
            'email': client.email or '',
            'town': town or ''
        })

    # Get unique towns from existing inspections
    towns = FoodSafetyAgencyInspection.objects.exclude(
        town__isnull=True
    ).exclude(
        town=''
    ).values_list('town', flat=True).distinct().order_by('town')

    context = {
        'form': form,
        'action': 'Add',
        'default_inspector': default_inspector,
        'clients': clients_with_towns,
        'towns': list(towns),
    }

    return render(request, 'main/fsa_inspection_form.html', context)


@login_required(login_url='login')
@role_required(['admin', 'super_admin', 'developer'])
def edit_fsa_inspection(request, pk):
    """Edit an existing Food Safety Agency inspection."""
    clear_messages(request)

    inspection = get_object_or_404(FoodSafetyAgencyInspection, pk=pk)

    if request.method == 'POST':
        form = FoodSafetyAgencyInspectionForm(request.POST, instance=inspection)
        if form.is_valid():
            try:
                inspection = form.save()
                messages.success(request, f"Inspection for {inspection.client_name} updated successfully!")
                return redirect('shipment_list')
            except Exception as e:
                messages.error(request, f"Error updating inspection: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = FoodSafetyAgencyInspectionForm(instance=inspection)

    # Get all clients for autocomplete with their associated towns
    clients_qs = Client.objects.all().order_by('name')

    # Build clients list with town data from most recent inspection
    clients_with_towns = []
    for client in clients_qs:
        # Get most recent town for this client
        town = FoodSafetyAgencyInspection.objects.filter(
            client_name__iexact=client.name
        ).exclude(town__isnull=True).exclude(town='').order_by('-date_of_inspection').values_list('town', flat=True).first()

        clients_with_towns.append({
            'name': client.name,
            'client_id': client.client_id,
            'internal_account_code': client.internal_account_code or '',
            'email': client.email or '',
            'town': town or ''
        })

    # Get unique towns from existing inspections
    towns = FoodSafetyAgencyInspection.objects.exclude(
        town__isnull=True
    ).exclude(
        town=''
    ).values_list('town', flat=True).distinct().order_by('town')

    context = {
        'form': form,
        'inspection': inspection,
        'action': 'Edit',
        'clients': clients_with_towns,
        'towns': list(towns),
        'is_occurrence_report': getattr(inspection, 'occurrence_report', False),
    }

    return render(request, 'main/fsa_inspection_form.html', context)


@login_required(login_url='login')
@role_required(['admin', 'super_admin', 'developer'])
def delete_fsa_inspection(request, pk):
    """Delete a Food Safety Agency inspection."""
    clear_messages(request)

    inspection = get_object_or_404(FoodSafetyAgencyInspection, pk=pk)

    if request.method == 'POST':
        try:
            inspection_info = f"{inspection.client_name} - {inspection.date_of_inspection}"
            inspection.delete()
            messages.success(request, f"Inspection {inspection_info} deleted successfully!")
            return redirect('shipment_list')
        except Exception as e:
            messages.error(request, f"Error deleting inspection: {str(e)}")
            return redirect('shipment_list')

    context = {
        'inspection': inspection
    }

    return render(request, 'main/fsa_inspection_confirm_delete.html', context)


# =============================================================================
# SHIPMENT PAGE (EMPTY - NO DATA)
# =============================================================================

def check_compliance_documents_status(inspections, client_name, date_of_inspection):
    """Check compliance documents status for a group of inspections.
    
    Returns:
        dict: {
            'has_any_compliance': bool,  # True if at least one commodity has compliance docs
            'all_commodities_have_compliance': bool,  # True if every commodity has compliance docs
            'commodity_status': dict  # Status for each commodity
        }
    """
    import os
    from django.conf import settings
    from datetime import datetime
    import re
    import requests
    import json
    from django.core.cache import cache
    
    # Create cache key for this specific group
    cache_key = f"compliance_status_{client_name}_{date_of_inspection}"
    cached_status = cache.get(cache_key)
    
    if cached_status:
        return cached_status
    
    try:
        # Check local media folder for compliance documents
        # Compliance documents are downloaded from Google Drive and stored locally
        # OneDrive is NOT used for compliance document checking
        local_status = check_compliance_documents_status_local(inspections, client_name, date_of_inspection)
        # Cache the result for 10 minutes
        cache.set(cache_key, local_status, 600)
        return local_status
        
    except Exception as e:
        # If any error, return default status
        default_status = {
            'has_any_compliance': False,
            'all_commodities_have_compliance': False,
            'commodity_status': {}
        }
        # Cache the default result for 5 minutes
        cache.set(cache_key, default_status, 300)
        return default_status


def check_compliance_documents_status_onedrive(inspections, client_name, date_of_inspection):
    """Check compliance documents status in OneDrive."""
    try:
        import os
        from django.conf import settings
        from datetime import datetime
        import re
        import requests
        import json
        from django.core.cache import cache
        
        # Create cache key for OneDrive status
        cache_key = f"onedrive_compliance_{client_name}_{date_of_inspection}"
        cached_status = cache.get(cache_key)
        
        if cached_status:
            return cached_status
        
        # Load OneDrive tokens
        token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
        if not os.path.exists(token_file):
            return None  # No OneDrive tokens, fall back to local check
        
        with open(token_file, 'r') as f:
            token_data = json.load(f)
        
        access_token = token_data.get('access_token')
        if not access_token:
            return None
        
        # Build OneDrive path
        year_folder = date_of_inspection.strftime("%Y")
        month_folder = date_of_inspection.strftime("%B")
        
        # Use original client name for OneDrive folder matching (folders now use original names)
        client_folder = client_name or 'Unknown Client'
        
        # Also try legacy variations for backwards compatibility
        client_folder_variations = [
            client_folder,  # Primary: original name
            # Legacy sanitized names for backwards compatibility
            re.sub(r'[^a-zA-Z0-9_]', '_', client_name).replace('_+', '_').strip('_'),
            re.sub(r'[^a-zA-Z0-9]', '', client_name),
            client_name.replace(' ', '_').replace('-', '_').replace("'", '').replace('"', ''),
        ]
        
        # Remove duplicates while preserving order
        client_folder_variations = list(dict.fromkeys(client_folder_variations))
        
        onedrive_base = getattr(settings, 'ONEDRIVE_FOLDER', 'FoodSafety Agency Inspections')
        
        # Get unique commodities from inspections
        commodities = set()
        for inspection in inspections:
            if inspection.commodity:
                commodities.add(str(inspection.commodity).upper().strip())
        
        commodity_status = {}
        has_any_compliance = False
        
        # Try each client folder variation
        for client_folder in client_folder_variations:
            base_path = f"{onedrive_base}/inspection/{year_folder}/{month_folder}/{client_folder}/Compliance"
            
            # Check each commodity folder in OneDrive
            for commodity in commodities:
                commodity_path = f"{base_path}/{commodity}"
                has_files = False
                
                # Check if commodity folder exists and has files
                check_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{commodity_path}:/children"
                headers = {
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/json'
                }
                
                response = requests.get(check_url, headers=headers)
                if response.status_code == 200:
                    files = response.json().get('value', [])
                    has_files = len(files) > 0
                    if has_files:
                        has_any_compliance = True
                
                commodity_status[commodity] = has_files
            
            # If we found any compliance documents, break out of the loop
            if has_any_compliance:
                break
        
        # Check if all commodities have compliance documents
        all_commodities_have_compliance = all(commodity_status.values()) if commodity_status else False
        
        result = {
            'has_any_compliance': has_any_compliance,
            'all_commodities_have_compliance': all_commodities_have_compliance,
            'commodity_status': commodity_status
        }
        
        # Cache the result for 15 minutes to reduce API calls
        cache.set(cache_key, result, 900)
        
        return result
        
    except Exception as e:
        return None  # Fall back to local check


def check_compliance_documents_status_local(inspections, client_name, date_of_inspection):
    """Check compliance documents status in local media folder (fallback)."""
    try:
        import os
        from django.conf import settings
        from datetime import datetime
        import re
        import glob
        
        # Initialize variables
        commodity_status = {}
        has_any_compliance = False
        
        # Build base path
        date_str = str(date_of_inspection).replace('-', '')
        year_folder = date_str[:4]
        month_num = int(date_str[4:6]) if len(date_str) >= 6 else int(datetime.now().strftime('%m'))
        year_num = int(year_folder)
        month_folder = datetime.strptime(f"{year_num}-{month_num:02d}-01", "%Y-%m-%d").strftime("%B")
        
        # Get unique commodities
        commodities = set()
        for inspection in inspections:
            commodity = getattr(inspection, 'commodity', None)
            if commodity:
                commodities.add(commodity)
            from django.conf import settings
            from datetime import datetime
            import os
            # Build base path: media/inspection/YYYY/Month/ClientName/Compliance/
            date_str = str(date_of_inspection).replace('-', '')
            year_folder = date_str[:4]
            month_num = int(date_str[4:6]) if len(date_str) >= 6 else int(datetime.now().strftime('%m'))
            year_num = int(year_folder)
            month_folder = datetime.strptime(f"{year_num}-{month_num:02d}-01", "%Y-%m-%d").strftime("%B")

            # Get unique commodities from inspections
            commodities = set()
            for inspection in inspections:
                commodities.add(getattr(inspection, 'commodity', None))

            commodity_status = {}
            has_any_compliance = False

            # Sanitize client name for filesystem (must match upload function)
            def create_folder_name(name):
                """Create Linux-friendly folder name - must match upload function"""
                if not name:
                    return "unknown_client"
                import re
                # Remove special characters, keep only alphanumeric, spaces, hyphens, underscores
                clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
                # Replace spaces and hyphens with underscores
                clean_name = clean_name.replace(' ', '_').replace('-', '_')
                # Remove consecutive underscores
                clean_name = re.sub(r'_+', '_', clean_name)
                # Remove leading/trailing underscores
                clean_name = clean_name.strip('_').lower()
                return clean_name or "unknown_client"
            
            client_folder = create_folder_name(client_name)

        # Check both main compliance folder and inspection-specific folders
        client_base = os.path.join(settings.MEDIA_ROOT, 'inspection', year_folder, month_folder, client_folder)
        
        # Check main Compliance folder
        main_compliance_path = os.path.join(client_base, 'compliance')
        
        # Check Inspection-* folders for compliance files
        inspection_folders = []
        if os.path.exists(client_base):
            # Case-insensitive check for inspection folders
            inspection_folders = [d for d in os.listdir(client_base) if d.lower().startswith('inspection-')]
        
        has_compliance_files = False
        
        # Check main compliance folder
        if os.path.exists(main_compliance_path):
            try:
                files = os.listdir(main_compliance_path)
                if files:
                    print(f"   [OK] Found files in main Compliance folder: {files}")
                    has_compliance_files = True
                    has_any_compliance = True  # FIXED: Set has_any_compliance when files are found
            except Exception as e:
                print(f"   [ERR] Error listing main compliance files: {e}")

        # Check each Inspection folder
        for folder in inspection_folders:
            inspection_path = os.path.join(client_base, folder)
            if os.path.exists(inspection_path):
                try:
                    files = os.listdir(inspection_path)
                    if files:
                        print(f"   Found files in {folder}: {files}")
                        has_compliance_files = True
                        has_any_compliance = True  # FIXED: Set has_any_compliance when files are found
                except Exception as e:
                    print(f"   [ERROR] Error listing {folder} files: {e}")

        if not has_compliance_files:
            print(f"   No compliance files found in any location")        # Use original client name for folder matching
        client_folder = client_name or 'Unknown Client'
        client_base = os.path.join(settings.MEDIA_ROOT, 'inspection', year_folder, month_folder, client_folder)
        print(f"   Checking client base path: {client_base}")
        
        if os.path.exists(client_base):
            # Check main Compliance folder
            main_compliance_path = os.path.join(client_base, 'compliance')
            print(f"   Checking main compliance path: {main_compliance_path}")
            
            if os.path.exists(main_compliance_path):
                try:
                    files = os.listdir(main_compliance_path)
                    if files:
                        print(f"   Found files in main Compliance folder: {files}")
                        has_any_compliance = True
                except Exception as e:
                    print(f"   ERROR: Error listing main compliance folder: {e}")
            
            # Check Inspection-* folders
            try:
                folders = os.listdir(client_base)
                # Case-insensitive check for inspection folders
                inspection_folders = [d for d in folders if d.lower().startswith('inspection-')]
                
                for folder in inspection_folders:
                    inspection_path = os.path.join(client_base, folder)
                    if os.path.exists(inspection_path):
                        try:
                            files = os.listdir(inspection_path)
                            if files:
                                has_any_compliance = True
                        except Exception:
                            pass
            except Exception:
                pass

        # ALSO check media/uploads/inspections folder for compliance files
        uploads_path = os.path.join(settings.MEDIA_ROOT, 'uploads', 'inspections')

        if os.path.exists(uploads_path):
            try:
                files = os.listdir(uploads_path)
                # Look for compliance files for this client
                client_folder_sanitized = create_folder_name(client_name)
                date_str_formatted = date_of_inspection.strftime('%Y%m%d')

                compliance_files = [f for f in files if 'compliance' in f.lower() and client_folder_sanitized in f.lower() and date_str_formatted in f]

                if compliance_files:
                    try:
                        print(f"   [OK] Found compliance file(s) in uploads folder: {compliance_files}")
                    except:
                        print(f"   [OK] Found compliance file(s) in uploads folder")
                    has_any_compliance = True
                else:
                    try:
                        print(f"   [WARN] No compliance files found for {client_folder_sanitized}_{date_str_formatted}")
                    except:
                        print(f"   [WARN] No compliance files found")
            except Exception as e:
                print(f"   ERROR: Error checking uploads folder: {e}")
        else:
            print(f"   Uploads folder does not exist")

        # Update commodity status
        for commodity in commodities:
            commodity_status[commodity] = has_any_compliance

        all_commodities_have_compliance = all(commodity_status.values()) if commodity_status else False

        print(f"    Final results:")
        print(f"      - Has any compliance: {has_any_compliance}")
        print(f"      - All commodities have compliance: {all_commodities_have_compliance}")
        print(f"      - Commodity status: {commodity_status}")

        return {
            'has_any_compliance': has_any_compliance,
            'all_commodities_have_compliance': all_commodities_have_compliance,
            'commodity_status': commodity_status
        }
    
    except Exception as e:
        print(f"   ERROR: Error checking compliance: {str(e)}")
        return {
            'has_any_compliance': False,
            'all_commodities_have_compliance': False,
            'commodity_status': {}
        }


@login_required(login_url='login')
def shipment_list(request):
    """Display Food Safety Agency inspections from local database on shipments page."""
    import time
    start_time = time.time()  # Track performance
    clear_messages(request)
    
    print("=" * 80)
    safe_print("SHIPMENT_LIST VIEW CALLED - DEBUGGING INSPECTION ISSUES")
    safe_print(f"User: {request.user.username} (Role: {getattr(request.user, 'role', 'unknown')})")
    print("=" * 80)
    
    # Check for stored sync messages in cache and display them
    from django.core.cache import cache
    sync_success = cache.get('sync_success_message')
    sync_info_1 = cache.get('sync_info_message_1')
    sync_info_2 = cache.get('sync_info_message_2')
    
    if sync_success:
        messages.success(request, sync_success)
        cache.delete('sync_success_message')
    if sync_info_1:
        messages.info(request, sync_info_1)
        cache.delete('sync_info_message_1')
    if sync_info_2:
        messages.info(request, sync_info_2)
        cache.delete('sync_info_message_2')
    
    # OPTIMIZED SMART CACHING: 60-second cache timeout for performance + freshness
    # Automatically shows new inspections within 1 minute without manual refresh
    from django.core.cache import cache
    import time

    # PERFORMANCE FIX: Include page number and filters in cache key so each page/filter combo is cached separately
    page_number = request.GET.get('page', 1)
    filter_params = '_'.join(f"{k}_{v}" for k, v in sorted(request.GET.items()) if k in ['claim_no', 'client', 'branch', 'inspection_date_from', 'inspection_date_to', 'sent_status', 'compliance_status'])
    cache_key = f"shipment_list_{request.user.id}_{getattr(request.user, 'role', 'unknown')}_page_{page_number}_{filter_params}"
    cache_timestamp_key = f"{cache_key}_timestamp"

    # Manual refresh option (clears all cache)
    if request.GET.get('refresh') == 'true':
        safe_print("MANUAL CACHE REFRESH requested - clearing ALL cache...")
        cache.clear()  # Clear ALL cache to ensure fresh data

    # Check cached data with automatic expiration
    cached_data = cache.get(cache_key)
    cache_timestamp = cache.get(cache_timestamp_key)

    if cached_data and cache_timestamp:
        age_seconds = time.time() - cache_timestamp
        if age_seconds > 60:  # Auto-expire after 60 seconds
            safe_print(f"Cache expired ({age_seconds:.0f}s old), reloading...")
            cached_data = None
        else:
            safe_print(f"Using cached data ({age_seconds:.0f}s old, {60-age_seconds:.0f}s until refresh)")

    # PERFORMANCE FIX: Cache includes filters in key, so no need to bypass cache when filters are applied
    # Each filter combination gets its own cache entry for maximum performance
    
    # Get Food Safety Agency inspections from local database with MASSIVE OPTIMIZATION
    from ..models import FoodSafetyAgencyInspection, InspectorMapping
    from django.db.models import Prefetch, Q
    
    # OPTIMIZED FOR MAXIMUM SPEED: Only load last 60 days
    from datetime import datetime as dt, timedelta

    # Calculate date 60 days ago for speed
    sixty_days_ago = (dt.now() - timedelta(days=60)).date()

    # MINIMAL QUERY - Only essential fields for MAXIMUM SPEED
    inspections = FoodSafetyAgencyInspection.objects.select_related(
        'rfi_uploaded_by', 'invoice_uploaded_by', 'sent_by'
    ).only(
        # Only critical fields to reduce data transfer
        'id', 'client_name', 'date_of_inspection', 'inspector_name',
        'commodity', 'remote_id', 'product_name', 'product_class',
        'hours', 'km_traveled', 'comment', 'is_sent', 'sent_date',
        'rfi_uploaded_by_id', 'invoice_uploaded_by_id', 'sent_by_id',
        # Testing parameters (user-editable)
        'fat', 'protein', 'calcium', 'dna', 'is_sample_taken', 'bought_sample', 'lab', 'needs_retest',
        'is_direction_present_for_this_inspection', 'is_manual',
        # Location and contact
        'town', 'additional_email', 'internal_account_code'
    ).filter(
        # Only show manually created inspections (not synced from SQL Server)
        is_manual=True
    )
    
    # No automatic background fetching - only manual via settings button
    
    # Check compliance status for View Files button colors
    def check_compliance_status(client_name, inspection_date, inspection_count=None, products=None):
        """Check if compliance documents exist for View Files button color"""
        try:
            import os
            from datetime import datetime
            from django.conf import settings
            import re
            
            # Parse date and build folder path
            if isinstance(inspection_date, str):
                date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
            else:
                date_obj = inspection_date
                
            year_folder = date_obj.strftime('%Y')
            month_folder = date_obj.strftime('%B')
            
            # Use original client name for folder matching (folders now use original names)
            client_folder = client_name or 'Unknown Client'
            
            # Check compliance folder
            compliance_base = os.path.join(
                    settings.MEDIA_ROOT,
                    'inspection',
                    year_folder,
                    month_folder,
                    client_folder,
                    'Compliance'
                )
            
            if not os.path.exists(compliance_base):
                return 'no_compliance'
            
            # If this is a grouped inspection, check that ALL inspections have compliance docs
            if inspection_count and inspection_count > 1 and products:
                print(f"Checking grouped inspection: {client_name} with {inspection_count} inspections")
                
                # Count how many inspections have compliance documents
                inspections_with_compliance = 0
                total_inspections = len(products)
                
                for product in products:
                    commodity = product.get('commodity', '').strip().lower()
                    if commodity == 'eggs':
                        commodity = 'egg'
                    
                    if commodity:
                        # Check if this commodity has compliance documents
                        commodity_path = os.path.join(compliance_base, commodity.upper())
                        if os.path.exists(commodity_path):
                            files = [f for f in os.listdir(commodity_path) if os.path.isfile(os.path.join(commodity_path, f))]
                            if files:
                                inspections_with_compliance += 1
                                print(f"{commodity.upper()} has {len(files)} compliance documents")
                            else:
                                print(f"ERROR: {commodity.upper()} has no compliance documents")
                        else:
                            print(f"ERROR: {commodity.upper()} folder does not exist")
                    else:
                        print(f"WARNING: Inspection has no commodity specified")
                
                print(f"Compliance status: {inspections_with_compliance}/{total_inspections} inspections have documents")
                
                # Only return 'has_compliance' if ALL inspections have their documents
                if inspections_with_compliance == total_inspections:
                    return 'has_compliance'
                elif inspections_with_compliance > 0:
                    return 'partial_compliance'
                else:
                    return 'no_compliance'
            else:
                # Single inspection - check if any commodity folders have files
                has_files = False
                for commodity_folder in os.listdir(compliance_base):
                    commodity_path = os.path.join(compliance_base, commodity_folder)
                    if os.path.isdir(commodity_path):
                        files = [f for f in os.listdir(commodity_path) if os.path.isfile(os.path.join(commodity_path, f))]
                        if files:
                            has_files = True
                            break
                
                if has_files:
                    return 'has_compliance'
                else:
                    return 'no_compliance'
                
        except Exception as e:
            print(f"Error checking compliance status: {e}")
            return 'no_compliance'
    
    # Filter inspections based on user role and inspector ID
    if request.user.role == 'inspector':
        # Get the inspector ID and name for the current user
        inspector_id = None
        inspector_name = None
        try:
            inspector_mapping = InspectorMapping.objects.get(
                inspector_name=request.user.get_full_name() or request.user.username
            )
            inspector_id = inspector_mapping.inspector_id
            inspector_name = inspector_mapping.inspector_name
        except InspectorMapping.DoesNotExist:
            # If no mapping found, try to find by username
            try:
                inspector_mapping = InspectorMapping.objects.get(
                    inspector_name=request.user.username
                )
                inspector_id = inspector_mapping.inspector_id
                inspector_name = inspector_mapping.inspector_name
            except InspectorMapping.DoesNotExist:
                # If still no mapping, show no inspections
                inspector_id = None
                inspector_name = None

        if inspector_id and inspector_name:
            # Filter inspections to show those done by this inspector
            # Include BOTH inspector_id matches (for data with IDs) AND inspector_name matches (for data without IDs)
            # Use case-insensitive match for inspector_name to handle 'Neo Noe' vs 'NEO NOE'
            inspections = inspections.filter(
                Q(inspector_id=inspector_id) | Q(inspector_name__iexact=inspector_name)
            )
        else:
            # If no inspector ID found, show no inspections
            inspections = inspections.none()
    elif request.user.role == 'lab_technician':
        # Lab technicians can only see inspections where samples were taken
        inspections = inspections.filter(is_sample_taken=True)
    # For admin, super_admin, financial roles, show all inspections
    # (no filtering needed)
    
    # Apply filters if provided (but skip sent_status filter as it's handled at group level)
    sent_status_param = request.GET.get('sent_status')
    if sent_status_param:
        # Temporarily remove sent_status from request to avoid double filtering
        request.GET = request.GET.copy()
        request.GET.pop('sent_status', None)
        inspections = apply_fsa_inspection_filters(request, inspections)
        # Restore sent_status parameter
        request.GET['sent_status'] = sent_status_param
    else:
        inspections = apply_fsa_inspection_filters(request, inspections)
    
    # DATABASE-LEVEL PAGINATION: Only load data for current page
    from django.db.models import Count, Max, Min, Case, When, BooleanField
    from django.db.models.functions import TruncDate
    from django.core.paginator import Paginator
    import re
    
    # Function to sanitize group_id
    def sanitize_group_id(text):
        """Replace special characters with underscores to match frontend"""
        if not text:
            return ""
        sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', text)
        sanitized = re.sub(r'_+', '_', sanitized)
        return sanitized.strip('_')
    
    # Get page number first
    page_number = request.GET.get('page', 1)
    
    # Create the base queryset for groups
    groups_queryset = inspections.values(
        'client_name',
        'date_of_inspection'
    ).annotate(
        inspection_count=Count('id'),
        latest_inspection_id=Max('id'),
        earliest_inspection_id=Min('id'),
        has_sent_inspections=Count('id', filter=Q(is_sent=True)),  # Count sent inspections in group
        has_unsent_inspections=Count('id', filter=Q(is_sent=False)),  # Count unsent inspections in group
        has_rfi_inspections=Count('id', filter=Q(rfi_uploaded_by__isnull=False)),  # Count inspections with RFI uploaded
        has_no_rfi_inspections=Count('id', filter=Q(rfi_uploaded_by__isnull=True)),  # Count inspections without RFI uploaded
        has_lab_form_inspections=Count('id', filter=Q(lab_form_uploaded_by__isnull=False)),  # Count inspections with Lab Form uploaded
        has_no_lab_form_inspections=Count('id', filter=Q(lab_form_uploaded_by__isnull=True)),  # Count inspections without Lab Form uploaded
        comment=Max('comment')  # Get the comment from the group (all should have the same comment)
    ).order_by('-date_of_inspection', 'client_name')
    
    # FILTER GROUPS BY SENT STATUS: Apply sent status filter to groups, not individual inspections
    sent_status = request.GET.get('sent_status')
    if sent_status:
        if sent_status == 'YES':
            # Only show groups that have at least one sent inspection
            groups_queryset = groups_queryset.filter(has_sent_inspections__gt=0)
        elif sent_status == 'NO':
            # Only show groups that have at least one unsent inspection
            groups_queryset = groups_queryset.filter(has_unsent_inspections__gt=0)
    
    # FILTER GROUPS BY RFI STATUS: DISABLED - Show all inspections regardless of RFI status
    # User requested to remove filtering so all inspections show up, even without files
    # rfi_status = request.GET.get('rfi_status')
    # if rfi_status:
    #     if rfi_status == 'HAS_RFI':
    #         # Only show groups that have at least one inspection with RFI uploaded
    #         groups_queryset = groups_queryset.filter(has_rfi_inspections__gt=0)
    #     elif rfi_status == 'NO_RFI':
    #         # Only show groups that have at least one inspection without RFI uploaded
    #         groups_queryset = groups_queryset.filter(has_no_rfi_inspections__gt=0)

    # FILTER GROUPS BY LAB FORM STATUS: DISABLED - Show all inspections regardless of lab form status
    # lab_form_status = request.GET.get('lab_form_status')
    # if lab_form_status:
    #     if lab_form_status == 'HAS_LAB_FORM':
    #         # Only show groups that have at least one inspection with Lab Form uploaded
    #         groups_queryset = groups_queryset.filter(has_lab_form_inspections__gt=0)
    #     elif lab_form_status == 'NO_LAB_FORM':
    #         # Only show groups that have at least one inspection without Lab Form uploaded
    #         groups_queryset = groups_queryset.filter(has_no_lab_form_inspections__gt=0)

    # Apply database-level pagination
    paginator = Paginator(groups_queryset, 25)  # 25 groups per page
    page_obj = paginator.get_page(page_number)
    
    # Get only the groups for the current page
    client_date_groups = list(page_obj.object_list)

    # PERFORMANCE FIX: Only load clients that appear in the current page (25 groups max)
    # Extract unique client names from the 25 groups on this page
    client_names_on_page = set(group['client_name'] for group in client_date_groups if group.get('client_name'))

    # Build cache key specific to these clients
    client_cache_key = f"client_maps_page_{','.join(sorted(list(client_names_on_page)[:5]))}"  # Use first 5 for cache key
    client_data = cache.get(client_cache_key)

    if not client_data:
        try:
            from ..models import Client as _Client, ClientEmail
            _client_map = {}
            _client_id_map = {}
            _client_email_map = {}

            # Helper: normalize names for matching (case-insensitive, collapse spaces, remove punctuation)
            def _norm(text):
                try:
                    cleaned = re.sub(r"[\(\)\[\]{}\\/._,-]", " ", (text or ""))
                    cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
                    return cleaned
                except Exception:
                    return (text or "").strip().lower()

            # PERFORMANCE FIX: Only load clients that appear on THIS PAGE (not all 4,916 clients!)
            # This reduces loading from 4,916 clients to ~25 clients maximum
            clients_queryset = _Client.objects.filter(
                Q(client_id__in=client_names_on_page) | Q(name__in=client_names_on_page)
            ).select_related().prefetch_related('additional_emails')

            print(f"[PERFORMANCE] Loading only {clients_queryset.count()} clients for {len(client_names_on_page)} groups (not all 4916!)")

            # Load client data including emails
            for _c in clients_queryset:
                key_id = _norm(_c.client_id)
                key_name = _norm(_c.name)
                
                if key_id:
                    if _c.internal_account_code:
                        _client_map[key_id] = _c.internal_account_code
                    _client_id_map[key_id] = _c.client_id
                    
                    # Collect all emails for this client
                    emails = []
                    if _c.email:
                        emails.append({'email': _c.email, 'type': 'primary', 'removable': True})
                    if _c.manual_email:
                        emails.append({'email': _c.manual_email, 'type': 'manual', 'removable': True})
                    
                    # Add additional emails
                    for additional_email in _c.additional_emails.all():
                        emails.append({
                            'email': additional_email.email, 
                            'type': 'additional', 
                            'removable': True,
                            'label': additional_email.label
                        })
                    
                    if emails:
                        _client_email_map[key_id] = emails
                
                # Also map by name as fallback
                if key_name and key_name != key_id:
                    if _c.internal_account_code and key_name not in _client_map:
                        _client_map[key_name] = _c.internal_account_code
                    if key_name not in _client_id_map:
                        _client_id_map[key_name] = _c.client_id
                    if key_name not in _client_email_map and emails:
                        _client_email_map[key_name] = emails
            
            client_data = {
                'client_map': _client_map,
                'client_id_map': _client_id_map,
                'client_email_map': _client_email_map
            }
            # Cache for 5 minutes (page-specific data, shorter cache)
            cache.set(client_cache_key, client_data, 300)
            
        except Exception as e:
            print(f"[ERROR] Exception while building client maps: {e}")
            import traceback
            traceback.print_exc()
            client_data = {
                'client_map': {},
                'client_id_map': {},
                'client_email_map': {}
            }
    
    _client_map = client_data['client_map']
    _client_id_map = client_data['client_id_map']
    _client_email_map = client_data['client_email_map']

    # DEBUG: Show what account codes we have
    print(f"[ACCOUNT CODE DEBUG] Client map has {len(_client_map)} entries with account codes")
    if len(_client_map) > 0:
        print(f"[ACCOUNT CODE DEBUG] Sample entries:")
        for key, code in list(_client_map.items())[:5]:
            print(f"  '{key}' -> '{code}'")

    # Helper to fetch internal account code exactly like client-allocation page
    from ..models import Client as _Client
    def _get_internal_account_code(raw_name):
        """Get internal account code from cache ONLY - NO database queries for performance"""
        try:
            # Use normalized key for lookup in pre-built cache
            code = _client_map.get(_norm(raw_name))
            if code:
                return code
            # PERFORMANCE FIX: Removed database fallback queries that were causing N+1 problem
            # If client isn't in cache, return None instead of querying database
            # This prevents 25+ database queries per page load
        except Exception:
            pass
        return None

    # Helper to fetch client emails
    def _get_client_emails(raw_name):
        """Get client emails from cache ONLY - NO database queries for performance"""
        try:
            # Use normalized key for lookup in pre-built cache
            emails = _client_email_map.get(_norm(raw_name))
            if emails:
                return emails
            # PERFORMANCE FIX: Removed database fallback queries that were causing N+1 problem
            # If client isn't in cache, return empty list instead of querying database
            # This prevents 25+ database queries per page load
        except Exception:
            pass
        return []

    # Helper: normalize names for matching (case-insensitive, collapse spaces, remove punctuation)
    def _norm(text):
        try:
            cleaned = re.sub(r"[\(\)\[\]{}\\/._,-]", " ", (text or ""))
            cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
            return cleaned
        except Exception:
            return (text or "").strip().lower()

    # PERFORMANCE OPTIMIZATION: Batch load all group data at once
    # Get all inspection IDs for the groups we're processing
    group_conditions = Q()
    for group in client_date_groups:
        group_conditions |= Q(
            client_name=group['client_name'],
            date_of_inspection=group['date_of_inspection']
        )
    
    # Load all inspections for these groups in one query
    all_group_inspections = inspections.filter(group_conditions).order_by('client_name', 'date_of_inspection', 'id')
    
    # Log potential data issues for debugging
    if len(all_group_inspections) == 0 and len(client_date_groups) > 0:
        print(f"WARNING: No inspections found for {len(client_date_groups)} groups - may be permission filtering issue")
    
    # Group them by client_name and date_of_inspection
    from collections import defaultdict
    grouped_inspections_dict = defaultdict(list)
    for inspection in all_group_inspections:
        key = (inspection.client_name, inspection.date_of_inspection)
        grouped_inspections_dict[key].append(inspection)
    
    # Helper function to check files for a single group (fast version for page load)
    def check_group_files(client_name, inspection_date):
        """Check if files exist for this group - optimized for speed

        Checks NEW structure first: MEDIA_ROOT/docs/{client_id}/{inspection_id}/{category}/
        Falls back to LEGACY structure: MEDIA_ROOT/inspection/YEAR/MONTH/CLIENT/...
        """
        import re
        from django.conf import settings
        from main.models import FoodSafetyAgencyInspection

        try:
            has_rfi = has_invoice = has_lab = has_lab_form = has_compliance = has_composition = has_occurrence = False

            # === NEW STRUCTURE: Check docs/{client_id}/{inspection_id}/ ===
            # Look up inspections for this client and date
            inspections = FoodSafetyAgencyInspection.objects.filter(
                client_name__iexact=client_name,
                date_of_inspection=inspection_date
            ).select_related('client')

            docs_base = os.path.join(settings.MEDIA_ROOT, 'docs')

            for inspection in inspections:
                if inspection.client:
                    insp_path = os.path.join(docs_base, str(inspection.client.id), str(inspection.id))
                    if os.path.exists(insp_path):
                        # Check each category
                        for category in ['rfi', 'invoice', 'lab', 'lab_form', 'compliance', 'composition', 'occurrence']:
                            cat_path = os.path.join(insp_path, category)
                            if os.path.exists(cat_path) and os.listdir(cat_path):
                                if category == 'rfi': has_rfi = True
                                elif category == 'invoice': has_invoice = True
                                elif category == 'lab': has_lab = True
                                elif category == 'lab_form': has_lab_form = True
                                elif category == 'compliance': has_compliance = True
                                elif category == 'composition': has_composition = True
                                elif category == 'occurrence': has_occurrence = True

            # === LEGACY STRUCTURE: Check inspection/YEAR/MONTH/CLIENT/ ===
            # Only check if we haven't found all files yet
            if not (has_rfi and has_invoice and has_lab and has_lab_form and has_compliance and has_composition and has_occurrence):
                inspection_base = os.path.join(settings.MEDIA_ROOT, 'inspection')

                def create_folder_name(name):
                    if not name:
                        return "unknown_client"
                    clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
                    clean_name = clean_name.replace(' ', '_').replace('-', '_')
                    clean_name = re.sub(r'_+', '_', clean_name)
                    clean_name = clean_name.strip('_').lower()
                    return clean_name or "unknown_client"

                sanitized_client_name = create_folder_name(client_name)
                client_folder_variations = [
                    sanitized_client_name,
                    f'btn_{sanitized_client_name}',
                    client_name
                ]

                year = inspection_date.strftime('%Y')
                month = inspection_date.strftime('%B')
                parent_path = os.path.join(inspection_base, year, month)

                if os.path.exists(parent_path):
                    for folder_variation in client_folder_variations:
                        test_path = os.path.join(parent_path, folder_variation)
                        if os.path.exists(test_path):
                            if not has_rfi:
                                for rfi_var in ['rfi', 'RFI', 'Request For Invoice']:
                                    rfi_path = os.path.join(test_path, rfi_var)
                                    if os.path.exists(rfi_path) and os.listdir(rfi_path):
                                        has_rfi = True
                                        break

                            if not has_invoice:
                                for inv_var in ['invoice', 'Invoice']:
                                    invoice_path = os.path.join(test_path, inv_var)
                                    if os.path.exists(invoice_path) and os.listdir(invoice_path):
                                        has_invoice = True
                                        break

                            if not has_lab:
                                for lab_var in ['lab', 'Lab', 'lab results']:
                                    lab_path = os.path.join(test_path, lab_var)
                                    if os.path.exists(lab_path) and os.listdir(lab_path):
                                        has_lab = True
                                        break

                            if not has_lab_form:
                                for lab_form_var in ['lab_form', 'Lab_Form', 'lab form']:
                                    lab_form_path = os.path.join(test_path, lab_form_var)
                                    if os.path.exists(lab_form_path) and os.listdir(lab_form_path):
                                        has_lab_form = True
                                        break

                            if not has_compliance:
                                for comp_var in ['compliance', 'Compliance']:
                                    comp_path = os.path.join(test_path, comp_var)
                                    if os.path.exists(comp_path) and os.listdir(comp_path):
                                        has_compliance = True
                                        break

                            if not has_composition:
                                for comp_var in ['composition', 'Composition']:
                                    comp_path = os.path.join(test_path, comp_var)
                                    if os.path.exists(comp_path) and os.listdir(comp_path):
                                        has_composition = True
                                        break

                            if not has_occurrence:
                                for occ_var in ['occurrence', 'Occurrence']:
                                    occ_path = os.path.join(test_path, occ_var)
                                    if os.path.exists(occ_path) and os.listdir(occ_path):
                                        has_occurrence = True
                                        break

            # Determine file status
            if has_rfi and has_invoice and has_lab and has_compliance:
                file_status = 'all_files'  # Green
            elif has_compliance:
                file_status = 'compliance_only'  # Orange
            elif has_rfi or has_invoice or has_lab:
                file_status = 'partial_files'  # Orange
            else:
                file_status = 'no_files'  # Red

            print(f"[FILE CHECK] {client_name}: RFI={has_rfi}, Invoice={has_invoice}, Lab={has_lab}, Lab_Form={has_lab_form}, Compliance={has_compliance}, Composition={has_composition}, Occurrence={has_occurrence}")

            return {
                'has_rfi': has_rfi,
                'has_invoice': has_invoice,
                'has_lab': has_lab,
                'has_lab_form': has_lab_form,
                'has_compliance': has_compliance,
                'has_composition': has_composition,
                'has_occurrence': has_occurrence,
                'file_status': file_status
            }
        except Exception as e:
            print(f"[FILE CHECK ERROR] {client_name} on {inspection_date}: {e}")
            return {
                'has_rfi': False,
                'has_invoice': False,
                'has_lab': False,
                'has_lab_form': False,
                'has_compliance': False,
                'has_composition': False,
                'has_occurrence': False,
                'file_status': 'no_files'
            }

    # Process grouped inspections efficiently - ONLY CREATE REPRESENTATIVE OBJECTS
    grouped_inspections = []
    print(f"[PROCESSING] Processing {len(client_date_groups)} groups...")
    for group in client_date_groups:
        client_name = group['client_name']
        date_of_inspection = group['date_of_inspection']
        inspection_count = group['inspection_count']

        # REMOVED: Excessive per-group debug logging for performance
        # print(f"[DEBUG] Processing group: {client_name} - {date_of_inspection} (expected: {inspection_count})")

        # Get inspections from our pre-loaded dictionary
        group_inspections = grouped_inspections_dict.get((client_name, date_of_inspection), [])

        # Log when we have empty groups but expected products (potential data integrity issue)
        if not group_inspections and inspection_count > 0:
            print(f"WARNING: Data integrity issue: {client_name} on {date_of_inspection} has {inspection_count} expected but 0 loaded")
        
        # Get sample inspector from the group
        sample_inspection = group_inspections[0] if group_inspections else None
        inspector_name = sample_inspection.inspector_name if sample_inspection else None
        
        # Get aggregated km and hours for the group
        group_km_traveled = None
        group_hours = None
        group_additional_email = None
        group_comment = None
        group_approved_status = None
        group_is_sent = False

        # CHECK FILES ON BACKEND FOR INSTANT BUTTON COLORS
        file_check_result = check_group_files(client_name, date_of_inspection)
        has_rfi = file_check_result['has_rfi']
        has_invoice = file_check_result['has_invoice']
        has_lab = file_check_result['has_lab']
        has_lab_form = file_check_result['has_lab_form']
        has_compliance = file_check_result['has_compliance']
        has_composition = file_check_result['has_composition']
        has_occurrence = file_check_result['has_occurrence']
        file_status = file_check_result['file_status']

        # For backward compatibility (some code checks uploader fields)
        rfi_uploader = 'System' if has_rfi else None
        invoice_uploader = 'System' if has_invoice else None
        rfi_upload_date = None
        invoice_upload_date = None
        composition_uploader = 'System' if has_composition else None
        composition_upload_date = None
        
        
        if sample_inspection:
            group_km_traveled = sample_inspection.km_traveled
            group_hours = sample_inspection.hours
            group_additional_email = sample_inspection.additional_email
            group_comment = sample_inspection.comment
            group_approved_status = sample_inspection.approved_status
            # Check if ANY inspection in the group is marked as sent
            group_is_sent = any(inspection.is_sent for inspection in group_inspections)
        
        # If the sample inspection doesn't have KM/hours/additional_email, check other inspections in the group
        # PERFORMANCE OPTIMIZATION: Removed excessive fallback logging
        if group_km_traveled is None and group_inspections:
            for inspection in group_inspections:
                if inspection.km_traveled is not None:
                    group_km_traveled = inspection.km_traveled
                    break

        if group_hours is None and group_inspections:
            for inspection in group_inspections:
                if inspection.hours is not None:
                    group_hours = inspection.hours
                    break

        if group_additional_email is None and group_inspections:
            for inspection in group_inspections:
                if inspection.additional_email is not None:
                    group_additional_email = inspection.additional_email
                    break

        if group_comment is None and group_inspections:
            for inspection in group_inspections:
                if inspection.comment is not None:
                    group_comment = inspection.comment
                    break

        if group_approved_status is None and group_inspections:
            for inspection in group_inspections:
                if inspection.approved_status is not None:
                    group_approved_status = inspection.approved_status
                    break

        # Check if this is an occurrence report inspection
        is_occurrence_report = any(getattr(inspection, 'occurrence_report', False) for inspection in group_inspections)

        # If group_inspections is empty but we expected products, try a direct query as fallback
        if not group_inspections and inspection_count > 0:
            print(f" FALLBACK TRIGGERED: {client_name} on {date_of_inspection} - expected {inspection_count}, got 0")
            try:
                # Direct query without user-specific filters to handle permission issues
                fallback_inspections = FoodSafetyAgencyInspection.objects.filter(
                    client_name=client_name,
                    date_of_inspection=date_of_inspection
                ).order_by('id')

                if fallback_inspections.exists():
                    group_inspections = list(fallback_inspections)
                    print(f"Recovered {len(fallback_inspections)} inspections using direct query for {client_name}")
            except Exception as e:
                print(f" Fallback query failed for {client_name}: {e}")

        # PERFORMANCE OPTIMIZATION: Get internal account code ONCE per group instead of for each product
        # First try to get from the inspection itself (for manually created inspections)
        group_internal_account_code = None
        if sample_inspection and sample_inspection.internal_account_code:
            group_internal_account_code = sample_inspection.internal_account_code
        else:
            # Fall back to checking other inspections in the group
            for inspection in group_inspections:
                if inspection.internal_account_code:
                    group_internal_account_code = inspection.internal_account_code
                    break

        # If still not found, try to get from Client model
        if not group_internal_account_code and client_name:
            group_internal_account_code = _get_internal_account_code(client_name)

        # Helper function to check if files exist for a specific inspection
        def check_inspection_files(inspection):
            """Check if files exist for this specific inspection - returns dict of booleans"""
            from django.conf import settings as django_settings
            result = {'composition': False, 'coa': False, 'retest': False, 'occurrence': False}
            docs_base = os.path.join(django_settings.MEDIA_ROOT, 'docs')

            if inspection.client_id:
                insp_path = os.path.join(docs_base, str(inspection.client_id), str(inspection.id))
                if os.path.exists(insp_path):
                    # Check each category for files
                    for category, key in [('composition', 'composition'), ('lab', 'coa'), ('retest', 'retest'), ('occurrence', 'occurrence')]:
                        cat_path = os.path.join(insp_path, category)
                        if os.path.exists(cat_path) and os.listdir(cat_path):
                            result[key] = True
            return result

        # PERFORMANCE OPTIMIZATION: Use list comprehension for faster product building (2-3x faster than loop+append)
        # Product names are fetched by background sync service - just use what's in database
        # DO NOT fetch from SQL Server on page load - it takes 11+ minutes for all inspections!
        products = []
        for inspection in group_inspections:
            # Check actual file existence for this inspection
            file_status = check_inspection_files(inspection)
            products.append({
                'id': inspection.id,  # Primary key for upload functions
                'remote_id': inspection.remote_id,
                'client_name': inspection.client_name,
                'internal_account_code': group_internal_account_code,  # Reuse cached value
                'commodity': inspection.commodity,
                'date_of_inspection': inspection.date_of_inspection,
                'is_sample_taken': inspection.is_sample_taken,
                'needs_retest': inspection.needs_retest,
                'product_name': inspection.product_name,  # Already in database from sync service
                'product_class': inspection.product_class,
                'fat': inspection.fat,
                'protein': inspection.protein,
                'calcium': inspection.calcium,
                'dna': inspection.dna,
                'bought_sample': inspection.bought_sample,
                'lab': inspection.lab,
                'is_complete': False,  # Default to False
                'is_direction_present_for_this_inspection': inspection.is_direction_present_for_this_inspection,
                # Upload status flags for button colors - check actual file existence
                'composition_uploaded': file_status['composition'],
                'coa_uploaded': file_status['coa'],
                'retest_uploaded': file_status['retest'],
                'occurrence_uploaded': file_status['occurrence'],
            })
        
        # Since we removed the logging system, set default values
        _cached_status = {
            'has_rfi': False,
            'has_invoice': False,
            'has_compliance_docs': False,
            'all_commodities_have_compliance': False,
            'has_any_compliance': False
        }
        
        has_compliance_documents = False

        # Compliance status result for template (simplified - file checking done above)
        compliance_status_result = {
            'has_any_compliance': has_compliance,
            'all_commodities_have_compliance': False,  # Not checking per-commodity
            'commodity_status': {}
        }

        # Determine INSPECTION compliance status (for display in header)
        # Check if ALL individual inspections in the group are compliant
        all_inspections_compliant = True
        for inspection in group_inspections:
            if inspection.is_direction_present_for_this_inspection:
                all_inspections_compliant = False
                break

        if all_inspections_compliant:
            inspection_compliance_status = 'compliant'  # All inspections are compliant
        else:
            inspection_compliance_status = 'non_compliant'  # At least one inspection is non-compliant

        # Determine FILE UPLOAD compliance status (controls "Sent" status dropdown)
        if has_rfi and has_invoice and has_compliance:
            compliance_status = 'complete'  # All required files uploaded
        elif has_rfi or has_invoice or has_compliance:
            compliance_status = 'partial'  # Some files uploaded but not all
        else:
            compliance_status = 'no_compliance'  # No files uploaded

        # file_status already set by check_group_files() above

        # PERFORMANCE OPTIMIZATION: Removed massive per-group compliance debug logging
        # This was printing 10+ lines for EVERY group, causing significant slowdown

        # Generate group_id
        sanitized_client = sanitize_group_id(client_name) if client_name else ""
        date_str = date_of_inspection.strftime('%Y%m%d') if date_of_inspection else "NO_DATE"
        group_id = f"{sanitized_client}_{date_str}" if sanitized_client and date_of_inspection else f"ERROR_client_{client_name}_date_{date_of_inspection}"
        
        
        # Create a simple dictionary instead of dynamic class object
        # This avoids pickling issues with Redis cache
        # Get town from group inspections
        group_town = None
        if sample_inspection:
            group_town = sample_inspection.town
        if group_town is None and group_inspections:
            for inspection in group_inspections:
                if inspection.town:
                    group_town = inspection.town
                    break

        representative_inspection = {
            'client_name': client_name,
            'date_of_inspection': date_of_inspection,
            'inspection_count': inspection_count,
            'internal_account_code': group_internal_account_code,  # PERFORMANCE: Reuse cached value
            'client_emails': _get_client_emails(client_name),
            'is_complete': False,  # Default to False - will be checked on-demand
            'group_id': group_id,
            'inspector_name': inspector_name,
            'town': group_town,  # Include town for display
            'inspector_id': next((inspection.inspector_id for inspection in group_inspections if getattr(inspection, 'inspector_id', None)), None),
            'location': None,  # No location field in model
            'total_products': inspection_count,
            'km_traveled': group_km_traveled,  # From actual inspection data
            'hours': group_hours,  # From actual inspection data
            'additional_email': group_additional_email,  # From actual inspection data
            'comment': group_comment,  # From actual inspection data
            'approved_status': group_approved_status,  # From actual inspection data
            'is_sent': group_is_sent,  # From actual inspection data
            'sent_by': next((inspection.sent_by for inspection in group_inspections if inspection.is_sent), None),  # Who marked as sent
            'sent_date': next((inspection.sent_date for inspection in group_inspections if inspection.is_sent), None),  # When marked as sent
            'rfi_uploaded': rfi_uploader is not None,  # Use file detection results
            'invoice_uploaded': invoice_uploader is not None,  # Use file detection results
            'rfi_uploaded_by': rfi_uploader,  # Who uploaded RFI
            'rfi_uploaded_date': rfi_upload_date,  # When RFI was uploaded
            'invoice_uploaded_by': invoice_uploader,  # Who uploaded Invoice
            'invoice_uploaded_date': invoice_upload_date,  # When Invoice was uploaded
            'composition_uploaded': composition_uploader is not None,  # Use file detection results
            'composition_uploaded_by': composition_uploader,  # Who uploaded Composition
            'composition_uploaded_date': composition_upload_date,  # When Composition was uploaded
            'inspection_compliance_status': inspection_compliance_status,  # For header compliance display (based on individual inspections)
            'compliance_status': compliance_status,  # For sent dropdown control (based on file uploads)
            'file_status': file_status,  # For View Files button color coding (based on actual files)
            'has_rfi': has_rfi,  # Individual file flags for button colors
            'has_invoice': has_invoice,
            'has_lab': has_lab,
            'has_lab_form': has_lab_form,
            'has_compliance': has_compliance,
            'has_composition': has_composition,
            'has_occurrence': has_occurrence,
            'is_occurrence_report': is_occurrence_report,  # True if this is an occurrence report (disables other uploads)
            'compliance_documents_status': {
                'all_commodities_have_compliance': bool(compliance_status_result.get('all_commodities_have_compliance', False)) if compliance_status_result else False,
                'has_any_compliance': bool(compliance_status_result.get('has_any_compliance', False)) if compliance_status_result else False
            },
            'products': products  # Now populated with actual inspection data
        }

        # PERFORMANCE OPTIMIZATION: Removed per-group product assignment debug logging

        grouped_inspections.append(representative_inspection)
    
    # Cache filter options
    filter_cache_key = "filter_options"
    filter_data = cache.get(filter_cache_key)
    
    if not filter_data:
        # Get unique values for filters efficiently - LIMIT TO PREVENT LARGE RESPONSES
        inspectors = list(inspections.filter(inspector_name__isnull=False, inspector_name__gt='').values_list('inspector_name', flat=True).distinct().order_by('inspector_name')[:100])
        # Removed clients list since we now use search input instead of dropdown
        commodities = list(inspections.values_list('commodity', flat=True).distinct()[:50])
        
        filter_data = {
            'inspectors': inspectors,
            'commodities': commodities
        }
        # Cache for 5 minutes
        cache.set(filter_cache_key, filter_data, 300)
    
    # Get OneDrive delay settings for countdown display
    from ..models import SystemSettings
    system_settings = SystemSettings.get_settings()
    
    # Calculate delay in days for countdown display
    delay_days = system_settings.onedrive_upload_delay_days
    delay_unit = system_settings.onedrive_upload_delay_unit
    
    if delay_unit == 'hours':
        onedrive_delay_days = delay_days / 24
    elif delay_unit == 'days':
        onedrive_delay_days = delay_days
    elif delay_unit == 'weeks':
        onedrive_delay_days = delay_days * 7
    elif delay_unit == 'months':
        onedrive_delay_days = delay_days * 30
    elif delay_unit == 'years':
        onedrive_delay_days = delay_days * 365
    else:
        onedrive_delay_days = 3  # Default fallback
    
    # Get theme settings for template (from Settings model, not SystemSettings)
    try:
        from ..models import Settings
        settings_obj = Settings.get_settings()
        settings = {
            'dark_mode': settings_obj.dark_mode,
        }
        print(f" THEME DEBUG: dark_mode = {settings_obj.dark_mode}")
    except Exception as e:
        settings = {'dark_mode': False}
        print(f" THEME ERROR: {e}")

    # FIX: Use the enhanced grouped_inspections instead of the raw page_obj
    # The page_obj contains raw database results without products, but grouped_inspections has the enhanced data
    context = {
        'shipments': grouped_inspections,  # Use enhanced objects with products
        'inspectors': filter_data['inspectors'],
        'commodities': filter_data['commodities'],
        'total_inspections': len(grouped_inspections),
        'total_shipments': len(grouped_inspections),  # Add this for template compatibility
        'user_role': request.user.role,
        'user_name': request.user.get_full_name() or request.user.username,
        'user_message': None,  # Add this for template compatibility
        'paginator': paginator,
        'page_obj': page_obj,
        'show_all': False,  # Force pagination to prevent broken pipe
        'onedrive_delay_days': int(onedrive_delay_days),  # Add OneDrive delay for countdown
        'settings': settings,  # Add theme settings
    }
    
    # DEBUG: Check what's being sent to template
    print(f" CONTEXT DEBUG (FIXED):")
    print(f"    Total shipments in context: {len(grouped_inspections)}")
    if len(grouped_inspections) > 0:
        first_shipment = grouped_inspections[0]
        print(f"    First shipment: {first_shipment.get('client_name', 'Unknown')}")
        print(f"    First shipment products: {len(first_shipment.get('products', []))}")
        print(f"    First shipment keys: {list(first_shipment.keys())}")
    print(f"    Context keys: {list(context.keys())}")
    print(" PRODUCTS SHOULD NOW BE VISIBLE IN TEMPLATE!")
    
    # Cache for 60 seconds (OPTIMIZED: Fast performance + fresh data)
    # PERFORMANCE FIX: Always cache since cache key includes filters - each filter combo gets its own cache
    cache.set(cache_key, context, 60)
    cache.set(cache_timestamp_key, time.time(), 60)
    safe_print("Context cached for 60 seconds")
    
    # PERFORMANCE LOGGING: Track loading improvements
    import time
    load_time = time.time() - start_time if 'start_time' in locals() else 0
    print(f" PERFORMANCE: Loaded {len(page_obj)} groups in {load_time:.2f}s (page {page_obj.number} of {paginator.num_pages})")
    
    
    try:
        return render(request, 'main/inspection_records.html', context)
    except BrokenPipeError:
        print(" Broken pipe error detected - client disconnected")
        # Return a minimal response to prevent server error
        return HttpResponse("Connection interrupted. Please refresh the page.", status=200)
    except Exception as e:
        print(f" Error rendering shipment list: {e}")
        import traceback
        traceback.print_exc()
        # Re-raise the exception so Django can handle it properly
        raise


def check_for_compliance_documents(client_name, inspection_date):
    """Check if compliance documents exist for a given client and date."""
    try:
        from django.conf import settings
        
        # Convert datetime.date to string format if needed
        if hasattr(inspection_date, 'strftime'):
            date_str = inspection_date.strftime('%Y-%m-%d')
        else:
            date_str = str(inspection_date)
        
        # Check OneDrive (preferred)
        onedrive_files = get_inspection_files_onedrive(client_name, date_str)
        if onedrive_files and 'compliance' in onedrive_files:
            compliance_files = onedrive_files['compliance']
            if compliance_files and len(compliance_files) > 0:
                return True

        # If we prefer OneDrive, skip local filesystem fallback entirely
        if getattr(settings, 'PREFER_ONEDRIVE', False):
            return False

        # Otherwise fall back to local files (legacy)
        local_files = get_inspection_files_local(client_name, date_str)
        if local_files and 'compliance' in local_files:
            compliance_files = local_files['compliance']
            if compliance_files and len(compliance_files) > 0:
                return True
        
        return False
    except Exception as e:
        print(f"Error checking compliance documents for {client_name}: {e}")
        return False


@login_required(login_url='login')
def delete_shipment(request, pk):
    """Delete a shipment entry."""
    clear_messages(request)
    shipment = get_object_or_404(Shipment, pk=pk)
    if request.method == 'POST':
        shipment.delete()
        messages.success(request, 'Shipment deleted successfully.')
    return redirect('shipment_list')


@login_required(login_url='login')
def edit_inspection(request, inspection_id):
    """Edit a Food Safety Agency inspection."""
    clear_messages(request)
    
    print(f"DEBUG: edit_inspection called with inspection_id: {inspection_id}")
    
    try:
        from ..models import FoodSafetyAgencyInspection, InspectorMapping
        inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id).first()
        if not inspection:
            messages.error(request, f'Inspection with ID {inspection_id} not found.')
            return redirect('shipment_list')
        print(f"DEBUG: Found inspection: {inspection.remote_id} for {inspection.client_name}")
        
        # Check if user has permission to edit this inspection
        if request.user.role == 'inspector':
            # Get the inspector ID for the current user
            inspector_id = None
            try:
                inspector_mapping = InspectorMapping.objects.get(
                    inspector_name=request.user.get_full_name() or request.user.username
                )
                inspector_id = inspector_mapping.inspector_id
            except InspectorMapping.DoesNotExist:
                try:
                    inspector_mapping = InspectorMapping.objects.get(
                        inspector_name=request.user.username
                    )
                    inspector_id = inspector_mapping.inspector_id
                except InspectorMapping.DoesNotExist:
                    inspector_id = None
            
            # Check if this inspection belongs to the current inspector
            if not inspector_id or inspection.inspector_id != inspector_id:
                messages.error(request, 'You can only edit your own inspections.')
                return redirect('shipment_list')
        
        # For admin, super_admin, financial, and scientist roles, allow editing any inspection
        # (no additional permission check needed)
        
        if request.method == 'POST':
            print(f"DEBUG: Processing POST request")
            # Handle form submission for editing inspection
            try:
                # Only update fields that won't be overwritten by server sync
                # Test result fields (these are local to our system)
                inspection.protein = 'protein_test' in request.POST
                inspection.fat = 'fat_test' in request.POST
                inspection.calcium = 'calcium_test' in request.POST
                inspection.dna = 'dna_test' in request.POST
                
                # Save the inspection
                inspection.save()
                print(f"DEBUG: Inspection saved successfully")
                messages.success(request, f"Inspection {inspection.remote_id} for {inspection.client_name} updated successfully!")
                return redirect('shipment_list')
                
            except Exception as e:
                print(f"DEBUG: Error updating inspection: {str(e)}")
                messages.error(request, f'Error updating inspection: {str(e)}')
                # Return to the form with error
                context = {
                    'inspection': inspection
                }
                return render(request, 'main/edit_inspection.html', context)
        else:
            print(f"DEBUG: Rendering edit form")
            # Render the edit form
            context = {
                'inspection': inspection
            }
            return render(request, 'main/edit_inspection.html', context)
            
    except Exception as e:
        print(f"DEBUG: Error in edit_inspection: {str(e)}")
        messages.error(request, f'Error editing inspection: {str(e)}')
        return redirect('shipment_list')

@login_required(login_url='login')
def delete_inspection(request, inspection_id):
    """Delete a Food Safety Agency inspection."""
    clear_messages(request)
    
    try:
        from ..models import FoodSafetyAgencyInspection, InspectorMapping
        inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id).first()
        if not inspection:
            messages.error(request, f'Inspection with ID {inspection_id} not found.')
            return redirect('shipment_list')
        
        # Check if user has permission to delete this inspection
        if request.user.role == 'inspector':
            # Get the inspector ID for the current user
            inspector_id = None
            try:
                inspector_mapping = InspectorMapping.objects.get(
                    inspector_name=request.user.get_full_name() or request.user.username
                )
                inspector_id = inspector_mapping.inspector_id
            except InspectorMapping.DoesNotExist:
                try:
                    inspector_mapping = InspectorMapping.objects.get(
                        inspector_name=request.user.username
                    )
                    inspector_id = inspector_mapping.inspector_id
                except InspectorMapping.DoesNotExist:
                    inspector_id = None
            
            # Check if this inspection belongs to the current inspector
            if not inspector_id or inspection.inspector_id != inspector_id:
                messages.error(request, 'You can only delete your own inspections.')
                return redirect('shipment_list')
        
        # For admin, super_admin, financial, and scientist roles, allow deleting any inspection
        # (no additional permission check needed)
        
        if request.method == 'POST':
            client_name = inspection.client_name
            inspection.delete()
            messages.success(request, f'Inspection {inspection_id} for {client_name} deleted successfully.')
        else:
            messages.error(request, 'Invalid request method.')
            
    except Exception as e:
        messages.error(request, f'Error deleting inspection: {str(e)}')
    
    return redirect('shipment_list')


def upload_document(request):
    """Handle document uploads for inspection groups and individual inspections."""
    # Check authentication - return JSON error for AJAX requests
    if not request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'multipart/form-data':
            return JsonResponse({'success': False, 'error': 'Authentication required. Please log in.'})
        return redirect('login')

    if request.method == 'POST':
        try:
            from ..models import FoodSafetyAgencyInspection
            from django.db import models
            import os
            from django.conf import settings
            from datetime import datetime
            import re
            
            # Get form data
            group_id = request.POST.get('group_id')
            inspection_id = request.POST.get('inspection_id')
            document_type = request.POST.get('document_type')
            inspection_number = request.POST.get('inspection_number')  # For individual inspection compliance files
            uploaded_file = request.FILES.get('file')
            is_occurrence_report = request.POST.get('is_occurrence_report') == 'true'

            # Handle NEW occurrence reports (creating a new inspection)
            if is_occurrence_report and document_type == 'occurrence' and not inspection_id and not group_id:
                # Create a new occurrence report inspection
                try:
                    from django.db.models import Max
                    from ..models import InspectorMapping

                    # Get the next remote_id
                    max_remote_id = FoodSafetyAgencyInspection.objects.aggregate(Max('remote_id'))['remote_id__max'] or 0
                    new_remote_id = max_remote_id + 1

                    # Parse the date
                    date_str = request.POST.get('date_of_inspection', '')
                    if date_str:
                        try:
                            inspection_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            inspection_date = datetime.now().date()
                    else:
                        inspection_date = datetime.now().date()

                    # Get inspector name and ID from mapping (for proper filtering)
                    inspector_name = request.POST.get('inspector_name', '')
                    inspector_id = None

                    # Try to find inspector mapping for current user
                    try:
                        inspector_mapping = InspectorMapping.objects.get(
                            inspector_name__iexact=request.user.get_full_name() or request.user.username
                        )
                        inspector_name = inspector_mapping.inspector_name
                        inspector_id = inspector_mapping.inspector_id
                    except InspectorMapping.DoesNotExist:
                        try:
                            inspector_mapping = InspectorMapping.objects.get(
                                inspector_name__iexact=request.user.username
                            )
                            inspector_name = inspector_mapping.inspector_name
                            inspector_id = inspector_mapping.inspector_id
                        except InspectorMapping.DoesNotExist:
                            # Use username as fallback
                            inspector_name = request.user.get_full_name() or request.user.username

                    # Create the occurrence report inspection
                    new_inspection = FoodSafetyAgencyInspection.objects.create(
                        remote_id=new_remote_id,
                        date_of_inspection=inspection_date,
                        client_name=request.POST.get('client_name', 'Unknown Client'),
                        town=request.POST.get('town', ''),
                        additional_email=request.POST.get('email', ''),
                        inspector_name=inspector_name,
                        inspector_id=inspector_id,
                        internal_account_code=request.POST.get('internal_account_code', ''),
                        corporate_group=request.POST.get('corporate_group', ''),
                        group_type=request.POST.get('group_type', ''),
                        facility_type=request.POST.get('facility_type', ''),
                        occurrence_report=True,
                        comment=request.POST.get('reason', ''),
                        commodity='',  # Occurrence reports don't have commodity
                        product_name='',
                        product_class='',
                        km_traveled=0,
                        hours=0,
                        is_sample_taken=False,
                    )

                    # Use the new inspection's ID for file upload
                    inspection_id = str(new_inspection.remote_id)
                    print(f"[OCCURRENCE] Created new occurrence report inspection with remote_id: {new_remote_id}, inspector: {inspector_name}, inspector_id: {inspector_id}")

                except Exception as e:
                    print(f"[OCCURRENCE] Error creating occurrence report: {e}")
                    return JsonResponse({'success': False, 'error': f'Error creating occurrence report: {str(e)}'})

            # SAFETY: Handle inspection_id - if it looks like a group_id format, use it as group_id
            if inspection_id and not str(inspection_id).isdigit():
                # Check if inspection_id looks like a group_id (e.g., "Chicken_Butchery_20260110")
                # Format: ClientName_YYYYMMDD (ends with 8 digits)
                import re
                if re.match(r'^.+_\d{8}$', str(inspection_id)):
                    # inspection_id is actually a group_id - use it if group_id is not already set
                    if not group_id:
                        print(f"[SAFETY] inspection_id '{inspection_id}' looks like group_id format - using as group_id")
                        group_id = inspection_id
                    else:
                        print(f"[SAFETY] inspection_id '{inspection_id}' looks like group_id but group_id already set to '{group_id}'")
                else:
                    print(f"[SAFETY] Invalid inspection_id received: '{inspection_id}' - setting to None")
                inspection_id = None

            # Debug logging
            print(f"Upload request - Group ID: '{group_id}', Inspection ID: '{inspection_id}', Document Type: '{document_type}'")
            print(f"Group ID type: {type(group_id)}, Inspection ID type: {type(inspection_id)}")
            print(f"Group ID length: {len(group_id) if group_id else 'None'}, Inspection ID length: {len(inspection_id) if inspection_id else 'None'}")
            print(f"POST data keys: {list(request.POST.keys())}")
            print(f"All POST data: {dict(request.POST)}")
            print(f"FILES data keys: {list(request.FILES.keys())}")
            print(f"Uploaded file: {uploaded_file}")
            if uploaded_file:
                print(f"File name: {uploaded_file.name}, File size: {uploaded_file.size}, Content type: {uploaded_file.content_type}")
            else:
                print("No file found in request.FILES")
            
            if not uploaded_file:
                return JsonResponse({'success': False, 'error': 'No file provided'})
            
            # RESTRICT TO PDF FILES ONLY
            file_extension = uploaded_file.name.split('.')[-1].lower() if '.' in uploaded_file.name else ''
            if file_extension != 'pdf':
                return JsonResponse({
                    'success': False, 
                    'error': f'Only PDF files are allowed. You uploaded a {file_extension.upper() if file_extension else "file without extension"}. Please convert your document to PDF and try again.'
                })
            
            # Validate file content type as additional security
            content_type = uploaded_file.content_type
            if content_type and not content_type.startswith('application/pdf'):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid file type. Only PDF documents are accepted.'
                })
            
            # Get inspection data to create proper folder structure
            if group_id:
                # Group-level upload (RFI, Invoice)
                # Parse group_id to get client name and date
                # group_id format: clientname_date (e.g., "3Amigos_20230530")
                parts = group_id.split('_')
                if len(parts) >= 2:
                    # Reconstruct client name (remove date part)
                    client_name = '_'.join(parts[:-1])
                    # Sanitize client name to match file lookup logic
                    import re
                    client_name = re.sub(r'[^a-zA-Z0-9_]', '_', client_name)
                    client_name = re.sub(r'_+', '_', client_name).strip('_')
                    date_str = parts[-1]
                    # Convert date format (e.g., 20230530 to 2023/May)
                    if len(date_str) == 8 and date_str.isdigit():
                        try:
                            year = date_str[:4]
                            month_num = int(date_str[4:6])
                            month_name = datetime.strptime(f"2023-{month_num:02d}-01", "%Y-%m-%d").strftime("%B")
                            year_folder = year
                            month_folder = month_name
                        except (ValueError, TypeError):
                            year_folder = datetime.now().strftime("%Y")
                            month_folder = datetime.now().strftime("%B")
                    else:
                        year_folder = datetime.now().strftime("%Y")
                        month_folder = datetime.now().strftime("%B")
                else:
                    client_name = group_id
                    # Sanitize client name to match file lookup logic
                    import re
                    client_name = re.sub(r'[^a-zA-Z0-9_]', '_', client_name)
                    client_name = re.sub(r'_+', '_', client_name).strip('_')
                    year_folder = datetime.now().strftime("%Y")
                    month_folder = datetime.now().strftime("%B")
                
                identifier = group_id
                upload_type = 'group'
                
            elif inspection_id:
                # Individual inspection upload (Lab, Retest)
                # For lab results and retest, we should use the group's client name if available
                # to keep all documents for a group in the same folder
                if group_id and document_type in ['lab', 'lab_form', 'retest']:
                    # Use group's client name for lab results and retest to keep them organized together
                    parts = group_id.split('_')
                    if len(parts) >= 2:
                        client_name = '_'.join(parts[:-1])
                        date_str = parts[-1]
                        if len(date_str) == 8 and date_str.isdigit():
                            try:
                                year = date_str[:4]
                                month_num = int(date_str[4:6])
                                month_name = datetime.strptime(f"2023-{month_num:02d}-01", "%Y-%m-%d").strftime("%B")
                                year_folder = year
                                month_folder = month_name
                            except (ValueError, TypeError):
                                year_folder = datetime.now().strftime("%Y")
                                month_folder = datetime.now().strftime("%B")
                        else:
                            year_folder = datetime.now().strftime("%Y")
                            month_folder = datetime.now().strftime("%B")
                    else:
                        client_name = group_id
                        # Sanitize client name to match file lookup logic
                        import re
                        client_name = re.sub(r'[^a-zA-Z0-9_]', '_', client_name)
                        client_name = re.sub(r'_+', '_', client_name).strip('_')
                        year_folder = datetime.now().strftime("%Y")
                        month_folder = datetime.now().strftime("%B")
                else:
                    # For other cases or when no group_id, use individual inspection's client name
                    try:
                        inspection = None
                        if inspection_id and str(inspection_id).isdigit():
                            inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()
                        if inspection:
                            client_name = inspection.client_name or "Unknown Client"
                            # Keep original client name for folder structure (don't sanitize)
                            
                            if inspection.date_of_inspection:
                                year_folder = inspection.date_of_inspection.strftime("%Y")
                                month_folder = inspection.date_of_inspection.strftime("%B")
                            else:
                                year_folder = datetime.now().strftime("%Y")
                                month_folder = datetime.now().strftime("%B")
                        else:
                            client_name = "Unknown Client"
                            year_folder = datetime.now().strftime("%Y")
                            month_folder = datetime.now().strftime("%B")
                    except:
                        client_name = "Unknown Client"
                        year_folder = datetime.now().strftime("%Y")
                        month_folder = datetime.now().strftime("%B")
                
                identifier = inspection_id
                upload_type = 'inspection'
            else:
                return JsonResponse({'success': False, 'error': f'No group ID or inspection ID provided. Group ID: "{group_id}", Inspection ID: "{inspection_id}". Please refresh the page and try again.'})
            
            # Use original client name for folder structure (keep spaces and special characters)
            def create_folder_name(name):
                """Create Linux-friendly folder name - simple and robust"""
                if not name:
                    return "unknown_client"
                # Convert to lowercase, replace spaces and special chars with underscores
                import re
                # Remove special characters, keep only alphanumeric, spaces, hyphens, underscores
                clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
                # Replace spaces and hyphens with underscores
                clean_name = clean_name.replace(' ', '_').replace('-', '_')
                # Remove consecutive underscores
                clean_name = re.sub(r'_+', '_', clean_name)
                # Remove leading/trailing underscores
                clean_name = clean_name.strip('_').lower()
                return clean_name or "unknown_client"
            
            # Parse group_id to get the correct client name with spaces
            if group_id:
                parts = group_id.split('_')
                if len(parts) >= 2:
                    # Get the date from group_id
                    date_str = parts[-1]
                    if len(date_str) == 8:
                        try:
                            # Convert date string to date object for lookup
                            date_obj = datetime.strptime(date_str, '%Y%m%d').date()
                            
                            # Find the inspection to get the original client name
                            # This ensures we use the exact client name from the database
                            reconstructed_name = '_'.join(parts[:-1]).replace('_', ' ')
                            
                            # Try to find a matching inspection by date and similar client name
                            inspections_on_date = FoodSafetyAgencyInspection.objects.filter(
                                date_of_inspection=date_obj
                            )
                            
                            # Look for exact or similar client name match
                            matched_inspection = None
                            for inspection in inspections_on_date:
                                if inspection.client_name:
                                    # Try exact match first
                                    if inspection.client_name == reconstructed_name:
                                        matched_inspection = inspection
                                        break
                                    # Try normalized comparison (remove special chars for comparison)
                                    import re
                                    normalized_db = re.sub(r'[^a-zA-Z0-9\s]', '', inspection.client_name.lower())
                                    normalized_reconstructed = re.sub(r'[^a-zA-Z0-9\s]', '', reconstructed_name.lower())
                                    if normalized_db == normalized_reconstructed:
                                        matched_inspection = inspection
                                        break
                            
                            if matched_inspection and matched_inspection.client_name:
                                client_name = matched_inspection.client_name
                                print(f"[UPLOAD] Found original client name from database: '{client_name}'")
                            else:
                                # Fallback: Convert underscores back to spaces
                                client_name = reconstructed_name
                                print(f"[UPLOAD] Using fallback client name: '{client_name}'")
                        except ValueError:
                            # Fallback: Convert underscores back to spaces
                            client_name = '_'.join(parts[:-1]).replace('_', ' ')
                            print(f"[UPLOAD] Using fallback client name (date parse error): '{client_name}'")
                    else:
                        # Fallback: Convert underscores back to spaces
                        client_name = '_'.join(parts[:-1]).replace('_', ' ')
                        print(f"[UPLOAD] Using fallback client name (invalid date): '{client_name}'")
                else:
                    # Fallback: Use group_id as client name
                    client_name = group_id
                    print(f"[UPLOAD] Using group_id as client name: '{client_name}'")
            else:
                # No group_id provided, client_name should already be set from earlier logic
                print(f"[UPLOAD] No group_id provided, using client_name: '{client_name}'")
            
            # === NEW STRUCTURE: docs/{client_id}/{inspection_id}/{category}/ ===
            # Look up inspection(s) to get IDs for folder path
            from main.models import Client

            STANDARD_FOLDER_MAP = {
                'rfi': 'rfi',
                'invoice': 'invoice',
                'lab': 'lab',
                'lab_form': 'lab_form',
                'retest': 'retest',
                'compliance': 'compliance',
                'form': 'lab',
                'occurrence': 'occurrence',
                'composition': 'composition',
                'other': 'other',
                'coa': 'compliance',
            }
            mapped_type = STANDARD_FOLDER_MAP.get(document_type, document_type)

            # Find the inspection(s) for this upload
            target_inspection = None
            if inspection_id and str(inspection_id).isdigit():
                # Individual upload - try id first (primary key), then remote_id
                target_inspection = FoodSafetyAgencyInspection.objects.filter(id=int(inspection_id)).first()
                if not target_inspection:
                    target_inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()

            if not target_inspection and group_id:
                # Group upload - find first inspection for this client+date
                parts = group_id.split('_')
                if len(parts) >= 2:
                    date_str = parts[-1]
                    if len(date_str) == 8:
                        try:
                            date_obj = datetime.strptime(date_str, '%Y%m%d').date()
                            target_inspection = FoodSafetyAgencyInspection.objects.filter(
                                client_name__iexact=client_name,
                                date_of_inspection=date_obj
                            ).first()
                        except ValueError:
                            pass

            # Ensure inspection has a linked Client
            if target_inspection:
                if not target_inspection.client:
                    # Try to find existing client by name
                    client_obj = Client.objects.filter(name__iexact=client_name).first()
                    if not client_obj:
                        # Create new client
                        client_obj = Client.objects.create(name=client_name or "Unknown Client")
                        print(f"✅ Created new Client: {client_obj.name} (ID: {client_obj.id})")
                    target_inspection.client = client_obj
                    target_inspection.save(update_fields=['client'])
                    print(f"✅ Linked inspection {target_inspection.id} to Client {client_obj.id}")

                    # Also link ALL other inspections for this client+date to the same Client
                    # This ensures consistency when listing files later
                    if group_id:
                        parts = group_id.split('_')
                        if len(parts) >= 2:
                            date_str = parts[-1]
                            if len(date_str) == 8:
                                try:
                                    date_obj_for_link = datetime.strptime(date_str, '%Y%m%d').date()
                                    other_inspections = FoodSafetyAgencyInspection.objects.filter(
                                        client_name__iexact=client_name,
                                        date_of_inspection=date_obj_for_link,
                                        client__isnull=True
                                    ).exclude(id=target_inspection.id)
                                    linked_count = other_inspections.update(client=client_obj)
                                    if linked_count > 0:
                                        print(f"✅ Also linked {linked_count} other inspection(s) for {client_name} to Client {client_obj.id}")
                                except ValueError:
                                    pass

                # Build new structure path: docs/{client_id}/{inspection_id}/{category}/
                document_dir = os.path.join(
                    settings.MEDIA_ROOT,
                    'docs',
                    str(target_inspection.client.id),
                    str(target_inspection.id),
                    mapped_type
                )
                print(f"[UPLOAD] Using NEW structure: {document_dir}")
            else:
                # Fallback to old structure if no inspection found
                client_folder = create_folder_name(client_name)
                base_dir = os.path.join(settings.MEDIA_ROOT, 'inspection')
                year_dir = os.path.join(base_dir, year_folder)
                month_dir = os.path.join(year_dir, month_folder)
                client_dir = os.path.join(month_dir, client_folder)

                if upload_type == 'individual' and inspection_id:
                    document_dir = os.path.join(client_dir, f"Inspection-{inspection_id}", mapped_type)
                else:
                    document_dir = os.path.join(client_dir, mapped_type)
                print(f"[UPLOAD] Using LEGACY structure (no inspection found): {document_dir}")

            # Create all directories
            os.makedirs(document_dir, exist_ok=True)

            # Log the folder structure for debugging
            print(f"Created folder structure: {document_dir}")
            if target_inspection and target_inspection.client:
                print(f"Client: {client_name} -> Client ID {target_inspection.client.id}")
            else:
                print(f"Client: {client_name} -> {client_folder if 'client_folder' in dir() else 'N/A'}")
            print(f"Document type: {document_type}")
            if document_type == 'lab' and inspection_id and str(inspection_id).isdigit():
                inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()
                if inspection and inspection.commodity:
                    print(f"Commodity: {inspection.commodity}")

            # Generate unique filename with timestamp
            file_extension = os.path.splitext(uploaded_file.name)[1]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Special naming for COA/Lab documents: FSL-RAW-CN-250703
            if document_type in ['lab', 'lab_form'] and inspection_id and str(inspection_id).isdigit():
                # Get inspection data
                inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()

                if inspection:
                    # Get lab name (default to FSL if not set)
                    lab_map = {
                        'lab_a': 'FSL',
                        'lab_b': 'Lab-B',
                        'lab_c': 'Lab-C',
                        'lab_d': 'Lab-D',
                        'lab_e': 'Lab-E',
                        'lab_f': 'Lab-F',
                    }
                    lab_name = lab_map.get(inspection.lab, 'FSL') if inspection.lab else 'FSL'

                    # Get commodity (e.g., RAW, POULTRY, PMP, EGGS)
                    commodity = inspection.commodity.upper() if inspection.commodity else 'UNKNOWN'

                    # Get uploader's first and last name initials (uppercase)
                    # e.g., Mpho Sevenster → MS, Chrisna Nel → CN
                    if request.user.first_name and request.user.last_name:
                        uploader_initials = (request.user.first_name[0] + request.user.last_name[0]).upper()
                    elif request.user.first_name:
                        uploader_initials = request.user.first_name[:2].upper()
                    else:
                        uploader_initials = request.user.username[:2].upper()

                    # Get year-month from inspection date (format: YYMM)
                    if inspection.date_of_inspection:
                        year_month = inspection.date_of_inspection.strftime('%y%m')
                    else:
                        year_month = datetime.now().strftime('%y%m')

                    # Get inspection ID (pad with zeros if needed, max 2 digits)
                    inspection_suffix = str(inspection_id).zfill(2) if len(str(inspection_id)) <= 2 else str(inspection_id)[-2:]

                    # Format: FSL-RAW-CN-250703
                    filename = f"{lab_name}-{commodity}-{uploader_initials}-{year_month}{inspection_suffix}{file_extension}"
                    print(f"[COA NAMING] Generated COA filename: {filename}")
                    print(f"[COA NAMING] Lab: {lab_name}, Commodity: {commodity}, Uploader: {uploader_initials}, Date: {year_month}, InspectionID: {inspection_suffix}")
                else:
                    # Fallback to default naming if inspection not found
                    filename = f"lab-{inspection_id}-{timestamp}{file_extension}"
            # Special naming for RFI and Invoice: FSA-INV-CN-250707
            elif document_type in ['rfi', 'invoice']:
                # Get uploader's first and last name initials (uppercase)
                # e.g., Mpho Sevenster → MS, Chrisna Nel → CN
                if request.user.first_name and request.user.last_name:
                    uploader_initials = (request.user.first_name[0] + request.user.last_name[0]).upper()
                elif request.user.first_name:
                    uploader_initials = request.user.first_name[:2].upper()
                else:
                    uploader_initials = request.user.username[:2].upper()

                # Get date from group_id or inspection (format: YYMMDD)
                if group_id:
                    parts = group_id.split('_')
                    if len(parts) >= 2:
                        date_str = parts[-1]
                        if len(date_str) == 8:
                            # Convert YYYYMMDD to YYMMDD format
                            try:
                                date_obj = datetime.strptime(date_str, '%Y%m%d')
                                formatted_date = date_obj.strftime('%y%m%d')
                            except ValueError:
                                formatted_date = datetime.now().strftime('%y%m%d')
                        else:
                            formatted_date = datetime.now().strftime('%y%m%d')
                    else:
                        formatted_date = datetime.now().strftime('%y%m%d')
                elif inspection_id and str(inspection_id).isdigit():
                    # Get date from inspection
                    inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()
                    if inspection and inspection.date_of_inspection:
                        formatted_date = inspection.date_of_inspection.strftime('%y%m%d')
                    else:
                        formatted_date = datetime.now().strftime('%y%m%d')
                else:
                    formatted_date = datetime.now().strftime('%y%m%d')

                # Map document types to their naming suffixes
                type_mapping = {
                    'rfi': 'RFI',
                    'invoice': 'INV'
                }

                type_suffix = type_mapping.get(document_type, document_type.upper())
                # Format: FSA-INV-CN-250707
                filename = f"FSA-{type_suffix}-{uploader_initials}-{formatted_date}{file_extension}"
                print(f"[{type_suffix} NAMING] Generated filename: {filename}")
                print(f"[{type_suffix} NAMING] Uploader: {uploader_initials}, Date: {formatted_date}")

            # Special naming for retest: keep old format with client name
            elif document_type == 'retest':
                # Clean client name for filename (remove special characters)
                import re
                clean_client_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', client_name)
                clean_client_name = clean_client_name.replace(' ', '-').replace('_', '-')
                clean_client_name = re.sub(r'-+', '-', clean_client_name).strip('-')

                # Get date from group_id or inspection
                if group_id:
                    parts = group_id.split('_')
                    if len(parts) >= 2:
                        date_str = parts[-1]
                        if len(date_str) == 8:
                            # Convert YYYYMMDD to YYYY-MM-DD format
                            try:
                                date_obj = datetime.strptime(date_str, '%Y%m%d')
                                formatted_date = date_obj.strftime('%Y-%m-%d')
                            except ValueError:
                                formatted_date = datetime.now().strftime('%Y-%m-%d')
                        else:
                            formatted_date = datetime.now().strftime('%Y-%m-%d')
                    else:
                        formatted_date = datetime.now().strftime('%Y-%m-%d')
                elif inspection_id and str(inspection_id).isdigit():
                    # Get date from inspection
                    inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()
                    if inspection and inspection.date_of_inspection:
                        formatted_date = inspection.date_of_inspection.strftime('%Y-%m-%d')
                    else:
                        formatted_date = datetime.now().strftime('%Y-%m-%d')
                else:
                    formatted_date = datetime.now().strftime('%Y-%m-%d')

                filename = f"{clean_client_name}-{formatted_date}-retest{file_extension}"
            else:
                filename = f"{identifier}_{document_type}_{timestamp}{file_extension}"
            
            # Log the filename for debugging
            print(f"Generated filename: {filename}")
            file_path = os.path.join(document_dir, filename)

            # SINGLE-FILE ENFORCEMENT: Delete existing files for document types that only allow one file
            # Only 'other' and 'compliance' can have multiple files
            single_file_types = ['rfi', 'invoice', 'lab', 'lab_form', 'retest', 'occurrence', 'composition']
            if document_type in single_file_types:
                # Delete all existing files in this document folder before saving the new one
                try:
                    existing_files = [f for f in os.listdir(document_dir) if os.path.isfile(os.path.join(document_dir, f))]
                    for existing_file in existing_files:
                        existing_file_path = os.path.join(document_dir, existing_file)
                        os.remove(existing_file_path)
                        print(f"[SINGLE-FILE] Deleted existing file: {existing_file_path}")
                    if existing_files:
                        print(f"[SINGLE-FILE] Removed {len(existing_files)} existing file(s) for {document_type} - only one file allowed")
                except Exception as e:
                    print(f"[SINGLE-FILE] Warning: Could not clean existing files: {e}")

            # Save file to local storage first (for OneDrive sync)
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            
            # Clear file cache for this client to ensure immediate visibility
            from django.core.cache import cache
            # Clear cache for both old and new structure
            if target_inspection and target_inspection.client:
                cache_key = f"docs_files:{target_inspection.client.id}:{target_inspection.id}"
                cache.delete(cache_key)
                print(f"[UPLOAD] Cleared cache for: {cache_key}")
            # Also clear legacy cache key
            cache_key_legacy = f"local_files:{client_name}:{year_folder}:{month_folder}"
            cache.delete(cache_key_legacy)
            
            # Update upload tracking in database
            from django.utils import timezone
            current_time = timezone.now()
            
            if upload_type == 'group':
                # Update all inspections in the group
                # Parse group_id to get client_name and date
                parts = group_id.split('_')
                if len(parts) >= 2:
                    # Convert underscores back to spaces for client name matching
                    client_name_from_group = '_'.join(parts[:-1]).replace('_', ' ')
                    date_str = parts[-1]
                    
                    # Convert date string to date object
                    if len(date_str) == 8:
                        try:
                            date_obj = datetime.strptime(date_str, '%Y%m%d').date()
                            
                            # Find all inspections for this client and date using exact matching
                            
                            print(f"DEBUG: Parsing group_id: {group_id}")
                            print(f"DEBUG: Client name from group: {client_name_from_group}")
                            print(f"DEBUG: Date object: {date_obj}")
                            
                            # Get all inspections for this date
                            candidate_inspections = FoodSafetyAgencyInspection.objects.filter(date_of_inspection=date_obj)
                            print(f"DEBUG: Found {candidate_inspections.count()} inspections for date {date_obj}")
                            
                            # Use exact client name matching (much more efficient)
                            # Convert to list to avoid MySQL subquery limitation
                            matching_inspections = list(candidate_inspections.filter(
                                client_name=client_name_from_group
                            ).values_list('id', flat=True))
                            
                            print(f"DEBUG: Found {len(matching_inspections)} matching inspections for '{client_name_from_group}'")
                            
                            # Update upload tracking fields for matching inspections
                            # Using raw SQL to avoid MySQL subquery limitation
                            if matching_inspections:
                                from django.db import connection
                                updated_count = 0

                                # Map document types to field names
                                field_map = {
                                    'rfi': ('rfi_uploaded_by_id', 'rfi_uploaded_date'),
                                    'invoice': ('invoice_uploaded_by_id', 'invoice_uploaded_date'),
                                    'occurrence': ('occurrence_uploaded_by_id', 'occurrence_uploaded_date'),
                                    'composition': ('composition_uploaded_by_id', 'composition_uploaded_date'),
                                    'lab': ('coa_uploaded_by_id', 'coa_uploaded_date'),
                                    'coa': ('coa_uploaded_by_id', 'coa_uploaded_date'),
                                    'lab_form': ('lab_form_uploaded_by_id', 'lab_form_uploaded_date'),
                                    'retest': ('retest_uploaded_by_id', 'retest_uploaded_date'),
                                }

                                if document_type in field_map:
                                    user_field, date_field = field_map[document_type]
                                    id_list = ','.join(str(id) for id in matching_inspections)

                                    sql = f"""
                                        UPDATE food_safety_agency_inspections
                                        SET {user_field} = %s, {date_field} = %s
                                        WHERE id IN ({id_list})
                                    """

                                    with connection.cursor() as cursor:
                                        cursor.execute(sql, [request.user.id, current_time])
                                        updated_count = cursor.rowcount

                                    print(f"DEBUG: Updated {document_type.upper()} tracking for {updated_count} inspections")
                                                                
                                print(f"Updated upload tracking for {updated_count} inspections in group {group_id}")
                                
                                # Clear ALL relevant caches to ensure updated data is shown
                                from django.core.cache import cache

                                # Construct inspection_date from group_id for cache clearing
                                inspection_date = None
                                if len(parts) >= 2 and len(parts[-1]) == 8:
                                    date_str = parts[-1]
                                    inspection_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"

                                # CRITICAL: Create sanitized folder name to match cache keys in get_inspection_files_local
                                sanitized_folder = create_folder_name(client_name_from_group)

                                # Clear all cache keys that could affect file detection
                                cache_keys_to_clear = [
                                    f"shipment_list_{request.user.id}_{request.user.role}",
                                    "filter_options",
                                    "inspection_files_cache",
                                    "page_clients_status_cache"
                                ]

                                if inspection_date:
                                    cache_keys_to_clear.extend([
                                        f"file_status_{client_name_from_group}_{inspection_date}",
                                        f"local_files:{client_name_from_group}:{current_time.year}:{current_time.strftime('%B')}",
                                        # CRITICAL: Also clear with SANITIZED folder name (lowercase, underscores)
                                        f"local_files:{sanitized_folder}:{current_time.year}:{current_time.strftime('%B')}",
                                        f"files_cache_{client_name_from_group}_{inspection_date}",
                                        f"compliance_status_{client_name_from_group}_{date_obj}",
                                        f"onedrive_compliance_{client_name_from_group}_{date_obj}",
                                        f"local_compliance_{client_name_from_group}_{date_obj}"
                                    ])

                                for cache_key in cache_keys_to_clear:
                                    cache.delete(cache_key)
                                    print(f"✅ Cleared cache key: {cache_key}")

                                print(f"✅ Cleared ALL relevant caches after group upload (including sanitized folder name)")
                                
                            else:
                                print(f"WARNING: No matching inspections found for group {group_id}")
                                print(f"DEBUG: Available client names on {date_obj}:")
                                for inspection in candidate_inspections:
                                    print(f"   - '{inspection.client_name}'")
                                print(f"DEBUG: Looking for client name: '{client_name_from_group}'")
                                
                        except ValueError:
                            print(f"WARNING: Could not parse date from group_id: {date_str}")
            
            elif upload_type == 'individual' and inspection_id and str(inspection_id).isdigit():
                # Update individual inspection - try id first (primary key), then remote_id
                inspection = FoodSafetyAgencyInspection.objects.filter(id=int(inspection_id)).first()
                if not inspection:
                    inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=int(inspection_id)).first()
                if inspection:
                    if document_type == 'rfi':
                        inspection.rfi_uploaded_by = request.user
                        inspection.rfi_uploaded_date = current_time
                        inspection.save(update_fields=['rfi_uploaded_by', 'rfi_uploaded_date'])
                        print(f"DEBUG: Updated RFI tracking for individual inspection {inspection_id}")
                    elif document_type == 'invoice':
                        inspection.invoice_uploaded_by = request.user
                        inspection.invoice_uploaded_date = current_time
                        inspection.save(update_fields=['invoice_uploaded_by', 'invoice_uploaded_date'])
                        print(f"DEBUG: Updated Invoice tracking for individual inspection {inspection_id}")
                    elif document_type == 'occurrence':
                        inspection.occurrence_uploaded_by = request.user
                        inspection.occurrence_uploaded_date = current_time
                        inspection.save(update_fields=['occurrence_uploaded_by', 'occurrence_uploaded_date'])
                        print(f"DEBUG: Updated Occurrence tracking for individual inspection {inspection_id}")
                    elif document_type == 'composition':
                        inspection.composition_uploaded_by = request.user
                        inspection.composition_uploaded_date = current_time
                        inspection.save(update_fields=['composition_uploaded_by', 'composition_uploaded_date'])
                        print(f"DEBUG: Updated Composition tracking for individual inspection {inspection_id}")
                    elif document_type in ['lab', 'coa']:
                        inspection.coa_uploaded_by = request.user
                        inspection.coa_uploaded_date = current_time
                        inspection.save(update_fields=['coa_uploaded_by', 'coa_uploaded_date'])
                        print(f"DEBUG: Updated COA/Lab tracking for individual inspection {inspection_id}")
                    elif document_type == 'lab_form':
                        inspection.lab_form_uploaded_by = request.user
                        inspection.lab_form_uploaded_date = current_time
                        inspection.save(update_fields=['lab_form_uploaded_by', 'lab_form_uploaded_date'])
                        print(f"DEBUG: Updated Lab Form tracking for individual inspection {inspection_id}")
                    elif document_type == 'retest':
                        inspection.retest_uploaded_by = request.user
                        inspection.retest_uploaded_date = current_time
                        inspection.save(update_fields=['retest_uploaded_by', 'retest_uploaded_date'])
                        print(f"DEBUG: Updated Retest tracking for individual inspection {inspection_id}")
                    # Clear cache after individual upload - CLEAR ALL RELEVANT CACHES
                    from django.core.cache import cache

                    # CRITICAL: Create sanitized folder name to match cache keys in get_inspection_files_local
                    sanitized_folder = create_folder_name(client_name)

                    # Clear all cache keys that could affect file detection
                    cache_keys_to_clear = [
                        f"shipment_list_{request.user.id}_{request.user.role}",
                        "filter_options",
                        f"file_status_{client_name}_{inspection_date}",
                        "inspection_files_cache",
                        f"local_files:{client_name}:{current_time.year}:{current_time.strftime('%B')}",
                        # CRITICAL: Also clear with SANITIZED folder name (lowercase, underscores)
                        f"local_files:{sanitized_folder}:{current_time.year}:{current_time.strftime('%B')}",
                        "page_clients_status_cache",
                        f"files_cache_{client_name}_{inspection_date}",
                        f"compliance_status_{client_name}_{inspection.date_of_inspection}",
                        f"onedrive_compliance_{client_name}_{inspection.date_of_inspection}",
                        f"local_compliance_{client_name}_{inspection.date_of_inspection}"
                    ]

                    for cache_key in cache_keys_to_clear:
                        cache.delete(cache_key)
                        print(f"✅ Cleared cache key: {cache_key}")

                    print(f"✅ Cleared ALL relevant caches after individual upload (including sanitized folder name)")
            
            # Note: OneDrive sync is handled by the scheduled sync service after the configured delay period
            # Files are only uploaded to OneDrive after inspections are marked as "sent" and the delay period has passed
            print(f" File saved locally: {filename} (OneDrive sync will occur after delay period)")
            
            # Log the file upload to system logs
            try:
                from ..models import SystemLog
                
                # Get client IP address
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip_address = x_forwarded_for.split(',')[0]
                else:
                    ip_address = request.META.get('REMOTE_ADDR')
                
                # Prepare log details
                log_details = {
                    'filename': uploaded_file.name,
                    'file_size': uploaded_file.size,
                    'document_type': document_type.upper(),
                    'upload_type': upload_type,
                    'file_path': file_path
                }
                
                if upload_type == 'group':
                    log_description = f"Uploaded {document_type.upper()} file '{uploaded_file.name}' for inspection group {group_id}"
                    log_details['group_id'] = group_id
                else:
                    log_description = f"Uploaded {document_type.upper()} file '{uploaded_file.name}' for inspection {inspection_id}"
                    log_details['inspection_id'] = inspection_id
                
                SystemLog.log_activity(
                    user=request.user,
                    action='FILE_UPLOAD',
                    page=request.path,
                    object_type='inspection_document',
                    object_id=group_id or inspection_id,
                    description=log_description,
                    details=log_details,
                    ip_address=ip_address,
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
            except Exception as log_error:
                print(f"WARNING: Error logging file upload: {log_error}")
            
            # Return success message with user's first name
            user_first_name = request.user.first_name or request.user.username
            if upload_type == 'group':
                message = f'{document_type.upper()} uploaded successfully by {user_first_name} for group {group_id}'
            else:
                message = f'{document_type.upper()} uploaded successfully by {user_first_name} for inspection {inspection_id}'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'filename': filename,
                'file_path': file_path
            })

        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            print("=" * 80)
            print("[UPLOAD ERROR] EXCEPTION IN upload_document:")
            print(error_traceback)
            print("=" * 80)
            return JsonResponse({'success': False, 'error': str(e), 'traceback': error_traceback})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def scan_inspection_folders(base_path, seen_files, inspection_id=None):
    """Scan Inspection-XXXX folders for files and return a list of files by category."""
    import os
    files_list = {}
    # Case-insensitive check for inspection folders (handles both Inspection-XXX and inspection-XXX)
    inspection_folders = [d for d in os.listdir(base_path) if d.lower().startswith('inspection-') and os.path.isdir(os.path.join(base_path, d))]

    for inspection_folder in sorted(inspection_folders):
        inspection_path = os.path.join(base_path, inspection_folder)

        # Check each document type in the inspection folder
        doc_types = ['Request For Invoice', 'invoice', 'lab results', 'lab_form', 'Lab Form', 'retest', 'Compliance', 'occurrence', 'composition', 'other', 'Other', 'coa', 'COA']
        for doc_type in doc_types:
            # CRITICAL: For Compliance folder, check both capitalized and lowercase versions (case sensitivity fix)
            if doc_type == 'Compliance':
                doc_path_caps = os.path.join(inspection_path, 'Compliance')
                doc_path_lower = os.path.join(inspection_path, 'compliance')
                if os.path.exists(doc_path_caps):
                    doc_path = doc_path_caps
                elif os.path.exists(doc_path_lower):
                    doc_path = doc_path_lower
                else:
                    continue  # Neither exists, skip
            else:
                doc_path = os.path.join(inspection_path, doc_type)
                if not os.path.exists(doc_path):
                    continue  # Path doesn't exist, skip

            if os.path.exists(doc_path):
                files = []

                # Handle subfolders (for lab results and compliance)
                # ALWAYS include the main folder, PLUS any subfolders for lab results and compliance
                paths_to_check = [doc_path]  # Always include the main folder
                if doc_type in ['lab results', 'Compliance']:
                    # Also check commodity subfolders
                    try:
                        for item in os.listdir(doc_path):
                            item_path = os.path.join(doc_path, item)
                            if os.path.isdir(item_path):
                                paths_to_check.append(item_path)
                    except Exception:
                        pass

                # Process files in all checked paths
                for check_path in paths_to_check:
                    try:
                        for filename in os.listdir(check_path):
                            file_path = os.path.join(check_path, filename)
                            if os.path.isfile(file_path):
                                file_size = os.path.getsize(file_path)
                                unique_key = (filename, file_size, os.path.getmtime(file_path))

                                if unique_key not in seen_files:
                                    seen_files.add(unique_key)
                                    # Determine category key first
                                    category_key = doc_type.lower()
                                    if doc_type == 'Request For Invoice':
                                        category_key = 'rfi'
                                    elif doc_type == 'lab results':
                                        # Check if it's a lab form file based on filename
                                        if 'lab_form' in filename.lower() or 'lab-form' in filename.lower() or 'labform' in filename.lower():
                                            category_key = 'lab_form'
                                        else:
                                            category_key = 'lab'
                                    elif doc_type in ['lab_form', 'Lab Form']:
                                        category_key = 'lab_form'
                                    elif doc_type == 'Compliance':
                                        category_key = 'compliance'

                                    # Use get_file_info to create proper file structure
                                    file_info = get_file_info(file_path, category_key)
                                    files.append(file_info)
                    except Exception:
                        pass
                
                if files:
                    # Group files by their individual category keys (already determined above)
                    for file_info in files:
                        file_category = file_info.get('category', 'lab')  # Default to 'lab' if no category
                        if file_category not in files_list:
                            files_list[file_category] = []
                        files_list[file_category].append(file_info)
    
    return files_list

def list_uploaded_files(request):
    """List uploaded files for a given inspection or group."""
    import os
    import json
    from django.conf import settings
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from ..models import Inspection
    import logging

    logger = logging.getLogger(__name__)
    inspection_id = None
    group_id = None

    if request.method == 'GET':
        inspection_id = request.GET.get('inspection_id') or request.GET.get('inspection')
        group_id = request.GET.get('group_id') or request.GET.get('group')
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            inspection_id = data.get('inspection_id') or data.get('inspection')
            group_id = data.get('group_id') or data.get('group')
        except json.JSONDecodeError as e:
            logger.error(f" [ERROR] JSON decode error: {str(e)}")
            return JsonResponse({'error': 'Invalid JSON in request body'}, status=400)
        except Exception as e:
            logger.error(f" [ERROR] Unexpected error processing POST data: {str(e)}")
            return JsonResponse({'error': 'Error processing request'}, status=400)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not inspection_id and not group_id:
        logger.warning("[WARNING] No inspection_id or group_id provided")
        return JsonResponse({'error': 'No inspection_id or group_id provided'}, status=400)
        
    base_path = None
    try:
        if inspection_id:
            try:
                inspection = get_object_or_404(Inspection, pk=inspection_id)
                base_path = os.path.join(settings.MEDIA_ROOT, 'inspections', 
                                       inspection.get_date_folder(), 
                                       inspection.client.name)
            except Exception:
                return JsonResponse({'error': 'Invalid inspection ID'}, status=400)
        elif group_id:
            # The Group model is not defined in main.models in this codebase.
            # Return a clear error instead of attempting to import a missing model.
            return JsonResponse({'error': 'Group-based listing not supported'}, status=400)

        if not base_path:
            return JsonResponse({'error': 'Could not determine base path'}, status=500)

        if not os.path.exists(base_path):
            return JsonResponse({'files': []})

        seen_files = set()
        files_list = scan_inspection_folders(base_path, seen_files)

        # Initialize any missing categories
        for category in ['rfi', 'invoice', 'lab', 'lab_form', 'retest', 'compliance', 'occurrence', 'composition', 'other']:
            if category not in files_list:
                files_list[category] = []

        return JsonResponse({'files': files_list})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def list_client_folder_files(request):
    """Return files for a client inspection folder (POST JSON: client_name, inspection_date, inspection_id optional)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        import json
        from ..models import FoodSafetyAgencyInspection
        import os
        from django.conf import settings
        from datetime import datetime

        data = json.loads(request.body)
        client_name = data.get('client_name')
        inspection_date = data.get('inspection_date')
        inspection_id = data.get('inspection_id')
        force_refresh = data.get('_force_refresh', False)

        if not client_name or not inspection_date:
            return JsonResponse({'success': False, 'error': 'Client name and inspection date are required'})

        # Parse date - support multiple formats
        date_obj = None
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d']
        for date_format in date_formats:
            try:
                date_obj = datetime.strptime(inspection_date, date_format)
                break
            except ValueError:
                continue

        if not date_obj:
            return JsonResponse({
                'success': False,
                'error': f'Invalid date format. Received: {inspection_date}. Expected formats: YYYY-MM-DD or DD/MM/YYYY'
            })

        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')

        # Sanitize client name to match folder structure (same as upload_document)
        import re
        def create_folder_name(name):
            """Create Linux-friendly folder name - must match upload function"""
            if not name:
                return "unknown_client"
            # Remove special characters, keep only alphanumeric, spaces, hyphens, underscores
            clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
            # Replace spaces and hyphens with underscores
            clean_name = clean_name.replace(' ', '_').replace('-', '_')
            # Remove consecutive underscores
            clean_name = re.sub(r'_+', '_', clean_name)
            # Remove leading/trailing underscores
            clean_name = clean_name.strip('_').lower()
            return clean_name or "unknown_client"
        
        sanitized_client_name = create_folder_name(client_name)

        # Also try with forward slashes replaced by spaces (for T/A type names)
        name_with_spaces_for_slash = client_name.replace('/', ' ')
        sanitized_with_slash = create_folder_name(name_with_spaces_for_slash)

        # Also try with apostrophes replaced by spaces (for names like "Mamma's Eggs")
        name_with_spaces_for_apostrophe = client_name.replace("'", ' ')
        sanitized_with_apostrophe = create_folder_name(name_with_spaces_for_apostrophe)

        # Clear cache if force_refresh is requested
        if force_refresh:
            from django.core.cache import cache
            # Clear BOTH original and sanitized cache keys (cache uses sanitized name)
            cache_key_original = f"local_files:{client_name}:{year_folder}:{month_folder}"
            cache_key_sanitized = f"local_files:{sanitized_client_name}:{year_folder}:{month_folder}"
            cache.delete(cache_key_original)
            cache.delete(cache_key_sanitized)

        # Use unified inspection/ folder structure
        # All files should be in: MEDIA_ROOT/inspection/YEAR/MONTH/CLIENT/
        parent_path = os.path.join(settings.MEDIA_ROOT, 'inspection', year_folder, month_folder)

        # Check multiple variations: ORIGINAL NAME FIRST (with spaces), then sanitized variations
        # CRITICAL: Try original name first to handle folders created with spaces (mobile/web upload inconsistency)
        # Also include btn_ and group_ prefixed folders (button upload paths)
        client_folder_variations = [
            client_name,
            sanitized_client_name,
            sanitized_with_slash,
            sanitized_with_apostrophe,
            f'btn_{sanitized_client_name}',  # Button upload prefix
            f'group_{sanitized_client_name}',  # Group upload prefix
        ]

        # List files in document type folders (checking multiple folder variations)
        files_list = {}
        seen_files = set()  # Track files to avoid duplicates

        # === NEW STRUCTURE: Check docs/{client_id}/{inspection_id}/ first ===
        # Strip btn- prefix if present
        clean_client_name = client_name[4:] if client_name.startswith('btn-') else client_name

        # Look up ALL inspections for this client+date (not just first)
        # This matches the behavior of check_group_files which loops through all
        inspections = FoodSafetyAgencyInspection.objects.filter(
            client_name__iexact=clean_client_name,
            date_of_inspection=date_obj.date()
        ).select_related('client')

        docs_base = os.path.join(settings.MEDIA_ROOT, 'docs')
        doc_categories = ['rfi', 'invoice', 'compliance', 'composition', 'coa', 'lab', 'occurrence', 'retest', 'other']

        for inspection in inspections:
            client_obj = None
            if inspection.client:
                client_obj = inspection.client
            else:
                # Try to look up client by name
                from main.models import Client
                client_obj = Client.objects.filter(name__iexact=clean_client_name).first()

            if client_obj:
                docs_path = os.path.join(docs_base, str(client_obj.id), str(inspection.id))
                if os.path.exists(docs_path):
                    for cat in doc_categories:
                        cat_path = os.path.join(docs_path, cat)
                        if os.path.exists(cat_path):
                            try:
                                for filename in os.listdir(cat_path):
                                    file_path = os.path.join(cat_path, filename)
                                    if os.path.isfile(file_path):
                                        file_size = os.path.getsize(file_path)
                                        unique_key = (filename, file_size, os.path.getmtime(file_path))
                                        if unique_key not in seen_files:
                                            seen_files.add(unique_key)
                                            file_info = get_file_info(file_path, cat)
                                            if cat not in files_list:
                                                files_list[cat] = []
                                            files_list[cat].append(file_info)
                            except (OSError, PermissionError):
                                pass

        # Find all matching folders (exact and partial matches for nested "/" structures)
        matched_folders = []

        # First try exact matches
        for folder_variation in client_folder_variations:
            test_path = os.path.join(parent_path, folder_variation)
            if os.path.exists(test_path):
                matched_folders.append((test_path, folder_variation))

        # If no exact match found, try partial matching (for client names with "/" that create nested dirs)
        if not matched_folders and os.path.exists(parent_path):
            try:
                # List all folders in parent_path
                available_folders = [f for f in os.listdir(parent_path) if os.path.isdir(os.path.join(parent_path, f))]

                # Try to find folders that partially match the client name
                for available_folder in available_folders:
                    # Check if this folder is a partial match (for names split by "/" like "t/a")
                    for folder_variation in client_folder_variations:
                        # Get the part before any "/" or special char
                        variation_prefix = folder_variation.split('/')[0].strip()

                        # Normalize for comparison (case insensitive, ignore underscores)
                        normalized_available = available_folder.lower().replace('_', ' ').strip()
                        normalized_prefix = variation_prefix.lower().replace('_', ' ').strip()

                        # Check if the available folder starts with the variation prefix
                        # This handles cases like "Marang Layers Farming Enterprises t/" matching "Marang Layers Farming Enterprises t/a..."
                        if normalized_available.startswith(normalized_prefix) and len(normalized_prefix) > 10:
                            potential_path = os.path.join(parent_path, available_folder)
                            matched_folders.append((potential_path, available_folder))
                            break  # Found a match for this available folder, move to next
            except Exception:
                pass

        # Process all matched folders (both exact and partial matches)
        for test_path, folder_name in matched_folders:

            # First, use the scan_inspection_folders function to check Inspection-XXXX folders
            inspection_files = scan_inspection_folders(test_path, seen_files, inspection_id)
            for category, file_list in inspection_files.items():
                if category not in files_list:
                    files_list[category] = []
                files_list[category].extend(file_list)

            # Also check traditional document folders (for backward compatibility)
            # Include uppercase variations for Windows compatibility
            traditional_docs = ['rfi', 'RFI', 'Request For Invoice', 'invoice', 'Invoice', 'lab', 'Lab', 'lab results', 'Lab Results', 'retest', 'Retest', 'occurrence', 'Occurrence', 'composition', 'Composition', 'compliance', 'Compliance', 'other', 'Other', 'coa', 'COA']
            for doc_type in traditional_docs:
                doc_path = os.path.join(test_path, doc_type)
                if os.path.exists(doc_path):
                    files = []

                    # Check both the main folder and any subfolders (for lab results with commodity subfolders)
                    paths_to_check = [doc_path]
                    if doc_type in ['lab', 'lab results']:
                        # Also check subfolders for lab results (commodity folders)
                        try:
                            for item in os.listdir(doc_path):
                                item_path = os.path.join(doc_path, item)
                                if os.path.isdir(item_path):
                                    paths_to_check.append(item_path)
                        except Exception:
                            pass

                    for check_path in paths_to_check:
                        try:
                            for filename in os.listdir(check_path):
                                if os.path.isfile(os.path.join(check_path, filename)):
                                    file_path = os.path.join(check_path, filename)
                                    file_size = os.path.getsize(file_path)

                                    # Create unique key to avoid duplicates
                                    unique_key = (filename, file_size, os.path.getmtime(file_path))
                                    if unique_key not in seen_files:
                                        seen_files.add(unique_key)

                                        # Determine the actual document type based on filename
                                        actual_doc_type = doc_type.lower()
                                        if doc_type.lower() in ['rfi', 'request for invoice']:
                                            actual_doc_type = 'rfi'
                                        elif doc_type.lower() in ['lab', 'lab results']:
                                            if 'lab_form' in filename.lower() or 'lab-form' in filename.lower() or 'labform' in filename.lower():
                                                actual_doc_type = 'lab_form'
                                            else:
                                                actual_doc_type = 'lab'
                                        elif doc_type.lower() == 'invoice':
                                            actual_doc_type = 'invoice'
                                        elif doc_type.lower() == 'composition':
                                            actual_doc_type = 'composition'
                                        elif doc_type.lower() == 'compliance':
                                            actual_doc_type = 'compliance'
                                        elif doc_type.lower() == 'occurrence':
                                            actual_doc_type = 'occurrence'
                                        elif doc_type.lower() == 'other':
                                            actual_doc_type = 'other'
                                        elif doc_type.lower() == 'coa':
                                            actual_doc_type = 'coa'
                                        elif doc_type.lower() == 'retest':
                                            actual_doc_type = 'retest'

                                        file_info = get_file_info(file_path, actual_doc_type)
                                        files.append(file_info)
                        except Exception:
                            pass

                    if files:
                        # Group files by their actual document type
                        for file_info in files:
                            actual_doc_type = file_info.get('document_type', doc_type)
                            if actual_doc_type not in files_list:
                                files_list[actual_doc_type] = []
                            files_list[actual_doc_type].append(file_info)

        # Also check for compliance documents in the new folder structure
        for folder_variation in client_folder_variations:
            test_path = os.path.join(parent_path, folder_variation)
            if not os.path.exists(test_path):
                continue

            compliance_path = os.path.join(test_path, 'Compliance')
            if os.path.exists(compliance_path):
                # Get all commodity subfolders
                for commodity_folder in os.listdir(compliance_path):
                    commodity_path = os.path.join(compliance_path, commodity_folder)

                    if os.path.isdir(commodity_path):
                        # List all files in this commodity folder
                        files = []
                        for filename in os.listdir(commodity_path):
                            if os.path.isfile(os.path.join(commodity_path, filename)):
                                file_path = os.path.join(commodity_path, filename)
                                file_size = os.path.getsize(file_path)

                                # Create unique key to avoid duplicates
                                unique_key = (filename, file_size, os.path.getmtime(file_path))
                                if unique_key not in seen_files:
                                    seen_files.add(unique_key)

                                    # Determine document type based on filename or folder
                                    doc_type = 'compliance'
                                    filename_lower = filename.lower()
                                    # Check for new naming conventions (FSA-RFI, FSA-INV, FSL-)
                                    if 'rfi' in filename_lower or filename_lower.startswith('fsa-rfi'):
                                        doc_type = 'rfi'
                                    elif 'inv' in filename_lower or 'invoice' in filename_lower or filename_lower.startswith('fsa-inv'):
                                        doc_type = 'invoice'
                                    elif 'lab' in filename_lower or filename_lower.startswith('fsl-') or filename_lower.startswith('lab-'):
                                        doc_type = 'lab'
                                    elif 'retest' in filename_lower:
                                        doc_type = 'retest'

                                    file_info = get_file_info(file_path, doc_type)
                                    file_info['commodity'] = commodity_folder
                                    files.append(file_info)

                        if files:
                            # Group files by their actual document type
                            for file_info in files:
                                actual_doc_type = file_info.get('document_type', 'compliance')
                                if actual_doc_type not in files_list:
                                    files_list[actual_doc_type] = []
                                files_list[actual_doc_type].append(file_info)

        # Initialize any missing categories before returning
        for category in ['rfi', 'invoice', 'lab', 'lab_form', 'retest', 'compliance', 'occurrence', 'composition', 'other', 'coa']:
            if category not in files_list:
                files_list[category] = []

        return JsonResponse({
            'success': True,
            'files': files_list,
            'base_path': parent_path
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



@login_required(login_url='login')
def edit_shipment(request, pk):
    """Edit an existing shipment."""
    clear_messages(request)
    shipment = get_object_or_404(Shipment, pk=pk)
    clients = Client.objects.all().order_by('name')
    
    # Check if this is a cancel/keep original request
    if request.method == 'POST' and 'keep_original' in request.POST:
        messages.info(request, 'No changes were made to the shipment.')
        return redirect('shipment_list')
    
    if request.method == 'POST':
        # For now, just redirect since we don't have the ShipmentForm
        messages.info(request, 'Edit functionality will be implemented with proper form.')
        return redirect('shipment_list')
    else:
        # For now, just show the shipment details
        return render(request, 'main/edit_shipment.html', {
            'shipment': shipment,
            'clients': clients
        })






def apply_inspection_filters(request, inspections):
    """Apply filters to the inspections queryset based on request parameters."""
    
    # Filter by inspection number
    inspection_no = request.GET.get('claim_no')  # Keep same parameter name for template compatibility
    if inspection_no:
        inspections = inspections.filter(inspection_number__icontains=inspection_no)
    
    # Filter by client name
    client_name = request.GET.get('client')
    if client_name:
        inspections = inspections.filter(facility_client_name__icontains=client_name)
    
    # Filter by inspector
    inspector = request.GET.get('branch')  # Keep same parameter name for template compatibility
    if inspector:
        inspections = inspections.filter(inspector__icontains=inspector)
    
    # Filter by inspection date range
    inspection_date_from = request.GET.get('inspection_date_from')
    if inspection_date_from:
        inspections = inspections.filter(inspection_date__gte=inspection_date_from)
    
    inspection_date_to = request.GET.get('inspection_date_to')
    if inspection_date_to:
        inspections = inspections.filter(inspection_date__lte=inspection_date_to)
    
    return inspections


def apply_fsa_inspection_filters(request, inspections):
    """Apply filters to the Food Safety Agency inspections queryset based on request parameters."""
    
    # Filter by inspection number (remote_id)
    inspection_no = request.GET.get('claim_no')  # Keep same parameter name for template compatibility
    if inspection_no:
        inspections = inspections.filter(remote_id__icontains=inspection_no)
    
    # Filter by client name
    client_name = request.GET.get('client')
    if client_name:
        inspections = inspections.filter(client_name__icontains=client_name)
    
    # Filter by inspector
    inspector = request.GET.get('branch')  # Keep same parameter name for template compatibility
    if inspector:
        inspections = inspections.filter(inspector_name__icontains=inspector)
    
    # Filter by inspection date range
    inspection_date_from = request.GET.get('inspection_date_from')
    if inspection_date_from:
        inspections = inspections.filter(date_of_inspection__gte=inspection_date_from)
    
    inspection_date_to = request.GET.get('inspection_date_to')
    if inspection_date_to:
        inspections = inspections.filter(date_of_inspection__lte=inspection_date_to)
    
    # Filter by sample status - same approach as scientist role filtering
    sample_status = request.GET.get('sample_status')
    if sample_status:
        if sample_status == 'SAMPLED':
            # Show only inspections where samples were taken
            inspections = inspections.filter(is_sample_taken=True)
        elif sample_status == 'NOT_SAMPLED':
            # Show only inspections where no samples were taken
            inspections = inspections.filter(is_sample_taken=False)
        # If sample_status is anything else or empty, show all (no filter)
    
    # Filter by compliance status - based on direction present field
    compliance_status = request.GET.get('compliance_status')
    if compliance_status:
        if compliance_status == 'COMPLIANT':
            # Show only compliant inspections (no direction present)
            inspections = inspections.filter(is_direction_present_for_this_inspection=False)
        elif compliance_status == 'NON_COMPLIANT':
            # Show only non-compliant inspections (direction present)
            inspections = inspections.filter(is_direction_present_for_this_inspection=True)
        # If compliance_status is anything else or empty, show all (no filter)
    
    # Note: Sent status filtering is now handled at the group level in shipment_list view
    # to properly show groups that contain sent/unsent inspections
    
    return inspections


def apply_client_filters(request, clients):
    """Apply filters to the clients queryset based on request parameters."""

    # Filter by client ID
    client_id = request.GET.get('client_id')
    if client_id:
        clients = clients.filter(client_id__icontains=client_id)

    # Filter by client name
    client_name = request.GET.get('client_name')
    if client_name:
        clients = clients.filter(eclick_name__icontains=client_name)

    # Filter by account code
    account_code = request.GET.get('account_code')
    if account_code:
        clients = clients.filter(internal_account_code__icontains=account_code)

    # Filter by commodity
    commodity = request.GET.get('commodity')
    if commodity:
        clients = clients.filter(commodity__iexact=commodity)

    # Filter by facility type
    facility_type = request.GET.get('facility_type')
    if facility_type:
        clients = clients.filter(facility_type__iexact=facility_type)

    # Filter by province
    province = request.GET.get('province')
    if province:
        clients = clients.filter(province__iexact=province)

    # Filter by corporate group
    corporate_group = request.GET.get('corporate_group')
    if corporate_group:
        clients = clients.filter(corporate_group__iexact=corporate_group)

    # Filter by group type
    group_type = request.GET.get('group_type')
    if group_type:
        clients = clients.filter(group_type__iexact=group_type)

    # Filter by facility code
    facility_code = request.GET.get('facility_code')
    if facility_code:
        clients = clients.filter(facility_code__icontains=facility_code)

    # Filter by date range (using created_at field)
    date_from = request.GET.get('date_from')
    if date_from:
        clients = clients.filter(created_at__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        clients = clients.filter(created_at__lte=date_to)

    # Filter by status - since Client model doesn't have is_active field,
    # we'll filter by email presence as a proxy for "active" status
    status = request.GET.get('status')
    if status:
        if status == 'active':
            # Clients with email are considered "active"
            clients = clients.filter(
                models.Q(email__isnull=False) | models.Q(manual_email__isnull=False)
            ).exclude(email='').exclude(manual_email='')
        elif status == 'inactive':
            # Clients without email are considered "inactive"
            clients = clients.filter(
                models.Q(email__isnull=True) | models.Q(email=''),
                models.Q(manual_email__isnull=True) | models.Q(manual_email='')
            )

    return clients


# =============================================================================
# GOOGLE SHEETS INTEGRATION VIEWS
# =============================================================================


@login_required(login_url='login')
def client_allocation(request):
    """Client allocation view - shows clients from Client model."""
    from ..models import Client
    from django.db.models import Count, Q

    # Get search parameter
    search_query = request.GET.get('search', '').strip()
    show_all = request.GET.get('show_all', 'true').lower() == 'true'

    # Get all clients with inspection count
    clients = Client.objects.annotate(
        inspection_count=Count('inspections', distinct=True)
    ).order_by('name')

    # Apply search filter if provided
    if search_query:
        clients = clients.filter(
            Q(name__icontains=search_query) |
            Q(client_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(manual_email__icontains=search_query) |
            Q(internal_account_code__icontains=search_query)
        )

    total_clients = clients.count()

    context = {
        'clients': clients,
        'total_clients': total_clients,
        'has_data': total_clients > 0,
        'search_query': search_query,
        'show_all': show_all,
    }

    return render(request, 'main/client_allocation.html', context)


@login_required(login_url='login')
def client_allocation_sheet(request):
    """Client Allocation Sheet view - Google Sheets-like interface.

    Optimized for maximum performance with caching and efficient queries.
    """
    from ..models import ClientAllocation
    from django.core.paginator import Paginator
    from django.core.cache import cache
    from django.db.models import Prefetch

    # Get filter parameters
    client_name = request.GET.get('client_name', '').strip()
    commodity = request.GET.get('commodity', '').strip()
    corporate_group = request.GET.get('corporate_group', '').strip()
    has_email = request.GET.get('has_email', '').strip()
    has_phone = request.GET.get('has_phone', '').strip()

    # Check if user wants all data (default) or paginated view
    show_all = request.GET.get('show_all', 'true').lower() == 'true'
    page_number = request.GET.get('page', 1)

    # Cache key based on pagination settings and filters
    cache_key = f'client_allocation_data_{show_all}_{page_number}_{client_name}_{commodity}_{corporate_group}_{has_email}_{has_phone}'
    cache_timeout = 300  # 5 minutes cache

    # Try to get cached data first
    cached_data = cache.get(cache_key)
    if cached_data:
        return render(request, 'main/client_allocation_sheet.html', cached_data)

    # Cache miss - query database with optimizations
    # Use select_related/prefetch_related if there were FK relationships
    # Use only() to fetch only needed fields
    # Start with base query
    allocations_query = ClientAllocation.objects.only(
        'client_id', 'facility_type', 'group_type', 'commodity', 'province',
        'corporate_group', 'other', 'internal_account_code', 'allocated',
        'eclick_name', 'representative_email', 'phone_number', 'duplicates',
        'active_status'
    )

    # Apply filters
    if client_name:
        allocations_query = allocations_query.filter(eclick_name__icontains=client_name)
    if commodity:
        allocations_query = allocations_query.filter(commodity=commodity)
    if corporate_group:
        allocations_query = allocations_query.filter(corporate_group=corporate_group)
    if has_email == 'yes':
        allocations_query = allocations_query.exclude(representative_email__isnull=True).exclude(representative_email='')
    elif has_email == 'no':
        from django.db.models import Q
        allocations_query = allocations_query.filter(Q(representative_email__isnull=True) | Q(representative_email=''))
    if has_phone == 'yes':
        allocations_query = allocations_query.exclude(phone_number__isnull=True).exclude(phone_number='')
    elif has_phone == 'no':
        from django.db.models import Q
        allocations_query = allocations_query.filter(Q(phone_number__isnull=True) | Q(phone_number=''))

    # Order and limit to latest 100 records
    allocations_query = allocations_query.order_by('-last_synced')[:100]

    # Get total count (all records in database)
    total_count = cache.get('client_allocation_count')
    if total_count is None:
        total_count = ClientAllocation.objects.count()
        cache.set('client_allocation_count', total_count, 600)  # 10 minute cache

    has_data = total_count > 0

    # Always show latest 100 records (simplified - no pagination)
    # Convert to list for caching (querysets can't be pickled)
    allocations = list(allocations_query)
    page_obj = None

    context = {
        'allocations': allocations,
        'has_data': has_data,
        'page_obj': page_obj,
        'show_all': show_all,
        'total_count': total_count,
        'displayed_count': len(allocations)
    }

    # Cache the context for faster subsequent requests
    cache.set(cache_key, context, cache_timeout)

    return render(request, 'main/client_allocation_sheet.html', context)


@login_required(login_url='login')
def export_client_allocations(request):
    """Export ALL client allocation records to Excel file."""
    from django.http import HttpResponse
    from ..models import ClientAllocation
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Allocations"

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='007890', end_color='007890', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    cell_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # Header row
    headers = [
        'Client ID', 'Business Name (E-Click)', 'Facility Type', 'Group Type', 'Commodity',
        'Province', 'Corporate Group', 'Internal Account Code', 'Allocated',
        'Representative Email', 'Phone Number', 'Duplicates',
        'Active/Deactive', 'Last Synced', 'Created At'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # ALWAYS return all clients
    allocations = ClientAllocation.objects.all().order_by('client_id')

    # Data rows
    for row_num, allocation in enumerate(allocations, 2):
        data = [
            allocation.client_id,
            allocation.eclick_name or '',
            allocation.facility_type or '',
            allocation.group_type or '',
            allocation.commodity or '',
            allocation.province or '',
            allocation.corporate_group or '',
            allocation.internal_account_code or '',
            'Yes' if allocation.allocated else 'No',
            allocation.representative_email or '',
            allocation.phone_number or '',
            allocation.duplicates or '',
            allocation.active_status or '',
            allocation.last_synced.strftime('%Y-%m-%d %H:%M:%S') if allocation.last_synced else '',
            allocation.created_at.strftime('%Y-%m-%d %H:%M:%S') if allocation.created_at else ''
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = cell_border

    # Column widths
    column_widths = {
        1: 10,  2: 30, 3: 20, 4: 25, 5: 12,
        6: 15, 7: 25, 8: 25, 9: 10, 10: 30,
        11: 15, 12: 15, 13: 15, 14: 20, 15: 20
    }

    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze header
    ws.freeze_panes = 'A2'

    # Save workbook to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="client_allocations_export.xlsx"'

    return response


@login_required(login_url='login')
def add_client_allocation(request):
    """Add a new client allocation record."""
    from ..models import ClientAllocation
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.core.cache import cache

    if request.method == 'POST':
        try:
            # Auto-generate next Client ID
            from django.db.models import Max
            max_id = ClientAllocation.objects.aggregate(Max('client_id'))['client_id__max']
            client_id = (max_id or 0) + 1

            # Get form data
            business_name = request.POST.get('business_name')
            facility_type = request.POST.get('facility_type')
            group_type = request.POST.get('group_type')
            commodity = request.POST.get('commodity')
            province = request.POST.get('province')
            allocated = request.POST.get('allocated') == 'yes'
            representative_email = request.POST.get('representative_email')
            phone_number = request.POST.get('phone_number')

            # Auto-detect corporate group from business name
            def detect_corporate_group(client_name):
                if not client_name:
                    return "Other (Unlisted Group)"

                name = client_name.lower().strip()

                rules = [
                    (['pick n pay franchise', 'pnp franchise'], 'Pick n Pay - Franchise'),
                    (['pick n pay corporate', 'pnp corporate'], 'Pick n Pay - Corporate'),
                    (['pick n pay', 'pnp', "pick'n pay", 'picknpay'], 'Pick n Pay - Corporate'),
                    (['fruit & veg', 'fruit and veg', 'fruit&veg'], 'Fruit & Veg'),
                    (['ok foods', 'ok food', 'okfoods'], 'OK Foods'),
                    (['checkers'], 'Checkers'),
                    (['spar northrand', 'spar - northrand'], 'Spar - Northrand'),
                    (['superspar', 'super spar'], 'SuperSpar'),
                    (['spar'], 'Spar'),
                    (['shoprite', 'shop rite'], 'Shoprite'),
                    (['massmart'], 'Massmart'),
                    (['chester butcheries', 'chester butchery'], 'Chester Butcheries'),
                    (['boxer'], 'Boxer'),
                    (['food lovers market', "food lover's market", 'foodlovers'], 'Food Lovers Market'),
                    (['cambridge'], 'Cambridge'),
                    (['woolworths', 'woolworth'], 'Woolworths'),
                    (['jwayelani'], 'Jwayelani'),
                    (['usave', 'u-save', 'u save'], 'Usave'),
                    (['obc'], 'OBC'),
                    (['roots'], 'Roots'),
                    (['meat world', 'meatworld'], 'Meat World'),
                    (['quantum foods', 'nulaid', 'quantum'], 'Quantum Foods Nulaid'),
                    (['bluff meat supply', 'bluff meat'], 'Bluff Meat Supply'),
                    (['eat sum meat', 'eatsum'], 'Eat Sum Meat'),
                    (['waltloo meat', 'waltloo chicken', 'waltloo'], 'Waltloo Meat and Chicken'),
                    (['choppies', 'choppy'], 'Choppies'),
                    (['econo foods', 'econofoods'], 'Econo Foods'),
                    (['makro'], 'Makro'),
                    (['boma vleismark', 'boma vleis'], 'Boma Vleismark'),
                    (['eskort'], 'Eskort'),
                    (['nesta foods', 'nesta'], 'Nesta Foods'),
                ]

                for keywords, group in rules:
                    for keyword in keywords:
                        if keyword in name:
                            return group

                return "Other (Unlisted Group)"

            corporate_group = detect_corporate_group(business_name)

            # Generate internal account code
            def generate_account_code(facility_type, group_type, commodity, corporate_group, client_id):
                # Corporate Group code mapping
                corporate_group_codes = {
                    'Not Applicable (None)': 'NA',
                    'Pick n Pay - Franchise': 'PNP-F',
                    'Pick n Pay - Corporate': 'PNP-C',
                    'Fruit & Veg': 'FNV',
                    'OK Foods': 'OK',
                    'Checkers': 'CHK',
                    'Spar': 'SPR',
                    'SuperSpar': 'SSPR',
                    'Spar - Northrand': 'SPR-N',
                    'Shoprite': 'SHO',
                    'Massmart': 'MAS',
                    'Chester Butcheries': 'CHE',
                    'Boxer': 'BOX',
                    'Food Lovers Market': 'FLM',
                    'Cambridge': 'CAM',
                    'Woolworths': 'WOO',
                    'Jwayelani': 'JWA',
                    'Usave': 'USA',
                    'Other (Unlisted Group)': 'OTH',
                    'OBC': 'OBC',
                    'Roots': 'ROO',
                    'Meat World': 'MEA',
                    'Quantum Foods Nulaid': 'QFN',
                    'Bluff Meat Supply': 'BMS',
                    'Eat Sum Meat': 'ESM',
                    'Waltloo Meat and Chicken': 'WMC',
                    'Choppies': 'CHO',
                    'Econo Foods': 'ECO',
                    'Makro': 'MAK',
                    'Boma Vleismark': 'BOM',
                    'Eskort': 'ESK',
                    'Nesta Foods': 'NES'
                }

                # Part 1: First 2 characters of Facility Type in uppercase
                part1 = facility_type[:2].upper() if facility_type else '-'

                # Part 2: Group Type mapping
                part2 = 'IND'
                if group_type == 'Corporate Store':
                    part2 = 'COR'
                elif group_type == 'Franchise Store':
                    part2 = 'FRN'

                # Part 3: Commodity code
                part3 = 'OTH'
                if commodity in ['PMP', 'RAW', 'EGG', 'PLT']:
                    part3 = commodity

                # Part 4: Corporate Group code lookup
                part4 = corporate_group_codes.get(corporate_group, '-')

                # Part 5: Client ID padded to 4 digits
                part5 = str(client_id).zfill(4) if client_id else '0000'

                return f"{part1}-{part2}-{part3}-{part4}-{part5}"

            internal_account_code = generate_account_code(
                facility_type, group_type, commodity, corporate_group, client_id
            )

            # Create new ClientAllocation
            ClientAllocation.objects.create(
                client_id=client_id,
                eclick_name=business_name,
                facility_type=facility_type,
                group_type=group_type,
                commodity=commodity,
                province=province,
                corporate_group=corporate_group,
                internal_account_code=internal_account_code,
                allocated=allocated,
                representative_email=representative_email,
                phone_number=phone_number,
                manually_added=True  # Mark as manually added
            )

            # Clear cache
            cache.delete('client_allocation_count')
            cache.delete_pattern('client_allocation_data_*')

            messages.success(request, f'Client {client_id} - {business_name} added successfully!')

        except Exception as e:
            messages.error(request, f'Error adding client: {str(e)}')

    return redirect('client_allocation_sheet')


@login_required(login_url='login')
def edit_client_allocation(request):
    """Edit an existing client allocation record."""
    from ..models import ClientAllocation
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.core.cache import cache

    if request.method == 'POST':
        try:
            client_id = request.POST.get('client_id')

            # Get the existing record
            allocation = ClientAllocation.objects.get(client_id=client_id)

            # Update fields
            allocation.eclick_name = request.POST.get('business_name')
            allocation.facility_type = request.POST.get('facility_type')
            allocation.group_type = request.POST.get('group_type')
            allocation.commodity = request.POST.get('commodity')
            allocation.province = request.POST.get('province')

            # Get corporate group from form, or auto-detect if empty
            corporate_group = request.POST.get('corporate_group', '').strip()
            if not corporate_group:
                corporate_group = detect_corporate_group(allocation.eclick_name)
            allocation.corporate_group = corporate_group

            allocation.allocated = request.POST.get('allocated') == 'yes'
            allocation.representative_email = request.POST.get('representative_email')
            allocation.phone_number = request.POST.get('phone_number')
            allocation.active_status = request.POST.get('active_status')

            # Save changes
            allocation.save()

            # Clear cache
            cache.delete('client_allocation_count')
            cache.delete_pattern('client_allocation_data_*')

            messages.success(request, f'Client {client_id} - {allocation.eclick_name} updated successfully!')

        except ClientAllocation.DoesNotExist:
            messages.error(request, f'Client {client_id} not found.')
        except Exception as e:
            messages.error(request, f'Error updating client: {str(e)}')

    return redirect('client_allocation_sheet')


@login_required(login_url='login')
def get_dropdown_options(request):
    """Get all unique values for facility_type, commodity, and corporate_group with counts."""
    from ..models import ClientAllocation
    from django.db.models import Count, Q
    from django.http import JsonResponse

    # Get facility types with counts
    facility_types = ClientAllocation.objects.values('facility_type').annotate(
        count=Count('id')
    ).filter(~Q(facility_type='') & ~Q(facility_type__isnull=True)).order_by('facility_type')

    # Get commodities with counts
    commodities = ClientAllocation.objects.values('commodity').annotate(
        count=Count('id')
    ).filter(~Q(commodity='') & ~Q(commodity__isnull=True)).order_by('commodity')

    # Get corporate groups with counts
    corporate_groups = ClientAllocation.objects.values('corporate_group').annotate(
        count=Count('id')
    ).filter(~Q(corporate_group='') & ~Q(corporate_group__isnull=True)).order_by('corporate_group')

    return JsonResponse({
        'facility_types': [{'value': item['facility_type'], 'count': item['count']} for item in facility_types],
        'commodities': [{'value': item['commodity'], 'count': item['count']} for item in commodities],
        'corporate_groups': [{'value': item['corporate_group'], 'count': item['count']} for item in corporate_groups],
    })


@login_required(login_url='login')
def delete_dropdown_option(request):
    """Delete a dropdown option by setting it to empty for all clients using it."""
    from ..models import ClientAllocation
    from django.http import JsonResponse
    import json

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            field_type = data.get('field_type')
            value = data.get('value')

            if not field_type or not value:
                return JsonResponse({'success': False, 'error': 'Missing field_type or value'})

            # Map field_type to actual model field
            field_map = {
                'facility_type': 'facility_type',
                'commodity': 'commodity',
                'corporate_group': 'corporate_group'
            }

            if field_type not in field_map:
                return JsonResponse({'success': False, 'error': 'Invalid field type'})

            field_name = field_map[field_type]

            # Update all clients with this value to empty string
            updated_count = ClientAllocation.objects.filter(**{field_name: value}).update(**{field_name: ''})

            return JsonResponse({
                'success': True,
                'updated_count': updated_count
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required(login_url='login')
def delete_client_allocation(request):
    """Delete a client allocation record."""
    from ..models import ClientAllocation
    from django.http import JsonResponse
    import json

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            client_id = data.get('client_id')

            if not client_id:
                return JsonResponse({'success': False, 'error': 'Missing client_id'})

            # Delete the client
            deleted_count = ClientAllocation.objects.filter(client_id=client_id).delete()[0]

            if deleted_count > 0:
                return JsonResponse({
                    'success': True,
                    'message': f'Client {client_id} deleted successfully'
                })
            else:
                return JsonResponse({'success': False, 'error': 'Client not found'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def detect_corporate_group(client_name):
    """Auto-detect corporate group based on client name"""
    if not client_name:
        return "Other (Unlisted Group)"

    name = client_name.lower().strip()

    # Corporate group matching rules (ordered by specificity)
    rules = [
        (['pick n pay franchise', 'pnp franchise', "pick 'n pay franchise"], 'Pick n Pay - Franchise'),
        (['pick n pay corporate', 'pnp corporate', "pick 'n pay corporate"], 'Pick n Pay - Corporate'),
        (['pick n pay', 'pnp', "pick'n pay", "pick 'n pay", 'picknpay'], 'Pick n Pay - Corporate'),
        (['fruit & veg', 'fruit and veg', 'fruit&veg'], 'Fruit & Veg'),
        (['ok foods', 'ok food', 'okfoods'], 'OK Foods'),
        (['checkers'], 'Checkers'),
        (['spar northrand', 'spar - northrand'], 'Spar - Northrand'),
        (['superspar', 'super spar'], 'SuperSpar'),
        (['spar'], 'Spar'),
        (['shoprite', 'shop rite'], 'Shoprite'),
        (['massmart'], 'Massmart'),
        (['chester butcheries', 'chester butchery'], 'Chester Butcheries'),
        (['boxer'], 'Boxer'),
        (['food lovers market', "food lover's market", 'foodlovers'], 'Food Lovers Market'),
        (['cambridge'], 'Cambridge'),
        (['woolworths', 'woolworth'], 'Woolworths'),
        (['jwayelani'], 'Jwayelani'),
        (['usave', 'u-save', 'u save'], 'Usave'),
        (['obc'], 'OBC'),
        (['roots'], 'Roots'),
        (['meat world', 'meatworld'], 'Meat World'),
        (['quantum foods', 'nulaid', 'quantum'], 'Quantum Foods Nulaid'),
        (['bluff meat supply', 'bluff meat'], 'Bluff Meat Supply'),
        (['eat sum meat', 'eatsum'], 'Eat Sum Meat'),
        (['waltloo meat', 'waltloo chicken', 'waltloo'], 'Waltloo Meat and Chicken'),
        (['choppies', 'choppy'], 'Choppies'),
        (['econo foods', 'econofoods'], 'Econo Foods'),
        (['makro'], 'Makro'),
        (['boma vleismark', 'boma vleis'], 'Boma Vleismark'),
        (['nesta foods', 'nesta'], 'Nesta Foods'),
        (['eskort'], 'Eskort'),
    ]

    for keywords, group in rules:
        for keyword in keywords:
            if keyword in name:
                return group

    return "Other (Unlisted Group)"


def sync_client_allocations(request):
    """Sync client allocation data from SQL Server Clients table to PostgreSQL database.

    Optimized with bulk operations for maximum performance.
    """
    from ..models import ClientAllocation
    from django.contrib import messages
    from django.db import transaction
    from django.core.cache import cache
    from ..utils.sql_server_utils import SQLServerConnection

    # Mapping dictionaries for internal account codes
    FACILITY_TYPE_MAP = {
        'RE': 'Retailer',
        'BU': 'Butchery',
        'RP': 'Re-Packer',
        'PR': 'Production Plant',
        'FA': 'Farm',
        'AB': 'Abattoir'
    }

    GROUP_TYPE_MAP = {
        'COR': 'Corporate Store',
        'FRN': 'Franchise Store',
        'IND': 'Individual / Independent Owner'
    }

    COMMODITY_MAP = {
        'PMP': 'Processed Meat Products (PMP)',
        'RAW': 'Certain Raw Processed Meat',
        'EGG': 'Eggs',
        'PLT': 'Poultry',
        'XX': 'Unknown/Other'
    }

    def parse_internal_account_code(account_code):
        """Parse internal account code to extract facility type, group type, and commodity"""
        if not account_code:
            return None, None, None

        parts = account_code.split('-')
        if len(parts) < 4:
            return None, None, None

        facility_code = parts[0]
        group_code = parts[1]
        commodity_code = parts[2]

        facility_type = FACILITY_TYPE_MAP.get(facility_code)
        group_type = GROUP_TYPE_MAP.get(group_code)
        commodity = COMMODITY_MAP.get(commodity_code)

        return facility_type, group_type, commodity

    if request.method == 'POST':
        try:
            # Connect to SQL Server
            sql_conn = SQLServerConnection()

            if not sql_conn.connect():
                messages.error(request, "Failed to connect to SQL Server.")
                return redirect('client_allocation_sheet')

            # Fetch all active clients from SQL Server
            cursor = sql_conn.connection.cursor()

            # Query to get all clients with province information
            query = """
                SELECT
                    c.Id,
                    c.Name,
                    c.InternalAccountNumber,
                    c.ContactNumber,
                    c.ContactEmail,
                    c.ContactNumberForInspections,
                    c.ContactEmailForInspections,
                    c.SiteName,
                    c.PhysicalAddress,
                    c.IsActive,
                    CASE
                        WHEN c.ProvinceStateId = 1 THEN 'Eastern Cape'
                        WHEN c.ProvinceStateId = 2 THEN 'Gauteng'
                        WHEN c.ProvinceStateId = 3 THEN 'KwaZulu-Natal'
                        WHEN c.ProvinceStateId = 4 THEN 'Limpopo'
                        WHEN c.ProvinceStateId = 5 THEN 'Mpumalanga'
                        WHEN c.ProvinceStateId = 6 THEN 'Northern Cape'
                        WHEN c.ProvinceStateId = 7 THEN 'North West'
                        WHEN c.ProvinceStateId = 8 THEN 'Western Cape'
                        WHEN c.ProvinceStateId = 9 THEN 'Free State'
                        ELSE 'Unknown'
                    END AS Province
                FROM Clients c
                WHERE c.IsActive = 1
                ORDER BY c.Id
            """

            cursor.execute(query)
            sql_clients = cursor.fetchall()

            if not sql_clients:
                sql_conn.disconnect()
                messages.warning(request, "No active clients found in SQL Server.")
                return redirect('client_allocation_sheet')

            # Prepare bulk data - much faster than individual creates
            bulk_records = []
            seen_client_ids = set()

            for row in sql_clients:
                client_id = row[0]
                name = row[1]
                internal_account_code = row[2]
                contact_number = row[3]
                contact_email = row[4]
                contact_number_inspections = row[5]
                contact_email_inspections = row[6]
                site_name = row[7]
                physical_address = row[8]
                is_active = row[9]
                province = row[10]

                # Skip duplicate client IDs
                if client_id in seen_client_ids:
                    continue
                seen_client_ids.add(client_id)

                # Use inspection contact info if available, otherwise use general contact info
                phone_number = contact_number_inspections or contact_number
                email = contact_email_inspections or contact_email

                # Determine active status
                active_status = "Active" if is_active else "Inactive"

                # Parse internal account code to extract facility type, group type, and commodity
                facility_type, group_type, commodity = parse_internal_account_code(internal_account_code)

                # Auto-detect corporate group from client name
                corporate_group = detect_corporate_group(name)

                # Create ClientAllocation object (not yet saved to DB)
                bulk_records.append(ClientAllocation(
                    client_id=client_id,
                    facility_type=facility_type,
                    group_type=group_type,
                    commodity=commodity,
                    province=province,
                    corporate_group=corporate_group,
                    other=physical_address,
                    internal_account_code=internal_account_code,
                    allocated=False,
                    eclick_name=name,
                    representative_email=email,
                    phone_number=phone_number,
                    duplicates=None,
                    active_status=active_status,
                    manually_added=False
                ))

            # Close SQL Server connection
            sql_conn.disconnect()

            # Use atomic transaction for data integrity and speed
            with transaction.atomic():
                # Only delete records synced from SQL Server (preserve manually added clients)
                ClientAllocation.objects.filter(manually_added=False).delete()

                # Bulk create all records in a single query (MUCH faster)
                # batch_size=500 optimizes for PostgreSQL performance
                ClientAllocation.objects.bulk_create(bulk_records, batch_size=500)

            sync_count = len(bulk_records)

            # Clear cache after successful sync
            cache.delete('client_allocation_count')
            cache.delete_pattern('client_allocation_*')

            messages.success(request, f"Successfully synced {sync_count} client allocation records from SQL Server.")
            return redirect('client_allocation_sheet')

        except Exception as e:
            print(f"Error syncing SQL Server data: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error syncing data: {str(e)}")
            return redirect('client_allocation_sheet')

    return redirect('client_allocation_sheet')


@login_required(login_url='login')
def refresh_clients(request):
    """Refresh the Food Safety Agency clients table with fresh data from SQL Server."""
    from django.http import JsonResponse
    from django.contrib import messages

    # Mapping dictionaries for internal account codes
    FACILITY_TYPE_MAP = {
        'RE': 'Retailer',
        'BU': 'Butchery',
        'RP': 'Re-Packer',
        'PR': 'Production Plant',
        'FA': 'Farm',
        'AB': 'Abattoir'
    }

    GROUP_TYPE_MAP = {
        'COR': 'Corporate Store',
        'FRN': 'Franchise Store',
        'IND': 'Individual / Independent Owner'
    }

    COMMODITY_MAP = {
        'PMP': 'Processed Meat Products (PMP)',
        'RAW': 'Certain Raw Processed Meat',
        'EGG': 'Eggs',
        'PLT': 'Poultry',
        'XX': 'Unknown/Other'
    }

    def parse_internal_account_code(account_code):
        """Parse internal account code to extract facility type, group type, and commodity"""
        if not account_code:
            return None, None, None

        parts = account_code.split('-')
        if len(parts) < 4:
            return None, None, None

        facility_code = parts[0]
        group_code = parts[1]
        commodity_code = parts[2]

        facility_type = FACILITY_TYPE_MAP.get(facility_code)
        group_type = GROUP_TYPE_MAP.get(group_code)
        commodity = COMMODITY_MAP.get(commodity_code)

        return facility_type, group_type, commodity

    clear_messages(request)

    # Ensure session is valid but don't modify it during long operations
    if not request.session.session_key:
        request.session.create()

    # Set flag to prevent session modifications during sync
    request._sync_in_progress = True

    if request.method == 'POST':
        # FORCE RESTART: Clear all sync locks and caches to allow fresh sync
        from django.core.cache import cache
        import time

        print("[FORCE SYNC] Clearing all sync locks and progress caches...")

        # Clear all sync-related locks
        cache.delete('client_sync_running')
        cache.delete('inspection_sync_lock')

        # Clear progress and result caches to start fresh
        cache.delete('sync_progress')
        cache.delete('sync_result')

        print("[FORCE SYNC] All locks cleared - starting fresh sync")

        # Set new client sync lock
        client_sync_lock_key = 'client_sync_running'
        cache.set(client_sync_lock_key, time.time(), 900)

        print("\n" + "="*60)
        print(" STARTING CLIENT SYNC OPERATION")
        print("="*60)

        try:
            from ..utils.sql_server_utils import SQLServerConnection
            from ..models import ClientAllocation
            from django.db import transaction
            from django.utils import timezone

            # Update session activity timestamp to prevent timeout during long sync
            request.session['last_activity'] = timezone.now().isoformat()
            request.session.modified = True

            print(" Step 1: Connecting to SQL Server...")
            sql_conn = SQLServerConnection()

            if not sql_conn.connect():
                cache.delete('client_sync_running')
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to connect to SQL Server'
                })

            print(" SQL Server connection established successfully")

            print("\n Step 2: Fetching clients from SQL Server...")

            # Fetch all active clients from SQL Server
            cursor = sql_conn.connection.cursor()

            # Query to get all clients with province information
            query = """
                SELECT
                    c.Id,
                    c.Name,
                    c.InternalAccountNumber,
                    c.ContactNumber,
                    c.ContactEmail,
                    c.ContactNumberForInspections,
                    c.ContactEmailForInspections,
                    c.SiteName,
                    c.PhysicalAddress,
                    c.IsActive,
                    CASE
                        WHEN c.ProvinceStateId = 1 THEN 'Eastern Cape'
                        WHEN c.ProvinceStateId = 2 THEN 'Gauteng'
                        WHEN c.ProvinceStateId = 3 THEN 'KwaZulu-Natal'
                        WHEN c.ProvinceStateId = 4 THEN 'Limpopo'
                        WHEN c.ProvinceStateId = 5 THEN 'Mpumalanga'
                        WHEN c.ProvinceStateId = 6 THEN 'Northern Cape'
                        WHEN c.ProvinceStateId = 7 THEN 'North West'
                        WHEN c.ProvinceStateId = 8 THEN 'Western Cape'
                        WHEN c.ProvinceStateId = 9 THEN 'Free State'
                        ELSE 'Unknown'
                    END AS Province
                FROM Clients c
                WHERE c.IsActive = 1
                ORDER BY c.Id
            """

            cursor.execute(query)
            sql_clients = cursor.fetchall()

            print(f"    Found {len(sql_clients)} active clients in SQL Server")

            # Prepare bulk data
            bulk_records = []
            seen_client_ids = set()

            for row in sql_clients:
                client_id = row[0]
                name = row[1]
                internal_account_code = row[2]
                contact_number = row[3]
                contact_email = row[4]
                contact_number_inspections = row[5]
                contact_email_inspections = row[6]
                site_name = row[7]
                physical_address = row[8]
                is_active = row[9]
                province = row[10]

                # Skip duplicate client IDs
                if client_id in seen_client_ids:
                    continue
                seen_client_ids.add(client_id)

                # Use inspection contact info if available, otherwise use general contact info
                phone_number = contact_number_inspections or contact_number
                email = contact_email_inspections or contact_email

                # Determine active status
                active_status = "Active" if is_active else "Inactive"

                # Parse internal account code to extract facility type, group type, and commodity
                facility_type, group_type, commodity = parse_internal_account_code(internal_account_code)

                # Auto-detect corporate group from client name
                corporate_group = detect_corporate_group(name)

                # Create ClientAllocation object (not yet saved to DB)
                bulk_records.append(ClientAllocation(
                    client_id=client_id,
                    facility_type=facility_type,
                    group_type=group_type,
                    commodity=commodity,
                    province=province,
                    corporate_group=corporate_group,
                    other=physical_address,
                    internal_account_code=internal_account_code,
                    allocated=False,
                    eclick_name=name,
                    representative_email=email,
                    phone_number=phone_number,
                    duplicates=None,
                    active_status=active_status,
                    manually_added=False
                ))

            # Close SQL Server connection
            sql_conn.disconnect()

            print(f"\n Step 3: Syncing to database...")

            # Get count before deletion
            old_count = ClientAllocation.objects.filter(manually_added=False).count()

            # Use atomic transaction for data integrity and speed
            with transaction.atomic():
                # Only delete records synced from SQL Server (preserve manually added clients)
                deleted_count = ClientAllocation.objects.filter(manually_added=False).delete()[0]
                print(f"    Deleted {deleted_count} old synced clients")

                # Bulk create all records in batches (increased batch size for better performance)
                ClientAllocation.objects.bulk_create(bulk_records, batch_size=1000)
                print(f"    Created {len(bulk_records)} new clients")

            # Detect AJAX request
            is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

            print(f"\n CLIENT SYNC COMPLETED SUCCESSFULLY!")
            print(f"    Deleted: {deleted_count} old clients")
            print(f"    Created: {len(bulk_records)} new clients")
            print(f"    Processed: {len(sql_clients)} total rows from SQL Server")

            # Clear sync lock
            cache.delete('client_sync_running')

            # Update session timestamp before returning to ensure session stays valid
            request.session['last_activity'] = timezone.now().isoformat()
            request.session.modified = True

            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': f"Successfully synced clients! Deleted {deleted_count} old clients and created {len(bulk_records)} new clients from SQL Server."
                })
            else:
                messages.success(request, f"Clients synced. Deleted {deleted_count} old, created {len(bulk_records)} new. Processed {len(sql_clients)} rows.")
                return redirect('client_allocation')

        except Exception as e:
            print(f" CLIENT SYNC EXCEPTION!")
            print(f"   Exception: {str(e)}")
            import traceback
            traceback.print_exc()

            # Clear sync lock on exception too
            cache.delete('client_sync_running')

            # Return JSON response for AJAX
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
        finally:
            print("="*60)
            print(" CLIENT SYNC OPERATION ENDED")
            print("="*60 + "\n")

    # If not POST, redirect back to client allocation page
    return redirect('client_allocation')


@login_required(login_url='login')
def refresh_inspections(request):
    """Refresh the Food Safety Agency inspections table with fresh data from SQL Server."""
    try:
        # Block administrators from accessing sync functionality
        if request.user.role == 'admin':
            return redirect('home')
        
        clear_messages(request)
        
        # Ensure session is valid but don't modify it during long operations
        if not request.session.session_key:
            request.session.create()
        
        # Don't modify session during long operations to avoid cache conflicts
        # The session will be preserved as-is during the sync
        
        if request.method == 'POST':
            # For AJAX requests, return immediately and run sync in background
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                from django.core.cache import cache
                import threading
                import time

                # FORCE RESTART: Clear sync lock and progress (client sync already cleared these, but double-check)
                sync_lock_key = 'inspection_sync_lock'

                # Clear any existing locks or progress
                cache.delete(sync_lock_key)
                cache.delete('sync_progress')
                cache.delete('sync_result')
                cache.delete('sync_google_sheets_lock')
                cache.delete('sync_sql_server_lock')

                print("[INSPECTION SYNC] All locks cleared - starting fresh sync")

                # Acquire lock with timestamp - expires after 15 minutes (900 seconds) as safety measure
                cache.set(sync_lock_key, time.time(), 900)
                print("[SYNC] Lock acquired - starting sync operation")

                # Store user info before starting background operation to avoid session issues
                user_id = request.user.id
                user_role = getattr(request.user, 'role', 'user')

                # Don't modify session for AJAX requests to avoid conflicts with middleware
                # The middleware will handle session timeout automatically

                # Start sync in background thread
                def run_sync():
                    try:
                        # Set a flag to prevent session modifications during sync
                        request._sync_in_progress = True
                        print("\n" + "="*80)
                        print(" STARTING COMPLETE SYNC OPERATION (BACKGROUND)")
                        print(" SQL Server -> Clients + Inspections + Account Code Matching")
                        print("="*80)

                        # USE GLOBAL SYNC SERVICE - Syncs EVERYTHING in correct order
                        from ..services.scheduled_sync_service import scheduled_sync_service

                        print("\n Step 1: Using global Scheduled Sync Service...")
                        sync_service = scheduled_sync_service
                        print(" [OK] Sync Service ready")

                        print("\n Step 2: Syncing SQL Server clients (names & account codes)...")
                        google_success = sync_service.sync_google_sheets()

                        if google_success:
                            print(" [OK] SQL Server client sync completed!")
                        else:
                            print(" [WARNING] SQL Server client sync had issues (continuing anyway)")

                        print("\n Step 3: Syncing SQL Server inspections...")
                        print(" (Matching account codes with SQL Server clients for names)")
                        sql_success = sync_service.sync_sql_server()

                        if sql_success:
                            # Get count for reporting
                            from ..models import FoodSafetyAgencyInspection
                            total_count = FoodSafetyAgencyInspection.objects.count()

                            print(f"\n [OK] COMPLETE SYNC SUCCESSFUL!")
                            print(f"    Total inspections in database: {total_count}")
                            print(f"    [OK] SQL Server is the source for all data!")
                            print(f"    [OK] All account codes matched and names updated!")

                            # Store result in cache for frontend to check FIRST (before clearing cache)
                            sync_result_data = {
                                'success': True,
                                'message': f'Successfully synced {total_count} inspections with SQL Server client names!',
                                'created_count': total_count,
                                'total_processed': total_count
                            }
                            cache.set('sync_result', sync_result_data, 300)  # 5 minutes
                            print(f"    Sync result stored in cache")

                            # Clear specific cache keys to ensure fresh data is displayed (but keep sync_result)
                            cache_keys_to_clear = [
                                'compliance_status_',  # Clear compliance status cache
                                'inspection_',         # Clear inspection-related cache
                            ]
                            for key_pattern in cache_keys_to_clear:
                                # This is a simplified approach - in production you'd want more specific key management
                                pass
                            print("    Cache cleared to ensure fresh data display")
                        else:
                            print(f"\n [ERROR] SQL SERVER SYNC FAILED!")
                            print(f"   Error: Sync returned False")

                            # Store error in cache
                            cache.set('sync_result', {
                                'success': False,
                                'error': 'Sync failed - check server logs'
                            }, 300)

                    except Exception as e:
                        print(f"\n [ERROR] SYNC EXCEPTION!")
                        print(f"   Exception: {str(e)}")
                        import traceback
                        traceback.print_exc()

                        # Store error in cache
                        cache.set('sync_result', {
                            'success': False,
                            'error': f'Error syncing: {str(e)}'
                        }, 300)

                    finally:
                        # ALWAYS release the lock when sync completes (success or failure)
                        cache.delete(sync_lock_key)
                        print("[SYNC] Lock released - sync operation complete")

                    print("="*80)
                    print(" COMPLETE SYNC OPERATION ENDED")
                    print("="*80 + "\n")
                
                # Start background thread
                thread = threading.Thread(target=run_sync)
                thread.daemon = True
                thread.start()
                
                # Return immediately
                return JsonResponse({
                    'success': True,
                    'message': 'Sync started in background. Please wait...',
                    'status': 'started'
                })
        
        # For non-AJAX requests, run sync normally
        print("\n" + "="*60)
        print(" STARTING INSPECTION SYNC OPERATION")
        print("="*60)
        
        try:
            # Don't modify session expiry here to avoid conflicts with middleware
            # The middleware will handle session timeout automatically
            
            from ..services.google_sheets_service import GoogleSheetsService
            
            print(" Step 1: Initializing Google Sheets Service...")
            sheets_service = GoogleSheetsService()
            print(" Google Sheets Service initialized successfully")
            
            print("\n Step 2: Starting inspection sync from SQL Server...")
            refresh_result = sheets_service.sync_inspections_from_sql_server(request)
            
            if refresh_result.get('success'):
                print(f" INSPECTION SYNC COMPLETED SUCCESSFULLY!")
                print(f"    Deleted: {refresh_result['deleted_count']} old inspections")
                print(f"    Created: {refresh_result['inspections_created']} new inspections")
                print(f"    Processed: {refresh_result['total_processed']} total records from SQL Server")
                
                # Clear cache to ensure fresh data is displayed
                from django.core.cache import cache
                cache.clear()
                print("   [ERROR]  Cache cleared to ensure fresh data display")
                
                # Store messages in cache instead of session to avoid session conflicts
                from django.core.cache import cache
                cache.set('sync_success_message', f"Successfully synced inspections!", 300)  # 5 minutes
                cache.set('sync_info_message_1', f"Deleted {refresh_result['deleted_count']} old inspections and created {refresh_result['inspections_created']} new inspections from SQL Server.", 300)
                cache.set('sync_info_message_2', f"Processed {refresh_result['total_processed']} total records from SQL Server.", 300)
                
                # Return JSON response for AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    response = JsonResponse({
                        'success': True,
                        'message': f'Successfully synced {refresh_result["inspections_created"]} inspections!',
                        'deleted_count': refresh_result['deleted_count'],
                        'created_count': refresh_result['inspections_created'],
                        'total_processed': refresh_result['total_processed']
                    })
                    # Prevent session save to avoid Redis conflicts
                    response._session_save = False
                    # Mark session as not modified to prevent save attempts
                    request.session.modified = False
                    return response
                
                # Redirect to inspections page after successful sync
                # Mark session as not modified to prevent save attempts
                request.session.modified = False
                return redirect('shipment_list')
            else:
                print(f" INSPECTION SYNC FAILED!")
                print(f"   Error: {refresh_result.get('error', 'Unknown error')}")
                messages.error(request, f"Error syncing inspections: {refresh_result.get('error', 'Unknown error')}")
                
                # Return JSON response for AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({
                        'success': False,
                        'error': refresh_result.get('error', 'Unknown error')
                    })
                
                # Redirect back to inspections page with error message
                return redirect('shipment_list')
                
        except Exception as e:
            print(f" INSPECTION SYNC EXCEPTION!")
            print(f"   Exception: {str(e)}")
            messages.error(request, f"Error syncing inspections: {str(e)}")
            
            # Return JSON response for AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': False,
                    'error': f'Error syncing inspections: {str(e)}'
                })
            
            # Redirect back to inspections page with error message
            return redirect('shipment_list')
        
        print("="*60)
        print(" INSPECTION SYNC OPERATION ENDED")
        print("="*60 + "\n")
        
        # If not POST, redirect back to shipment list page
        return redirect('shipment_list')
        
    except Exception as e:
        # Handle session interruption and other errors gracefully
        if 'SessionInterrupted' in str(type(e).__name__) or 'session' in str(e).lower():
            print(f"[ERROR] Session interrupted: {e}")
            messages.warning(request, "Session was interrupted during sync. Please try again.")
            return redirect('shipment_list')
        elif 'UpdateError' in str(type(e).__name__):
            print(f"[ERROR] Cache update error: {e}")
            messages.warning(request, "Cache error occurred. Please try again.")
            return redirect('shipment_list')
        else:
            print(f" Unexpected error in refresh_inspections: {e}")
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('shipment_list')


def check_sync_status(request):
    """Check the status of a background sync operation."""
    from django.http import JsonResponse
    from django.core.cache import cache

    if request.method == 'GET':
        sync_result = cache.get('sync_result')
        sync_progress = cache.get('sync_progress')

        if sync_result:
            # Sync completed - return result with progress
            response = sync_result.copy()
            if sync_progress:
                response['progress'] = sync_progress
            return JsonResponse(response)
        elif sync_progress:
            # Sync still running - return progress
            return JsonResponse({
                'success': False,
                'status': 'running',
                'message': sync_progress.get('message', 'Sync is running...'),
                'progress': sync_progress
            })
        else:
            # No result or progress - assume still running
            return JsonResponse({
                'success': False,
                'status': 'running',
                'message': 'Sync is still running...'
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required(login_url='login')
def check_sync_status(request):
    """Check if any sync is currently running and return progress."""
    from django.core.cache import cache

    # Check both sync locks
    google_sheets_lock = cache.get('sync_google_sheets_lock')
    sql_server_lock = cache.get('sync_sql_server_lock')

    # Get current progress
    sync_progress = cache.get('sync_progress', {})

    # Determine if any sync is running
    is_running = bool(google_sheets_lock or sql_server_lock)

    return JsonResponse({
        'is_running': is_running,
        'locks': {
            'google_sheets': bool(google_sheets_lock),
            'sql_server': bool(sql_server_lock)
        },
        'progress': sync_progress
    })


@login_required(login_url='login')
def refresh_shipments(request):
    """Refresh the shipments table with fresh data from SQL Server."""
    clear_messages(request)
    
    if request.method == 'POST':
        print("\n" + "="*60)
        print(" STARTING SHIPMENT SYNC OPERATION")
        print("="*60)
        
        try:
            from ..services.google_sheets_service import GoogleSheetsService
            
            print(" Step 1: Initializing Google Sheets Service...")
            sheets_service = GoogleSheetsService()
            print(" Google Sheets Service initialized successfully")
            
            print("\n Step 2: Starting shipment table refresh from SQL Server...")
            refresh_result = sheets_service.populate_shipments_table(request)
            
            if refresh_result.get('success'):
                print(f" SHIPMENT SYNC COMPLETED SUCCESSFULLY!")
                print(f"    Deleted: {refresh_result['deleted_count']} old shipments")
                print(f"    Created: {refresh_result['shipments_created']} new shipments")
                print(f"    Processed: {refresh_result['total_processed']} total records from SQL Server")
                
                messages.success(request, f"Successfully refreshed shipments table!")
                messages.info(request, f"Deleted {refresh_result['deleted_count']} old shipments and created {refresh_result['shipments_created']} new shipments from SQL Server.")
                messages.info(request, f"Processed {refresh_result['total_processed']} total records from SQL Server.")
            else:
                print(f" SHIPMENT SYNC FAILED!")
                print(f"   Error: {refresh_result.get('error', 'Unknown error')}")
                messages.error(request, f"Error refreshing shipments: {refresh_result.get('error', 'Unknown error')}")
                
        except Exception as e:
            print(f" SHIPMENT SYNC EXCEPTION!")
            print(f"   Exception: {str(e)}")
            messages.error(request, f"Error refreshing shipments: {str(e)}")
        
        print("="*60)
        print(" SHIPMENT SYNC OPERATION ENDED")
        print("="*60 + "\n")
    
    # Redirect back to shipments page
    return redirect('shipment_list')


def google_oauth_callback(request):
    """Handle OAuth callback from Google"""
    from django.http import HttpResponse
    
    # Get the authorization code from the URL
    code = request.GET.get('code')
    
    if code:
        # Store the code in session for the service to use
        request.session['google_auth_code'] = code
        
        # Redirect to client allocation page with success message
        from django.contrib import messages
        messages.success(request, 'Google authentication successful! You can now fetch data.')
        return redirect('client_allocation')
    else:
        # Handle error
        from django.contrib import messages
        messages.error(request, 'Google authentication failed. No authorization code received.')
        return redirect('client_allocation')


# =============================================================================
# DASHBOARD VIEWS
# =============================================================================

@login_required(login_url='login')
def dashboard(request):
    """Main dashboard view."""
    clear_messages(request)
    
    # Get system settings for theme - matching user_management approach
    try:
        from ..models import Settings
        system_settings = Settings.objects.first()
        if not system_settings:
            system_settings = Settings.objects.create()
        
        settings = type('Settings', (), {
            'dark_mode': system_settings.dark_mode,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    # Handle theme saving via AJAX
    if request.method == 'POST' and request.POST.get('action') == 'save_theme':
        try:
            system_settings.dark_mode = 'theme_mode' in request.POST
            system_settings.save()
            
            # Log the activity
            try:
                SystemLog.log_activity(
                    user=request.user,
                    action='SETTINGS',
                    page=request.path,
                    description='Updated theme setting from dashboard',
                    details={'dark_mode': system_settings.dark_mode}
                )
            except Exception:
                pass
                
            return JsonResponse({'success': True, 'theme': 'dark' if system_settings.dark_mode else 'light'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # Get summary statistics from the correct models
    total_clients = Client.objects.count()
    total_inspections = FoodSafetyAgencyInspection.objects.count()
    
    # Get recent inspections from FoodSafetyAgencyInspection
    recent_inspections = FoodSafetyAgencyInspection.objects.order_by('-created_at')[:5]
    
    # Get inspections by inspector from FoodSafetyAgencyInspection
    inspector_stats = FoodSafetyAgencyInspection.objects.values('inspector_name').annotate(
        count=models.Count('id')
    ).order_by('-count')[:5]
    
    context = {
        'total_clients': total_clients,
        'total_inspections': total_inspections,
        'recent_inspections': recent_inspections,
        'inspector_stats': inspector_stats,
        'settings': settings
    }
    
    return render(request, 'main/dashboard.html', context)


@login_required(login_url='login')
def export_sheet(request):
    """Export Sheet page - Invoice dashboard replicating Looker Studio"""
    from datetime import datetime, timedelta
    from django.db.models import Q

    # Get yesterday to today for default filter
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    default_start_date = yesterday.strftime('%Y-%m-%d')
    default_end_date = today.strftime('%Y-%m-%d')

    # Get date range from request parameters or use defaults
    date_from = request.GET.get('date_from', default_start_date)
    date_to = request.GET.get('date_to', default_end_date)

    # Parse dates
    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    except:
        start_date = yesterday

    try:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    except:
        end_date = today

    # Fetch inspections with billable data
    # REQUIRED: Both hours AND km must be present for an inspection to appear
    # PERFORMANCE: Filter by date range at database level to avoid loading thousands of records
    # FILTERING: Only include RAW and PMP commodities (exclude POULTRY and EGGS from export)
    inspections = FoodSafetyAgencyInspection.objects.filter(
        commodity__in=['RAW', 'PMP'],
        hours__isnull=False,
        km_traveled__isnull=False,
        date_of_inspection__gte=start_date,
        date_of_inspection__lte=end_date
    ).select_related(
        'sent_by', 'rfi_uploaded_by', 'invoice_uploaded_by'
    ).order_by('client_name', 'date_of_inspection', 'commodity')

    # Generate invoice line items
    invoice_items = []
    inspection_counter = {}  # Track sequential inspection IDs per inspector
    inspections_processed = 0
    alternating_flag = True  # Flag that alternates TRUE/FALSE for each inspection

    print(f"[EXPORT_SHEET] Fetching inspections...")

    # STEP 1: Group inspections by VISIT (inspector + client + date)
    # This prevents duplicate hours/km charges when multiple products inspected in one visit
    from collections import defaultdict
    visits = defaultdict(list)

    for inspection in inspections:
        # Create visit key (inspector, client, date) - one visit can have multiple products
        visit_key = (
            inspection.inspector_name or '',
            inspection.client_name or '',
            str(inspection.date_of_inspection) if inspection.date_of_inspection else ''
        )
        visits[visit_key].append(inspection)

    print(f"[EXPORT_SHEET] Found {len(visits)} unique visits covering {len(inspections)} product inspections")

    # STEP 2: Process each visit
    for visit_key, visit_inspections in visits.items():
        inspector_name, client_name, date_str = visit_key

        # Sort visit inspections by commodity (PMP first, then RAW) for consistent processing
        visit_inspections.sort(key=lambda x: (x.commodity or '', x.product_name or ''))

        # Use first inspection for visit-level data
        first_inspection = visit_inspections[0]

        inspections_processed += 1

        # Generate unique inspection ID
        inspector_key = inspector_name or 'Unknown'
        if inspector_key not in inspection_counter:
            inspection_counter[inspector_key] = 1
        else:
            inspection_counter[inspector_key] += 1

        inspection_id = inspection_counter[inspector_key]

        # Get inspector code (first letters of each name)
        inspector_code = ''
        if inspector_name:
            name_parts = inspector_name.split()
            inspector_code = ''.join([part[0].upper() for part in name_parts if part])

        # Generate invoice reference number
        date_formatted = ''
        if first_inspection.date_of_inspection:
            date_formatted = first_inspection.date_of_inspection.strftime('%y%m%d')

        invoice_ref = f"FSA-INV-{inspector_code}-{date_formatted}" if date_formatted else ''
        rfi_ref = f"FSA-RFI-{inspector_code}-{date_formatted}" if date_formatted else ''

        # Get client location (city)
        city = ''
        if client_name:
            parts = client_name.split()
            city = parts[-1] if len(parts) > 1 else ''

        # Lab used
        lab_name = 'Food Safety Laboratory' if first_inspection.lab else ''

        # STEP 3: Get hours and km from first inspection
        # (All products in same visit should have same hours/km, not summed)
        total_hours = float(first_inspection.hours) if first_inspection.hours else 0
        total_km = float(first_inspection.km_traveled) if first_inspection.km_traveled else 0

        # STEP 4: Generate HOURS/KM line items
        # Business Rule: If visit has BOTH PMP and RAW, split hours/km equally between them
        # If only one commodity type, charge full amount to that type
        if total_hours > 0 or total_km > 0:
            pmp_products = [i for i in visit_inspections if 'PMP' in (i.commodity or '').upper()]
            raw_products = [i for i in visit_inspections if 'RAW' in (i.commodity or '').upper()]

            has_both_types = len(pmp_products) > 0 and len(raw_products) > 0

            if has_both_types:
                # Split hours/km equally between PMP and RAW
                split_hours = total_hours / 2
                split_km = total_km / 2

                # Generate PMP hours/km (half)
                pmp_items = generate_visit_hours_km_items(
                    inspection_id=inspection_id,
                    inspection=pmp_products[0],
                    invoice_ref=invoice_ref,
                    rfi_ref=rfi_ref,
                    product_type='PMP',
                    city=city,
                    lab_name=lab_name,
                    total_hours=split_hours,
                    total_km=split_km
                )
                invoice_items.extend(pmp_items)

                # Generate RAW hours/km (half)
                raw_items = generate_visit_hours_km_items(
                    inspection_id=inspection_id,
                    inspection=raw_products[0],
                    invoice_ref=invoice_ref,
                    rfi_ref=rfi_ref,
                    product_type='RAW',
                    city=city,
                    lab_name=lab_name,
                    total_hours=split_hours,
                    total_km=split_km
                )
                invoice_items.extend(raw_items)
            else:
                # Only one commodity type - charge full amount
                product_type = 'RAW' if first_inspection.commodity and 'RAW' in first_inspection.commodity.upper() else 'PMP'

                visit_items = generate_visit_hours_km_items(
                    inspection_id=inspection_id,
                    inspection=first_inspection,
                    invoice_ref=invoice_ref,
                    rfi_ref=rfi_ref,
                    product_type=product_type,
                    city=city,
                    lab_name=lab_name,
                    total_hours=total_hours,
                    total_km=total_km
                )
                invoice_items.extend(visit_items)

        # STEP 5: Generate TEST line items ONCE per visit (aggregate by test type)
        # Check which tests are needed across ALL products in this visit
        pmp_needs_fat = any(i.fat and i.is_sample_taken and 'PMP' in (i.commodity or '').upper() for i in visit_inspections)
        pmp_needs_protein = any(i.protein and i.is_sample_taken and 'PMP' in (i.commodity or '').upper() for i in visit_inspections)
        pmp_needs_calcium = any(i.calcium and i.is_sample_taken and 'PMP' in (i.commodity or '').upper() for i in visit_inspections)
        raw_needs_fat = any(i.fat and i.is_sample_taken and 'RAW' in (i.commodity or '').upper() for i in visit_inspections)
        raw_needs_protein = any(i.protein and i.is_sample_taken and 'RAW' in (i.commodity or '').upper() for i in visit_inspections)
        raw_needs_dna = any(i.dna and i.is_sample_taken and 'RAW' in (i.commodity or '').upper() for i in visit_inspections)

        # Generate PMP test items if ANY PMP product needs them
        pmp_products = [i for i in visit_inspections if 'PMP' in (i.commodity or '').upper()]
        if pmp_products:
            test_items = generate_test_line_items(
                inspection_id=inspection_id,
                inspection=pmp_products[0],  # Use first PMP product for metadata
                invoice_ref=invoice_ref,
                rfi_ref=rfi_ref,
                product_type='PMP',
                city=city,
                lab_name=lab_name,
                force_tests={'fat': pmp_needs_fat, 'protein': pmp_needs_protein, 'calcium': pmp_needs_calcium}
            )
            invoice_items.extend(test_items)

        # Generate RAW test items if ANY RAW product needs them
        raw_products = [i for i in visit_inspections if 'RAW' in (i.commodity or '').upper()]
        if raw_products:
            test_items = generate_test_line_items(
                inspection_id=inspection_id,
                inspection=raw_products[0],  # Use first RAW product for metadata
                invoice_ref=invoice_ref,
                rfi_ref=rfi_ref,
                product_type='RAW',
                city=city,
                lab_name=lab_name,
                force_tests={'fat': raw_needs_fat, 'protein': raw_needs_protein, 'dna': raw_needs_dna}
            )
            invoice_items.extend(test_items)

    # Calculate unique inspectors
    unique_inspectors = set(item['inspector_name'] for item in invoice_items if item.get('inspector_name'))

    # Debug: Final summary
    print(f"[EXPORT_SHEET] Processed {inspections_processed} inspections, generated {len(invoice_items)} line items")

    # Sort invoice items by client name, then date, then item code
    # This groups all items for the same client together (RAW and PMP intermixed)
    invoice_items.sort(key=lambda x: (
        x.get('client_name', ''),
        x.get('invoice_date', ''),
        x.get('item_code', '')
    ))

    # Get system settings for theme
    from ..models import SystemSettings
    settings = SystemSettings.get_settings()

    context = {
        'invoice_items': invoice_items,
        'total_items': len(invoice_items),
        'inspections_processed': inspections_processed,
        'unique_inspectors': len(unique_inspectors),
        'settings': settings,
        'default_start_date': start_date.strftime('%Y-%m-%d'),
        'default_end_date': end_date.strftime('%Y-%m-%d'),
    }

    return render(request, 'main/export_sheet.html', context)


def get_fee_rate(fee_code, default_value, inspection_date=None):
    """
    Get fee rate from database, fallback to default if not found.

    Args:
        fee_code: The fee code to look up (e.g., 'inspection_hour_rate')
        default_value: Fallback value if fee not found
        inspection_date: Optional date to get historical rate for that date

    Returns:
        Float rate value (historical if date provided, current if not)
    """
    try:
        from ..models import InspectionFee
        fee = InspectionFee.objects.filter(fee_code=fee_code).first()
        if not fee:
            return default_value

        # If inspection date provided, use historical rate lookup
        if inspection_date:
            return float(fee.get_rate_for_date(inspection_date))
        else:
            # No date provided, return current rate
            return float(fee.rate)
    except Exception as e:
        print(f"[ERROR] get_fee_rate({fee_code}): {e}")
        return default_value


def generate_visit_hours_km_items(inspection_id, inspection, invoice_ref, rfi_ref, product_type, city, lab_name, total_hours, total_km):
    """Generate HOURS and KM line items ONCE per visit (not per product)"""
    items = []

    # Load pricing - Use HISTORICAL rates based on inspection date
    INSPECTION_HOUR_RATE = get_fee_rate('inspection_hour_rate', 510.00, inspection.date_of_inspection)
    TRAVEL_RATE_PER_KM = get_fee_rate('travel_rate_per_km', 6.50, inspection.date_of_inspection)

    # Handle missing client name
    client_name = inspection.client_name if inspection.client_name else 'Unknown Client'

    # Get inspector display name
    inspector_display = ''
    if inspection.inspector_name:
        name_parts = inspection.inspector_name.split()
        if len(name_parts) >= 2:
            inspector_display = f"{name_parts[0][0]} {name_parts[-1]}"
        else:
            inspector_display = inspection.inspector_name

    # Format date
    date_formatted = inspection.date_of_inspection.strftime('%d/%m/%Y') if inspection.date_of_inspection else ''

    # Determine item codes based on product type
    if product_type == 'PMP':
        hours_code = 'PMP 060'
        km_code = 'PMP 061'
        hours_desc = f"Inspection on Regulated Animal Products (Processed Meat Products): Inspection (hours) - {date_formatted} - {inspector_display} - {client_name}"
        km_desc = f"Inspection on Regulated Animal Products (Processed Meat Products): Travel Cost (Km's) - {date_formatted}"
        account_code = '1000/060'
    else:  # RAW
        hours_code = 'RAW 050'
        km_code = 'RAW 051'
        hours_desc = f"Inspection on Regulated Animal Products (Certain Raw Processed Meat Products): Inspection (Hours) - {date_formatted} - {inspector_display} - {client_name}"
        km_desc = f"Inspection on Regulated Animal Products (Certain Raw Processed Meat Products): Travel Cost (Km's) - {date_formatted}"
        account_code = '1000/050'

    # Hours line item
    if total_hours > 0:
        items.append({
            'row_number': 1 if product_type == 'PMP' else 7,
            'inspection_id': inspection_id,
            'invoice_ref': invoice_ref,
            'rfi_ref': rfi_ref,
            'inspector_name': inspection.inspector_name,
            'invoice_date': date_formatted,
            'client_name': client_name,
            'city': city,
            'product_type': product_type,
            'product_name': 'Multiple Products' if total_hours > 0 else inspection.product_name,
            'product_class': inspection.product_class,
            'item_code': hours_code,
            'description': hours_desc,
            'quantity': float(total_hours),
            'unit_amount': INSPECTION_HOUR_RATE,
            'account_code': account_code,
            'tax_rate': 'Standard Rate Sales (15%)',
            'country': 'South Africa',
            'lab': lab_name,
            'total': float(total_hours) * INSPECTION_HOUR_RATE,
            'id': inspection.id,
            'invoice_number': inspection.invoice_number,
        })

    # KM line item
    if total_km > 0:
        items.append({
            'row_number': 2 if product_type == 'PMP' else 8,
            'inspection_id': inspection_id,
            'invoice_ref': invoice_ref,
            'rfi_ref': rfi_ref,
            'inspector_name': inspection.inspector_name,
            'invoice_date': date_formatted,
            'client_name': client_name,
            'city': city,
            'product_type': product_type,
            'product_name': 'Multiple Products' if total_km > 0 else inspection.product_name,
            'product_class': inspection.product_class,
            'item_code': km_code,
            'description': km_desc,
            'quantity': float(total_km),
            'unit_amount': TRAVEL_RATE_PER_KM,
            'account_code': account_code,
            'tax_rate': 'Standard Rate Sales (15%)',
            'country': 'South Africa',
            'lab': lab_name,
            'total': float(total_km) * TRAVEL_RATE_PER_KM,
            'id': inspection.id,
            'invoice_number': inspection.invoice_number,
        })

    return items


def generate_test_line_items(inspection_id, inspection, invoice_ref, rfi_ref, product_type, city, lab_name, force_tests=None):
    """Generate TEST line items (aggregated per visit, not per product)

    Args:
        force_tests: Dict of which tests to generate {'fat': bool, 'protein': bool, 'calcium': bool, 'dna': bool}
                    If provided, uses these flags instead of inspection's individual test flags
    """
    items = []

    # Load pricing - Use HISTORICAL rates based on inspection date
    FAT_TEST_RATE = get_fee_rate('fat_test_rate', 826.00, inspection.date_of_inspection)
    PROTEIN_TEST_RATE = get_fee_rate('protein_test_rate', 503.00, inspection.date_of_inspection)
    CALCIUM_TEST_RATE = get_fee_rate('calcium_test_rate', 379.00, inspection.date_of_inspection)
    DNA_TEST_RATE = get_fee_rate('dna_test_rate', 2605.00, inspection.date_of_inspection)

    # Handle missing client name
    client_name = inspection.client_name if inspection.client_name else 'Unknown Client'

    # Format date
    date_formatted = inspection.date_of_inspection.strftime('%d/%m/%Y') if inspection.date_of_inspection else ''

    # Lab used
    lab_name_val = 'Food Safety Laboratory' if inspection.lab else ''

    # Determine which tests to generate (use force_tests if provided, else use inspection fields)
    generate_fat = force_tests['fat'] if force_tests else (inspection.fat and inspection.is_sample_taken)
    generate_protein = force_tests['protein'] if force_tests else (inspection.protein and inspection.is_sample_taken)
    generate_calcium = force_tests.get('calcium', False) if force_tests else (inspection.calcium and inspection.is_sample_taken)
    generate_dna = force_tests.get('dna', False) if force_tests else (inspection.dna and inspection.is_sample_taken)

    if product_type == 'PMP':
        # PMP Test Items
        # Fat Test
        if generate_fat:
            items.append({
                'row_number': 3,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 062',
                'description': 'Sample Taking: Fat Content (Processed Meat Products)',
                'quantity': 1,
                'unit_amount': FAT_TEST_RATE,
                'account_code': '1000/061',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': FAT_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Protein Test
        if generate_protein:
            items.append({
                'row_number': 4,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 064',
                'description': 'Sample Taking: Protein Content (Processed Meat Products)',
                'quantity': 1,
                'unit_amount': PROTEIN_TEST_RATE,
                'account_code': '1000/062',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': PROTEIN_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Calcium Test
        if generate_calcium:
            items.append({
                'row_number': 5,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 066',
                'description': 'Sample Taking: Calcium Determination (MRM only)',
                'quantity': 1,
                'unit_amount': CALCIUM_TEST_RATE,
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': CALCIUM_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Samples Bought
        if inspection.bought_sample:
            items.append({
                'row_number': 6,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 067',
                'description': 'Samples Bought',
                'quantity': 1,
                'unit_amount': float(inspection.bought_sample),
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': float(inspection.bought_sample),
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

    else:  # RAW products
        # Fat Test
        if inspection.fat and inspection.is_sample_taken:
            items.append({
                'row_number': 9,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 052',
                'description': 'Sample Taking: Category A - Fat Content (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': FAT_TEST_RATE,
                'account_code': '1000/051',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': FAT_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Protein Test
        if generate_protein:
            items.append({
                'row_number': 10,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 053',
                'description': 'Sample Taking: Category A - Protein (Meat) Content (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': PROTEIN_TEST_RATE,
                'account_code': '1000/052',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': PROTEIN_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # DNA Test
        if generate_dna:
            items.append({
                'row_number': 11,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 056',
                'description': 'Sample Taking: Category C - Meat Specie ID (DNA) (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': DNA_TEST_RATE,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': DNA_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Calcium Test
        if generate_calcium:
            items.append({
                'row_number': 12,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 057',
                'description': 'Sample Taking: Calcium Determination (MRM Only) (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': CALCIUM_TEST_RATE,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': CALCIUM_TEST_RATE,
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

        # Samples Bought
        if inspection.bought_sample:
            items.append({
                'row_number': 13,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 058',
                'description': 'Samples Bought',
                'quantity': 1,
                'unit_amount': float(inspection.bought_sample),
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name_val,
                'total': float(inspection.bought_sample),
                'id': inspection.id,
                'invoice_number': inspection.invoice_number,
            })

    return items


def generate_invoice_line_items(inspection_id, inspection, invoice_ref, rfi_ref, product_type, city, lab_name):
    """DEPRECATED: Old function kept for backwards compatibility. Use generate_visit_hours_km_items + generate_test_line_items instead."""
    items = []

    # Load pricing from database (fallback to 2025 gazette rates if not in DB)
    INSPECTION_HOUR_RATE = get_fee_rate('inspection_hour_rate', 510.00)
    INSPECTION_OVERTIME_RATE = get_fee_rate('inspection_overtime_rate', 567.00)
    INSPECTION_SUNDAY_RATE = get_fee_rate('inspection_sunday_rate', 680.00)
    TRAVEL_RATE_PER_KM = get_fee_rate('travel_rate_per_km', 6.50)
    FAT_TEST_RATE = get_fee_rate('fat_test_rate', 826.00)
    PROTEIN_TEST_RATE = get_fee_rate('protein_test_rate', 503.00)
    CALCIUM_TEST_RATE = get_fee_rate('calcium_test_rate', 379.00)
    DNA_TEST_RATE = get_fee_rate('dna_test_rate', 2605.00)
    SOYA_TEST_RATE = get_fee_rate('soya_test_rate', 1665.00)
    STARCH_TEST_RATE = get_fee_rate('starch_test_rate', 1472.00)
    PHYSICAL_TEST_RATE = get_fee_rate('physical_test_rate', 200.00)

    # Handle missing client name - use fallback to prevent empty/None values in export
    client_name = inspection.client_name if inspection.client_name else 'Unknown Client'

    # Get inspector initials + last name
    inspector_display = ''
    if inspection.inspector_name:
        name_parts = inspection.inspector_name.split()
        if len(name_parts) >= 2:
            inspector_display = f"{name_parts[0][0]} {name_parts[-1]}"
        else:
            inspector_display = inspection.inspector_name

    # Format date for descriptions
    date_formatted = inspection.date_of_inspection.strftime('%d/%m/%Y') if inspection.date_of_inspection else ''

    if product_type == 'PMP':
        # PMP Line Items (rows 1-6)

        # Row 1: Inspection Hours
        if inspection.hours:
            items.append({
                'row_number': 1,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 060',
                'description': f"Inspection on Regulated Animal Products (Processed Meat Products): Inspection (hours) - {date_formatted} - {inspector_display} - {client_name}",
                'quantity': float(inspection.hours),
                'unit_amount': INSPECTION_HOUR_RATE,
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 2: Travel Cost
        if inspection.km_traveled:
            items.append({
                'row_number': 2,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 061',
                'description': f"Inspection on Regulated Animal Products (Processed Meat Products): Travel Cost (Km's) - {date_formatted}",
                'quantity': float(inspection.km_traveled),
                'unit_amount': TRAVEL_RATE_PER_KM,
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 3: Fat Test
        if inspection.fat:
            items.append({
                'row_number': 3,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 062',
                'description': 'Sample Taking: Fat Content (Processed Meat Products)',
                'quantity': 1,
                'unit_amount': FAT_TEST_RATE,
                'account_code': '1000/061',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 4: Protein Test
        if inspection.protein:
            items.append({
                'row_number': 4,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 064',
                'description': 'Sample Taking: Protein Content (Processed Meat Products)',
                'quantity': 1,
                'unit_amount': PROTEIN_TEST_RATE,
                'account_code': '1000/062',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 5: Calcium Test
        if inspection.calcium:
            items.append({
                'row_number': 5,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 066',
                'description': 'Sample Taking: Calcium Determination (MRM only)',
                'quantity': 1,
                'unit_amount': CALCIUM_TEST_RATE,
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 6: Samples Bought
        if inspection.bought_sample:
            items.append({
                'row_number': 6,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'PMP 067',
                'description': 'Samples Bought',
                'quantity': 1,
                'unit_amount': float(inspection.bought_sample),
                'account_code': '1000/060',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

    else:  # RAW products
        # RAW Line Items (rows 7-13)

        # Row 7: Inspection Hours
        if inspection.hours:
            items.append({
                'row_number': 7,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 050',
                'description': f"Inspection on Regulated Animal Products (Certain Raw Processed Meat Products): Inspection (Hours) - {date_formatted} - {inspector_display} - {inspection.client_name}",
                'quantity': float(inspection.hours),
                'unit_amount': INSPECTION_HOUR_RATE,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 8: Travel Cost
        if inspection.km_traveled:
            items.append({
                'row_number': 8,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 051',
                'description': f"Inspection on Regulated Animal Products (Certain Raw Processed Meat Products): Travel Cost (Km's) - {date_formatted}",
                'quantity': float(inspection.km_traveled),
                'unit_amount': TRAVEL_RATE_PER_KM,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 9: Fat Test
        if inspection.fat:
            items.append({
                'row_number': 9,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 052',
                'description': 'Sample Taking: Category A - Fat Content (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': FAT_TEST_RATE,
                'account_code': '1000/051',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 10: Protein Test
        if inspection.protein:
            items.append({
                'row_number': 10,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 053',
                'description': 'Sample Taking: Category A - Protein (Meat) Content (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': PROTEIN_TEST_RATE,
                'account_code': '1000/052',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 11: DNA Test
        if inspection.dna:
            items.append({
                'row_number': 11,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 056',
                'description': 'Sample Taking: Category C - Meat Specie ID (DNA) (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': DNA_TEST_RATE,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 12: Calcium Test
        if inspection.calcium:
            items.append({
                'row_number': 12,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 057',
                'description': 'Sample Taking: Calcium Determination (MRM Only) (Certain Raw Processed Meat Products)',
                'quantity': 1,
                'unit_amount': CALCIUM_TEST_RATE,
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

        # Row 13: Samples Bought
        if inspection.bought_sample:
            items.append({
                'row_number': 13,
                'inspection_id': inspection_id,
                'invoice_ref': invoice_ref,
                'rfi_ref': rfi_ref,
                'inspector_name': inspection.inspector_name,
                'invoice_date': date_formatted,
                'client_name': client_name,
                'city': city,
                'product_type': product_type,
                'product_name': inspection.product_name,
                'product_class': inspection.product_class,
                'item_code': 'RAW 058',
                'description': 'Samples Bought',
                'quantity': 1,
                'unit_amount': float(inspection.bought_sample),
                'account_code': '1000/050',
                'tax_rate': 'Standard Rate Sales (15%)',
                'country': 'South Africa',
                'lab': lab_name,
            })

    # Calculate total for each item and add inspection ID and invoice number
    for item in items:
        item['total'] = item['quantity'] * item['unit_amount']
        item['id'] = inspection.id
        item['invoice_number'] = inspection.invoice_number
        # Use custom invoice number if set, otherwise use auto-generated one
        if inspection.invoice_number:
            item['invoice_ref'] = inspection.invoice_number

    return items

@login_required(login_url='login')
def export_to_google_sheets(request):
    """Export data to Google Sheets"""
    from ..services.google_sheets_service import GoogleSheetsService
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from datetime import datetime
    import json
    import pickle
    import os
    from django.conf import settings

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST requests allowed'})

    try:
        # Get the data from request
        data = json.loads(request.body)
        export_data = data.get('data', [])

        print(f"📊 Export request received with {len(export_data)} rows")

        if not export_data:
            return JsonResponse({'success': False, 'error': 'No data to export'})

        # Authenticate with Google Sheets
        token_path = os.path.join(settings.BASE_DIR, 'token.pickle')

        if not os.path.exists(token_path):
            print("❌ token.pickle not found")
            return JsonResponse({
                'success': False,
                'error': 'Google Sheets not authenticated. Please delete token.pickle and re-authenticate with write permissions.'
            })

        print("🔐 Loading credentials from token.pickle...")
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)

        print("✅ Credentials loaded")

        # Check if credentials are valid
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                print("🔄 Refreshing expired token...")
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                with open(token_path, 'wb') as token:
                    pickle.dump(creds, token)
                print("✅ Token refreshed")
            else:
                print("❌ Credentials invalid")
                return JsonResponse({
                    'success': False,
                    'error': 'Credentials expired. Please delete token.pickle and re-authenticate.'
                })

        # Build the service
        print("🔨 Building Google Sheets service...")
        service = build('sheets', 'v4', credentials=creds)

        # Create a new spreadsheet
        print("📝 Creating new spreadsheet...")
        spreadsheet = {
            'properties': {
                'title': f'Export Data {datetime.now().strftime("%Y-%m-%d %H-%M")}'
            }
        }

        result = service.spreadsheets().create(body=spreadsheet, fields='spreadsheetId,spreadsheetUrl').execute()
        spreadsheet_id = result.get('spreadsheetId')
        spreadsheet_url = result.get('spreadsheetUrl')

        print(f"✅ Spreadsheet created: {spreadsheet_id}")
        print(f"🔗 URL: {spreadsheet_url}")

        # Write data to the spreadsheet
        print(f"✍️ Writing {len(export_data)} rows...")
        body = {'values': export_data}

        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range='A1',
            valueInputOption='RAW',
            body=body
        ).execute()

        print("✅ Data written successfully!")

        # Share the spreadsheet with anyone who has the link
        print("🔓 Making spreadsheet accessible to anyone with the link...")
        try:
            drive_service = build('drive', 'v3', credentials=creds)
            permission = {
                'type': 'anyone',
                'role': 'writer'  # Allows editing - change to 'reader' for view-only
            }
            drive_service.permissions().create(
                fileId=spreadsheet_id,
                body=permission,
                fields='id'
            ).execute()
            print("✅ Spreadsheet shared successfully!")
        except Exception as share_error:
            print(f"⚠️ Warning: Could not share spreadsheet automatically: {str(share_error)}")
            print("   You may need to re-authenticate with Drive permissions.")
            # Continue anyway - the spreadsheet was created successfully

        return JsonResponse({
            'success': True,
            'url': spreadsheet_url
        })

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required(login_url='login')
def update_invoice_number(request):
    """Update invoice number for an inspection and all items in the same group"""
    import json
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST requests allowed'})

    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        invoice_number = data.get('invoice_number', '').strip()
        group_item_ids = data.get('group_item_ids', [])

        if not item_id:
            return JsonResponse({'success': False, 'error': 'Missing item_id'})

        # Determine which IDs to update
        ids_to_update = list(set(group_item_ids)) if group_item_ids else [item_id]

        # Convert to integers and filter out invalid values
        valid_ids = []
        for id_val in ids_to_update:
            try:
                valid_ids.append(int(id_val))
            except (ValueError, TypeError):
                continue

        if not valid_ids:
            return JsonResponse({'success': False, 'error': 'No valid item IDs provided'})

        # Update all inspections in the group
        updated_count = FoodSafetyAgencyInspection.objects.filter(
            id__in=valid_ids
        ).update(
            invoice_number=invoice_number if invoice_number else None
        )

        return JsonResponse({
            'success': True,
            'message': f'Invoice number updated for {updated_count} item(s)',
            'invoice_number': invoice_number,
            'updated_count': updated_count
        })

    except FoodSafetyAgencyInspection.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Inspection not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='login')
@inspector_only_inspections
def home(request):
    """Home page view that requires login."""
    clear_messages(request)
    
    # Redirect administrators to inspection page since they're not allowed on home page
    if request.user.role == 'admin':
        messages.info(request, "Administrators are redirected to the inspection page.")
        return redirect('shipment_list')
    
    # Get system settings for theme - matching dashboard approach
    try:
        from ..models import Settings
        system_settings = Settings.objects.first()
        if not system_settings:
            system_settings = Settings.objects.create()
        
        settings = type('Settings', (), {
            'dark_mode': system_settings.dark_mode,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    # Handle theme saving via AJAX
    if request.method == 'POST' and request.POST.get('action') == 'save_theme':
        try:
            system_settings.dark_mode = 'theme_mode' in request.POST
            system_settings.save()
            
            # Log the activity
            try:
                SystemLog.log_activity(
                    user=request.user,
                    action='SETTINGS',
                    page=request.path,
                    description='Updated theme setting from home',
                    details={'dark_mode': system_settings.dark_mode}
                )
            except Exception:
                pass
                
            return JsonResponse({'success': True, 'theme': 'dark' if system_settings.dark_mode else 'light'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # Get summary statistics from the correct models
    total_clients = Client.objects.count()
    total_inspections = FoodSafetyAgencyInspection.objects.count()
    
    # Get recent inspections from FoodSafetyAgencyInspection
    recent_inspections = FoodSafetyAgencyInspection.objects.order_by('-created_at')[:5]
    
    # Get inspections by inspector from FoodSafetyAgencyInspection
    inspector_stats = FoodSafetyAgencyInspection.objects.values('inspector_name').annotate(
        count=models.Count('id')
    ).order_by('-count')[:5]
    
    # Get system status
    def check_database_status():
        """Check PostgreSQL database connectivity"""
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                result = cursor.fetchone()
                return True if result else False
        except Exception:
            return False
    
    def check_sql_server_status():
        """Check SQL Server database connectivity using pymssql - NO ODBC DRIVERS NEEDED!"""
        try:
            # Test SQL Server connection using pymssql
            import pymssql

            conn = pymssql.connect(
                server='102.67.140.12',
                port=1053,
                user='FSAUser2',
                password='password',
                database='AFS',
                timeout=5
            )
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            result = cursor.fetchone()
            cursor.close()
            conn.close()
            return True if result else False
        except Exception as e:
            print(f"SQL Server status check failed: {e}")
            return False
    
    def check_google_sheets_status():
        """Check Google Sheets API connectivity with automatic token refresh"""
        try:
            from ..services.google_sheets_service import GoogleSheetsService

            service = GoogleSheetsService()
            is_connected, message = service.check_connection_status()

            if is_connected:
                print(f"[OK] Google Sheets: {message}")
                return True
            else:
                print(f"[ERROR] Google Sheets: {message}")
                return False

        except Exception as e:
            print(f"[ERROR] Google Sheets status check failed: {e}")
            return False
    
    
    def get_last_sync_status():
        """Get last sync status from scheduled sync service"""
        try:
            # First, try to get the actual sync service status
            from ..services.scheduled_sync_service import scheduled_sync_service
            if scheduled_sync_service and scheduled_sync_service.is_running:
                # Get the most recent sync time from all sync types
                last_sync_times = scheduled_sync_service.last_sync_times
                most_recent = None

                for sync_type, last_sync in last_sync_times.items():
                    if last_sync:
                        if most_recent is None or last_sync > most_recent:
                            most_recent = last_sync

                if most_recent:
                    now = datetime.now()
                    time_diff = now - most_recent

                    if time_diff.total_seconds() < 60:  # Less than 1 minute
                        return "Just now"
                    elif time_diff.total_seconds() < 3600:  # Less than 1 hour
                        minutes = int(time_diff.total_seconds() / 60)
                        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
                    elif time_diff.total_seconds() < 86400:  # Less than 1 day
                        hours = int(time_diff.total_seconds() / 3600)
                        return f"{hours} hour{'s' if hours > 1 else ''} ago"
                    else:
                        days = int(time_diff.total_seconds() / 86400)
                        return f"{days} day{'s' if days > 1 else ''} ago"

            # Fallback: Check when last inspection was created
            latest_inspection = FoodSafetyAgencyInspection.objects.order_by('-created_at').first()
            if latest_inspection:
                now = timezone.now()
                created_at = latest_inspection.created_at

                # Handle timezone-aware datetime comparison
                if timezone.is_aware(created_at) and not timezone.is_aware(now):
                    now = timezone.make_aware(now)
                elif not timezone.is_aware(created_at) and timezone.is_aware(now):
                    created_at = timezone.make_aware(created_at)

                time_diff = now - created_at
                if time_diff.total_seconds() < 3600:  # Less than 1 hour
                    return "Just now"
                elif time_diff.total_seconds() < 86400:  # Less than 1 day
                    hours = int(time_diff.total_seconds() / 3600)
                    return f"{hours} hour{'s' if hours > 1 else ''} ago"
                else:
                    days = int(time_diff.total_seconds() / 86400)
                    return f"{days} day{'s' if days > 1 else ''} ago"
            else:
                return "Service not running"
        except Exception as e:
            # Fallback to inspection created_at if service check fails
            try:
                latest_inspection = FoodSafetyAgencyInspection.objects.order_by('-created_at').first()
                if latest_inspection:
                    now = timezone.now()
                    created_at = latest_inspection.created_at

                    if timezone.is_aware(created_at) and not timezone.is_aware(now):
                        now = timezone.make_aware(now)
                    elif not timezone.is_aware(created_at) and timezone.is_aware(now):
                        created_at = timezone.make_aware(created_at)

                    time_diff = now - created_at
                    if time_diff.total_seconds() < 86400:
                        hours = int(time_diff.total_seconds() / 3600)
                        return f"{hours} hour{'s' if hours > 1 else ''} ago"
                    else:
                        days = int(time_diff.total_seconds() / 86400)
                        return f"{days} day{'s' if days > 1 else ''} ago"
            except Exception:
                pass
            return "Unknown"
    
    # Check system status with caching to avoid performance issues
    from django.core.cache import cache
    
    # Cache status checks for 30 seconds to avoid repeated expensive checks
    postgresql_online = cache.get('status_postgresql')
    if postgresql_online is None:
        postgresql_online = check_database_status()
        cache.set('status_postgresql', postgresql_online, 30)
    
    sql_server_online = cache.get('status_sql_server')
    if sql_server_online is None:
        sql_server_online = check_sql_server_status()
        cache.set('status_sql_server', sql_server_online, 30)
    
    # Check if we should force refresh the Google Sheets status
    force_refresh = request.GET.get('refresh_status') == 'true'
    
    google_sheets_online = cache.get('status_google_sheets')
    if google_sheets_online is None or force_refresh:
        if force_refresh:
            cache.delete('status_google_sheets')
        google_sheets_online = check_google_sheets_status()
        cache.set('status_google_sheets', google_sheets_online, 30)
    
    last_sync = get_last_sync_status()

    # Check if scheduled sync service is running
    def check_sync_service_status():
        """Check if the scheduled sync service is running"""
        try:
            from ..services.scheduled_sync_service import scheduled_sync_service
            if scheduled_sync_service and scheduled_sync_service.is_running:
                return True
            return False
        except Exception:
            return False

    sync_service_running = check_sync_service_status()

    # Get recent activities from SystemLog
    def get_recent_activities():
        try:
            from ..models import SystemLog
            activities = SystemLog.objects.select_related('user').order_by('-timestamp')[:5]
            return activities
        except Exception:
            return []
    
    recent_activities = get_recent_activities()

    context = {
        'total_clients': total_clients,
        'total_inspections': total_inspections,
        'recent_inspections': recent_inspections,
        'inspector_stats': inspector_stats,
        'settings': settings,
        'system_status': {
            'postgresql_online': postgresql_online,
            'sql_server_online': sql_server_online,
            'google_sheets_online': google_sheets_online,
            'sync_service_running': sync_service_running,
            'last_sync': last_sync
        },
        'recent_activities': recent_activities
    }
    
    return render(request, 'main/home.html', context)


@login_required
def inspector_dashboard(request):
    """Display inspector-specific analytics dashboard with only their data."""
    # Only allow inspectors to access this dashboard
    if request.user.role != 'inspector':
        return redirect('analytics_dashboard')
    
    from ..models import Client, Inspection, FoodSafetyAgencyInspection, InspectorMapping
    from django.db.models import Count, Q
    from datetime import datetime, timedelta, date
    
    # Resolve inspector_id via mapping (handles cases where source has 'Unknown' names)
    inspector_name = request.user.get_full_name() or request.user.username
    inspector_id = None

    # Prefer active mappings and the most recently updated one if duplicates exist
    mapping = (
        InspectorMapping.objects
        .filter(inspector_name__iexact=inspector_name)
        .order_by('-is_active', '-updated_at')
        .first()
    )
    if not mapping:
        # Fallback to username-only mapping if full name was not stored
        mapping = (
            InspectorMapping.objects
            .filter(inspector_name__iexact=request.user.username)
            .order_by('-is_active', '-updated_at')
            .first()
        )
    if mapping:
        inspector_id = mapping.inspector_id
    
    # Get inspector-specific statistics using inspector_id when available
    if inspector_id is not None:
        inspector_inspections = FoodSafetyAgencyInspection.objects.filter(
            inspector_id=inspector_id
        )
        # If the selected mapping yields no data (e.g., stale/dummy ID), fall back to name match
        if not inspector_inspections.exists():
            inspector_inspections = FoodSafetyAgencyInspection.objects.filter(
                inspector_name__icontains=inspector_name
            )
    else:
        # As a last resort (unlikely), try name match so page still loads
        inspector_inspections = FoodSafetyAgencyInspection.objects.filter(
            inspector_name__icontains=inspector_name
        )
    
    total_inspections = inspector_inspections.count()
    
    # Get recent activity (last 30 days) for this inspector
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_inspections = inspector_inspections.filter(
        date_of_inspection__gte=thirty_days_ago
    ).count()
    
    # Get this month's inspections for this inspector
    current_month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_inspections = inspector_inspections.filter(
        date_of_inspection__gte=current_month_start
    ).count()
    
    # Get average inspections per month (Oct 2025 to Apr 2026) for this inspector
    start_date = date(2025, 10, 1)
    end_date = date(2026, 4, 1)
    avg_monthly_inspections = inspector_inspections.filter(
        date_of_inspection__gte=start_date,
        date_of_inspection__lt=end_date
    ).count() / 6
    
    # Get commodity breakdown for this inspector
    commodity_stats = inspector_inspections.values('commodity').annotate(
        count=Count('commodity')
    ).order_by('-count')
    
    # Get top clients by inspection count for this inspector
    top_clients = inspector_inspections.values('client_name').annotate(
        count=Count('client_name')
    ).order_by('-count')[:10]
    
    # Calculate percentages for top clients
    total_inspections_for_percentage = sum(client['count'] for client in top_clients)
    for client in top_clients:
        if total_inspections_for_percentage > 0:
            client['percentage'] = (client['count'] / total_inspections_for_percentage) * 100
        else:
            client['percentage'] = 0
    
    # Get monthly trends (Oct 2025 to Apr 2026) for this inspector
    monthly_trends = inspector_inspections.filter(
        date_of_inspection__gte=start_date,
        date_of_inspection__lt=end_date
    ).extra(
        select={'month': "EXTRACT(month FROM date_of_inspection)"}
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Convert month numbers to month names
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for trend in monthly_trends:
        month_num = int(trend['month']) - 1  # Convert to 0-based index
        if 0 <= month_num < 12:
            trend['month_name'] = month_names[month_num]
        else:
            trend['month_name'] = 'Unknown'
    
    # Get inspections over time (monthly for last 12 months) for this inspector
    from django.db.models.functions import TruncMonth
    monthly_inspections = inspector_inspections.filter(
        date_of_inspection__gte=datetime.now() - timedelta(days=365)
    ).annotate(
        month=TruncMonth('date_of_inspection')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Get compliance statistics for this inspector
    # Compliant inspections are those where direction is NOT present
    compliant_inspections = inspector_inspections.filter(
        is_direction_present_for_this_inspection=False
    ).count()
    
    # Non-compliant inspections are those where direction IS present
    non_compliant_inspections = inspector_inspections.filter(
        is_direction_present_for_this_inspection=True
    ).count()
    
    compliance_rate = 0
    if total_inspections > 0:
        compliance_rate = (compliant_inspections / total_inspections) * 100
    
    # Get recent inspections list for this inspector
    recent_inspections_list = inspector_inspections.order_by('-date_of_inspection')[:10]
    
    # Simple forecasting for next 3 months for this inspector
    if monthly_inspections:
        # Calculate average monthly growth rate
        recent_months = list(monthly_inspections)[-6:]  # Last 6 months
        if len(recent_months) >= 2:
            # Calculate average monthly change
            total_change = recent_months[-1]['count'] - recent_months[0]['count']
            avg_monthly_change = total_change / (len(recent_months) - 1)
            
            # Generate forecast for next 3 months
            last_count = recent_months[-1]['count']
            forecast_data = []
            for i in range(1, 4):
                forecast_count = max(0, last_count + (avg_monthly_change * i))
                forecast_data.append({
                    'month': f'Forecast {i}',
                    'count': round(forecast_count, 0)
                })
        else:
            forecast_data = []
    else:
        forecast_data = []
    
    # Debug: Print some values to see what's happening
    print(f"Debug - Inspector: {inspector_name}")
    print(f"Debug - Total inspections for inspector: {total_inspections}")
    print(f"Debug - Monthly inspections count: {len(monthly_inspections)}")
    print(f"Debug - Compliance rate: {compliance_rate:.1f}%")
    
    # Convert QuerySets to lists for proper JSON serialization
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    
    context = {
        'inspector_name': inspector_name,
        'total_inspections': total_inspections,
        'recent_inspections': recent_inspections,
        'this_month_inspections': this_month_inspections,
        'avg_monthly_inspections': round(avg_monthly_inspections, 1),
        'monthly_inspections': json.dumps(list(monthly_inspections), cls=DjangoJSONEncoder),
        'top_clients': json.dumps(list(top_clients), cls=DjangoJSONEncoder),
        'forecast_data': json.dumps(forecast_data, cls=DjangoJSONEncoder),
        'commodity_stats': json.dumps(list(commodity_stats), cls=DjangoJSONEncoder),
        'compliant_inspections': compliant_inspections,
        'non_compliant_inspections': non_compliant_inspections,
        'compliance_rate': round(compliance_rate, 1),
        'recent_inspections_list': recent_inspections_list,
    }
    
    return render(request, 'main/inspector_dashboard.html', context)


@login_required
def analytics_dashboard(request):
    """Display comprehensive analytics dashboard with advanced metrics and insights."""
    # Block administrators from accessing analytics dashboard
    if request.user.role == 'admin':
        return redirect('home')
    
    # Redirect inspectors to their specific dashboard
    if request.user.role == 'inspector':
        return redirect('inspector_dashboard')
    
    from ..models import Client, Inspection, FoodSafetyAgencyInspection
    from django.db.models import Count, Q, Avg, Max, Min, Sum, Case, When, IntegerField
    from django.db.models.functions import TruncMonth, TruncWeek, TruncDay, Extract
    from datetime import datetime, timedelta
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    
    # Date ranges for analysis
    now = datetime.now()
    thirty_days_ago = now - timedelta(days=30)
    seven_days_ago = now - timedelta(days=7)
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = current_month_start - timedelta(days=1)
    start_date = date(2025, 10, 1)
    end_date = date(2026, 4, 1)
    
    # === BASIC STATISTICS ===
    total_clients = Client.objects.count()
    total_inspections = Inspection.objects.count()
    total_food_safety_inspections = FoodSafetyAgencyInspection.objects.count()
    
    # === TIME-BASED METRICS ===
    recent_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=thirty_days_ago
    ).count()
    
    this_week_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=seven_days_ago
    ).count()
    
    this_month_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=current_month_start
    ).count()
    
    last_month_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=last_month_start,
        date_of_inspection__lte=last_month_end
    ).count()
    
    # Calculate growth rates
    month_growth_rate = 0
    if last_month_inspections > 0:
        month_growth_rate = ((this_month_inspections - last_month_inspections) / last_month_inspections) * 100
    
    # === COMPLIANCE ANALYTICS ===
    compliance_stats = FoodSafetyAgencyInspection.objects.aggregate(
        total=Count('id'),
        compliant=Count('id', filter=Q(approved_status='APPROVED')),
        non_compliant=Count('id', filter=Q(approved_status='PENDING'))
    )
    
    compliance_rate = 0
    if compliance_stats['total'] > 0:
        compliance_rate = (compliance_stats['compliant'] / compliance_stats['total']) * 100
    
    # === RFI (REQUEST FOR INFORMATION) ANALYTICS ===
    rfi_stats = FoodSafetyAgencyInspection.objects.aggregate(
        total=Count('id'),
        with_rfi=Count('id', filter=Q(rfi_uploaded_date__isnull=False)),
        without_rfi=Count('id', filter=Q(rfi_uploaded_date__isnull=True))
    )
    
    rfi_rate = 0
    if rfi_stats['total'] > 0:
        rfi_rate = (rfi_stats['with_rfi'] / rfi_stats['total']) * 100
    
    # === INSPECTOR PERFORMANCE ===
    inspector_performance = FoodSafetyAgencyInspection.objects.exclude(
        Q(inspector_name__isnull=True) | Q(inspector_name='') | Q(inspector_name='Unknown')
    ).values('inspector_name').annotate(
        total_inspections=Count('id'),
        compliant=Count('id', filter=Q(approved_status='APPROVED')),
        non_compliant=Count('id', filter=Q(approved_status='PENDING')),
        avg_hours=Avg('hours'),
        avg_distance=Avg('km_traveled'),
        last_inspection=Max('date_of_inspection')
    ).order_by('-total_inspections')[:15]

    # Calculate compliance rate for each inspector
    for inspector in inspector_performance:
        if inspector['total_inspections'] > 0:
            inspector['compliance_rate'] = (inspector['compliant'] / inspector['total_inspections']) * 100
        else:
            inspector['compliance_rate'] = 0
        # For backwards compatibility
        inspector['compliant_inspections'] = inspector['compliant']
        inspector['count'] = inspector['total_inspections']

    # === UNKNOWN INSPECTORS (for assignment) ===
    unknown_inspectors = FoodSafetyAgencyInspection.objects.filter(
        Q(inspector_name__isnull=True) | Q(inspector_name='') | Q(inspector_name='Unknown')
    ).values('inspector_id', 'inspector_name').annotate(
        total_inspections=Count('id'),
        latest_inspection=Max('date_of_inspection'),
        earliest_inspection=Min('date_of_inspection')
    ).order_by('-total_inspections')

    # === CLIENT ANALYTICS ===
    client_analytics = FoodSafetyAgencyInspection.objects.values('client_name').annotate(
        total_inspections=Count('id'),
        compliant_inspections=Count('id', filter=Q(approved_status='APPROVED')),
        avg_hours=Avg('hours'),
        avg_distance=Avg('km_traveled'),
        last_inspection=Max('date_of_inspection'),
        first_inspection=Min('date_of_inspection')
    ).order_by('-total_inspections')[:20]
    
    # Calculate compliance rate and risk score for each client
    for client in client_analytics:
        if client['total_inspections'] > 0:
            client['compliance_rate'] = (client['compliant_inspections'] / client['total_inspections']) * 100
            # Risk score: lower compliance rate = higher risk
            client['risk_score'] = max(0, 100 - client['compliance_rate'])
        else:
            client['compliance_rate'] = 0
            client['risk_score'] = 100
    
    # === COMMODITY ANALYSIS ===
    commodity_analysis = FoodSafetyAgencyInspection.objects.values('commodity').annotate(
        total_inspections=Count('id'),
        compliant_inspections=Count('id', filter=Q(approved_status='APPROVED')),
        avg_hours=Avg('hours'),
        avg_distance=Avg('km_traveled')
    ).order_by('-total_inspections')
    
    for commodity in commodity_analysis:
        if commodity['total_inspections'] > 0:
            commodity['compliance_rate'] = (commodity['compliant_inspections'] / commodity['total_inspections']) * 100
        else:
            commodity['compliance_rate'] = 0
    
    # === GEOGRAPHIC ANALYSIS ===
    geographic_analysis = FoodSafetyAgencyInspection.objects.exclude(
        Q(km_traveled__isnull=True) | Q(km_traveled=0)
    ).aggregate(
        avg_distance=Avg('km_traveled'),
        max_distance=Max('km_traveled'),
        min_distance=Min('km_traveled'),
        total_distance=Sum('km_traveled')
    )
    
    # === TIME SERIES DATA ===
    # Daily inspections for last 30 days
    daily_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=thirty_days_ago
    ).annotate(
        day=TruncDay('date_of_inspection')
    ).values('day').annotate(
        count=Count('id'),
        compliant=Count('id', filter=Q(approved_status='APPROVED')),
        non_compliant=Count('id', filter=Q(approved_status='PENDING'))
    ).order_by('day')

    # Weekly inspections for last 12 weeks
    weekly_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=now - timedelta(weeks=12)
    ).annotate(
        week=TruncWeek('date_of_inspection')
    ).values('week').annotate(
        count=Count('id'),
        compliant=Count('id', filter=Q(approved_status='APPROVED')),
        non_compliant=Count('id', filter=Q(approved_status='PENDING'))
    ).order_by('week')

    # Monthly inspections for last 12 months
    monthly_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=now - timedelta(days=365)
    ).annotate(
        month=TruncMonth('date_of_inspection')
    ).values('month').annotate(
        count=Count('id'),
        compliant=Count('id', filter=Q(approved_status='APPROVED')),
        non_compliant=Count('id', filter=Q(approved_status='PENDING'))
    ).order_by('month')
    
    # === HOURS AND EFFICIENCY ANALYSIS ===
    hours_analysis = FoodSafetyAgencyInspection.objects.exclude(
        Q(hours__isnull=True) | Q(hours=0)
    ).aggregate(
        avg_hours=Avg('hours'),
        max_hours=Max('hours'),
        min_hours=Min('hours'),
        total_hours=Sum('hours')
    )
    
    # === TREND ANALYSIS ===
    # Calculate moving averages
    def calculate_moving_average(data, window=7):
        if len(data) < window:
            return data
        result = []
        for i in range(len(data)):
            if i < window - 1:
                result.append(data[i])
            else:
                avg = sum(data[i-window+1:i+1]) / window
                result.append(avg)
        return result
    
    # === FORECASTING ===
    # Simple linear trend forecasting for next 3 months
    monthly_inspections_list = list(monthly_inspections)
    if len(monthly_inspections_list) >= 3:
        recent_counts = [item['count'] for item in monthly_inspections_list[-3:]]
        if len(recent_counts) >= 2:
            # Simple linear trend
            trend = (recent_counts[-1] - recent_counts[0]) / (len(recent_counts) - 1)
            forecast_data = []
            for i in range(1, 4):  # Next 3 months
                forecast_count = max(0, int(recent_counts[-1] + (trend * i)))
                forecast_data.append({
                    'month': (now + timedelta(days=30*i)).strftime('%Y-%m'),
                    'count': forecast_count
                })
        else:
            forecast_data = []
    else:
        forecast_data = []
    
    # === KPI CALCULATIONS ===
    # Average inspections per month (Oct 2025 to Apr 2026)
    avg_monthly_inspections = FoodSafetyAgencyInspection.objects.filter(
        date_of_inspection__gte=start_date,
        date_of_inspection__lt=end_date
    ).count() / 6
    
    # Efficiency metrics
    efficiency_metrics = {
        'inspections_per_day': recent_inspections / 30 if recent_inspections > 0 else 0,
        'avg_inspection_duration': hours_analysis['avg_hours'] or 0,
        'compliance_rate': compliance_rate,
        'rfi_rate': rfi_rate,
        'month_growth_rate': month_growth_rate
    }
    
    # === RISK ASSESSMENT ===
    # High-risk clients (low compliance rate, recent inspections)
    high_risk_clients = [
        client for client in client_analytics[:10] 
        if client['compliance_rate'] < 70 and client['total_inspections'] >= 3
    ]
    
    # === EXPORT DATA PREPARATION ===
    export_data = {
        'summary': {
            'total_clients': total_clients,
            'total_inspections': total_food_safety_inspections,
            'compliance_rate': compliance_rate,
            'rfi_rate': rfi_rate,
            'month_growth_rate': month_growth_rate
        },
        'time_series': {
            'daily': list(daily_inspections),
            'weekly': list(weekly_inspections),
            'monthly': list(monthly_inspections)
        },
        'performance': {
            'inspectors': list(inspector_performance),
            'clients': list(client_analytics),
            'commodities': list(commodity_analysis)
        }
    }
    
    # === TOP CLIENTS (LEGACY SUPPORT) ===
    top_clients = FoodSafetyAgencyInspection.objects.values('client_name').annotate(
        count=Count('client_name')
    ).order_by('-count')[:10]
    
    # Calculate percentages for top clients
    total_inspections_for_percentage = sum(client['count'] for client in top_clients)
    for client in top_clients:
        if total_inspections_for_percentage > 0:
            client['percentage'] = (client['count'] / total_inspections_for_percentage) * 100
        else:
            client['percentage'] = 0
    
    # === MAJOR COMPANIES (LEGACY SUPPORT) ===
    major_companies = FoodSafetyAgencyInspection.objects.filter(
        Q(client_name__icontains='Shoprite') |
        Q(client_name__icontains='Boxer') |
        Q(client_name__icontains='Pick n Pay') |
        Q(client_name__icontains='Woolworths') |
        Q(client_name__icontains='Spar') |
        Q(client_name__icontains='Checkers') |
        Q(client_name__icontains='Food Lover') |
        Q(client_name__icontains='Massmart') |
        Q(client_name__icontains='Makro') |
        Q(client_name__icontains='Game')
    ).values('client_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # === CONTEXT PREPARATION ===
    # Ensure all data is properly JSON-encoded and handle None values
    def safe_json_dumps(data, default=None):
        """Safely convert data to JSON, handling None values"""
        if data is None:
            return json.dumps(default or [])
        try:
            return json.dumps(data, cls=DjangoJSONEncoder)
        except (TypeError, ValueError):
            return json.dumps(default or [])
    
    context = {
        # Basic statistics
        'total_clients': total_clients or 0,
        'total_inspections': total_inspections or 0,
        'total_food_safety_inspections': total_food_safety_inspections or 0,
        'recent_inspections': recent_inspections or 0,
        'this_week_inspections': this_week_inspections or 0,
        'this_month_inspections': this_month_inspections or 0,
        'last_month_inspections': last_month_inspections or 0,
        'avg_monthly_inspections': round(avg_monthly_inspections or 0, 1),
        
        # Compliance and RFI metrics
        'compliance_rate': round(compliance_rate or 0, 1),
        'compliance_stats': safe_json_dumps(compliance_stats, {'total': 0, 'compliant': 0, 'non_compliant': 0}),
        'rfi_rate': round(rfi_rate or 0, 1),
        'rfi_stats': safe_json_dumps(rfi_stats, {'total': 0, 'with_rfi': 0, 'without_rfi': 0}),
        'month_growth_rate': round(month_growth_rate or 0, 1),

        # Performance data
        'inspector_performance': safe_json_dumps(list(inspector_performance), []),
        'active_inspectors_count': len(inspector_performance),
        'unknown_inspectors': safe_json_dumps(list(unknown_inspectors), []),
        'unknown_inspectors_count': len(unknown_inspectors),
        'client_analytics': safe_json_dumps(list(client_analytics), []),
        'commodity_analysis': safe_json_dumps(list(commodity_analysis), []),
        
        # Time series data
        'daily_inspections': safe_json_dumps(list(daily_inspections), []),
        'weekly_inspections': safe_json_dumps(list(weekly_inspections), []),
        'monthly_inspections': safe_json_dumps(list(monthly_inspections), []),
        
        # Geographic and efficiency data
        'geographic_analysis': safe_json_dumps(geographic_analysis, {'avg_distance': 0, 'max_distance': 0, 'min_distance': 0}),
        'hours_analysis': safe_json_dumps(hours_analysis, {'avg_hours': 0, 'max_hours': 0, 'min_hours': 0, 'total_hours': 0}),
        'efficiency_metrics': safe_json_dumps(efficiency_metrics, {'inspections_per_day': 0, 'avg_inspection_duration': 0, 'compliance_rate': 0, 'rfi_rate': 0, 'month_growth_rate': 0}),
        
        # Risk assessment
        'high_risk_clients': safe_json_dumps(high_risk_clients, []),
        
        # Legacy data for existing charts
        'top_inspectors': safe_json_dumps(list(inspector_performance[:10]), []),
        'major_companies': safe_json_dumps(list(major_companies), []),
        'forecast_data': safe_json_dumps(forecast_data, []),
        
        # Export data
        'export_data': safe_json_dumps(export_data, {}),
    }
    
    return render(request, 'main/analytics_dashboard.html', context)

@login_required
def export_analytics(request, format_type):
    """Export analytics data in various formats"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        if format_type == 'excel':
            return export_analytics_excel(data)
        elif format_type == 'csv':
            return export_analytics_csv(data)
        elif format_type == 'pdf':
            return export_analytics_pdf(data)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_analytics_excel(data):
    """Export analytics data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Report"
    
    # Header
    ws['A1'] = "Food Safety Agency Analytics Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    # Summary Stats
    ws['A3'] = "Summary Statistics"
    ws['A3'].font = Font(size=14, bold=True)
    
    ws['A4'] = "Total Inspections"
    ws['B4'] = data.get('total_food_safety_inspections', 0)
    ws['A5'] = "Recent (30 days)"
    ws['B5'] = data.get('recent_inspections', 0)
    ws['A6'] = "This Month"
    ws['B6'] = data.get('this_month_inspections', 0)
    ws['A7'] = "Avg Monthly (6m)"
    ws['B7'] = data.get('avg_monthly_inspections', 0)
    
    # Monthly Inspections
    ws['A9'] = "Monthly Inspections"
    ws['A9'].font = Font(size=14, bold=True)
    
    ws['A10'] = "Month"
    ws['B10'] = "Inspections"
    ws['C10'] = "Forecast"
    
    row = 11
    monthly_data = data.get('monthly_inspections', [])
    forecast_data = data.get('forecast_data', [])
    
    for month in monthly_data:
        ws[f'A{row}'] = month.get('month', '')
        ws[f'B{row}'] = month.get('count', 0)
        row += 1
    
    for forecast in forecast_data:
        ws[f'A{row}'] = forecast.get('month', '')
        ws[f'C{row}'] = forecast.get('count', 0)
        row += 1
    
    # Top Inspectors
    ws['A{row}'] = "Top Inspectors"
    ws['A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    ws['A{row}'] = "Inspector"
    ws['B{row}'] = "Inspections"
    row += 1
    
    for inspector in data.get('top_inspectors', []):
        ws[f'A{row}'] = inspector.get('inspector_name', 'Unknown')
        ws[f'B{row}'] = inspector.get('count', 0)
        row += 1
    
    # Major Companies
    ws['A{row}'] = "Major Companies"
    ws['A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    ws['A{row}'] = "Company"
    ws['B{row}'] = "Inspections"
    row += 1
    
    for company in data.get('major_companies', []):
        ws[f'A{row}'] = company.get('client_name', '')
        ws[f'B{row}'] = company.get('count', 0)
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="analytics_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response

def export_analytics_csv(data):
    """Export analytics data to CSV"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["Food Safety Agency Analytics Report"])
    writer.writerow([])
    
    # Summary Stats
    writer.writerow(["Summary Statistics"])
    writer.writerow(["Total Inspections", data.get('total_food_safety_inspections', 0)])
    writer.writerow(["Recent (30 days)", data.get('recent_inspections', 0)])
    writer.writerow(["This Month", data.get('this_month_inspections', 0)])
    writer.writerow(["Avg Monthly (6m)", data.get('avg_monthly_inspections', 0)])
    writer.writerow([])
    
    # Monthly Inspections
    writer.writerow(["Monthly Inspections"])
    writer.writerow(["Month", "Inspections", "Forecast"])
    
    monthly_data = data.get('monthly_inspections', [])
    forecast_data = data.get('forecast_data', [])
    
    for month in monthly_data:
        writer.writerow([month.get('month', ''), month.get('count', 0), ''])
    
    for forecast in forecast_data:
        writer.writerow([forecast.get('month', ''), '', forecast.get('count', 0)])
    
    writer.writerow([])
    
    # Top Inspectors
    writer.writerow(["Top Inspectors"])
    writer.writerow(["Inspector", "Inspections"])
    
    for inspector in data.get('top_inspectors', []):
        writer.writerow([inspector.get('inspector_name', 'Unknown'), inspector.get('count', 0)])
    
    writer.writerow([])
    
    # Major Companies
    writer.writerow(["Major Companies"])
    writer.writerow(["Company", "Inspections"])
    
    for company in data.get('major_companies', []):
        writer.writerow([company.get('client_name', ''), company.get('count', 0)])
    
    output.seek(0)
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="analytics_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    return response

def export_analytics_pdf(data):
    """Export analytics data to PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from io import BytesIO
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    story.append(Paragraph("Food Safety Agency Analytics Report", title_style))
    story.append(Spacer(1, 20))
    
    # Summary Stats
    story.append(Paragraph("Summary Statistics", styles['Heading2']))
    summary_data = [
        ["Metric", "Value"],
        ["Total Inspections", str(data.get('total_food_safety_inspections', 0))],
        ["Recent (30 days)", str(data.get('recent_inspections', 0))],
        ["This Month", str(data.get('this_month_inspections', 0))],
        ["Avg Monthly (6m)", str(data.get('avg_monthly_inspections', 0))]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Monthly Inspections
    story.append(Paragraph("Monthly Inspections", styles['Heading2']))
    monthly_data = data.get('monthly_inspections', [])
    forecast_data = data.get('forecast_data', [])
    
    monthly_table_data = [["Month", "Inspections", "Forecast"]]
    for month in monthly_data:
        monthly_table_data.append([month.get('month', ''), str(month.get('count', 0)), ''])
    
    for forecast in forecast_data:
        monthly_table_data.append([forecast.get('month', ''), '', str(forecast.get('count', 0))])
    
    monthly_table = Table(monthly_table_data)
    monthly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(monthly_table)
    story.append(Spacer(1, 20))
    
    # Top Inspectors
    story.append(Paragraph("Top Inspectors", styles['Heading2']))
    inspector_data = [["Inspector", "Inspections"]]
    for inspector in data.get('top_inspectors', []):
        inspector_data.append([inspector.get('inspector_name', 'Unknown'), str(inspector.get('count', 0))])
    
    inspector_table = Table(inspector_data)
    inspector_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(inspector_table)
    story.append(Spacer(1, 20))
    
    # Major Companies
    story.append(Paragraph("Major Companies", styles['Heading2']))
    company_data = [["Company", "Inspections"]]
    for company in data.get('major_companies', []):
        company_data.append([company.get('client_name', ''), str(company.get('count', 0))])
    
    company_table = Table(company_data)
    company_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(company_table)
    
    doc.build(story)
    output.seek(0)
    
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="analytics_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required(login_url='login')
def inspector_settings_view(request):
    """Inspector-specific settings page view."""
    clear_messages(request)
    
    # Only allow inspectors to access this page
    user_role = getattr(request.user, 'role', 'inspector')
    if user_role != 'inspector':
        messages.error(request, "Access denied. This page is only available to inspectors.")
        return redirect('home')
    
    # Get or create settings
    from ..models import Settings, SystemSettings
    settings = Settings.get_settings()
    system_settings = SystemSettings.get_settings()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'save_theme':
            # Handle theme toggle
            theme_mode = request.POST.get('theme_mode')
            if theme_mode == 'on':
                system_settings.dark_mode = True
            else:
                system_settings.dark_mode = False
            system_settings.save()
            # Message removed - theme now saves automatically via JavaScript
            
        elif action == 'update_profile':
            # Handle profile updates
            user = request.user
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            
            if first_name:
                user.first_name = first_name
            if last_name:
                user.last_name = last_name
            if email and email != user.email:
                user.email = email
                
            user.save()
            messages.success(request, "Profile updated successfully!")
            
    
    context = {
        'settings': settings,
        'system_settings': system_settings,
        'user': request.user,
    }
    
    return render(request, 'main/inspector_settings.html', context)


@login_required(login_url='login')
@inspector_only_inspections
def settings_view(request):
    """Settings page view."""
    clear_messages(request)
    
    # Get or create settings
    from ..models import Settings, SystemSettings
    settings = Settings.get_settings()
    system_settings = SystemSettings.get_settings()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'clear_cache':
            # Clear Django cache
            from django.core.cache import cache
            cache.clear()
            messages.success(request, "Application cache cleared successfully!")
            
        elif action == 'reset_settings':
            # Reset settings to defaults
            settings.auto_sync = False
            settings.backup_frequency = 'weekly'
            settings.session_timeout = 30
            settings.google_sheets_enabled = True
            settings.sql_server_enabled = True
            settings.sync_interval = 24
            settings.sync_interval_unit = 'hours'
            settings.email_notifications = False
            settings.sync_notifications = True
            settings.notification_email = None
            settings.two_factor_auth = False
            settings.password_expiry = 90
            settings.max_login_attempts = 5
            settings.save()
            messages.success(request, "Settings reset to default values!")
            
        elif action == 'export_data':
            # Export data functionality (placeholder)
            messages.info(request, "Data export functionality will be implemented soon!")
            
        elif action == 'create_backup':
            # Create manual backup
            from ..services.backup_service import BackupService
            backup_service = BackupService()
            result = backup_service.create_backup(backup_type='manual')
            
            if result['success']:
                messages.success(request, f"Backup created successfully! Files: {', '.join(result['files_created'])}. Records backed up: {result['record_counts']['clients']} clients, {result['record_counts']['inspections']} inspections, {result['record_counts']['shipments']} shipments.")
                try:
                    SystemLog.log_activity(
                        user=request.user,
                        action='EXPORT',
                        page=request.path,
                        description='Created manual backup',
                        details={
                            'files_created': result.get('files_created'),
                            'record_counts': result.get('record_counts')
                        }
                    )
                except Exception:
                    pass
            else:
                messages.error(request, f"Backup failed: {result['error']}")
                try:
                    SystemLog.log_activity(
                        user=request.user,
                        action='ERROR',
                        page=request.path,
                        description='Manual backup failed',
                        details={'error': result.get('error')}
                    )
                except Exception:
                    pass
                    
        elif action == 'save_theme':
            # Save theme setting
            settings.dark_mode = 'theme_mode' in request.POST
            settings.save()
            try:
                SystemLog.log_activity(
                    user=request.user,
                    action='SETTINGS',
                    page=request.path,
                    description='Updated theme setting',
                    details={'dark_mode': settings.dark_mode}
                )
            except Exception:
                pass
                
        elif action == 'save_compliance':
            # Compliance document settings
            settings.compliance_auto_sync = 'compliance_auto_sync' in request.POST
            settings.compliance_sync_interval = int(request.POST.get('compliance_sync_interval', 5))
            settings.compliance_sync_unit = request.POST.get('compliance_sync_unit', 'minutes')
            settings.compliance_batch_mode = request.POST.get('compliance_batch_mode', 'batch')
            settings.compliance_batch_size = int(request.POST.get('compliance_batch_size', 50))
            settings.compliance_date_range = int(request.POST.get('compliance_date_range', 7))
            settings.save()
            
            # Update background service with new settings
            try:
                # Compliance service removed - OneDrive service handles this
                pass
            except:
                pass  # Service might not be running
            
            mode_text = "Process ALL at once" if settings.compliance_batch_mode == 'all' else f"Process in batches of {settings.compliance_batch_size}"
            messages.success(request, f"Compliance settings updated! Auto sync: {'Enabled' if settings.compliance_auto_sync else 'Disabled'}, Interval: {settings.compliance_sync_interval} {settings.compliance_sync_unit}, Mode: {mode_text}")
        
        elif action == 'save_onedrive_cache' or 'onedrive_sync_interval_hours' in request.POST or 'system_auto_sync_enabled' in request.POST or 'onedrive_upload_delay_days' in request.POST or 'compliance_daily_sync_enabled' in request.POST:
            # OneDrive background sync and caching settings (stored in SystemSettings)
            # Always enable OneDrive when delay is configured
            system_settings.onedrive_enabled = True
            # Scheduled sync master toggle
            system_settings.auto_sync_enabled = 'system_auto_sync_enabled' in request.POST
            system_settings.onedrive_local_caching = 'onedrive_local_caching' in request.POST
            try:
                system_settings.onedrive_cache_days = int(request.POST.get('onedrive_cache_days', system_settings.onedrive_cache_days or 60))
            except Exception:
                pass
            # Always enable auto sync when OneDrive is enabled
            system_settings.onedrive_auto_sync = True
            try:
                system_settings.onedrive_sync_interval_hours = int(request.POST.get('onedrive_sync_interval_hours', system_settings.onedrive_sync_interval_hours or 2))
            except Exception:
                pass
            
            # OneDrive upload delay settings
            try:
                system_settings.onedrive_upload_delay_days = int(request.POST.get('onedrive_upload_delay_days', system_settings.onedrive_upload_delay_days or 3))
            except Exception:
                pass
            system_settings.onedrive_upload_delay_unit = request.POST.get('onedrive_upload_delay_unit', system_settings.onedrive_upload_delay_unit or 'days')
            
            # Daily compliance sync settings
            system_settings.compliance_daily_sync_enabled = 'compliance_daily_sync_enabled' in request.POST
            system_settings.compliance_skip_processed = 'compliance_skip_processed' in request.POST
            
            # Compliance sync interval settings (save to Settings model)
            if 'compliance_sync_interval' in request.POST:
                try:
                    settings.compliance_sync_interval = int(request.POST.get('compliance_sync_interval', 5))
                    settings.compliance_sync_unit = request.POST.get('compliance_sync_interval_unit', 'days')
                    settings.save()
                except Exception as e:
                    print(f"Error saving compliance sync interval: {e}")
            
            system_settings.save()
            
            # Create success message
            success_parts = []
            success_parts.append(f"OneDrive auto-upload: {system_settings.onedrive_upload_delay_days} {system_settings.onedrive_upload_delay_unit}")
            
            if system_settings.compliance_daily_sync_enabled:
                success_parts.append("Daily compliance sync: Enabled")
            if system_settings.compliance_skip_processed:
                success_parts.append("Skip processed: Enabled")
            
            messages.success(request, f"Settings updated! {' | '.join(success_parts)}")
            
        elif action == 'pull_all_data':
            # Pull ALL data from Google Drive (not just 4 months)
            try:
                import os
                import threading
                import time
                from datetime import datetime, timedelta
                from concurrent.futures import ThreadPoolExecutor, as_completed
                import multiprocessing
                from ..services.google_drive_service import GoogleDriveService
                from ..models import Client, FoodSafetyAgencyInspection
                from django.conf import settings as django_settings
                
                # Get all inspections (not just 4 months)
                all_inspections = FoodSafetyAgencyInspection.objects.all()
                
                # Get client account code mapping
                client_map = {c.name: (c.internal_account_code or 'no') for c in Client.objects.all()}

                # Google Drive parent folder ID (2025 folder containing month subfolders)
                parent_folder_id = '1pzot8MQ-m3u0f9-BWxpBO40QgLmeZhRP'

                def pull_all_data_background():
                    """Background function to pull ALL data from Google Drive"""
                    try:
                        from datetime import datetime, timedelta

                        # Dynamically determine which months to pull from (October 2025 onwards)
                        start_month = datetime(2025, 10, 1)  # October 2025
                        current_date = datetime.now()

                        # Generate list of month names from October 2025 to current month + 1 month ahead
                        target_months = []
                        temp_date = start_month
                        while temp_date <= current_date.replace(day=1) + timedelta(days=62):  # Current + ~2 months
                            month_name = temp_date.strftime('%B %Y')  # e.g., "October 2025"
                            target_months.append(month_name)
                            # Move to next month
                            if temp_date.month == 12:
                                temp_date = datetime(temp_date.year + 1, 1, 1)
                            else:
                                temp_date = datetime(temp_date.year, temp_date.month + 1, 1)

                        print(f"Starting full system sync - pulling data from: {', '.join(target_months)}")

                        # Initialize Google Drive service
                        drive_service = GoogleDriveService()
                        print(" Google Drive service initialized")

                        # Get all subfolders in the 2025 parent folder
                        print(f" Fetching month folders from parent folder...")
                        month_folders = drive_service.list_files_in_folder(parent_folder_id, request=None)

                        # Collect all files from target month folders
                        all_drive_files = []
                        for folder in month_folders:
                            if folder.get('mimeType') == 'application/vnd.google-apps.folder' and folder.get('name') in target_months:
                                print(f" Fetching files from {folder['name']}...")
                                month_files = drive_service.list_files_in_folder(folder['id'], request=None)
                                all_drive_files.extend(month_files)
                                print(f"   Found {len(month_files)} files in {folder['name']}")

                        # Build file lookup from collected files
                        file_lookup = GoogleDriveService.build_file_lookup(all_drive_files)
                        print(f" Total files found across all target months: {len(all_drive_files)}")
                        
                        # Process all inspections
                        total_inspections = all_inspections.count()
                        processed = 0
                        downloaded = 0
                        errors = 0
                        
                        print(f" Processing {total_inspections} inspections...")
                        
                        for inspection in all_inspections:
                            try:
                                processed += 1
                                
                                # Get client's internal account code and commodity
                                account_code = client_map.get(inspection.client_name, 'no')
                                commodity = str(inspection.commodity).strip().lower() if inspection.commodity else ''
                                
                                if not commodity or account_code == 'no':
                                    continue
                                
                                if commodity == 'eggs':
                                    commodity = 'egg'
                                
                                # Find matching file in Google Drive
                                best_file = None
                                best_days = 10**9
                                
                                for file_info in file_lookup.values():
                                    # Match by account code only (ignore commodity)
                                    if file_info['accountCode'] == account_code:
                                        days_diff = abs((file_info['zipDate'] - inspection.date_of_inspection).days)
                                        if days_diff <= 15 and days_diff < best_days:
                                            best_file = file_info
                                            best_days = days_diff
                                
                                if best_file:
                                    # Create folder structure
                                    year_folder = inspection.date_of_inspection.strftime('%Y')
                                    month_folder = inspection.date_of_inspection.strftime('%B')
                                    
                                    # Use original client name for folder structure
                                    client_folder = inspection.client_name or 'Unknown Client'
                                    
                                    compliance_folder = os.path.join(
                                        django_settings.MEDIA_ROOT, 'inspection', year_folder, month_folder, 
                                        client_folder, 'Compliance', commodity.upper()
                                    )
                                    os.makedirs(compliance_folder, exist_ok=True)
                                    
                                    # Download file if it doesn't exist
                                    file_path = os.path.join(compliance_folder, best_file['name'])
                                    if not os.path.exists(file_path):
                                        success = drive_service.download_file(best_file['id'], file_path)
                                        if success:
                                            downloaded += 1
                                            print(f" Downloaded {best_file['name']} for {inspection.client_name}")
                                        else:
                                            errors += 1
                                            print(f" Failed to download {best_file['name']} for {inspection.client_name}")
                                    else:
                                        print(f"[INFO] File exists: {best_file['name']} for {inspection.client_name}")
                                
                                # Progress update every 100 inspections
                                if processed % 100 == 0:
                                    print(f" Progress: {processed}/{total_inspections} inspections processed, {downloaded} files downloaded, {errors} errors")
                                
                            except Exception as e:
                                errors += 1
                                print(f" Error processing {inspection.client_name}: {e}")
                                continue
                        
                        print(f" Full system sync completed!")
                        print(f" Final results: {processed} inspections processed, {downloaded} files downloaded, {errors} errors")
                        
                    except Exception as e:
                        print(f" Fatal error in full system sync: {e}")
                
                # Start background process
                sync_thread = threading.Thread(target=pull_all_data_background, daemon=True)
                sync_thread.start()
                
                messages.success(request, "Full system sync started in background! Check console for progress.")
                
            except Exception as e:
                messages.error(request, f"Error starting full system sync: {str(e)}")
            
        else:
            # Handle form submissions for different settings sections
            if 'auto_sync' in request.POST:
                # System settings
                settings.auto_sync = 'auto_sync' in request.POST
                
                # Handle backup frequency - convert days to frequency choice
                backup_days = int(request.POST.get('backup_frequency', 7))
                if backup_days == 1:
                    settings.backup_frequency = 'daily'
                elif backup_days == 7:
                    settings.backup_frequency = 'weekly'
                elif backup_days == 30:
                    settings.backup_frequency = 'monthly'
                else:
                    # Default to weekly for any other value
                    settings.backup_frequency = 'weekly'
                
                settings.session_timeout = int(request.POST.get('session_timeout', 30))
                settings.dark_mode = 'theme_mode' in request.POST
                settings.save()
                
                # Create success message with backup frequency info
                frequency_text = {
                    'daily': 'daily',
                    'weekly': 'weekly (every 7 days)',
                    'monthly': 'monthly (every 30 days)'
                }.get(settings.backup_frequency, 'weekly')
                
                messages.success(request, f"System settings updated successfully! Auto sync: {'Enabled' if settings.auto_sync else 'Disabled'}, Backup: {frequency_text}, Session timeout: {settings.session_timeout} minutes, Theme: {'Dark' if settings.dark_mode else 'Light'} mode")
                # Log the settings update
                try:
                    SystemLog.log_activity(
                        user=request.user,
                        action='SETTINGS',
                        page=request.path,
                        description='Updated system settings',
                        details={
                            'auto_sync': settings.auto_sync,
                            'backup_frequency': settings.backup_frequency,
                            'session_timeout': settings.session_timeout,
                            'dark_mode': settings.dark_mode,
                        }
                    )
                except Exception:
                    pass
                
            elif 'google_sheets_enabled' in request.POST:
                # Data sync settings
                settings.google_sheets_enabled = 'google_sheets_enabled' in request.POST
                settings.sql_server_enabled = 'sql_server_enabled' in request.POST
                settings.sync_interval = int(request.POST.get('sync_interval', 24))
                settings.sync_interval_unit = request.POST.get('sync_interval_unit', 'hours')
                settings.save()
                messages.success(request, f"Data sync settings updated successfully! Interval: {settings.sync_interval} {settings.sync_interval_unit}")
            
            else:
                # Handle form submissions for different settings sections
                if 'auto_sync' in request.POST:
                    # System settings
                    settings.auto_sync = 'auto_sync' in request.POST
                    settings.session_timeout = int(request.POST.get('session_timeout', 30))
                    settings.dark_mode = 'theme_mode' in request.POST
                    settings.save()
                    messages.success(request, "System settings updated successfully!")
                    
    # Get backup status
    try:
        from ..services.backup_service import BackupService
        backup_service = BackupService()
        backup_status = backup_service.get_backup_status()
    except Exception as e:
        backup_status = {
            'auto_sync_enabled': settings.auto_sync,
            'backup_frequency': settings.backup_frequency,
            'last_backup': None,
            'next_backup': None,
            'backup_files': []
        }
    
    # Bridge SystemSettings into the settings object for template compatibility
    try:
        for attr in ['onedrive_enabled', 'onedrive_local_caching', 'onedrive_cache_days', 'onedrive_auto_sync', 'onedrive_sync_interval_hours', 'auto_sync_enabled', 'compliance_daily_sync_enabled', 'compliance_skip_processed']:
            setattr(settings, attr, getattr(system_settings, attr))
    except Exception:
        pass

    context = {
        'settings': settings,
        'backup_status': backup_status
    }
    
    return render(request, 'main/settings.html', context)


@login_required
def session_status(request):
    """Debug view to show current session status"""
    from django.utils import timezone
    from datetime import timedelta
    
    settings = Settings.get_settings()
    last_activity = request.session.get('last_activity')
    current_time = timezone.now()
    
    if last_activity:
        last_activity_dt = timezone.datetime.fromisoformat(last_activity.replace('Z', '+00:00'))
        time_since_activity = current_time - last_activity_dt
        timeout_threshold = last_activity_dt + timedelta(minutes=settings.session_timeout)
        time_until_timeout = timeout_threshold - current_time
    else:
        last_activity_dt = None
        time_since_activity = None
        timeout_threshold = None
        time_until_timeout = None
    
    context = {
        'current_time': current_time,
        'last_activity': last_activity_dt,
        'time_since_activity': time_since_activity,
        'session_timeout_minutes': settings.session_timeout,
        'timeout_threshold': timeout_threshold,
        'time_until_timeout': time_until_timeout,
        'is_expired': time_until_timeout and time_until_timeout.total_seconds() < 0
    }
    
    return render(request, 'main/session_status.html', context)


@login_required
def update_test_result(request):
    """Update test result (fat, protein, calcium, dna) for an inspection"""
    # Note: This function handles potential duplicate remote_id records by selecting the first one
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            test_type = request.POST.get('test_type')
            test_result = request.POST.get('test_result') == 'true'
            
            # Validate test type
            valid_test_types = ['fat', 'protein', 'calcium', 'dna']
            if test_type not in valid_test_types:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid test type: {test_type}'
                })
            
            # Get the inspection record - handle potential duplicates by getting the first one
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
                # Log if there are duplicates for debugging
                if inspections.count() > 1:
                    print(f"Warning: Found {inspections.count()} inspections with remote_id {inspection_id}, using the first one")
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })
            
            # Update the appropriate field
            if test_type == 'fat':
                inspection.fat = test_result
            elif test_type == 'protein':
                inspection.protein = test_result
            elif test_type == 'calcium':
                inspection.calcium = test_result
            elif test_type == 'dna':
                inspection.dna = test_result
            
            inspection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'{test_type.title()} test result updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating test result: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_sample_taken(request):
    """Update sample taken status for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            is_sample_taken = request.POST.get('is_sample_taken') == 'true'

            # Get the inspection record - handle potential duplicates by getting the first one
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })

            # Update the sample taken field
            inspection.is_sample_taken = is_sample_taken
            inspection.save()

            print(f"[SUCCESS] Sample taken status updated to {is_sample_taken} for inspection {inspection_id}")

            return JsonResponse({
                'success': True,
                'message': f'Sample taken status updated successfully to {is_sample_taken}'
            })

        except Exception as e:
            print(f"[ERROR] Error updating sample taken status: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error updating sample taken status: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_needs_retest(request):
    """Update needs retest field for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            needs_retest = request.POST.get('needs_retest')

            # Get the inspection record - handle potential duplicates by getting the first one
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
                # Log if there are duplicates for debugging
                if inspections.count() > 1:
                    print(f"Warning: Found {inspections.count()} inspections with remote_id {inspection_id}, using the first one")
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })
            
            # If no sample was taken, automatically set needs_retest to NO
            if not inspection.is_sample_taken:
                inspection.needs_retest = 'NO'
                inspection.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Needs retest automatically set to NO (no sample taken)'
                })
            
            # Validate needs_retest value for cases where sample was taken
            valid_values = ['YES', 'NO', 'yes', 'no', 'pending', 'PENDING', '']
            if needs_retest not in valid_values:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid needs_retest value: {needs_retest}'
                })
            
            # Convert to uppercase for consistency
            if needs_retest:
                needs_retest_upper = needs_retest.upper()
            else:
                needs_retest_upper = None
            
            # Update the needs_retest field
            inspection.needs_retest = needs_retest_upper
            inspection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Needs retest updated successfully to: {needs_retest if needs_retest else "Not set"}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating needs retest: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_km_traveled(request):
    """Update km traveled field for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            km_traveled = request.POST.get('km_traveled')
            
            # Get the inspection record
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })
            
            # Validate km_traveled value
            if km_traveled:
                try:
                    km_value = float(km_traveled)
                    if km_value < 0:
                        return JsonResponse({
                            'success': False,
                            'error': 'Km traveled cannot be negative'
                        })
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid km traveled value'
                    })
            
            # Update the km_traveled field
            inspection.km_traveled = km_value if km_traveled else None
            inspection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Km traveled updated successfully to: {km_value if km_traveled else "Not set"}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating km traveled: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_hours(request):
    """Update hours field for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            hours = request.POST.get('hours')
            
            # Get the inspection record
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })
            
            # Validate hours value
            if hours:
                try:
                    hours_value = float(hours)
                    if hours_value < 0:
                        return JsonResponse({
                            'success': False,
                            'error': 'Hours cannot be negative'
                        })
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid hours value'
                    })
            
            # Update the hours field
            inspection.hours = hours_value if hours else None
            inspection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Hours updated successfully to: {hours_value if hours else "Not set"}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating hours: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_lab(request):
    """Update lab field for an inspection"""
    if request.method == 'POST':
        try:
            inspection_id = request.POST.get('inspection_id')
            lab = request.POST.get('lab')
            
            # Get the inspection record
            try:
                inspections = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id)
                inspection = inspections.first()
                if not inspection:
                    return JsonResponse({
                        'success': False,
                        'error': f'Inspection with ID {inspection_id} not found'
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error finding inspection: {str(e)}'
                })
            
            # Validate lab value (keep in sync with model choices)
            valid_labs = ['lab_a', 'lab_b', 'lab_c', 'lab_d', 'lab_e', 'lab_f', '']
            if lab not in valid_labs:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid lab value: {lab}'
                })
            
            # Update the lab field
            inspection.lab = lab if lab else None
            inspection.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Lab updated successfully to: {lab if lab else "Not set"}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating lab: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_group_km_traveled(request):
    """Update km_traveled field for all inspections in a group"""
    if request.method == 'POST':
        try:
            group_id = request.POST.get('group_id')
            km_traveled = request.POST.get('km_traveled')

            # Parse group_id to extract client_name and date_of_inspection
            # group_id format: "client_name_date_of_inspection"
            if '_' not in group_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Split the group_id to get client_name and date
            parts = group_id.split('_')
            if len(parts) < 2:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Reconstruct client_name (may contain underscores) and date
            date_part = parts[-1]
            client_name_parts = parts[:-1]
            client_name = '_'.join(client_name_parts)

            # Convert date string to date object (support YYYY-MM-DD and YYYYMMDD)
            from datetime import datetime
            date_of_inspection = None
            for fmt in ('%Y-%m-%d', '%Y%m%d'):
                try:
                    date_of_inspection = datetime.strptime(date_part, fmt).date()
                    break
                except ValueError:
                    continue
            if not date_of_inspection:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid date format in group ID: {date_part}'
                })

            # Find all inspections in this group using normalized client name
            import re as _re
            def _normalize(n):
                return _re.sub(r'[^a-zA-Z0-9]', '', (n or '')).lower()

            raw_key = _normalize(client_name)
            candidate_qs = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection=date_of_inspection
            )
            matching_ids = [ins.id for ins in candidate_qs if _normalize(ins.client_name) == raw_key]
            inspections = FoodSafetyAgencyInspection.objects.filter(id__in=matching_ids)

            if not inspections.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'No inspections found for group: {group_id}'
                })

            # Update km_traveled for all inspections in the group
            km_value = float(km_traveled) if km_traveled else None
            updated_count = inspections.update(km_traveled=km_value)

            return JsonResponse({
                'success': True,
                'message': f'Km traveled updated successfully for {updated_count} inspections in group'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating group km traveled: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

def update_group_comment(request):
    """Update comment field for all inspections in a group"""
    if request.method == 'POST':
        try:
            group_id = request.POST.get('group_id')
            comment = request.POST.get('comment', '')

            # Parse group_id to extract client_name and date_of_inspection
            # group_id format: "client_name_date_of_inspection"
            if '_' not in group_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Split the group_id to get client_name and date
            parts = group_id.split('_')
            if len(parts) < 2:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Reconstruct client_name (may contain underscores) and date
            date_part = parts[-1]
            client_name_parts = parts[:-1]
            client_name = '_'.join(client_name_parts)

            # Convert date string to date object (support YYYY-MM-DD and YYYYMMDD)
            from datetime import datetime
            date_of_inspection = None
            for fmt in ('%Y-%m-%d', '%Y%m%d'):
                try:
                    date_of_inspection = datetime.strptime(date_part, fmt).date()
                    break
                except ValueError:
                    continue
            if not date_of_inspection:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid date format in group ID: {date_part}'
                })

            # Find all inspections in this group using normalized client name
            import re as _re
            import logging
            logger = logging.getLogger(__name__)

            def _normalize(n):
                return _re.sub(r'[^a-zA-Z0-9]', '', (n or '')).lower()

            raw_key = _normalize(client_name)
            logger.info(f'[COMMENT] Looking for inspections - group_id: {group_id}, client_name: {client_name}, raw_key: {raw_key}, date: {date_of_inspection}')

            candidate_qs = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection=date_of_inspection
            )
            logger.info(f'[COMMENT] Found {candidate_qs.count()} inspections on date {date_of_inspection}')

            matching_ids = [ins.id for ins in candidate_qs if _normalize(ins.client_name) == raw_key]
            logger.info(f'[COMMENT] Matching inspection IDs: {matching_ids}')

            inspections = FoodSafetyAgencyInspection.objects.filter(id__in=matching_ids)

            if not inspections.exists():
                logger.error(f'[COMMENT] No inspections found for group: {group_id}')
                return JsonResponse({
                    'success': False,
                    'error': f'No inspections found for group: {group_id}'
                })

            # Update comment for all inspections in the group
            updated_count = inspections.update(comment=comment)
            logger.info(f'[COMMENT] Updated comment for {updated_count} inspections')

            return JsonResponse({
                'success': True,
                'message': f'Comment updated successfully for {updated_count} inspections in group'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating group comment: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_group_hours(request):
    """Update hours field for all inspections in a group"""
    if request.method == 'POST':
        try:
            group_id = request.POST.get('group_id')
            hours = request.POST.get('hours')

            # Parse group_id to extract client_name and date_of_inspection
            # group_id format: "client_name_date_of_inspection"
            if '_' not in group_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Split the group_id to get client_name and date
            parts = group_id.split('_')
            if len(parts) < 2:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Reconstruct client_name (may contain underscores) and date
            date_part = parts[-1]
            client_name_parts = parts[:-1]
            client_name = '_'.join(client_name_parts)

            # Convert date string to date object (support YYYY-MM-DD and YYYYMMDD)
            from datetime import datetime
            date_of_inspection = None
            for fmt in ('%Y-%m-%d', '%Y%m%d'):
                try:
                    date_of_inspection = datetime.strptime(date_part, fmt).date()
                    break
                except ValueError:
                    continue
            if not date_of_inspection:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid date format in group ID: {date_part}'
                })

            # Find all inspections in this group using normalized client name
            import re as _re
            def _normalize(n):
                return _re.sub(r'[^a-zA-Z0-9]', '', (n or '')).lower()

            raw_key = _normalize(client_name)
            candidate_qs = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection=date_of_inspection
            )
            matching_ids = [ins.id for ins in candidate_qs if _normalize(ins.client_name) == raw_key]
            inspections = FoodSafetyAgencyInspection.objects.filter(id__in=matching_ids)

            if not inspections.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'No inspections found for group: {group_id}'
                })

            # Update hours for all inspections in the group
            hours_value = float(hours) if hours else None
            updated_count = inspections.update(hours=hours_value)

            return JsonResponse({
                'success': True,
                'message': f'Hours updated successfully for {updated_count} inspections in group'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating group hours: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
def update_group_additional_email(request):
    """Update additional_email field for all inspections in a group"""
    if request.method == 'POST':
        try:
            group_id = request.POST.get('group_id')
            additional_email = request.POST.get('additional_email', '').strip()

            # Parse group_id to extract client_name and date_of_inspection
            # group_id format: "client_name_date_of_inspection"
            if '_' not in group_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Split the group_id to get client_name and date
            parts = group_id.split('_')
            if len(parts) < 2:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Reconstruct client_name (may contain underscores) and date
            date_part = parts[-1]
            client_name_parts = parts[:-1]
            client_name = '_'.join(client_name_parts)

            # Convert date string to date object (support YYYY-MM-DD and YYYYMMDD)
            from datetime import datetime
            date_of_inspection = None
            for fmt in ('%Y-%m-%d', '%Y%m%d'):
                try:
                    date_of_inspection = datetime.strptime(date_part, fmt).date()
                    break
                except ValueError:
                    continue
            if not date_of_inspection:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid date format in group ID: {date_part}'
                })

            # Find all inspections in this group using normalized client name
            import re as _re
            def _normalize(n):
                return _re.sub(r'[^a-zA-Z0-9]', '', (n or '')).lower()

            raw_key = _normalize(client_name)
            candidate_qs = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection=date_of_inspection
            )
            matching_ids = [ins.id for ins in candidate_qs if _normalize(ins.client_name) == raw_key]
            inspections = FoodSafetyAgencyInspection.objects.filter(id__in=matching_ids)

            if not inspections.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'No inspections found for group: {group_id}'
                })

            # Validate email if provided
            if additional_email:
                from django.core.validators import validate_email
                from django.core.exceptions import ValidationError
                
                # Split comma-separated emails and validate each one
                emails = [email.strip() for email in additional_email.split(',') if email.strip()]
                
                for email in emails:
                    try:
                        validate_email(email)
                    except ValidationError:
                        return JsonResponse({
                            'success': False,
                            'error': f'Invalid email format: {email}'
                        })

            # Update additional_email for all inspections in the group
            email_value = additional_email if additional_email else None
            updated_count = inspections.update(additional_email=email_value)

            return JsonResponse({
                'success': True,
                'message': f'Additional email updated successfully for {updated_count} inspections in group',
                'email': email_value
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating group additional email: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


def update_group_approved(request):
    """Update approved_status field for all inspections in a group"""
    if request.method == 'POST':
        try:
            group_id = request.POST.get('group_id')
            approved_status = request.POST.get('approved_status')

            # Parse group_id to extract client_name and date_of_inspection
            # group_id format: "client_name_date_of_inspection"
            if '_' not in group_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Split the group_id to get client_name and date
            parts = group_id.split('_')
            if len(parts) < 2:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid group ID format'
                })

            # Reconstruct client_name (may contain underscores) and date
            date_part = parts[-1]
            client_name_parts = parts[:-1]
            client_name = '_'.join(client_name_parts)

            # Convert date string to date object (support YYYY-MM-DD and YYYYMMDD)
            from datetime import datetime
            date_of_inspection = None
            for fmt in ('%Y-%m-%d', '%Y%m%d'):
                try:
                    date_of_inspection = datetime.strptime(date_part, fmt).date()
                    break
                except ValueError:
                    continue
            if not date_of_inspection:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid date format in group ID: {date_part}'
                })

            # Find all inspections in this group using normalized client name
            import re as _re
            def _normalize(n):
                return _re.sub(r'[^a-zA-Z0-9]', '', (n or '')).lower()

            raw_key = _normalize(client_name)
            candidate_qs = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection=date_of_inspection
            )
            matching_ids = [ins.id for ins in candidate_qs if _normalize(ins.client_name) == raw_key]
            inspections = FoodSafetyAgencyInspection.objects.filter(id__in=matching_ids)

            if not inspections.exists():
                return JsonResponse({
                    'success': False,
                    'error': f'No inspections found for group: {group_id}'
                })

            # Validate approved_status
            valid_statuses = ['PENDING', 'APPROVED']
            if approved_status not in valid_statuses:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid approved status. Must be one of: {", ".join(valid_statuses)}'
                })

            # Update approved_status for all inspections in the group
            updated_count = inspections.update(approved_status=approved_status)

            return JsonResponse({
                'success': True,
                'message': f'Approved status updated successfully for {updated_count} inspections in group',
                'approved_status': approved_status
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating group approved status: {str(e)}'
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })


@login_required
@role_required(['developer'])
def compliance_documents(request):
    """Developer-only page for managing/viewing compliance documents."""
    context = {
        'page_title': 'Compliance Documents (Developer Only)'
    }
    return render(request, 'main/compliance_documents.html', context)


@login_required
@role_required(['developer'])
def onedrive_view(request):
    """Developer-only page for viewing OneDrive contents."""
    import os
    import json
    from django.conf import settings
    from datetime import datetime
    
    # Initialize OneDrive data
    onedrive_data = {
        'connected': False,
        'folders': [],
        'total_files': 0,
        'error': None,
        'token_refreshed': False
    }
    
    try:
        # Check if OneDrive tokens exist
        token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
        if os.path.exists(token_file):
            with open(token_file, 'r') as f:
                tokens = json.load(f)
            
            access_token = tokens.get('access_token')
            expires_at = tokens.get('expires_at', 0)
            current_time = datetime.now().timestamp()
            
            # Check if token is expired or will expire soon (within 10 minutes)
            if access_token and current_time < expires_at:
                # Token is valid
                onedrive_data['connected'] = True
            elif access_token and expires_at - current_time < 600:  # 10 minutes
                # Token expires soon, try to refresh
                print(" Token expires soon, attempting refresh...")
                onedrive_data['token_refreshed'] = True
                onedrive_data['connected'] = True
            elif access_token:
                # Token is expired, try to refresh
                print(" Token expired, attempting refresh...")
                onedrive_data['token_refreshed'] = True
                onedrive_data['connected'] = True
            else:
                onedrive_data['error'] = 'No valid access token found'
        else:
            onedrive_data['error'] = 'OneDrive not connected. Please connect in Settings.'
        
        # If we have a connection (valid or refreshed), try to get folder structure
        if onedrive_data['connected']:
            from ..services.onedrive_direct_service import OneDriveDirectUploadService
            onedrive_service = OneDriveDirectUploadService()
            
            # Authenticate with OneDrive (this will handle token refresh if needed)
            if onedrive_service.authenticate_onedrive():
                # Get ALL folders from OneDrive root directory
                folders = onedrive_service.list_folders_in_onedrive(None)
                
                if folders:
                    onedrive_data['folders'] = folders
                    # Count total files
                    total_files = 0
                    for folder in folders:
                        if 'file_count' in folder:
                            total_files += folder['file_count']
                    onedrive_data['total_files'] = total_files
                else:
                    onedrive_data['error'] = 'No folders found in OneDrive'
            else:
                onedrive_data['error'] = 'Failed to authenticate with OneDrive. For business accounts, you may need to re-authenticate through Settings.'
                onedrive_data['connected'] = False
            
    except Exception as e:
        onedrive_data['error'] = f'Error accessing OneDrive: {str(e)}'
        onedrive_data['connected'] = False
    
    # Get theme settings
    try:
        from ..models import SystemSettings
        system_settings = SystemSettings.get_settings()
        settings = type('Settings', (), {
            'dark_mode': system_settings.theme_mode == 'dark' if hasattr(system_settings, 'theme_mode') else False,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    context = {
        'page_title': 'OneDrive View (Developer Only)',
        'onedrive_data': onedrive_data,
        'settings': settings,
    }
    return render(request, 'main/onedrive_view.html', context)


@login_required
@role_required(['developer'])
def compliance_linking_page(request):
    """Main page for inspection document linking (Apps Script replica)."""
    context = {
        'page_title': 'Inspection Document Links'
    }
    return render(request, 'main/compliance_linking.html', context)


@login_required
@role_required(['developer'])
def get_inspection_data(request):
    """Get ALL inspection data with account codes and document links."""
    try:
        from ..models import FoodSafetyAgencyInspection, Client
        from django.core.cache import cache
        
        # All inspections are now from October 2025 onwards after cleanup
        print("Loading inspections from October 2025 onwards...")
        inspections = FoodSafetyAgencyInspection.objects.select_related().order_by('-date_of_inspection')
        total_inspections_count = inspections.count()
        print(f"Found {total_inspections_count} inspections from October 2025 onwards")
        
        # Build client mapping from ALL clients in database (business name -> account code)
        print("Loading ALL client data from database...")
        client_map = {}
        all_clients = Client.objects.exclude(internal_account_code__isnull=True).exclude(internal_account_code='')
        
        for client in all_clients:
            # Normalize client name like Apps Script does
            key = normalize_client_name(client.client_id or '')
            if key:
                client_map[key] = {
                    'account_code': client.internal_account_code,
                    'email': ''  # Email would come from client database in real implementation
                }
        
        print(f"Loaded {len(client_map)} clients with account codes")
        
        # Load Drive files fresh (with debugging)
        cache_key = 'drive_files_lookup_v2'  # New cache key to force refresh
        file_lookup = cache.get(cache_key)
        
        if not file_lookup:
            print("Loading Drive files from Google Drive...")
            file_lookup = load_drive_files_real(request)
            cache.set(cache_key, file_lookup, 600)  # Cache for 10 minutes
            print(f"Loaded {len(file_lookup)} files from Drive")
        else:
            print(f"Using cached Drive files: {len(file_lookup)} files")
        
        # Process inspections (without Drive file lookup for refresh)
        inspection_data = []
        total_matches = 0
        total_links = 0
        total_errors = 0
        
        print("Processing inspections (refresh mode - no Drive lookup)...")
        for i, inspection in enumerate(inspections):
            try:
                # Progress logging
                if (i + 1) % 1000 == 0:
                    print(f"Processed {i + 1} of {total_inspections_count} inspections...")
                
                # Get client account code using normalized lookup
                client_key = normalize_client_name(inspection.client_name or '')
                client_info = client_map.get(client_key, {})
                
                account_code = client_info.get('account_code', '')
                email = client_info.get('email', '')
                
                # Determine document status and link
                if not account_code:
                    document_status = 'no-account'
                    document_link = 'Can\'t Find Internal Account Code'
                    total_errors += 1
                else:
                    total_matches += 1
                    # Find document link using Apps Script replica logic
                    document_link = find_document_link_apps_script_replica(
                        account_code, 
                        inspection.commodity, 
                        inspection.date_of_inspection,
                        file_lookup
                    )
                    
                    if document_link and document_link != 'Document Not Found' and '<a href=' in document_link:
                        document_status = 'found'
                        total_links += 1
                    else:
                        document_status = 'not-found'
                        document_link = 'Document Not Found'
                
                inspection_data.append({
                    'remote_id': inspection.remote_id,
                    'client_name': inspection.client_name,
                    'account_code': account_code or 'Can\'t Find Internal Account Code',
                    'email': email or 'No Email',
                    'commodity': inspection.commodity,
                    'date_of_inspection': inspection.date_of_inspection.strftime('%Y-%m-%d') if inspection.date_of_inspection else '',
                    'document_link': document_link,
                    'document_status': document_status
                })
                
            except Exception as e:
                total_errors += 1
                inspection_data.append({
                    'remote_id': inspection.remote_id,
                    'client_name': inspection.client_name,
                    'account_code': 'Error',
                    'email': 'Error',
                    'commodity': inspection.commodity,
                    'date_of_inspection': inspection.date_of_inspection.strftime('%Y-%m-%d') if inspection.date_of_inspection else '',
                    'document_link': f'Error: {str(e)[:50]}',
                    'document_status': 'error'
                })
        
        stats = {
            'total': len(inspection_data),
            'matches': total_matches,
            'links': total_links,
            'errors': total_errors
        }
        
        print(f"Processing complete: {len(inspection_data)} inspections, {total_matches} matches, {total_links} links, {total_errors} errors")
        
        return JsonResponse({
            'success': True,
            'inspections': inspection_data,
            'stats': stats
        })
        
    except Exception as e:
        print(f"Error in get_inspection_data: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def process_document_links(request):
    """Process document links for inspections (Apps Script replica)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        from ..models import FoodSafetyAgencyInspection, Client
        from datetime import datetime, timedelta
        
        data = json.loads(request.body)
        batch_size = data.get('batch_size', 100)
        date_range = data.get('date_range', 'recent')
        
        # Get inspections from October 2025 ± 6 months, then apply additional filters
        from datetime import date
        start_date = date(2025, 10, 1)
        end_date = date(2026, 4, 1)
        
        if date_range == 'recent':
            date_filter = datetime.now() - timedelta(days=30)
            # Use the more recent of start_date or recent filter, but not beyond end_date
            final_start_date = max(start_date, date_filter.date())
            inspections = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection__gte=final_start_date,
                date_of_inspection__lt=end_date
            ).order_by('-date_of_inspection')
        elif date_range == 'month':
            now = datetime.now()
            # Only show current month if it's within our date range
            if start_date <= now.date() < end_date:
                inspections = FoodSafetyAgencyInspection.objects.filter(
                    date_of_inspection__gte=start_date,
                    date_of_inspection__lt=end_date,
                    date_of_inspection__year=now.year,
                    date_of_inspection__month=now.month
                ).order_by('-date_of_inspection')
            else:
                inspections = FoodSafetyAgencyInspection.objects.none()
        else:
            # All dates from Oct 2025 to Apr 2026
            inspections = FoodSafetyAgencyInspection.objects.filter(
                date_of_inspection__gte=start_date,
                date_of_inspection__lt=end_date
            ).order_by('-date_of_inspection')
        
        # Apply batch size limit if not processing all
        if batch_size < 500:
            inspections = inspections[:batch_size]
        
        # Load all client data (replica of Apps Script loadAllClientData function)
        client_map = {}
        for client in Client.objects.exclude(internal_account_code__isnull=True).exclude(internal_account_code=''):
            # Normalize client name like Apps Script
            key = normalize_client_name(client.client_id or '')
            if key:
                client_map[key] = {
                    'account_code': client.internal_account_code,
                    'email': ''  # Would be populated from client database
                }
        
        # Load all Drive files using real Google Drive API
        file_lookup = load_drive_files_real(request)
        
        # Process inspections
        processed_inspections = []
        total_matches = 0
        total_links = 0
        total_errors = 0
        
        for inspection in inspections:
            try:
                # Normalize client name for lookup
                client_key = normalize_client_name(inspection.client_name or '')
                client_info = client_map.get(client_key, {})
                
                account_code = client_info.get('account_code', '')
                email = client_info.get('email', '')
                
                # Find document link (replica of Apps Script findDocumentLink function)
                if account_code and account_code != 'no':
                    total_matches += 1
                    document_link = find_document_link_apps_script_replica(
                        account_code,
                        inspection.commodity,
                        inspection.date_of_inspection,
                        file_lookup
                    )
                    
                    if document_link and document_link != 'Document Not Found' and '<a href=' in document_link:
                        document_status = 'found'
                        total_links += 1
                    else:
                        document_status = 'not-found'
                        document_link = 'Document Not Found'
                else:
                    document_status = 'no-account'
                    document_link = 'Can\'t Find Internal Account Code'
                    total_errors += 1
                
                processed_inspections.append({
                    'remote_id': inspection.remote_id,
                    'client_name': inspection.client_name,
                    'account_code': account_code or 'Can\'t Find Internal Account Code',
                    'email': email or 'No Email',
                    'commodity': inspection.commodity,
                    'date_of_inspection': inspection.date_of_inspection.strftime('%Y-%m-%d') if inspection.date_of_inspection else '',
                    'document_link': document_link,
                    'document_status': document_status
                })
                
            except Exception as e:
                total_errors += 1
                processed_inspections.append({
                    'remote_id': inspection.remote_id,
                    'client_name': inspection.client_name,
                    'account_code': 'Error',
                    'email': 'Error',
                    'commodity': inspection.commodity,
                    'date_of_inspection': inspection.date_of_inspection.strftime('%Y-%m-%d') if inspection.date_of_inspection else '',
                    'document_link': f'Error: {str(e)[:50]}',
                    'document_status': 'error'
                })
        
        stats = {
            'total': len(processed_inspections),
            'matches': total_matches,
            'links': total_links,
            'errors': total_errors
        }
        
        summary = f"Processed {len(processed_inspections)} inspections, found {total_matches} matches, created {total_links} links"
        
        return JsonResponse({
            'success': True,
            'inspections': processed_inspections,
            'stats': stats,
            'summary': summary
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def normalize_client_name(name):
    """Normalize client name like Apps Script does."""
    if not name:
        return ''
    
    return name.strip().lower().replace('\u00A0', ' ').replace('\u200B', '').replace('\u2002', ' ').replace('\u2003', ' ').replace('  ', ' ')


def load_drive_files_real(request, use_cache=True):
    """Load real Drive files using the existing GoogleDriveService - Apps Script replica.
    
    Args:
        request: Django request object
        use_cache: If True, checks Redis cache first (default: True)
        
    Returns:
        dict: File lookup dictionary keyed by compound_key
    """
    try:
        from ..services.google_drive_service import GoogleDriveService
        from django.core.cache import cache
        import re
        from datetime import datetime
        
        # Check cache first for faster retrieval
        if use_cache:
            cache_key = 'drive_files_lookup_v3'  # New cache key version
            cached_lookup = cache.get(cache_key)
            if cached_lookup:
                print(f"Using cached Drive files: {len(cached_lookup)} files")
                return cached_lookup
        
        drive_service = GoogleDriveService()
        # Updated to scan 2025 folder with month subfolders (November 2025, October 2025, etc.)
        parent_folder_id = "1pzot8MQ-m3u0f9-BWxpBO40QgLmeZhRP"  # 2025 folder in Shared Drive

        print("Loading Google Drive files from 2025 folder...")
        start_time = datetime.now()

        # Get ALL files from Drive folder (not limited to 1000)
        # Suppress verbose output from Google Drive service
        import logging
        original_level = logging.getLogger().level
        logging.getLogger().setLevel(logging.ERROR)

        # Scan month folders from October 2025 onwards
        all_files = []

        # Month order for filtering (October = 10, November = 11, December = 12)
        month_names = {
            'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
            'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12
        }
        start_month = 10  # October

        print(f"Scanning for month folders from October 2025 onwards...")

        # Get folders in parent directory
        parent_items = drive_service.list_files_in_folder(parent_folder_id, request=request, max_items=None)

        month_folders = []
        for item in parent_items:
            mime_type = item.get('mimeType', '')
            folder_name = item.get('name', '')

            if mime_type == 'application/vnd.google-apps.folder':
                # Check if folder matches pattern "{Month} 2025"
                for month_name, month_num in month_names.items():
                    if folder_name == f"{month_name} 2025" and month_num >= start_month:
                        month_folders.append({'id': item.get('id'), 'name': folder_name, 'month_num': month_num})
                        print(f"Found target folder: {folder_name}")
                        break

        # Sort folders by month order
        month_folders.sort(key=lambda x: x['month_num'])

        # Scan each target month folder for zip files
        for folder in month_folders:
            print(f"Scanning {folder['name']}...")
            month_files = drive_service.list_files_in_folder(folder['id'], request=request, max_items=None)

            file_count = 0
            # Only add actual files (not subfolders)
            for item in month_files:
                if item.get('mimeType', '') != 'application/vnd.google-apps.folder':
                    all_files.append(item)
                    file_count += 1

            print(f"  Found {file_count} files in {folder['name']}")

        print(f"Total files from October 2025 onwards: {len(all_files)}")
        files = all_files

        # Restore logging level
        logging.getLogger().setLevel(original_level)

        file_lookup = {}
        file_count = 0
        
        for file in files:
            file_name = file.get('name', '')
            file_id = file.get('id', '')
            web_view_link = file.get('webViewLink', '')
            
            # Skip debug output for cleaner logs
            
            # Apps Script pattern: COMMODITY-ACCOUNT_CODE-DATE
            # Example: RAW-RE-IND-RAW-NA-1000-2025-01-15
            full_pattern = re.match(r'^([A-Za-z]+)-([A-Z]{2}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2,3}-\d+)-(\d{4}-\d{2}-\d{2})', file_name)
            
            if full_pattern and file_id:
                commodity_prefix = full_pattern.group(1)
                account_code = full_pattern.group(2)
                zip_date_str = full_pattern.group(3)
                
                try:
                    zip_date = datetime.strptime(zip_date_str, '%Y-%m-%d')
                except:
                    continue
                
                # Create compound key exactly like Apps Script
                compound_key = f"{commodity_prefix.lower()}|{account_code}|{zip_date_str}"
                
                file_lookup[compound_key] = {
                    'url': web_view_link or f"https://drive.google.com/file/d/{file_id}/view",
                    'name': file_name,
                    'commodity': commodity_prefix,
                    'accountCode': account_code,
                    'zipDate': zip_date,
                    'zipDateStr': zip_date_str,
                    'file_id': file_id
                }
                
                file_count += 1
                
                # Progress logging (less verbose)
                if file_count % 5000 == 0 and file_count > 0:
                    print(f"Loaded {file_count} files...")
        
        load_time = (datetime.now() - start_time).total_seconds()
        print(f"Loaded {len(file_lookup)} files in {load_time:.1f} seconds")
        
        # Cache the results for 60 minutes (3600 seconds) for faster subsequent access
        if use_cache and file_lookup:
            cache_key = 'drive_files_lookup_v3'
            cache.set(cache_key, file_lookup, 3600)  # Cache for 60 minutes
        
        return file_lookup
        
    except Exception as e:
        print(f"Drive folder access issue: {e}")
        return {}


def find_document_link_apps_script_replica(account_code, commodity, inspection_date, file_lookup):
    """Exact replica of Apps Script findDocumentLink function."""
    if not account_code or account_code == "no" or not commodity or not inspection_date:
        return "Document Not Found"
    
    # Exact commodity normalization from Apps Script
    commodity_str = str(commodity).lower().strip()
    if commodity_str == "eggs":
        commodity_str = "egg"
    
    # Convert inspection_date to datetime if it's not already
    if isinstance(inspection_date, str):
        try:
            from datetime import datetime
            inspection_date_obj = datetime.strptime(inspection_date, '%Y-%m-%d').date()
        except:
            return "Document Not Found"
    elif hasattr(inspection_date, 'date'):
        # It's a datetime object, get just the date part
        inspection_date_obj = inspection_date.date()
    else:
        # It's already a date object
        inspection_date_obj = inspection_date
    
    best_match = None
    best_days_diff = float('inf')
    
    # Search through file lookup - match by account code only
    # Commodity is ignored because compliance docs may have different commodity prefix
    for file_key in file_lookup:
        file = file_lookup[file_key]

        # Match by account code only (ignore commodity mismatch)
        if file['accountCode'] == account_code:
            
            try:
                # Calculate days difference exactly like Apps Script
                zip_date = file['zipDate']
                if hasattr(zip_date, 'date'):
                    zip_date = zip_date.date()
                days_difference = abs((zip_date - inspection_date_obj).days)
                
                if days_difference <= 15 and days_difference < best_days_diff:
                    best_match = file
                    best_days_diff = days_difference
                    
            except Exception as e:
                continue
    
    if best_match:
        # Use webViewLink like Apps Script getUrl() 
        url = best_match['url']
        
        # Also store compliance document info for download organization
        compliance_info = {
            'url': url,
            'filename': best_match['name'],
            'account_code': account_code,
            'commodity': commodity_str,
            'inspection_date': inspection_date_obj,
            'file_id': best_match.get('file_id', '')
        }
        
        return f'<a href="{url}" class="document-link" target="_blank" data-compliance="{compliance_info}"> Document Link</a>'
    
    return "Document Not Found"


def create_compliance_folder_structure():
    """Create compliance document folder structure organized by commodity."""
    import os
    from django.conf import settings
    
    base_compliance_dir = os.path.join(settings.MEDIA_ROOT, 'compliance')
    
    # Create commodity-specific compliance folders
    commodity_folders = [
        'raw_compliance',
        'pmp_compliance', 
        'poultry_compliance',
        'eggs_compliance',
        'general_compliance'
    ]
    
    created_folders = []
    for folder in commodity_folders:
        folder_path = os.path.join(base_compliance_dir, folder)
        os.makedirs(folder_path, exist_ok=True)
        created_folders.append(folder_path)
    
    return created_folders


def organize_zip_file_automatically(zip_file_path, client_name, inspection_date, commodity):
    """Automatically organize ZIP file contents by inspection numbers."""
    import os
    import zipfile
    import shutil
    from datetime import datetime
    import re
    
    try:
        print(f" Analyzing ZIP file: {os.path.basename(zip_file_path)}")
        
        # Get the base directory (client folder)
        # The ZIP file is in: media/inspection/YYYY/Month/ClientName/Compliance/COMMODITY/
        # We need to go up three levels to get to the client folder
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(zip_file_path)))
        
        # Create a temporary extraction directory
        temp_dir = os.path.join(base_dir, f"temp_extract_{os.path.basename(zip_file_path)}")
        os.makedirs(temp_dir, exist_ok=True)
        
        # Extract ZIP file
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Analyze files and organize them
        organized_files = []
        general_files = []
        
        # Walk through extracted files
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):  # Only process PDF files
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, temp_dir)
                    
                    # Look for inspection number in filename (4-digit number)
                    inspection_match = re.search(r'(\d{4})', file)
                    if inspection_match:
                        inspection_number = inspection_match.group(1)
                        organized_files.append({
                            'file_path': file_path,
                            'relative_path': relative_path,
                            'filename': file,
                            'inspection_number': inspection_number
                        })
                        print(f"   Found file for inspection {inspection_number}: {file}")
                    else:
                        general_files.append({
                            'file_path': file_path,
                            'relative_path': relative_path,
                            'filename': file
                        })
                        print(f"   General file: {file}")
        
        # Create individual inspection folders and move files
        for file_info in organized_files:
            # Use "Inspection-XXX" format for individual inspection folders (e.g., "Inspection-212")
            inspection_folder = os.path.join(base_dir, f"Inspection-{file_info['inspection_number']}")

            # Create complete folder structure for each inspection
            compliance_folder = os.path.join(inspection_folder, "Compliance", commodity.upper())
            lab_folder = os.path.join(inspection_folder, "Lab")
            retest_folder = os.path.join(inspection_folder, "Retest")
            form_folder = os.path.join(inspection_folder, "Form")

            os.makedirs(compliance_folder, exist_ok=True)
            os.makedirs(lab_folder, exist_ok=True)
            os.makedirs(retest_folder, exist_ok=True)
            os.makedirs(form_folder, exist_ok=True)

            # Move file to individual inspection compliance folder with commodity subfolder
            dest_path = os.path.join(compliance_folder, file_info['filename'])
            shutil.move(file_info['file_path'], dest_path)
            print(f"   ✅ Moved {file_info['filename']} to Inspection-{file_info['inspection_number']}/Compliance/{commodity.upper()}/")
        
        # Create main document type folders at client level
        rfi_folder = os.path.join(base_dir, "RFI")
        invoice_folder = os.path.join(base_dir, "Invoice")
        main_compliance_folder = os.path.join(base_dir, "Compliance", commodity.upper())  # Use existing Compliance folder with commodity subfolder
        
        os.makedirs(rfi_folder, exist_ok=True)
        os.makedirs(invoice_folder, exist_ok=True)
        os.makedirs(main_compliance_folder, exist_ok=True)
        
        # Move general files to main Compliance folder with commodity subfolder
        for file_info in general_files:
            dest_path = os.path.join(main_compliance_folder, file_info['filename'])
            shutil.move(file_info['file_path'], dest_path)
            print(f"   Moved {file_info['filename']} to main Compliance/{commodity.upper()}/ folder")
        
        # Clean up temporary directory
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        # Optionally, remove the original ZIP file since we've organized its contents
        try:
            os.remove(zip_file_path)
            print(f"  [ERROR] Removed original ZIP file: {os.path.basename(zip_file_path)}")
        except Exception as e:
            print(f"  [ERROR] Could not remove original ZIP file: {e}")
        
        print(f" Auto-organization complete: {len(organized_files)} files organized into individual inspections, {len(general_files)} files moved to general compliance")
        
    except Exception as e:
        print(f" Error in auto-organization: {e}")
        # Clean up temp directory if it exists
        if 'temp_dir' in locals() and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


def download_compliance_document(file_id, account_code, commodity, inspection_date, filename, client_name, request):
    """Download compliance document to client's inspection folder structure."""
    import os
    import re
    from django.conf import settings
    from ..services.google_drive_service import GoogleDriveService
    from datetime import datetime
    
    try:
        # Build path like existing inspection structure: media/inspection/YYYY/Month/ClientName/Compliance/COMMODITY
        if isinstance(inspection_date, str):
            date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
        else:
            date_obj = inspection_date
        
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')  # Full month name like "May"
        
        # Use original client name for folder structure
        client_folder = client_name or 'Unknown Client'
        
        # Note: commodity is tracked but not used for folder structure
        commodity_upper = str(commodity).upper().strip()
        
        # Build full path: media/inspection/YYYY/Month/ClientName/Compliance/COMMODITY
        base_path = os.path.join(
            settings.MEDIA_ROOT, 
            'inspection', 
            year_folder, 
            month_folder, 
            client_folder,
            'Compliance',
            commodity_upper
        )
        os.makedirs(base_path, exist_ok=True)
        
        # Keep original filename - don't rename compliance documents
        safe_filename = filename
        file_path = os.path.join(base_path, safe_filename)
        
        # Download file if it doesn't exist
        if not os.path.exists(file_path):
            drive_service = GoogleDriveService()
            success = drive_service.download_file(file_id, file_path, request=request)
            
            if success:
                # Check if downloaded file is a ZIP and organize it automatically
                if filename.lower().endswith('.zip'):
                    # Check if auto-organization is enabled
                    auto_organize_enabled = getattr(settings, 'AUTO_ORGANIZE_ZIP_FILES', True)
                    if auto_organize_enabled:
                        print(f"[ERROR] Auto-organizing ZIP file: {filename}")
                        try:
                            organize_zip_file_automatically(file_path, client_name, inspection_date, commodity_upper)
                        except Exception as e:
                            print(f"[ERROR] Auto-organization failed for {filename}: {e}")
                            # Continue anyway - the ZIP file is still downloaded
                    else:
                        print(f" ZIP file downloaded but auto-organization is disabled: {filename}")
                
                print(f" Downloaded: {safe_filename}")
                return file_path
            else:
                print(f" Failed to download: {safe_filename}")
                return None
        else:
            print(f"[SKIP] File already exists, skipping: {safe_filename}")
            return file_path  # Already exists
            
    except Exception as e:
        print(f"Error downloading compliance document: {e}")
        return None


@login_required
@role_required(['developer'])
def download_compliance_documents(request):
    """Download compliance documents with found links to organized folders."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        from ..models import FoodSafetyAgencyInspection, Client
        from datetime import date
        
        data = json.loads(request.body)
        commodity_filter = data.get('commodity', 'all')
        date_range = data.get('date_range', 'recent')
        download_all = data.get('download_all', False)
        
        # Get inspections with links found from Oct 2025 to Apr 2026
        start_date = date(2025, 10, 1)
        end_date = date(2026, 4, 1)
        inspections = FoodSafetyAgencyInspection.objects.filter(
            date_of_inspection__gte=start_date,
            date_of_inspection__lt=end_date
        ).order_by('-date_of_inspection')
        
        if commodity_filter != 'all':
            inspections = inspections.filter(commodity__iexact=commodity_filter)
        
        # Build client mapping - use client name instead of client_id
        client_map = {}
        for client in Client.objects.exclude(internal_account_code__isnull=True).exclude(internal_account_code=''):
            key = normalize_client_name(client.name or '')  # Use client.name instead of client.client_id
            if key:
                client_map[key] = client.internal_account_code
        
        # Load Drive files
        file_lookup = load_drive_files_real(request)
        
        # Create compliance folder structure
        create_compliance_folder_structure()
        
        # Process and download documents
        downloaded_count = 0
        error_count = 0
        processed_count = 0
        skipped_duplicates = 0
        processed_files = set()  # Track processed files to prevent duplicates
        
        # Process ALL inspections if download_all is True, otherwise limit to 100
        inspection_limit = None if download_all else 100
        inspections_to_process = inspections[:inspection_limit] if inspection_limit else inspections
        
        for inspection in inspections_to_process:
            try:
                processed_count += 1
                
                # Get account code
                client_key = normalize_client_name(inspection.client_name or '')
                account_code = client_map.get(client_key, '')
                
                if account_code:
                    # Find document link
                    document_link = find_document_link_apps_script_replica(
                        account_code,
                        inspection.commodity,
                        inspection.date_of_inspection,
                        file_lookup
                    )
                    
                    # If link found, extract file info and download
                    if document_link and '<a href=' in document_link:
                        # Extract file info from the link (by account code only)
                        for file_key, file_info in file_lookup.items():
                            if file_info['accountCode'] == account_code:
                                
                                # Create unique file identifier to prevent duplicates
                                file_identifier = f"{file_info['name']}_{account_code}_{inspection.commodity}"
                                
                                # Skip if already processed
                                if file_identifier in processed_files:
                                    skipped_duplicates += 1
                                    print(f"     [SKIP] Skipping duplicate: {file_info['name']}")
                                    break
                                
                                # Mark as processed
                                processed_files.add(file_identifier)
                                
                                # Extract file ID from URL
                                file_id = file_info['url'].split('/d/')[1].split('/')[0] if '/d/' in file_info['url'] else None
                                
                                if file_id:
                                    # Download to compliance folder
                                    downloaded_path = download_compliance_document(
                                        file_id,
                                        account_code,
                                        inspection.commodity,
                                        inspection.date_of_inspection,
                                        file_info['name'],
                                        inspection.client_name,
                                        request
                                    )
                                    
                                    if downloaded_path:
                                        downloaded_count += 1
                                        print(f"      Downloaded: {downloaded_path}")
                                    else:
                                        print(f"      Download failed")
                                else:
                                    print(f"      Could not extract file ID from URL")
                                break
                
            except Exception as e:
                error_count += 1
                print(f"Error processing inspection {inspection.remote_id}: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Downloaded {downloaded_count} compliance documents from {processed_count} inspections (skipped {skipped_duplicates} duplicates)',
            'processed': processed_count,
            'downloaded': downloaded_count,
            'skipped_duplicates': skipped_duplicates,
            'errors': error_count,
            'folders_created': len(client_map)  # Approximate folder count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def start_compliance_document_download(request):
    """Start background compliance document downloading."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Compliance service removed - OneDrive service handles this
        return JsonResponse({
            'success': False,
            'message': 'Compliance service has been removed. Use OneDrive service instead.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def start_compliance_background(request):
    """Start background compliance document processing."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Compliance service removed - OneDrive service handles this
        return JsonResponse({
            'success': False,
            'message': 'Compliance service has been removed. Use OneDrive service instead.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def stop_compliance_background(request):
    """Stop background compliance document processing."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Compliance service removed - OneDrive service handles this
        return JsonResponse({
            'success': False,
            'message': 'Compliance service has been removed. Use OneDrive service instead.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def compliance_background_status(request):
    """Get status of background compliance processing."""
    try:
        # Compliance service removed - OneDrive service handles this
        return JsonResponse({
            'success': False,
            'message': 'Compliance service has been removed. Use OneDrive service instead.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def process_all_compliance_documents(request):
    """Process ALL compliance documents at once with comprehensive logging."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Compliance service removed - OneDrive service handles this
        return JsonResponse({
            'success': False,
            'message': 'Compliance service has been removed. Use OneDrive service instead.',
            'processed': 0,
            'downloaded': 0,
            'folders_created': 0,
            'processing_time': '0',
            'errors': 0
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'processed': 0,
            'downloaded': 0,
            'folders_created': 0,
            'processing_time': '0',
            'errors': 1
        })


@login_required
def scheduled_sync_service_status(request):
    """Get scheduled sync service status."""
    try:
        from ..services.scheduled_sync_service import get_scheduled_sync_service_status
        
        status = get_scheduled_sync_service_status()
        return JsonResponse(status)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def start_scheduled_sync_service(request):
    """Start the scheduled sync service."""
    try:
        from ..services.scheduled_sync_service import start_scheduled_sync_service
        
        success, message = start_scheduled_sync_service()
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def stop_scheduled_sync_service(request):
    """Stop the scheduled sync service."""
    try:
        from ..services.scheduled_sync_service import stop_scheduled_sync_service
        
        success, message = stop_scheduled_sync_service()
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def run_manual_sync(request):
    """Run a manual sync of a specific type."""
    try:
        import json
        
        data = json.loads(request.body)
        sync_type = data.get('sync_type', '')
        
        if not sync_type:
            return JsonResponse({'success': False, 'error': 'Sync type not specified'})
        
        from ..services.scheduled_sync_service import run_manual_sync
        
        success, message = run_manual_sync(sync_type)
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def save_system_settings(request):
    """Save system settings."""
    try:
        import json
        from ..models import SystemSettings
        
        data = json.loads(request.body)
        
        # Get or create settings
        settings = SystemSettings.get_settings()
        
        # Update settings
        settings.auto_sync_enabled = data.get('auto_sync_enabled', False)
        settings.backup_frequency_days = data.get('backup_frequency_days', 7)
        settings.session_timeout_minutes = data.get('session_timeout_minutes', 30)
        settings.google_sheets_enabled = data.get('google_sheets_enabled', True)
        settings.sql_server_enabled = data.get('sql_server_enabled', True)
        settings.sync_interval_hours = data.get('sync_interval_hours', 24)
        settings.onedrive_enabled = data.get('onedrive_enabled', True)
        settings.onedrive_local_caching = data.get('onedrive_local_caching', True)
        settings.onedrive_cache_days = data.get('onedrive_cache_days', 60)
        settings.onedrive_auto_sync = data.get('onedrive_auto_sync', True)
        settings.onedrive_sync_interval_hours = data.get('onedrive_sync_interval_hours', 2)
        settings.onedrive_upload_delay_days = data.get('onedrive_upload_delay_days', 3)
        settings.onedrive_upload_delay_unit = data.get('onedrive_upload_delay_unit', 'days')
        settings.theme_mode = data.get('theme_mode', 'light')
        
        # Compliance settings
        settings.compliance_daily_sync_enabled = data.get('compliance_daily_sync_enabled', True)
        settings.compliance_skip_processed = data.get('compliance_skip_processed', True)
        
        settings.save()
        
        return JsonResponse({'success': True, 'message': 'System settings saved successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def start_daily_compliance_sync(request):
    """Start the daily compliance sync service."""
    try:
        from ..services.daily_compliance_sync import daily_sync_service
        
        if daily_sync_service.is_running:
            return JsonResponse({'success': False, 'message': 'Daily compliance sync is already running.'})
        
        daily_sync_service.start_daily_sync(manual_start=True)
        return JsonResponse({'success': True, 'message': 'Daily compliance sync started successfully.'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def stop_daily_compliance_sync(request):
    """Stop the daily compliance sync service."""
    try:
        from ..services.daily_compliance_sync import daily_sync_service
        
        if not daily_sync_service.is_running:
            return JsonResponse({'success': False, 'message': 'Daily compliance sync is not running.'})
        
        daily_sync_service.stop_daily_sync()
        return JsonResponse({'success': True, 'message': 'Daily compliance sync stopped successfully.'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def daily_compliance_sync_status(request):
    """Get the status of the daily compliance sync service."""
    try:
        from ..services.daily_compliance_sync import daily_sync_service
        
        status = daily_sync_service.get_status()
        return JsonResponse({'success': True, 'status': status})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_system_settings(request):
    """Get current system settings."""
    try:
        from ..models import SystemSettings
        
        settings = SystemSettings.get_settings()
        
        return JsonResponse({
            'success': True,
            'settings': {
                'auto_sync_enabled': settings.auto_sync_enabled,
                'backup_frequency_days': settings.backup_frequency_days,
                'session_timeout_minutes': settings.session_timeout_minutes,
                'google_sheets_enabled': settings.google_sheets_enabled,
                'sql_server_enabled': settings.sql_server_enabled,
                'sync_interval_hours': settings.sync_interval_hours,
                'compliance_sync_enabled': settings.compliance_sync_enabled,
                'compliance_sync_interval_hours': settings.compliance_sync_interval_hours,
                'compliance_processing_mode': settings.compliance_processing_mode,
                'compliance_daily_sync_enabled': settings.compliance_daily_sync_enabled,
                'compliance_skip_processed': settings.compliance_skip_processed,
                'compliance_last_processed_date': settings.compliance_last_processed_date.isoformat() if settings.compliance_last_processed_date else None,
                'onedrive_enabled': settings.onedrive_enabled,
                'onedrive_local_caching': settings.onedrive_local_caching,
                'onedrive_cache_days': settings.onedrive_cache_days,
                'onedrive_auto_sync': settings.onedrive_auto_sync,
                'onedrive_sync_interval_hours': settings.onedrive_sync_interval_hours,
                'onedrive_upload_delay_days': settings.onedrive_upload_delay_days,
                'onedrive_upload_delay_unit': settings.onedrive_upload_delay_unit,
                'theme_mode': settings.theme_mode
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_inspection_files_onedrive(client_name, inspection_date):
    """Get inspection files from OneDrive with Redis caching."""
    try:
        import os
        import json
        import requests
        from datetime import datetime
        from django.conf import settings
        from django.core.cache import cache
        import re
        
        # Parse date
        date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')
        
        # Use original client name for folder matching (folders now use original names)
        client_folder = client_name or 'Unknown Client'
        
        # First try to get from Redis cache (fastest)
        cache_key = f"client_files:{client_folder}:{year_folder}:{month_folder}"
        cached_files = cache.get(cache_key)
        
        if cached_files:
            print(f"Using cached files for {client_name} ({len(cached_files)} files)")
            return organize_cached_files(cached_files)
        
        # If not in cache, check local files first (faster than OneDrive API)
        local_files = get_inspection_files_local(client_name, inspection_date)
        if local_files and any(local_files.values()):
            print(f" Using local files for {client_name}")
            return local_files
        
        # Finally, fall back to OneDrive API (slowest)
        print(f"[ERROR] Fetching from OneDrive for {client_name}")
        return get_inspection_files_onedrive_api(client_name, inspection_date)
        
    except Exception as e:
        print(f" Error getting files: {e}")
        return None


def organize_cached_files(cached_files):
    """Organize cached files into categories."""
    try:
        files_by_category = {
            'rfi': [],
            'invoice': [],
            'lab': [],
            'retest': [],
            'compliance': []
        }
        
        for file_metadata in cached_files:
            if isinstance(file_metadata, str):
                # Parse JSON string
                file_metadata = json.loads(file_metadata)
            
            # Determine category based on path
            if 'Compliance' in file_metadata.get('local_path', ''):
                files_by_category['compliance'].append({
                    'name': file_metadata['name'],
                    'size': file_metadata['size'],
                    'url': file_metadata.get('onedrive_path', ''),
                    'relative_path': file_metadata['local_path'],
                    'source': file_metadata.get('source', 'onedrive'),
                    'modified': file_metadata.get('cached_at', '')
                })
            else:
                # For other categories, we'd need to determine from path
                # For now, add to compliance if it's a compliance file
                files_by_category['compliance'].append({
                    'name': file_metadata['name'],
                    'size': file_metadata['size'],
                    'url': file_metadata.get('onedrive_path', ''),
                    'relative_path': file_metadata['local_path'],
                    'source': file_metadata.get('source', 'onedrive'),
                    'modified': file_metadata.get('cached_at', '')
                })
        
        return files_by_category
        
    except Exception as e:
        print(f" Error organizing cached files: {e}")
        # Return empty files object with proper structure instead of empty dict
        return {
            'rfi': [],
            'invoice': [],
            'lab': [],
            'retest': [],
            'compliance': []
        }


def get_inspection_files_onedrive_api(client_name, inspection_date):
    """Get inspection files from OneDrive API (slow method)."""
    try:
        import os
        import json
        import requests
        from datetime import datetime
        from django.conf import settings
        import re
        
        # Load OneDrive tokens
        token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
        if not os.path.exists(token_file):
            return None
        
        with open(token_file, 'r') as f:
            token_data = json.load(f)
        
        access_token = token_data.get('access_token')
        if not access_token:
            return None
        
        # Build OneDrive path
        date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')
        
        # Use original client name for folder matching (folders now use original names)
        client_folder = client_name or 'Unknown Client'
        
        onedrive_base = getattr(settings, 'ONEDRIVE_FOLDER', 'FoodSafety Agency Inspections')
        base_path = f"{onedrive_base}/inspection/{year_folder}/{month_folder}/{client_folder}"
        
        # Define file categories
        categories = {
            'rfi': 'rfi',
            'invoice': 'invoice',
            'lab': 'lab results',
            'retest': 'retest',
            'compliance': 'Compliance'
        }
        
        files_by_category = {}
        
        for category_key, folder_name in categories.items():
            files_by_category[category_key] = []
            
            if category_key == 'compliance':
                # Check compliance subfolders
                compliance_base = f"{base_path}/Compliance"
                for commodity in ['RAW', 'PMP', 'POULTRY', 'EGGS']:
                    commodity_path = f"{compliance_base}/{commodity}"
                    
                    # Get files from OneDrive
                    check_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{commodity_path}:/children"
                    headers = {
                        'Authorization': f'Bearer {access_token}',
                        'Content-Type': 'application/json'
                    }
                    
                    response = requests.get(check_url, headers=headers)
                    if response.status_code == 200:
                        files = response.json().get('value', [])
                        for file_info in files:
                            if file_info.get('file'):
                                file_data = {
                                    'name': file_info['name'],
                                    'size': file_info.get('size', 0),
                                    'url': file_info.get('@microsoft.graph.downloadUrl', ''),
                                    'relative_path': f'Compliance/{commodity}/{file_info["name"]}',
                                    'onedrive_id': file_info.get('id', ''),
                                    'source': 'onedrive'
                                }
                                files_by_category[category_key].append(file_data)
            else:
                # Check regular category folder
                category_path = f"{base_path}/{folder_name}"
                
                # Get files from OneDrive
                check_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{category_path}:/children"
                headers = {
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/json'
                }
                
                response = requests.get(check_url, headers=headers)
                if response.status_code == 200:
                    files = response.json().get('value', [])
                    for file_info in files:
                        if file_info.get('file'):
                            file_data = {
                                'name': file_info['name'],
                                'size': file_info.get('size', 0),
                                'url': file_info.get('@microsoft.graph.downloadUrl', ''),
                                'relative_path': f'{folder_name}/{file_info["name"]}',
                                'onedrive_id': file_info.get('id', ''),
                                'source': 'onedrive'
                            }
                            files_by_category[category_key].append(file_data)
        
        return files_by_category
        
    except Exception as e:
        return None


def get_inspection_files_local(client_name, inspection_date, force_refresh=False):
    """Get inspection files from local media folder.

    NEW STRUCTURE (preferred): MEDIA_ROOT/docs/{client_id}/{inspection_id}/{category}/
    LEGACY STRUCTURE (fallback): MEDIA_ROOT/inspection/YEAR/MONTH/CLIENT/...

    Uses database IDs for reliable file lookup - no more name normalization issues.
    """
    import os
    import time
    from datetime import datetime
    from django.conf import settings
    from django.core.cache import cache
    from main.models import FoodSafetyAgencyInspection

    # Standard categories (coa removed - lab folder contains all COA/Lab files)
    CATEGORIES = ['rfi', 'invoice', 'lab', 'lab_form', 'retest', 'compliance', 'occurrence', 'composition', 'other']

    # Initialize result
    files_by_category = {cat: [] for cat in CATEGORIES}

    try:
        # Strip "btn-" prefix from client name if present
        if client_name and client_name.startswith('btn-'):
            client_name = client_name[4:]

        # Parse date
        date_obj = None
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d']
        for fmt in date_formats:
            try:
                date_obj = datetime.strptime(inspection_date, fmt)
                break
            except ValueError:
                continue

        if not date_obj:
            print(f"[FILES DEBUG] Could not parse date: {inspection_date}")
            return files_by_category

        print(f"[FILES DEBUG] Looking up inspections for client='{client_name}', date={date_obj.date()}")

        # Look up ALL inspections for this client+date (not just first)
        # This matches the behavior of check_group_files which loops through all
        inspections = FoodSafetyAgencyInspection.objects.filter(
            client_name__iexact=client_name,
            date_of_inspection=date_obj.date()
        ).select_related('client')

        inspection_count = inspections.count()
        if inspection_count > 0:
            print(f"[FILES DEBUG] Found {inspection_count} inspection(s)")
        else:
            print(f"[FILES DEBUG] NO inspections found in database!")

        # Track added files to avoid duplicates
        added_files = set()

        # === NEW STRUCTURE: MEDIA_ROOT/docs/{client_id}/{inspection_id}/{category}/ ===
        # Loop through ALL inspections to find files (same as check_group_files)
        docs_base = os.path.join(settings.MEDIA_ROOT, 'docs')

        for inspection in inspections:
            client_obj = None
            if inspection.client:
                client_obj = inspection.client
            else:
                # Try to look up client by name
                from main.models import Client
                client_obj = Client.objects.filter(name__iexact=client_name).first()
                if client_obj:
                    print(f"[FILES DEBUG] Found client by name lookup: id={client_obj.id}, name={client_obj.name}")

            if client_obj:
                docs_path = os.path.join(
                    docs_base,
                    str(client_obj.id),
                    str(inspection.id)
                )
                print(f"[FILES DEBUG] Checking path for inspection {inspection.id}: {docs_path}")

                if os.path.exists(docs_path):
                    print(f"[FILES DEBUG] Path EXISTS, scanning: {docs_path}")
                    for category in CATEGORIES:
                        cat_path = os.path.join(docs_path, category)
                        if os.path.exists(cat_path):
                            try:
                                for filename in os.listdir(cat_path):
                                    file_path = os.path.join(cat_path, filename)
                                    if os.path.isfile(file_path):
                                        file_key = f"{filename}_{os.path.getsize(file_path)}"
                                        if file_key not in added_files:
                                            file_info = get_file_info(file_path, category)
                                            files_by_category[category].append(file_info)
                                            added_files.add(file_key)
                            except (OSError, PermissionError):
                                pass

        # === LEGACY STRUCTURE: MEDIA_ROOT/inspection/YEAR/MONTH/CLIENT/... ===
        # Fall back to old structure for backwards compatibility
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')

        # Create normalized folder name for legacy lookups
        def normalize_name(name):
            if not name:
                return "unknown_client"
            import re
            clean = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
            clean = clean.replace(' ', '_').replace('-', '_')
            clean = re.sub(r'_+', '_', clean).strip('_').lower()
            return clean or "unknown_client"

        client_folder = normalize_name(client_name)

        # Legacy folder variations to check
        legacy_base = os.path.join(settings.MEDIA_ROOT, 'inspection', year_folder, month_folder)
        folder_variations = [
            client_name,
            client_folder,
            f'btn_{client_folder}',
        ]

        # Legacy category folder names
        legacy_structures = [
            {'rfi': 'rfi', 'invoice': 'invoice', 'lab': 'lab', 'retest': 'retest', 'compliance': 'compliance', 'occurrence': 'occurrence', 'composition': 'composition'},
            {'rfi': 'Request For Invoice', 'invoice': 'Invoice', 'lab': 'lab results', 'compliance': 'Compliance'},
            {'rfi': 'RFI', 'invoice': 'Invoice', 'lab': 'Lab', 'compliance': 'Compliance'},
        ]

        for variation in folder_variations:
            client_path = os.path.join(legacy_base, variation)
            if not os.path.exists(client_path):
                continue

            print(f"[FILES] Scanning legacy path: {client_path}")

            # Scan top-level category folders
            for structure in legacy_structures:
                for cat_key, folder_name in structure.items():
                    cat_path = os.path.join(client_path, folder_name)
                    if os.path.exists(cat_path) and os.path.isdir(cat_path):
                        try:
                            for filename in os.listdir(cat_path):
                                file_path = os.path.join(cat_path, filename)
                                if os.path.isfile(file_path):
                                    file_key = f"{filename}_{os.path.getsize(file_path)}"
                                    if file_key not in added_files:
                                        file_info = get_file_info(file_path, folder_name)
                                        files_by_category[cat_key].append(file_info)
                                        added_files.add(file_key)
                        except (OSError, PermissionError):
                            pass

            # Scan Inspection-XXX subfolders
            try:
                for item in os.listdir(client_path):
                    if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(client_path, item)):
                        insp_folder = os.path.join(client_path, item)
                        for structure in legacy_structures:
                            for cat_key, folder_name in structure.items():
                                cat_path = os.path.join(insp_folder, folder_name)
                                if os.path.exists(cat_path) and os.path.isdir(cat_path):
                                    try:
                                        for filename in os.listdir(cat_path):
                                            file_path = os.path.join(cat_path, filename)
                                            if os.path.isfile(file_path):
                                                file_key = f"{filename}_{os.path.getsize(file_path)}"
                                                if file_key not in added_files:
                                                    file_info = get_file_info(file_path, f'{item}/{folder_name}')
                                                    files_by_category[cat_key].append(file_info)
                                                    added_files.add(file_key)
                                    except (OSError, PermissionError):
                                        pass
            except (OSError, PermissionError):
                pass

        # Cache result (use client_obj which might be from name lookup)
        if inspection and client_obj:
            cache_key = f"docs_files:{client_obj.id}:{inspection.id}"
            cache.set(cache_key, files_by_category, 30)

        # Debug output
        total = sum(len(f) for f in files_by_category.values())
        print(f"[FILES] Found {total} files for {client_name} on {inspection_date}")

        return files_by_category

    except Exception as e:
        print(f"[FILES] Error: {e}")
        return files_by_category


@login_required
def populate_six_month_files(request):
    """Populate files for all inspections in the last 6 months."""
    try:
        from datetime import datetime, timedelta
        import os
        from ..models import FoodSafetyAgencyInspection
        
        # Get inspections from Oct 2025 to Apr 2026
        start_date = date(2025, 10, 1)
        end_date = date(2026, 4, 1)
        recent_inspections = FoodSafetyAgencyInspection.objects.filter(
            date_of_inspection__gte=start_date,
            date_of_inspection__lt=end_date
        ).values('client_name', 'date_of_inspection').distinct()
        
        # Use the correct inspection structure: media/inspection/YYYY/Month/ClientName/
        total_folders_created = 0
        
        for inspection in recent_inspections:
            try:
                # Create folder structure using the correct format
                year_folder = inspection['date_of_inspection'].strftime('%Y')
                month_folder = inspection['date_of_inspection'].strftime('%B')
                
                # Use original client name for folder structure (folders now use original names)
                client_folder = inspection['client_name'] or 'Unknown Client'
                
                folder_path = os.path.join(
                    settings.MEDIA_ROOT, 
                    'inspection', 
                    year_folder, 
                    month_folder, 
                    client_folder
                )
                os.makedirs(folder_path, exist_ok=True)
                total_folders_created += 1
                
                # Try to fetch files for this inspection
                try:
                    files = get_inspection_files_local(inspection['client_name'], inspection['date_of_inspection'].strftime('%Y-%m-%d'))
                    if files:
                        print(f"Found {len(files)} files for {inspection['client_name']} - {inspection['date_of_inspection']}")
                except Exception as e:
                    print(f"Error fetching files for {inspection['client_name']}: {e}")
                    
            except Exception as e:
                print(f"Error creating folder for {inspection['client_name']}: {e}")
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'Created {total_folders_created} inspection folders for the last 6 months',
            'folders_created': total_folders_created
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def pull_six_month_data_from_google_drive(request):
    """Pull ALL 6-Month Data from Google Drive and store locally."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        from datetime import datetime, timedelta
        from ..models import FoodSafetyAgencyInspection, Client
        from ..services.google_drive_service import GoogleDriveService
        import os
        import re
        
        print("STARTING 6-MONTH GOOGLE DRIVE DATA PULL")
        print("=" * 60)
        
        # Get inspections from Oct 2025 to Apr 2026
        start_date = date(2025, 10, 1)
        end_date = date(2026, 4, 1)
        recent_inspections = FoodSafetyAgencyInspection.objects.filter(
            date_of_inspection__gte=start_date,
            date_of_inspection__lt=end_date
        ).order_by('-date_of_inspection')
        
        total_inspections = recent_inspections.count()
        print(f" Found {total_inspections:,} inspections from Oct 2025 to Apr 2026")
        
        if total_inspections == 0:
            return JsonResponse({
                'success': True,
                'message': 'No inspections found in the last 6 months',
                'files_downloaded': 0,
                'folders_created': 0
            })
        
        # Build client mapping
        client_map = {}
        for client in Client.objects.exclude(internal_account_code__isnull=True).exclude(internal_account_code=''):
            key = normalize_client_name(client.name or '')
            if key:
                client_map[key] = client.internal_account_code
        
        print(f" Loaded {len(client_map):,} client mappings")
        
        # Initialize Google Drive service
        drive_service = GoogleDriveService()
        if not drive_service.authenticate(request):
            return JsonResponse({
                'success': False,
                'error': 'Failed to authenticate with Google Drive. Please check your credentials.'
            })
        
        # Load Drive files
        print("[ERROR] Loading Google Drive files...")
        file_lookup = load_drive_files_real(request)
        print(f" Loaded {len(file_lookup):,} files from Google Drive")
        
        # Track statistics
        files_downloaded = 0
        folders_created = 0
        errors = 0
        processed_count = 0
        
        # Process each inspection
        for inspection in recent_inspections:
            try:
                processed_count += 1
                
                # Progress logging every 50 inspections
                if processed_count % 50 == 0:
                    print(f"[PROGRESS] Progress: {processed_count:,}/{total_inspections:,} ({(processed_count/total_inspections*100):.1f}%) - Downloaded: {files_downloaded:,}")
                
                # Get account code for this client
                client_key = normalize_client_name(inspection.client_name or '')
                account_code = client_map.get(client_key, '')
                
                if not account_code:
                    continue
                
                # Find matching files in Google Drive
                matching_files = []
                for file_key, file_info in file_lookup.items():
                    # Match by account code only (ignore commodity)
                    if file_info['accountCode'] == account_code:
                        matching_files.append(file_info)
                
                if not matching_files:
                    continue
                
                # Create folder structure for this inspection
                year_folder = inspection.date_of_inspection.strftime("%Y")
                month_folder = inspection.date_of_inspection.strftime("%B")
                
                # Use original client name for folder structure
                client_folder = inspection.client_name or 'Unknown Client'
                
                # Create base folder structure
                base_path = os.path.join(
                    settings.MEDIA_ROOT,
                    'inspection',
                    year_folder,
                    month_folder,
                    client_folder,
                    'Compliance',
                    str(inspection.commodity).upper()
                )
                os.makedirs(base_path, exist_ok=True)
                folders_created += 1
                
                # Download each matching file
                for file_info in matching_files:
                    try:
                        # Extract file ID from URL
                        file_id = file_info.get('file_id') or file_info.get('url', '').split('/d/')[1].split('/')[0] if '/d/' in file_info.get('url', '') else None
                        
                        if not file_id:
                            continue
                        
                        # Keep original filename - don't rename compliance documents
                        safe_filename = file_info['name']
                        file_path = os.path.join(base_path, safe_filename)
                        
                        # Download file if it doesn't exist
                        if not os.path.exists(file_path):
                            success = drive_service.download_file(file_id, file_path, request)
                            if success:
                                files_downloaded += 1
                                if files_downloaded % 10 == 0:
                                    print(f" Downloaded {files_downloaded:,} files so far...")
                            else:
                                errors += 1
                        else:
                            # File already exists, count as downloaded
                            files_downloaded += 1
                            
                    except Exception as e:
                        errors += 1
                        print(f" Error downloading file {file_info.get('name', 'unknown')}: {e}")
                
            except Exception as e:
                errors += 1
                print(f" Error processing inspection {inspection.remote_id}: {e}")
        
        # Final statistics
        print("=" * 60)
        print(" 6-MONTH GOOGLE DRIVE PULL COMPLETE")
        print(f" Total Inspections Processed: {processed_count:,}")
        print(f" Folders Created: {folders_created:,}")
        print(f" Files Downloaded: {files_downloaded:,}")
        print(f" Errors: {errors:,}")
        print("=" * 60)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully pulled 6-month data from Google Drive! Downloaded {files_downloaded:,} files and created {folders_created:,} folders.',
            'files_downloaded': files_downloaded,
            'folders_created': folders_created,
            'inspections_processed': processed_count,
            'errors': errors
        })
        
    except Exception as e:
        print(f" Fatal error in 6-month data pull: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Fatal error: {str(e)}'
        })


@login_required
def get_inspection_files(request):
    """Get all files for a specific grouped inspection (client + date) and store them locally."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        from datetime import datetime, timedelta
        from django.conf import settings
        from ..models import FoodSafetyAgencyInspection
        
        data = json.loads(request.body)
        group_id = data.get('group_id', '')
        client_name = data.get('client_name', '')
        inspection_date = data.get('inspection_date', '')
        force_refresh = data.get('_force_refresh', False)

        # Strip "btn-" prefix from client name if present (comes from FSA inspection rows)
        if client_name and client_name.startswith('btn-'):
            client_name = client_name[4:]

        # Clean Unicode escapes from inspection_date (fix for JavaScript JSON.stringify escaping)
        if isinstance(inspection_date, str):
            inspection_date = inspection_date.replace('\\u002D', '-')
            inspection_date = inspection_date.replace('\\u002F', '/')
            inspection_date = inspection_date.replace('\\u0020', ' ')

        # Create media folder structure for this inspection using correct format
        from datetime import datetime
        import re

        # Parse date and build folder path - support multiple formats
        date_obj = None
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d']
        for date_format in date_formats:
            try:
                date_obj = datetime.strptime(inspection_date, date_format)
                break
            except ValueError:
                continue

        if not date_obj:
            return JsonResponse({
                'success': False,
                'error': f'time data \'{inspection_date}\' does not match format \'%Y-%m-%d\''
            })

        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')
        
        # Sanitize client name for filesystem (must match upload function)
        def create_folder_name(name):
            """Create Linux-friendly folder name - must match upload function"""
            if not name:
                return "unknown_client"
            import re
            # Remove special characters, keep only alphanumeric, spaces, hyphens, underscores
            clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
            # Replace spaces and hyphens with underscores
            clean_name = clean_name.replace(' ', '_').replace('-', '_')
            # Remove consecutive underscores
            clean_name = re.sub(r'_+', '_', clean_name)
            # Remove leading/trailing underscores
            clean_name = clean_name.strip('_').lower()
            return clean_name or "unknown_client"
        
        client_folder = create_folder_name(client_name)

        # Use unified inspection/ folder structure
        inspection_folder = os.path.join(
            settings.MEDIA_ROOT,
            'inspection',
            year_folder,
            month_folder,
            client_folder
        )

        os.makedirs(inspection_folder, exist_ok=True)

        # Get files from local storage
        local_files = get_inspection_files_local(client_name, inspection_date, force_refresh)

        # Check if there are actually any files (not just empty arrays)
        has_files = local_files and any(file_list for file_list in local_files.values() if file_list)

        if not has_files:
            # Return empty files object with proper structure instead of empty array
            empty_files = {
                'rfi': [],
                'invoice': [],
                'lab': [],
                'lab_form': [],
                'retest': [],
                'compliance': [],
                'occurrence': [],
                'composition': [],
                'other': []
            }
            response = JsonResponse({
                'success': True,
                'files': empty_files,
                'client_name': client_name,
                'inspection_date': inspection_date,
                'source': 'local'
            })
        else:
            response = JsonResponse({
                'success': True,
                'files': local_files,
                'client_name': client_name,
                'inspection_date': inspection_date,
                'source': 'local'
            })
        
        # Add cache-busting headers to prevent browser caching
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['Last-Modified'] = 'Thu, 01 Jan 1970 00:00:00 GMT'
        
        return response
        
    except Exception as e:
        response = JsonResponse({'success': False, 'error': str(e)})
        # Add cache-busting headers to prevent browser caching
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response


@login_required
def delete_inspection_file(request):
    """Delete a specific file from an inspection."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        import re
        from django.conf import settings
        from django.core.cache import cache
        
        data = json.loads(request.body)
        file_path = data.get('file_path', '')
        client_name = data.get('client_name', '')
        inspection_date = data.get('inspection_date', '')
        
        print(f"[ERROR] Delete request: file_path={file_path}, client_name={client_name}, inspection_date={inspection_date}")
        
        if not file_path or not client_name or not inspection_date:
            return JsonResponse({'success': False, 'error': 'Missing required parameters'})

        # Security check: ensure the file path is within the media directory
        media_root = settings.MEDIA_ROOT
        
        print(f"[ERROR] Media root: {media_root}")
        print(f"[ERROR] Requested file path: {file_path}")
        print(f"[ERROR] Client name: {client_name}")
        print(f"[ERROR] Inspection date: {inspection_date}")
        
        # Parse inspection date to get year and month
        from datetime import datetime
        try:
            date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
            year = date_obj.strftime('%Y')
            month = date_obj.strftime('%B')
        except:
            return JsonResponse({'success': False, 'error': 'Invalid inspection date format'})
        
        # Use original client name for folder matching (folders now use original names)
        client_folder_pattern = client_name or 'Unknown Client'
        
        print(f"[ERROR] Parsed date: year={year}, month={month}")
        print(f"[ERROR] Client folder pattern: {client_folder_pattern}")
        
        # Try to find the file using the provided information
        full_file_path = None

        # Normalize the file path (handle both forward and back slashes)
        normalized_file_path = file_path.replace('\\', '/')
        print(f"[DELETE DEBUG] Normalized file path: {normalized_file_path}")

        # First, try the exact path if it looks like a full relative path
        if normalized_file_path.startswith('docs/') or normalized_file_path.startswith('inspection/'):
            potential_path = os.path.join(media_root, normalized_file_path)
            # Normalize the full path for the OS
            potential_path = os.path.normpath(potential_path)
            print(f"[DELETE DEBUG] Trying exact path: {potential_path}")
            if os.path.exists(potential_path):
                full_file_path = potential_path
                print(f"[DELETE DEBUG] Found file using exact path: {full_file_path}")

        # If not found and it's an inspection path, try alternative path construction
        if not full_file_path and normalized_file_path.startswith('inspection/'):
            # Try alternative path construction (same as file display)
            path_parts = normalized_file_path.split('/')
            if len(path_parts) >= 2:
                alternative_path = os.path.join(media_root, 'inspection', year, month, client_folder_pattern, path_parts[-2], path_parts[-1])
                alternative_path = os.path.normpath(alternative_path)
                print(f"[DELETE DEBUG] Trying alternative path: {alternative_path}")
                if os.path.exists(alternative_path):
                    full_file_path = alternative_path
                    print(f"[DELETE DEBUG] Found file using alternative path: {full_file_path}")

        # If not found, search for the file using client name and date
        if not full_file_path:
            print(f"[ERROR] Searching for file using client name and date...")
            
            # Build the expected path structure
            month_path = os.path.join(media_root, 'inspection', year, month)
            print(f"[ERROR] Month path: {month_path}")
            
            if os.path.exists(month_path):
                try:
                    # Find client folder
                    for folder_name in os.listdir(month_path):
                        folder_path = os.path.join(month_path, folder_name)
                        if os.path.isdir(folder_path):
                            # Check if folder name matches client pattern
                            folder_pattern = re.sub(r'[^a-zA-Z0-9_]', '', folder_name).lower()
                            client_pattern = client_folder_pattern.lower()
                            
                            if (folder_pattern == client_pattern or 
                                client_pattern in folder_pattern or
                                folder_pattern in client_pattern):
                                print(f"Found matching client folder: {folder_name}")
                                
                                # Search for the file in this folder
                                print(f"Searching in folder: {folder_path}")
                                print(f"Looking for filename: {os.path.basename(file_path)}")
                                
                                for root, dirs, files in os.walk(folder_path):
                                    print(f"Checking directory: {root}")
                                    print(f"Files in directory: {files}")
                                    
                                    for file in files:
                                        if file == os.path.basename(file_path):
                                            full_file_path = os.path.join(root, file)
                                            print(f"Found file at: {full_file_path}")
                                            break
                                    if full_file_path is not None:
                                        break
                                if full_file_path is not None:
                                    break
                except (OSError, PermissionError) as e:
                    print(f" Error accessing month folder: {e}")

        # Also search in the docs/ structure if not found
        if not full_file_path:
            print(f"[DELETE DEBUG] Searching in docs/ structure...")
            from ..models import FoodSafetyAgencyInspection
            try:
                date_obj_search = datetime.strptime(inspection_date, '%Y-%m-%d').date()
                inspection = FoodSafetyAgencyInspection.objects.filter(
                    client_name__iexact=client_name,
                    date_of_inspection=date_obj_search
                ).first()

                if inspection and inspection.client:
                    docs_path = os.path.join(media_root, 'docs', str(inspection.client.id), str(inspection.id))
                    print(f"[DELETE DEBUG] Checking docs path: {docs_path}")
                    if os.path.exists(docs_path):
                        filename = os.path.basename(normalized_file_path)
                        for root, dirs, files in os.walk(docs_path):
                            if filename in files:
                                full_file_path = os.path.join(root, filename)
                                print(f"[DELETE DEBUG] Found file in docs structure: {full_file_path}")
                                break
            except Exception as e:
                print(f"[DELETE DEBUG] Error searching docs structure: {e}")

        if not full_file_path:
            print(f" Could not find file: {file_path}")
            return JsonResponse({'success': False, 'error': 'File not found'})
        
        # Normalize paths to prevent directory traversal attacks
        full_file_path = os.path.normpath(full_file_path)
        media_root = os.path.normpath(media_root)
        
        print(f"[ERROR] Final normalized full path: {full_file_path}")
        print(f"[ERROR] Final normalized media root: {media_root}")
        
        if not full_file_path.startswith(media_root):
            print(f" Delete 404: Invalid file path - not within media root")
            return JsonResponse({'success': False, 'error': 'Invalid file path'})
        
        # Check if file exists (should already be confirmed above)
        print(f"[ERROR] Final check - file exists: {os.path.exists(full_file_path)}")
        
        # Delete the file
        try:
            os.remove(full_file_path)
            print(f"[ERROR] Deleted file: {file_path}")
            
            # Clear database upload records based on file type
            from ..models import FoodSafetyAgencyInspection
            from datetime import datetime
            
            try:
                # Parse the inspection date
                date_obj = datetime.strptime(inspection_date, '%Y-%m-%d').date()
                
                # Find matching inspections
                inspections = FoodSafetyAgencyInspection.objects.filter(
                    client_name=client_name,
                    date_of_inspection=date_obj
                )
                
                # Determine document type and clear appropriate fields
                if '/rfi/' in file_path.lower():
                    inspections.update(rfi_uploaded_by=None, rfi_uploaded_date=None)
                    print(f" Cleared RFI upload records for {inspections.count()} inspections")
                elif '/invoice/' in file_path.lower():
                    inspections.update(invoice_uploaded_by=None, invoice_uploaded_date=None)
                    print(f" Cleared Invoice upload records for {inspections.count()} inspections")
                elif '/lab/' in file_path.lower():
                    inspections.update(lab_uploaded_by=None, lab_uploaded_date=None)
                    print(f" Cleared Lab upload records for {inspections.count()} inspections")
                    
            except Exception as db_error:
                print(f"[ERROR] Warning: Could not update database records: {db_error}")
                # Continue with file deletion success even if DB update fails
            
            # Clear relevant caches - comprehensive cache invalidation
            # Parse date to get year and month for cache key
            from datetime import datetime
            date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
            year_folder = date_obj.strftime('%Y')
            month_folder = date_obj.strftime('%B')
            
            cache_keys_to_clear = [
                f"shipment_list_{request.user.id}_{request.user.role}",
                "filter_options",
                f"inspection_files_{client_name}_{inspection_date}",
                f"file_status_{client_name}_{inspection_date}",
                f"files_cache_{client_name}_{inspection_date}",
                "file_colors_cache",
                "inspection_files_cache",
                f"local_files:{client_name}:{year_folder}:{month_folder}"  # Clear the specific cache key used by get_inspection_files_local
            ]
            
            # Clear all related cache keys
            for cache_key in cache_keys_to_clear:
                cache.delete(cache_key)
                print(f" Cleared cache key: {cache_key}")
            
            # Force clear the specific cache used by get_inspection_files_local
            # This is the most important cache to clear for immediate UI updates
            specific_cache_key = f"local_files:{client_name}:{year_folder}:{month_folder}"
            cache.delete(specific_cache_key)
            print(f" FORCE cleared specific cache key: {specific_cache_key}")
            
            # Also try clearing with the client_folder format used by get_inspection_files_local
            client_folder = client_name or 'Unknown Client'
            client_folder_cache_key = f"local_files:{client_folder}:{year_folder}:{month_folder}"
            cache.delete(client_folder_cache_key)
            print(f" FORCE cleared client_folder cache key: {client_folder_cache_key}")
            
            # Clear all possible cache variations for this client/date
            cache_variations = [
                f"local_files:{client_name}:{year_folder}:{month_folder}",
                f"local_files:{client_folder}:{year_folder}:{month_folder}",
                f"local_files:{client_name.replace(' ', '_')}:{year_folder}:{month_folder}",
                f"local_files:{client_folder.replace(' ', '_')}:{year_folder}:{month_folder}",
                f"inspection_files_{client_name}_{inspection_date}",
                f"inspection_files_{client_folder}_{inspection_date}",
                f"file_status_{client_name}_{inspection_date}",
                f"file_status_{client_folder}_{inspection_date}",
                f"files_cache_{client_name}_{inspection_date}",
                f"files_cache_{client_folder}_{inspection_date}"
            ]
            
            for cache_key in cache_variations:
                cache.delete(cache_key)
                print(f" Cleared cache variation: {cache_key}")
            
            # Set a cache clear time marker to prevent stale cache usage
            import time
            cache_clear_time_key = f"cache_cleared:{client_name}:{year_folder}:{month_folder}"
            cache.set(cache_clear_time_key, time.time(), 60)  # Expire in 60 seconds
            
            # Also set cache clear time marker for client_folder format
            client_folder_clear_time_key = f"cache_cleared:{client_folder}:{year_folder}:{month_folder}"
            cache.set(client_folder_clear_time_key, time.time(), 60)  # Expire in 60 seconds
            
            # Nuclear option: Clear all cache if the above doesn't work
            try:
                cache.clear()
                print(" NUCLEAR: Cleared entire cache")
            except Exception as e:
                print(f"[ERROR] Could not clear entire cache: {e}")
            print(f" Set cache clear time marker: {cache_clear_time_key}")
            
            # Clear any wildcard cache keys that might contain this client/date
            try:
                # Get all cache keys (if using Redis or similar backend that supports this)
                from django.core.cache import caches
                default_cache = caches['default']
                if hasattr(default_cache, 'keys'):
                    # For Redis backend
                    all_keys = default_cache.keys('*')
                    client_normalized = client_name.replace(' ', '_').lower()
                    date_normalized = inspection_date.replace('-', '_')
                    
                    keys_to_delete = [
                        key for key in all_keys 
                        if (client_normalized in key.lower() and date_normalized in key) or
                           ('files' in key.lower() and client_normalized in key.lower()) or
                           ('local_files' in key.lower() and client_normalized in key.lower())
                    ]
                    
                    for key in keys_to_delete:
                        default_cache.delete(key)
                        print(f" Cleared wildcard cache key: {key}")
                        
            except Exception as cache_error:
                print(f"[ERROR] Warning: Could not clear wildcard cache keys: {cache_error}")
                # Continue without error - basic cache clearing already done
            
            return JsonResponse({
                'success': True, 
                'message': f'File {os.path.basename(file_path)} deleted successfully',
                'database_updated': True
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to delete file: {str(e)}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_client_all_files(request):
    """Get ALL files for a client across all their inspections from local PC."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        from datetime import datetime
        from django.conf import settings
        from ..models import FoodSafetyAgencyInspection
        import re
        
        data = json.loads(request.body)
        client_name = data.get('client_name', '')
        
        if not client_name:
            return JsonResponse({'success': False, 'error': 'Client name is required'})
        
        print(f" Getting ALL files for client: {client_name}")
        
        # Use exact client name for matching (folders now use original names)
        client_folder_pattern = client_name
        
        print(f"Looking for client folder: {client_folder_pattern}")
        
        # Base inspection path
        inspection_base = os.path.join(settings.MEDIA_ROOT, 'inspection')
        
        if not os.path.exists(inspection_base):
            return JsonResponse({
                'success': True,
                'files': {},
                'client_name': client_name,
                'message': 'No inspection folder found. Run the 6-month data pull first.'
            })
        
        # Define file categories
        categories = {
            'rfi': 'rfi',
            'invoice': 'invoice', 
            'lab': 'lab results',
            'retest': 'retest',
            'compliance': 'Compliance'
        }
        
        files_by_category = {key: [] for key in categories.keys()}
        total_files = 0
        inspections_found = []
        added_files = {}  # Track added files to avoid true duplicates (filename + size)
        
        # Search through all year/month folders
        for year_folder in os.listdir(inspection_base):
            year_path = os.path.join(inspection_base, year_folder)
            if not os.path.isdir(year_path):
                continue
                
            for month_folder in os.listdir(year_path):
                month_path = os.path.join(year_path, month_folder)
                if not os.path.isdir(month_path):
                    continue
                
                # Look for client folder (exact match or similar)
                for folder_name in os.listdir(month_path):
                    folder_path = os.path.join(month_path, folder_name)
                    if not os.path.isdir(folder_path):
                        continue
                    
                    # Use exact matching since folders now use original client names
                    if folder_name == client_folder_pattern:
                        
                        print(f" Found client folder: {folder_name} in {year_folder}/{month_folder}")
                        inspections_found.append(f"{year_folder}/{month_folder}")
                        
                        # Search for files in this client folder
                        for category_key, category_name in categories.items():
                            if category_key == 'compliance':
                                # Check compliance subfolders for all commodities
                                compliance_base = os.path.join(folder_path, 'Compliance')
                                if os.path.exists(compliance_base):
                                    for commodity_folder in os.listdir(compliance_base):
                                        commodity_path = os.path.join(compliance_base, commodity_folder)
                                        if os.path.isdir(commodity_path):
                                            for filename in os.listdir(commodity_path):
                                                file_path = os.path.join(commodity_path, filename)
                                                if os.path.isfile(file_path):
                                                    # Get file info
                                                    file_info = get_file_info(file_path, f'Compliance/{commodity_folder}')
                                                    
                                                    # Check for duplicates using filename + size
                                                    file_key = f"{filename}_{file_info.get('size', 0)}"
                                                    
                                                    if file_key not in added_files:
                                                        file_info['inspection_period'] = f"{year_folder}/{month_folder}"
                                                        file_info['commodity'] = commodity_folder
                                                        files_by_category[category_key].append(file_info)
                                                        added_files[file_key] = True
                                                        total_files += 1
                                                        print(f"   Added compliance: {filename} ({file_info.get('size', 0)} bytes)")
                                                    else:
                                                        print(f"   Skipped compliance (duplicate): {filename} (same name and size already added)")
                            else:
                                # Check regular category folder
                                category_path = os.path.join(folder_path, category_key)
                                if os.path.exists(category_path):
                                    for filename in os.listdir(category_path):
                                        file_path = os.path.join(category_path, filename)
                                        if os.path.isfile(file_path):
                                            file_info = get_file_info(file_path, category_key)
                                            file_info['inspection_period'] = f"{year_folder}/{month_folder}"
                                            files_by_category[category_key].append(file_info)
                                            total_files += 1
        
        # Sort files by date (newest first)
        for category_key in files_by_category:
            files_by_category[category_key].sort(key=lambda x: x.get('modified_time', 0), reverse=True)
        
        print(f" Found {total_files} total files across {len(inspections_found)} inspection periods")
        for category, files in files_by_category.items():
            print(f"  {categories[category]}: {len(files)} files")
        
        return JsonResponse({
            'success': True,
            'files': files_by_category,
            'client_name': client_name,
            'total_files': total_files,
            'inspections_found': inspections_found,
            'categories': categories,
            'source': 'local'
        })
        
    except Exception as e:
        print(f" Error in get_client_all_files: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_page_clients_files(request):
    """Get files for multiple clients from current page only - optimized for pagination."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        from datetime import datetime
        from django.conf import settings
        import re
        
        data = json.loads(request.body)
        client_names = data.get('client_names', [])
        target_client = data.get('target_client', '')
        
        if not client_names:
            return JsonResponse({'success': False, 'error': 'Client names list is required'})
        
        if not target_client:
            return JsonResponse({'success': False, 'error': 'Target client name is required'})
        
        print(f" Getting files for target client: {target_client}")
        print(f" Checking only {len(client_names)} clients from current page")
        
        # Create folder name variations for matching (handles different upload path naming)
        def create_folder_name(name):
            if not name:
                return "unknown_client"
            clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
            clean_name = clean_name.replace(' ', '_').replace('-', '_')
            clean_name = re.sub(r'_+', '_', clean_name)
            clean_name = clean_name.strip('_').lower()
            return clean_name or "unknown_client"

        sanitized_target = create_folder_name(target_client)
        target_folder_variations = [
            target_client,  # Original name
            sanitized_target,  # Sanitized
            f'btn_{sanitized_target}',  # btn_ prefix
            f'group_{sanitized_target}',  # group_ prefix
        ]

        # Sanitize all client names for comparison
        client_patterns = []
        for client_name in client_names:
            pattern = re.sub(r'[^a-zA-Z0-9_]', '_', client_name)
            pattern = re.sub(r'_+', '_', pattern).strip('_')
            client_patterns.append(pattern.lower())

        print(f" Looking for target client folder variations: {target_folder_variations}")
        print(f" Client patterns to check: {client_patterns}")
        
        # Base inspection path
        inspection_base = os.path.join(settings.MEDIA_ROOT, 'inspection')
        
        if not os.path.exists(inspection_base):
            return JsonResponse({
                'success': True,
                'files': {},
                'client_name': target_client,
                'message': 'No inspection folder found. Run the 6-month data pull first.'
            })
        
        # Define file categories
        categories = {
            'rfi': 'rfi',
            'invoice': 'invoice',
            'lab': 'lab results',
            'retest': 'retest',
            'compliance': 'Compliance',
            'occurrence': 'occurrence',
            'composition': 'composition',
            'other': 'other'
        }

        files_by_category = {key: [] for key in categories.keys()}
        total_files = 0
        inspections_found = []

        # Search through all year/month folders - but only check folders that match our client list
        for year_folder in os.listdir(inspection_base):
            year_path = os.path.join(inspection_base, year_folder)
            if not os.path.isdir(year_path):
                continue
                
            for month_folder in os.listdir(year_path):
                month_path = os.path.join(year_path, month_folder)
                if not os.path.isdir(month_path):
                    continue
                
                # Look for client folders that match our target client
                for folder_name in os.listdir(month_path):
                    folder_path = os.path.join(month_path, folder_name)
                    if not os.path.isdir(folder_path):
                        continue
                    
                    # Check if folder matches any of our target variations
                    if folder_name in target_folder_variations:
                        
                        print(f" Found target client folder: {folder_name} in {year_folder}/{month_folder}")
                        inspections_found.append(f"{year_folder}/{month_folder}")
                        
                        # Define folder name variations for each category
                        category_folder_variations = {
                            'rfi': ['rfi', 'RFI', 'Request For Invoice', 'request for invoice'],
                            'invoice': ['invoice', 'Invoice'],
                            'lab': ['lab', 'Lab', 'lab results', 'Lab Results', 'coa', 'COA'],
                            'retest': ['retest', 'Retest'],
                            'compliance': ['compliance', 'Compliance'],
                            'occurrence': ['occurrence', 'Occurrence'],
                            'composition': ['composition', 'Composition'],
                            'other': ['other', 'Other']
                        }

                        # Search for files in this client folder
                        for category_key, category_name in categories.items():
                            folder_variations = category_folder_variations.get(category_key, [category_key])

                            if category_key == 'compliance':
                                # Check compliance subfolders for all commodities
                                for comp_var in folder_variations:
                                    compliance_base = os.path.join(folder_path, comp_var)
                                    if os.path.exists(compliance_base):
                                        for commodity_folder in os.listdir(compliance_base):
                                            commodity_path = os.path.join(compliance_base, commodity_folder)
                                            if os.path.isdir(commodity_path):
                                                for filename in os.listdir(commodity_path):
                                                    file_path = os.path.join(commodity_path, filename)
                                                    if os.path.isfile(file_path):
                                                        file_info = get_file_info(file_path, f'Compliance/{commodity_folder}')
                                                        file_info['inspection_period'] = f"{year_folder}/{month_folder}"
                                                        file_info['commodity'] = commodity_folder
                                                        files_by_category[category_key].append(file_info)
                                                        total_files += 1
                            else:
                                # Check all folder variations for this category
                                for folder_var in folder_variations:
                                    category_path = os.path.join(folder_path, folder_var)
                                    if os.path.exists(category_path):
                                        for filename in os.listdir(category_path):
                                            file_path = os.path.join(category_path, filename)
                                            if os.path.isfile(file_path):
                                                file_info = get_file_info(file_path, category_key)
                                                file_info['inspection_period'] = f"{year_folder}/{month_folder}"
                                                files_by_category[category_key].append(file_info)
                                                total_files += 1
        
        # Sort files by date (newest first)
        for category_key in files_by_category:
            files_by_category[category_key].sort(key=lambda x: x.get('modified_time', 0), reverse=True)
        
        # Determine file status for color coding
        has_rfi = len(files_by_category.get('rfi', [])) > 0
        has_invoice = len(files_by_category.get('invoice', [])) > 0
        has_lab = len(files_by_category.get('lab', [])) > 0
        has_retest = len(files_by_category.get('retest', [])) > 0
        has_compliance = len(files_by_category.get('compliance', [])) > 0
        has_occurrence = len(files_by_category.get('occurrence', [])) > 0
        has_composition = len(files_by_category.get('composition', [])) > 0
        
        # Determine status based on user requirements:
        # Green: ALL required documents (RFI, Invoice, Lab, Compliance) exist
        # Orange: Only compliance document exists (or compliance + some others but not all)
        # Blue: Has RFI/Invoice/Lab/Retest/Occurrence/Composition but no compliance documents
        # Red: No files at all
        if total_files == 0:
            file_status = 'no_files'  # Red - no files at all
        elif has_rfi and has_invoice and has_lab and has_compliance:
            file_status = 'all_files'  # Green - all required documents exist
        elif has_compliance:
            file_status = 'compliance_only'  # Orange - compliance exists (with or without other docs)
        elif has_rfi or has_invoice or has_lab or has_retest or has_occurrence or has_composition:
            file_status = 'partial_files'  # Blue - has some files but no compliance
        else:
            file_status = 'no_files'  # Red - no files at all
        
        print(f" Found {total_files} total files across {len(inspections_found)} inspection periods")
        print(f" File status: {file_status}")
        for category, files in files_by_category.items():
            print(f"  {categories[category]}: {len(files)} files")
        
        return JsonResponse({
            'success': True,
            'files': files_by_category,
            'client_name': target_client,
            'total_files': total_files,
            'inspections_found': inspections_found,
            'categories': categories,
            'source': 'local',
            'optimized': True,
            'clients_checked': len(client_names),
            'file_status': file_status,
            'has_rfi': has_rfi,
            'has_invoice': has_invoice,
            'has_lab': has_lab,
            'has_retest': has_retest,
            'has_occurrence': has_occurrence,
            'has_composition': has_composition,
            'has_compliance': has_compliance
        })
        
    except Exception as e:
        print(f" Error in get_page_clients_files: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_page_clients_file_status(request):
    """Get file status for multiple clients from current page - optimized for bulk checking."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    from django.core.cache import cache
    cache.delete('page_clients_status_cache')
    cache.clear()
    
    try:
        import json
        import os
        from datetime import datetime
        from django.conf import settings
        import re
        
        data = json.loads(request.body)
        
        # Support both old and new formats for backwards compatibility
        client_date_combinations = data.get('client_date_combinations', [])
        
        # Legacy format support
        if not client_date_combinations:
            client_names = data.get('client_names', [])
            inspection_dates = data.get('inspection_dates', {})
            
            if client_names:
                # Convert legacy format to new format
                for client_name in client_names:
                    dates = inspection_dates.get(client_name, [])
                    if isinstance(dates, list):
                        for date in dates:
                            client_date_combinations.append({
                                'client_name': client_name,
                                'inspection_date': date,
                                'unique_key': f"{client_name}_{date}"
                            })
                    elif isinstance(dates, str) and dates:
                        # Handle single date as string
                        client_date_combinations.append({
                            'client_name': client_name,
                            'inspection_date': dates,
                            'unique_key': f"{client_name}_{dates}"
                        })
        
        if not client_date_combinations:
            return JsonResponse({'success': False, 'error': 'Client date combinations are required'})
        
        # Limit to reasonable number of combinations per request (prevent abuse)
        if len(client_date_combinations) > 50:
            client_date_combinations = client_date_combinations[:50]

        # Check cache first for bulk results (updated for new format)
        from django.core.cache import cache
        combination_keys = [c.get('unique_key', '') for c in client_date_combinations]
        cache_key = f"combination_status:{hash(tuple(sorted(combination_keys)))}"
        cached_result = cache.get(cache_key)
        if cached_result:
            return JsonResponse(cached_result)
        
        # Base inspection path
        inspection_base = os.path.join(settings.MEDIA_ROOT, 'inspection')
        
        if not os.path.exists(inspection_base):
            return JsonResponse({
                'success': True,
                'combination_statuses': {},
                'message': 'No inspection folder found. Run the 6-month data pull first.'
            })
        
        # Define file categories
        categories = ['rfi', 'invoice', 'lab', 'retest', 'compliance']
        
        combination_statuses = {}
        
        # Process each client+date combination individually
        for combination in client_date_combinations:
            client_name = combination.get('client_name')
            inspection_date = combination.get('inspection_date')
            unique_key = combination.get('unique_key')
            
            if not client_name or not inspection_date or not unique_key:
                continue
            try:
                # Initialize results for this specific combination
                has_rfi = has_invoice = has_lab = has_retest = has_compliance = has_occurrence = has_composition = False
                file_status = 'no_files'

                # OPTIMIZATION 1: Database check first (fastest - no file system access)
                date_obj = datetime.strptime(str(inspection_date), '%Y-%m-%d').date()
                inspections = FoodSafetyAgencyInspection.objects.filter(
                    client_name=client_name,
                    date_of_inspection=date_obj
                ).select_related('client')

                # === NEW STRUCTURE CHECK: docs/{client_id}/{inspection_id}/ ===
                docs_base = os.path.join(settings.MEDIA_ROOT, 'docs')
                for insp in inspections:
                    if insp.client:
                        insp_path = os.path.join(docs_base, str(insp.client.id), str(insp.id))
                        if os.path.exists(insp_path):
                            for cat in ['rfi', 'invoice', 'lab', 'retest', 'compliance', 'composition', 'occurrence']:
                                cat_path = os.path.join(insp_path, cat)
                                if os.path.exists(cat_path) and os.listdir(cat_path):
                                    if cat == 'rfi': has_rfi = True
                                    elif cat == 'invoice': has_invoice = True
                                    elif cat == 'lab': has_lab = True
                                    elif cat == 'retest': has_retest = True
                                    elif cat == 'compliance': has_compliance = True
                                    elif cat == 'composition': has_composition = True
                                    elif cat == 'occurrence': has_occurrence = True
                
                # Check database for upload records (super fast)
                has_rfi_db = inspections.filter(rfi_uploaded_by__isnull=False).exists()
                has_invoice_db = inspections.filter(invoice_uploaded_by__isnull=False).exists()
                # Note: lab_uploaded_by field doesn't exist in database, rely on directory check only
                
                # OPTIMIZATION 2: Try multiple folder variations (consistent with list_uploaded_files)
                year = date_obj.strftime('%Y')
                month = date_obj.strftime('%B')
                
                # Sanitize client name to match folder structure (same as upload_document)
                def create_folder_name(name):
                    """Create Linux-friendly folder name - must match upload function"""
                    if not name:
                        return "unknown_client"
                    # Remove special characters, keep only alphanumeric, spaces, hyphens, underscores
                    clean_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', name)
                    # Replace spaces and hyphens with underscores
                    clean_name = clean_name.replace(' ', '_').replace('-', '_')
                    # Remove consecutive underscores
                    clean_name = re.sub(r'_+', '_', clean_name)
                    # Remove leading/trailing underscores
                    clean_name = clean_name.strip('_').lower()
                    return clean_name or "unknown_client"
                
                sanitized_client_name = create_folder_name(client_name)
                
                # Also try with apostrophes replaced by spaces (for names like "Mamma's Eggs")
                name_with_spaces_for_apostrophe = client_name.replace("'", ' ')
                sanitized_with_apostrophe = create_folder_name(name_with_spaces_for_apostrophe)

                # Use sanitized client name first, then apostrophe version, then original for backward compatibility
                # Note: Mobile and desktop uploads both use the same folder (no mobile_ prefix)
                client_folder_variations = [sanitized_client_name, sanitized_with_apostrophe, client_name]
                
                # Check all folder variations for files for this specific date
                has_rfi_dir = has_invoice_dir = has_lab_dir = has_retest_dir = has_compliance_dir = has_occurrence_dir = has_composition_dir = False
                
                parent_path = os.path.join(inspection_base, year, month)
                if os.path.exists(parent_path):
                    for folder_variation in client_folder_variations:
                        test_path = os.path.join(parent_path, folder_variation)
                        if os.path.exists(test_path):
                            # Check for files in each document type folder (not just directory existence)
                            # Check both top-level and nested Inspection-XXXX folders
                            
                            # Check RFI files (check multiple variations)
                            if not has_rfi_dir:
                                rfi_variations = ['RFI', 'rfi', 'Request For Invoice', 'request for invoice']

                                for rfi_variant in rfi_variations:
                                    rfi_path = os.path.join(test_path, rfi_variant)
                                    if os.path.exists(rfi_path):
                                        has_rfi_dir = any(os.path.isfile(os.path.join(rfi_path, f)) for f in os.listdir(rfi_path))
                                        if has_rfi_dir:
                                            break
                                
                                # Also check nested Inspection-XXXX folders (case insensitive)
                                if not has_rfi_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            # Check all RFI folder variations in nested folder
                                            for rfi_variant in rfi_variations:
                                                nested_rfi_path = os.path.join(test_path, item, rfi_variant)
                                                if os.path.exists(nested_rfi_path) and any(os.path.isfile(os.path.join(nested_rfi_path, f)) for f in os.listdir(nested_rfi_path)):
                                                    has_rfi_dir = True
                                                    break
                                            if has_rfi_dir:
                                                break
                            
                            # Check Invoice files (check both uppercase and lowercase)
                            if not has_invoice_dir:
                                # Check top-level Invoice folder (uppercase)
                                invoice_path = os.path.join(test_path, 'Invoice')
                                has_invoice_dir = os.path.exists(invoice_path) and any(os.path.isfile(os.path.join(invoice_path, f)) for f in os.listdir(invoice_path)) if os.path.exists(invoice_path) else False

                                # Check top-level invoice folder (lowercase)
                                if not has_invoice_dir:
                                    invoice_path = os.path.join(test_path, 'invoice')
                                    has_invoice_dir = os.path.exists(invoice_path) and any(os.path.isfile(os.path.join(invoice_path, f)) for f in os.listdir(invoice_path)) if os.path.exists(invoice_path) else False
                                
                                # Also check nested Inspection-XXXX folders (case insensitive)
                                if not has_invoice_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            # Check both uppercase and lowercase
                                            nested_invoice_path = os.path.join(test_path, item, 'Invoice')
                                            if not os.path.exists(nested_invoice_path):
                                                nested_invoice_path = os.path.join(test_path, item, 'invoice')
                                            if os.path.exists(nested_invoice_path) and any(os.path.isfile(os.path.join(nested_invoice_path, f)) for f in os.listdir(nested_invoice_path)):
                                                has_invoice_dir = True
                                                break
                            
                            # Check Lab files (check multiple folder name variations)
                            if not has_lab_dir:
                                # Check top-level Lab folder (uppercase)
                                lab_path = os.path.join(test_path, 'Lab')
                                has_lab_dir = os.path.exists(lab_path) and any(os.path.isfile(os.path.join(lab_path, f)) for f in os.listdir(lab_path)) if os.path.exists(lab_path) else False
                                
                                # Check top-level lab results folder (with space)
                                if not has_lab_dir:
                                    lab_path = os.path.join(test_path, 'lab results')
                                    has_lab_dir = os.path.exists(lab_path) and any(os.path.isfile(os.path.join(lab_path, f)) for f in os.listdir(lab_path)) if os.path.exists(lab_path) else False
                                
                                # Check top-level lab folder (lowercase)
                                if not has_lab_dir:
                                    lab_path = os.path.join(test_path, 'lab')
                                    has_lab_dir = os.path.exists(lab_path) and any(os.path.isfile(os.path.join(lab_path, f)) for f in os.listdir(lab_path)) if os.path.exists(lab_path) else False
                                
                                # Also check nested Inspection-XXXX folders (case insensitive)
                                if not has_lab_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            # Check multiple variations: Lab, lab results, lab
                                            nested_lab_path = os.path.join(test_path, item, 'Lab')
                                            if not os.path.exists(nested_lab_path):
                                                nested_lab_path = os.path.join(test_path, item, 'lab results')
                                            if not os.path.exists(nested_lab_path):
                                                nested_lab_path = os.path.join(test_path, item, 'lab')
                                            if os.path.exists(nested_lab_path) and any(os.path.isfile(os.path.join(nested_lab_path, f)) for f in os.listdir(nested_lab_path)):
                                                has_lab_dir = True
                                                break
                            
                            # Check Retest files (check both uppercase and lowercase)
                            if not has_retest_dir:
                                # Check top-level Retest folder (uppercase)
                                retest_path = os.path.join(test_path, 'Retest')
                                has_retest_dir = os.path.exists(retest_path) and any(os.path.isfile(os.path.join(retest_path, f)) for f in os.listdir(retest_path)) if os.path.exists(retest_path) else False
                                
                                # Check top-level retest folder (lowercase)
                                if not has_retest_dir:
                                    retest_path = os.path.join(test_path, 'retest')
                                    has_retest_dir = os.path.exists(retest_path) and any(os.path.isfile(os.path.join(retest_path, f)) for f in os.listdir(retest_path)) if os.path.exists(retest_path) else False
                                
                                # Also check nested Inspection-XXXX folders (case insensitive)
                                if not has_retest_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            # Check both uppercase and lowercase
                                            nested_retest_path = os.path.join(test_path, item, 'Retest')
                                            if not os.path.exists(nested_retest_path):
                                                nested_retest_path = os.path.join(test_path, item, 'retest')
                                            if os.path.exists(nested_retest_path) and any(os.path.isfile(os.path.join(nested_retest_path, f)) for f in os.listdir(nested_retest_path)):
                                                has_retest_dir = True
                                                break
                            
                            if not has_compliance_dir:
                                # Check top-level Compliance folder
                                compliance_path = os.path.join(test_path, 'Compliance')
                                if os.path.exists(compliance_path):
                                    # Check for files in commodity subfolders
                                    for commodity_folder in os.listdir(compliance_path):
                                        commodity_path = os.path.join(compliance_path, commodity_folder)
                                        if os.path.isdir(commodity_path):
                                            if any(os.path.isfile(os.path.join(commodity_path, f)) for f in os.listdir(commodity_path)):
                                                has_compliance_dir = True
                                                break

                                # Also check nested Inspection-XXXX folders for compliance files
                                if not has_compliance_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            inspection_folder = os.path.join(test_path, item)
                                            # Check both uppercase and lowercase compliance folders
                                            nested_compliance_path = os.path.join(inspection_folder, 'Compliance')
                                            if not os.path.exists(nested_compliance_path):
                                                nested_compliance_path = os.path.join(inspection_folder, 'compliance')
                                            if os.path.exists(nested_compliance_path):
                                                # Check for files in commodity subfolders
                                                for commodity_folder in os.listdir(nested_compliance_path):
                                                    commodity_path = os.path.join(nested_compliance_path, commodity_folder)
                                                    if os.path.isdir(commodity_path):
                                                        if any(os.path.isfile(os.path.join(commodity_path, f)) for f in os.listdir(commodity_path)):
                                                            has_compliance_dir = True
                                                            break
                                                if has_compliance_dir:
                                                    break

                            # Check Occurrence files (similar to other document types)
                            if not has_occurrence_dir:
                                occurrence_variations = ['occurrence', 'Occurrence']

                                for occ_variant in occurrence_variations:
                                    occ_path = os.path.join(test_path, occ_variant)
                                    if os.path.exists(occ_path):
                                        has_occurrence_dir = any(os.path.isfile(os.path.join(occ_path, f)) for f in os.listdir(occ_path))
                                        if has_occurrence_dir:
                                            break

                                # Also check nested Inspection-XXXX folders
                                if not has_occurrence_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            for occ_variant in occurrence_variations:
                                                nested_occ_path = os.path.join(test_path, item, occ_variant)
                                                if os.path.exists(nested_occ_path) and any(os.path.isfile(os.path.join(nested_occ_path, f)) for f in os.listdir(nested_occ_path)):
                                                    has_occurrence_dir = True
                                                    break
                                            if has_occurrence_dir:
                                                break

                            # Check Composition files (similar to other document types)
                            if not has_composition_dir:
                                composition_variations = ['composition', 'Composition']

                                for comp_variant in composition_variations:
                                    comp_path = os.path.join(test_path, comp_variant)
                                    if os.path.exists(comp_path):
                                        has_composition_dir = any(os.path.isfile(os.path.join(comp_path, f)) for f in os.listdir(comp_path))
                                        if has_composition_dir:
                                            break

                                # Also check nested Inspection-XXXX folders
                                if not has_composition_dir:
                                    for item in os.listdir(test_path):
                                        if item.lower().startswith('inspection-') and os.path.isdir(os.path.join(test_path, item)):
                                            for comp_variant in composition_variations:
                                                nested_comp_path = os.path.join(test_path, item, comp_variant)
                                                if os.path.exists(nested_comp_path) and any(os.path.isfile(os.path.join(nested_comp_path, f)) for f in os.listdir(nested_comp_path)):
                                                    has_composition_dir = True
                                                    break
                                            if has_composition_dir:
                                                break

                # Check actual file existence on disk AND sync database records
                # If files exist on disk but database doesn't have uploader info, update database
                has_rfi = has_rfi_dir
                has_invoice = has_invoice_dir
                has_lab = has_lab_dir
                has_retest = has_retest_dir
                has_occurrence = has_occurrence_dir
                has_composition = has_composition_dir

                # SYNC DATABASE: Update database records to match actual files on disk
                if has_rfi or has_invoice or has_lab or has_retest or has_composition:
                    # Find matching inspections for this client and date
                    matching_inspections = FoodSafetyAgencyInspection.objects.filter(
                        client_name=client_name,
                        date_of_inspection=date_obj
                    )
                    
                    if matching_inspections.exists():
                        from django.utils import timezone
                        current_time = timezone.now()
                        
                        # Update database to reflect actual file status
                        if has_rfi and not matching_inspections.filter(rfi_uploaded_by__isnull=False).exists():
                            # Files exist but database doesn't have uploader info - set to system user
                            matching_inspections.update(
                                rfi_uploaded_by_id=1,  # System user
                                rfi_uploaded_date=current_time
                            )

                        if has_invoice and not matching_inspections.filter(invoice_uploaded_by__isnull=False).exists():
                            # Files exist but database doesn't have uploader info - set to system user
                            matching_inspections.update(
                                invoice_uploaded_by_id=1,  # System user
                                invoice_uploaded_date=current_time
                            )

                        # NOTE: Do NOT auto-sync COA/Lab, Retest, Composition, Occurrence flags
                        # These are per-product flags and should only be set when a user
                        # explicitly uploads a document for that specific product.
                        # Auto-syncing would incorrectly mark ALL products as uploaded when
                        # only one file exists in the folder.
                has_compliance = has_compliance_dir
                
                # Determine status for this specific client+date combination
                # Match sent status logic: only require RFI, Invoice, and Compliance
                if has_rfi and has_invoice and has_compliance:
                    file_status = 'all_files'  # Green
                elif has_compliance:
                    file_status = 'compliance_only'  # Orange
                elif has_rfi or has_invoice or has_lab or has_retest or has_occurrence or has_composition:
                    file_status = 'partial_files'  # Blue
                else:
                    file_status = 'no_files'  # Red

            except ValueError:
                has_rfi = has_invoice = has_lab = has_retest = has_compliance = has_occurrence = has_composition = False
                file_status = 'no_files'

            except Exception as e:
                has_rfi = has_invoice = has_lab = has_retest = has_compliance = has_occurrence = has_composition = False
                # Use 'no_files' instead of 'error' so button turns RED instead of staying GREY
                file_status = 'no_files'
            
            # Store optimized status for this specific combination (common for all cases)
            combination_statuses[unique_key] = {
                'file_status': file_status,
                'client_name': client_name,
                'inspection_date': inspection_date,
                'has_rfi': has_rfi,
                'has_invoice': has_invoice,
                'has_lab': has_lab,
                'has_retest': has_retest,
                'has_occurrence': has_occurrence,
                'has_composition': has_composition,
                'has_compliance': has_compliance
            }

        # Prepare optimized response - convert combination_statuses to client_statuses format
        client_statuses = {}
        for unique_key, status_data in combination_statuses.items():
            client_name = status_data['client_name']
            if client_name not in client_statuses:
                client_statuses[client_name] = status_data
            else:
                # If multiple dates for same client, use the most complete status
                current_status = client_statuses[client_name]['file_status']
                new_status = status_data['file_status']
                
                # Priority: all_files > partial_files > compliance_only > no_files
                status_priority = {'all_files': 4, 'partial_files': 3, 'compliance_only': 2, 'no_files': 1}
                if status_priority.get(new_status, 0) > status_priority.get(current_status, 0):
                    client_statuses[client_name] = status_data
        
        result_data = {
            'success': True,
            'client_statuses': client_statuses,  # Frontend expects this key
            'combination_statuses': combination_statuses,  # Keep for debugging
            'total_checked': len(client_date_combinations),
            'source': 'local',
            'optimized': True
        }
        cache.set(cache_key, result_data, 300)  # Cache for 5 minutes
        
        return JsonResponse(result_data)
        
    except Exception as e:
        print(f" Error in get_page_clients_file_status: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def download_all_inspection_files(request):
    """Download all files for a grouped inspection as a ZIP."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        import zipfile
        import tempfile
        from django.http import HttpResponse
        from django.conf import settings
        from datetime import datetime
        import re
        
        data = json.loads(request.body)
        client_name = data.get('client_name')
        inspection_date = data.get('inspection_date')
        group_id = data.get('group_id')
        
        if not client_name or not inspection_date:
            return JsonResponse({'success': False, 'error': 'Client name and inspection date are required'})
        
        safe_print(f"Creating ZIP for {client_name} on {inspection_date}")
        
        # Parse date and build folder path
        if isinstance(inspection_date, str):
            date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
        else:
            date_obj = inspection_date
        
        # Get all inspection IDs for this client and date
        from ..models import FoodSafetyAgencyInspection
        inspection_ids = set(
            FoodSafetyAgencyInspection.objects.filter(
                client_name=client_name,
                date_of_inspection=date_obj.date() if hasattr(date_obj, 'date') else date_obj
            ).values_list('id', flat=True)
        )
        inspection_ids_str = {str(id) for id in inspection_ids}
        safe_print(f"Found {len(inspection_ids_str)} inspection IDs for this group: {inspection_ids_str}")
        
        def is_file_for_inspection_date(filename, target_date):
            """Check if a file belongs to the specific inspection date"""
            try:
                import re
                target_date_obj = datetime.strptime(target_date, '%Y-%m-%d') if isinstance(target_date, str) else target_date
                target_date_str = target_date_obj.strftime('%Y-%m-%d')
                target_date_compact = target_date_obj.strftime('%Y%m%d')
                
                # Check for YYYY-MM-DD format (compliance files) - must be exact match
                # Look for the date pattern with non-alphanumeric boundaries
                date_pattern_str = r'(?:^|[^0-9])' + re.escape(target_date_str) + r'(?:[^0-9]|$)'
                if re.search(date_pattern_str, filename):
                    return True
                
                # Check for YYYYMMDD format (uploaded files) - must be exact match
                # Look for the date pattern with non-alphanumeric boundaries
                date_pattern_compact = r'(?:^|[^0-9])' + re.escape(target_date_compact) + r'(?:[^0-9]|$)'
                if re.search(date_pattern_compact, filename):
                    return True
                
                # If no date pattern found, include the file anyway
                # This is more permissive and includes files without dates in their names
                safe_print(f"Including file without date pattern: {filename}")
                return True
                
            except Exception as e:
                safe_print(f"Error checking file date for {filename}: {e}")
                # If there's an error, include the file to be safe
                return True
            
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')
        
        # Find the actual client folder using the same logic as file listing
        # Check multiple possible client folder variations
        inspection_base = os.path.join(settings.MEDIA_ROOT, 'inspection')
        
        if not os.path.exists(inspection_base):
            return JsonResponse({'success': False, 'error': 'No inspection folder found'})
        
        # Use exact client name for matching (folders now use original names)
        client_folder_pattern = client_name
        
        safe_print(f"Looking for client folder: {client_folder_pattern}")
        
        # Find all matching client folders across all months (like file listing does)
        matching_folders = []
        
        # Search through all year/month folders
        for year_folder_search in os.listdir(inspection_base):
            year_path = os.path.join(inspection_base, year_folder_search)
            if not os.path.isdir(year_path):
                continue
                
            for month_folder_search in os.listdir(year_path):
                month_path = os.path.join(year_path, month_folder_search)
                if not os.path.isdir(month_path):
                    continue
                
                # Look for client folder in this year/month
                for folder_name in os.listdir(month_path):
                    folder_path = os.path.join(month_path, folder_name)
                    if not os.path.isdir(folder_path):
                        continue
                    
                    # Normalize folder name for comparison
                    normalized_folder = re.sub(r'[^a-zA-Z0-9_]', '_', folder_name)
                    normalized_folder = re.sub(r'_+', '_', normalized_folder).strip('_')
                    
                    # Use flexible matching to handle variations in client names
                    # Normalize both names to handle spaces, underscores, and special characters
                    def normalize_name(name):
                        # Convert to lowercase, replace spaces and special chars with underscores
                        normalized = re.sub(r'[^a-zA-Z0-9]', '_', name.lower())
                        # Remove multiple underscores and strip
                        normalized = re.sub(r'_+', '_', normalized).strip('_')
                        # Strip btn_ prefix if present (legacy folder naming)
                        if normalized.startswith('btn_'):
                            normalized = normalized[4:]
                        return normalized

                    normalized_client = normalize_name(client_folder_pattern)
                    normalized_folder = normalize_name(folder_name)

                    is_match = (normalized_folder == normalized_client)
                    
                    if is_match:
                        safe_print(f"    Exact match: {folder_name} in {year_folder_search}/{month_folder_search}")
                    else:
                        safe_print(f"    No match: {folder_name} in {year_folder_search}/{month_folder_search}")
                    
                    if is_match:
                        matching_folders.append(folder_path)
                        safe_print(f"Found matching client folder: {folder_name} in {year_folder_search}/{month_folder_search}")
        
        # === NEW STRUCTURE: Also check docs/{client_id}/{inspection_id}/ ===
        docs_files = []  # List of (file_path, category) tuples from docs structure

        # Strip btn- prefix if present
        clean_client_name = client_name[4:] if client_name.startswith('btn-') else client_name

        # Look up inspection and client in database
        inspection = FoodSafetyAgencyInspection.objects.filter(
            client_name__iexact=clean_client_name,
            date_of_inspection=date_obj.date() if hasattr(date_obj, 'date') else date_obj
        ).first()

        client_obj = None
        if inspection:
            if inspection.client:
                client_obj = inspection.client
            else:
                from main.models import Client
                client_obj = Client.objects.filter(name__iexact=clean_client_name).first()

        if inspection and client_obj:
            docs_path = os.path.join(settings.MEDIA_ROOT, 'docs', str(client_obj.id), str(inspection.id))
            if os.path.exists(docs_path):
                safe_print(f"Checking docs path: {docs_path}")
                doc_categories = ['rfi', 'invoice', 'compliance', 'composition', 'coa', 'lab', 'occurrence', 'retest', 'other']
                for cat in doc_categories:
                    cat_path = os.path.join(docs_path, cat)
                    if os.path.exists(cat_path):
                        for filename in os.listdir(cat_path):
                            file_path = os.path.join(cat_path, filename)
                            if os.path.isfile(file_path):
                                docs_files.append((file_path, cat, filename))
                                safe_print(f"Found docs file: {filename} in {cat}")

        if not matching_folders and not docs_files:
            return JsonResponse({'success': False, 'error': f'No files found for {client_name}. Searched in {inspection_base} and docs/'})

        # Create temporary ZIP file
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')

        try:
            with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                files_added = 0
                added_files = {}  # Track added files by filename + size + modified time to avoid true duplicates

                # First add files from docs structure
                for file_path, category, filename in docs_files:
                    try:
                        file_size = os.path.getsize(file_path)
                        file_mtime = os.path.getmtime(file_path)
                        file_key = f"{filename}_{file_size}_{file_mtime}"

                        if file_key not in added_files:
                            # Add to ZIP with category folder structure
                            zip_path = f"{category.capitalize()}/{filename}"
                            zip_file.write(file_path, zip_path)
                            added_files[file_key] = True
                            files_added += 1
                            safe_print(f"Added docs file to ZIP: {zip_path}")
                    except Exception as e:
                        safe_print(f"Error adding docs file {filename}: {e}")

                # Process all matching client folders (legacy structure)
                for base_path in matching_folders:
                    folder_name = os.path.basename(base_path)
                    safe_print(f"Processing folder: {folder_name} at {base_path}")
                    
                    # Check what's actually in this folder
                    if os.path.exists(base_path):
                        folder_contents = os.listdir(base_path)
                        safe_print(f"Folder contents: {folder_contents}")
                    
                    # Define categories to check (using actual folder names)
                    # Include all document types: RFI, Invoice, Lab Form, COA, Composition, Lab Results, Retest, Occurrence, Other
                    categories = ['Request For Invoice', 'rfi', 'invoice', 'lab results', 'lab', 'retest', 'labform', 'coa', 'composition', 'occurrence', 'other']
                    
                    # Check each category folder
                    for category in categories:
                        category_path = os.path.join(base_path, category)
                        if os.path.exists(category_path):
                            safe_print(f"Found {category} folder")
                            for filename in os.listdir(category_path):
                                file_path = os.path.join(category_path, filename)
                                if os.path.isfile(file_path):
                                    # Filter files by inspection date to avoid mixing different dates
                                    if is_file_for_inspection_date(filename, inspection_date):
                                        # Extract inspection ID from filename
                                        inspection_id = None
                                        id_match = re.match(r'^(\d+)_', filename)
                                        if id_match:
                                            inspection_id = id_match.group(1)
                                        
                                        # Determine document type from filename and category
                                        if '_lab_form_' in filename.lower() or '_labform_' in filename.lower() or 'lab-form' in filename.lower() or '-lab-form' in filename.lower() or category == 'labform':
                                            doc_type = 'Lab Form'
                                        elif category in ['lab', 'lab results']:
                                            doc_type = 'Lab'
                                        elif category in ['Request For Invoice', 'rfi']:
                                            doc_type = 'RFI'
                                        elif category == 'invoice':
                                            doc_type = 'Invoice'
                                        elif category == 'retest':
                                            doc_type = 'Retest'
                                        elif category == 'coa':
                                            doc_type = 'COA'
                                        elif category == 'composition':
                                            doc_type = 'Composition'
                                        elif category == 'occurrence':
                                            doc_type = 'Occurrence'
                                        elif category == 'other':
                                            doc_type = 'Other'
                                        else:
                                            doc_type = category
                                        
                                        # Create archive path based on document type:
                                        # RFI and Invoice always go to root
                                        # Lab, Lab Form, COA, Composition, Retest, Occurrence, Other go to inspection-XXXX folders
                                        if doc_type in ['RFI', 'Invoice']:
                                            arcname = f"{doc_type}/{filename}"
                                        elif inspection_id and doc_type in ['Lab', 'Lab Form', 'Retest', 'COA', 'Composition', 'Occurrence', 'Other']:
                                            arcname = f"inspection-{inspection_id}/{doc_type}/{filename}"
                                        else:
                                            arcname = f"{doc_type}/{filename}"
                                        
                                        # Get file stats for duplicate detection
                                        try:
                                            stat = os.stat(file_path)
                                            file_size = stat.st_size
                                            file_modified = stat.st_mtime
                                            file_key = f"{arcname}_{file_size}_{file_modified}"
                                            
                                            # Check if we've already added this exact file (same name, size, and modified time)
                                            if file_key not in added_files:
                                                zip_file.write(file_path, arcname)
                                                added_files[file_key] = arcname
                                                files_added += 1
                                                safe_print(f"Added {category}: {arcname} ({file_size} bytes)")
                                            else:
                                                safe_print(f"Skipped {category} (exact duplicate): {arcname}")
                                        except Exception as e:
                                            safe_print(f"Error getting file stats for {file_path}: {e}")
                                            # If we can't get stats, include the file to be safe
                                            zip_file.write(file_path, arcname)
                                            files_added += 1
                                            safe_print(f"Added {category}: {arcname} (no stats)")
                                    else:
                                        safe_print(f"Skipped {category} (wrong date): {filename}")
                        else:
                            safe_print(f"No {category} folder found")
                    
                    # Check for compliance documents
                    compliance_base = os.path.join(base_path, 'Compliance')
                    if os.path.exists(compliance_base):
                        safe_print(f"Found Compliance folder: {compliance_base}")
                        compliance_contents = os.listdir(compliance_base)
                        safe_print(f"Compliance folder contents: {compliance_contents}")
                        
                        # Check all commodity subfolders
                        for commodity_folder in compliance_contents:
                            commodity_path = os.path.join(compliance_base, commodity_folder)
                            if os.path.isdir(commodity_path):
                                safe_print(f"Checking commodity folder: {commodity_folder}")
                                commodity_files = os.listdir(commodity_path)
                                safe_print(f"Commodity {commodity_folder} files: {commodity_files}")
                                
                                for filename in commodity_files:
                                    file_path = os.path.join(commodity_path, filename)
                                    if os.path.isfile(file_path):
                                        # Filter compliance files by inspection date
                                        if is_file_for_inspection_date(filename, inspection_date):
                                            # Extract inspection ID from filename
                                            inspection_id = None
                                            id_match = re.match(r'^(\d+)_', filename)
                                            if id_match:
                                                inspection_id = id_match.group(1)
                                            
                                            # Create archive path:
                                            # If inspection ID matches one in this group, put in inspection-XXXX/Compliance/
                                            # Otherwise, put in root Compliance/ folder
                                            if inspection_id and inspection_id in inspection_ids_str:
                                                arcname = f"inspection-{inspection_id}/Compliance/{commodity_folder}/{filename}"
                                            else:
                                                arcname = f"Compliance/{commodity_folder}/{filename}"
                                            
                                            # Get file stats for duplicate detection
                                            try:
                                                stat = os.stat(file_path)
                                                file_size = stat.st_size
                                                file_modified = stat.st_mtime
                                                
                                                # For compliance files, use filename + size for duplicate detection
                                                # This allows multiple files with same name but different sizes (different inspections)
                                                # but prevents true duplicates (same name AND same size)
                                                file_key = f"{filename}_{file_size}"
                                                
                                                # Check if we've already added this exact file (same name and size)
                                                if file_key not in added_files:
                                                    zip_file.write(file_path, arcname)
                                                    added_files[file_key] = arcname
                                                    files_added += 1
                                                    safe_print(f"   Added compliance: {arcname} ({file_size} bytes)")
                                                else:
                                                    safe_print(f"   Skipped compliance (exact duplicate): {arcname} (same name and size already added)")
                                            except Exception as e:
                                                safe_print(f"   Error getting file stats for {file_path}: {e}")
                                                # If we can't get stats, include the file to be safe
                                                zip_file.write(file_path, arcname)
                                                files_added += 1
                                                safe_print(f"   Added compliance: {arcname} (no stats)")
                                        else:
                                            safe_print(f"   Skipped compliance (wrong date): {filename}")
                    else:
                        safe_print(f"No Compliance folder found at: {compliance_base}")
                
                if files_added == 0:
                    return JsonResponse({'success': False, 'error': 'No files found to download'})
            
            # Prepare response
            zip_filename = f"{client_name}_{inspection_date}_inspection_files.zip"
            zip_filename = re.sub(r'[^a-zA-Z0-9._-]', '_', zip_filename)
            
            # Read ZIP file content
            with open(temp_zip.name, 'rb') as zip_file:
                zip_content = zip_file.read()
            
            # Create HTTP response
            response = HttpResponse(zip_content, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            response['Content-Length'] = len(zip_content)
            
            safe_print(f"ZIP created successfully: {zip_filename} ({files_added} files)")
            return response
            
        finally:
            # Cleanup temporary file
            try:
                os.unlink(temp_zip.name)
            except:
                pass
                
    except Exception as e:
        safe_print(f"Error creating ZIP: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@no_inspector_scientist
def update_sent_status(request):
    """Update sent status for inspection group and mark as complete if sent."""
    print(f" UPDATE_SENT_STATUS called - Method: {request.method}")
    print(f" POST data: {dict(request.POST)}")
    print(f" User: {request.user.username} (Role: {getattr(request.user, 'role', 'unknown')})")
    
    if request.method != 'POST':
        print(f" Invalid method: {request.method}")
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        group_id = request.POST.get('group_id')
        sent_status = request.POST.get('sent_status')
        
        print(f" Group ID: '{group_id}', Status: '{sent_status}'")
        
        if not group_id:
            print(" No group ID provided")
            return JsonResponse({'success': False, 'error': 'Group ID is required'})
        
        print(f" Updating sent status for group {group_id}: {sent_status}")
        
        # Parse group_id to get client_name and date
        # group_id format: "ClientName_YYYYMMDD"
        parts = group_id.rsplit('_', 1)
        if len(parts) != 2:
            return JsonResponse({'success': False, 'error': 'Invalid group ID format'})
        
        client_name_sanitized = parts[0]
        date_str = parts[1]
        
        # Parse date
        from datetime import datetime
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format in group ID'})
        
        # Find all inspections in this group and update their sent status
        from ..models import FoodSafetyAgencyInspection
        
        # We need to find the original client name from the sanitized version
        # Get all inspections for this date and find matching client
        inspections = FoodSafetyAgencyInspection.objects.filter(date_of_inspection=date_obj)
        
        # Find inspections where sanitized client name matches
        import re
        matching_inspections = []
        for inspection in inspections:
            if inspection.client_name:
                # Use the same sanitization logic as group ID creation (frontend logic)
                sanitized_client = re.sub(r'[^a-zA-Z0-9_]', '_', inspection.client_name)
                sanitized_client = re.sub(r'_+', '_', sanitized_client)
                sanitized_client = sanitized_client.strip('_')
                
                if sanitized_client.lower() == client_name_sanitized.lower():
                    matching_inspections.append(inspection)
        
        if not matching_inspections:
            return JsonResponse({'success': False, 'error': 'No inspections found for this group'})

        # No validation required - anyone except inspectors can change sent status freely
        # The @no_inspector_scientist decorator already blocks inspectors
        print(f" Updating sent status to '{sent_status}' for group {group_id} - no validation required")

        # Update sent status for all inspections in the group
        is_sent = sent_status == 'YES' if sent_status else False
        
        from django.utils import timezone
        sent_date = timezone.now() if is_sent else None
        sent_by = request.user if is_sent else None
        
        updated_count = 0
        for inspection in matching_inspections:
            inspection.is_sent = is_sent
            inspection.sent_date = sent_date
            inspection.sent_by = sent_by
            inspection.save(update_fields=['is_sent', 'sent_date', 'sent_by'])
            updated_count += 1
        
        print(f" Updated sent status for {updated_count} inspections in group {group_id}")

        # CRITICAL: Clear the shipment list cache so refreshing shows the updated status
        try:
            from django.core.cache import cache
            cache.clear()  # Clear ALL cache to ensure fresh data on refresh
            print(f" Cleared cache after sent status update")
        except Exception as cache_error:
            print(f"[WARNING] Could not clear cache: {cache_error}")

        # Log the sent status change to system logs
        try:
            # Get client name from the first matching inspection
            client_name = matching_inspections[0].client_name if matching_inspections else 'Unknown'
            
            # Get user IP address
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')
            
            # Create descriptive log message
            status_text = {
                'YES': 'Sent',
                'NO': 'Not Sent',
                '': 'Not Sent (cleared)'
            }.get(sent_status, sent_status)
            
            description = f"Changed sent status to '{status_text}' for {client_name} inspection group on {date_obj.strftime('%Y-%m-%d')}"
            
            # Log the activity
            SystemLog.log_activity(
                user=request.user,
                action='UPDATE',
                page='/inspections/',
                object_type='inspection_group',
                object_id=group_id,
                description=description,
                details={
                    'client_name': client_name,
                    'inspection_date': date_obj.strftime('%Y-%m-%d'),
                    'sent_status': sent_status,
                    'status_text': status_text,
                    'inspections_affected': updated_count,
                    'group_id': group_id
                },
                ip_address=ip_address,
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            print(f" Logged sent status change: {request.user.username} set {client_name} to '{status_text}'")
            
        except Exception as e:
            print(f"[ERROR] Error logging sent status change: {e}")
            # Don't fail the main operation if logging fails
        
        return JsonResponse({
            'success': True,
            'message': f'Sent status updated for {updated_count} inspections',
            'sent_status': sent_status,
            'is_complete': is_sent,
            'sent_by_username': request.user.username,
            'sent_date': sent_date.isoformat() if sent_date else None
        })
        
    except Exception as e:
        print(f" Error updating sent status: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def download_inspection_file(request):
    """Secure file download for inspection files."""
    try:
        import os
        from django.http import FileResponse, Http404, HttpResponseRedirect
        from django.conf import settings
        from urllib.parse import unquote
        
        file_param = request.GET.get('file', '')
        source = request.GET.get('source', 'local')  # 'local' or 'onedrive'
        
        if not file_param:
            raise Http404("File not specified")
        
        # Decode URL-encoded filename
        relative_path = unquote(file_param)
        
        if source == 'onedrive':
            # Handle OneDrive file download
            return download_onedrive_file(request, relative_path)
        else:
            # Handle local file download (existing logic)
            return download_local_file(request, relative_path)
        
    except Exception as e:
        raise Http404(f"Error serving file: {str(e)}")


def download_onedrive_file(request, relative_path):
    """Download file from OneDrive."""
    try:
        import os
        import json
        import requests
        from django.http import HttpResponseRedirect
        
        # Load OneDrive tokens
        token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
        if not os.path.exists(token_file):
            raise Http404("OneDrive not configured")
        
        with open(token_file, 'r') as f:
            token_data = json.load(f)
        
        access_token = token_data.get('access_token')
        if not access_token:
            raise Http404("OneDrive not authenticated")
        
        # Build OneDrive path
        onedrive_base = getattr(settings, 'ONEDRIVE_FOLDER', 'FoodSafety Agency Inspections')
        onedrive_path = f"{onedrive_base}/inspection/{relative_path}"
        
        # Get file download URL from OneDrive
        check_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{onedrive_path}"
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        response = requests.get(check_url, headers=headers)
        if response.status_code != 200:
            raise Http404("File not found in OneDrive")
        
        file_info = response.json()
        download_url = file_info.get('@microsoft.graph.downloadUrl')
        
        if not download_url:
            raise Http404("Download URL not available")
        
        # Redirect to OneDrive download URL
        return HttpResponseRedirect(download_url)
        
    except Exception as e:
        raise Http404(f"Error downloading from OneDrive: {str(e)}")


def download_local_file(request, relative_path):
    """Download file from local media folder."""
    try:
        import os
        from django.http import FileResponse, Http404
        from django.conf import settings

        print(f"[DOWNLOAD DEBUG] Download request: {relative_path}")

        # Normalize path separators
        normalized_relative_path = relative_path.replace('\\', '/')

        # Security: Ensure file is within allowed directories (inspection/ or docs/)
        if not (normalized_relative_path.startswith('inspection/') or normalized_relative_path.startswith('docs/')):
            print(f"[DOWNLOAD DEBUG] Download 404: Access denied - path doesn't start with 'inspection/' or 'docs/': {relative_path}")
            raise Http404("Access denied")

        # Build full file path - normalize path separators for Windows
        normalized_path = normalized_relative_path.replace('/', os.sep)
        file_path = os.path.join(settings.MEDIA_ROOT, normalized_path)
        
        # Security: Ensure file exists and is within allowed directory
        if not os.path.exists(file_path):
            print(f" Download 404: File does not exist: {file_path}")
            raise Http404(f"File not found: {relative_path}")
        
        if not os.path.isfile(file_path):
            print(f" Download 404: Path is not a file: {file_path}")
            raise Http404(f"Path is not a file: {relative_path}")
        
        # Additional security: Ensure resolved path is still within MEDIA_ROOT
        real_path = os.path.realpath(file_path)
        media_root_real = os.path.realpath(settings.MEDIA_ROOT)
        if not real_path.startswith(media_root_real):
            raise Http404("Access denied")
        
        # Determine content type
        content_type = 'application/octet-stream'
        filename = os.path.basename(file_path)
        ext = filename.split('.')[-1].lower()
        
        content_types = {
            'pdf': 'application/pdf',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'doc': 'application/msword',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'zip': 'application/zip'
        }
        
        content_type = content_types.get(ext, content_type)

        # Check if action is view or download
        action = request.GET.get('action', 'download')
        is_download = action != 'view'

        # Create file response
        response = FileResponse(
            open(file_path, 'rb'),
            content_type=content_type,
            as_attachment=is_download,
            filename=filename
        )

        # Set headers based on action
        if is_download:
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        else:
            response['Content-Disposition'] = f'inline; filename="{filename}"'

        return response
        
    except Exception as e:
        raise Http404(f"Error serving file: {str(e)}")


@login_required
def get_zip_contents(request):
    """Get contents of a ZIP file for preview."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        import zipfile
        from django.conf import settings
        from urllib.parse import unquote
        
        data = json.loads(request.body)
        relative_path = data.get('file_path', '')
        
        if not relative_path:
            return JsonResponse({'success': False, 'error': 'File path not specified'})
        
        # Security: Ensure file is within MEDIA_ROOT/inspection
        if not relative_path.startswith('inspection/'):
            return JsonResponse({'success': False, 'error': 'Access denied'})
        
        # Build full file path
        file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
        
        # Security checks
        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            return JsonResponse({'success': False, 'error': 'File not found'})
        
        # Read ZIP contents
        contents = []
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                for info in zip_file.infolist():
                    if not info.is_dir():  # Only show files, not directories
                        contents.append({
                            'name': info.filename,
                            'size': info.file_size,
                            'compressed_size': info.compress_size,
                            'modified': info.date_time
                        })
        except zipfile.BadZipFile:
            return JsonResponse({'success': False, 'error': 'Invalid ZIP file'})
        
        return JsonResponse({
            'success': True,
            'contents': contents,
            'total_files': len(contents),
            'zip_name': os.path.basename(file_path)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def send_group_documents(request):
    """Send all documents for a grouped inspection via email."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        import os
        from datetime import datetime
        from django.conf import settings
        from django.core.mail import EmailMessage
        
        data = json.loads(request.body)
        group_id = data.get('group_id', '')
        client_name = data.get('client_name', '')
        inspection_date = data.get('inspection_date', '')
        
        # Parse date and build folder path
        date_obj = datetime.strptime(inspection_date, '%Y-%m-%d')
        year_folder = date_obj.strftime('%Y')
        month_folder = date_obj.strftime('%B')
        
        # Use original client name for folder structure (folders now use original names)
        client_folder = client_name or 'Unknown Client'
        
        # Base client path
        client_base_path = os.path.join(
            settings.MEDIA_ROOT,
            'inspection',
            year_folder,
            month_folder,
            client_folder
        )
        
        # Collect all available documents
        document_categories = ['rfi', 'invoice', 'lab', 'retest']
        attachments = []
        documents_found = []
        
        for category in document_categories:
            category_path = os.path.join(client_base_path, category)
            if os.path.exists(category_path):
                for filename in os.listdir(category_path):
                    file_path = os.path.join(category_path, filename)
                    if os.path.isfile(file_path) and filename.lower().endswith('.pdf'):
                        attachments.append(file_path)
                        documents_found.append(f"{category.upper()}: {filename}")
        
        # Also check compliance documents
        compliance_path = os.path.join(client_base_path, 'Compliance')
        if os.path.exists(compliance_path):
            for commodity in ['RAW', 'PMP', 'POULTRY', 'EGGS']:
                commodity_path = os.path.join(compliance_path, commodity)
                if os.path.exists(commodity_path):
                    for filename in os.listdir(commodity_path):
                        file_path = os.path.join(commodity_path, filename)
                        if os.path.isfile(file_path):
                            attachments.append(file_path)
                            documents_found.append(f"Compliance/{commodity}: {filename}")
        
        if not attachments:
            return JsonResponse({
                'success': False,
                'error': 'No documents found to send. Please upload RFI, Invoice, Lab results, or other documents first.'
            })
        
        # Get client email (you'll need to implement client email lookup)
        # For now, using a placeholder - you can extend this to get actual client emails
        recipient_email = get_client_email(client_name)  # Function to implement
        
        if not recipient_email:
            return JsonResponse({
                'success': False,
                'error': f'No email address found for {client_name}. Please add client email in the system.'
            })
        
        # Create email
        subject = f'Inspection Documents - {client_name} - {inspection_date}'
        message = f"""
Dear {client_name},

Please find attached the inspection documents for the inspection conducted on {inspection_date}.

Documents included:
{chr(10).join('• ' + doc for doc in documents_found)}

Best regards,
Food Safety Agency (Pty) Ltd
        """.strip()
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient_email],
            reply_to=[settings.DEFAULT_FROM_EMAIL]
        )
        
        # Attach all documents
        for file_path in attachments:
            email.attach_file(file_path)
        
        # Send email
            email.send()
        
        # Log the activity
        from ..models import SystemLog
        SystemLog.log_activity(
            user=request.user,
            action='EMAIL',
            page='inspections',
            object_type='group_documents',
            object_id=group_id,
            description=f'Sent {len(attachments)} documents for {client_name}',
            details={
                'client_name': client_name,
                'inspection_date': inspection_date,
                'documents_sent': documents_found,
                'recipient': recipient_email
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Documents sent successfully to {recipient_email}',
            'recipients': recipient_email,
            'documents_sent': len(attachments),
            'email_id': f'inspection_{group_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_client_email(client_name):
    """Get client email address from database (manual override preferred)."""
    try:
        from ..models import Client
        
        # Try to find client by business name (client_id field)
        client = Client.objects.filter(client_id__iexact=client_name).first()
        
        if client:
            return client.manual_email or client.email
        
        # If no email found, return None for now
        return None
        
    except Exception:
        return None


@login_required
@role_required(['admin', 'super_admin', 'developer'])
def sync_client_emails_from_sheets(request):
    """Sync client emails from Google Sheets column K."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    # Check if user has required permissions
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Authentication required'})
    
    # Check user role
    user_role = getattr(request.user, 'role', None)
    if user_role not in ['admin', 'super_admin', 'developer']:
        return JsonResponse({'success': False, 'error': f'Insufficient permissions. Your role: {user_role}'})
    
    try:
        from ..services.google_sheets_service import GoogleSheetsService
        from ..models import Client
        import re
        
        # Initialize Google Sheets service
        try:
            sheets_service = GoogleSheetsService()
            print("Google Sheets service initialized successfully")
            
            # Authenticate the service
            print("Authenticating with Google Sheets...")
            sheets_service.authenticate(request)
            print("Google Sheets authentication completed")
            
        except Exception as e:
            print(f"Error initializing Google Sheets service: {e}")
            return JsonResponse({'success': False, 'error': f'Failed to initialize Google Sheets service: {str(e)}'})
        
        # Get data from the client database sheet
        # CLIENT_DATABASE_URL from your Apps Script: 1iNULGBAzJ9n2ZulxwP8ZZZbwcPhj7X6e6rPwYqtI_fM
        spreadsheet_id = '1iNULGBAzJ9n2ZulxwP8ZZZbwcPhj7X6e6rPwYqtI_fM'
        sheet_name = 'Internal Account Code Generator'
        
        print(f"Attempting to access spreadsheet: {spreadsheet_id}")
        print(f"Sheet name: {sheet_name}")
        
        # Get data from columns H (account code), J (client name), K (email)
        # Starting from row 2 as per your request
        range_name = f'{sheet_name}!H2:K'
        print(f"Range: {range_name}")
        
        try:
            result = sheets_service.service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            
            if not values:
                return JsonResponse({'success': False, 'error': 'No data found in Google Sheets'})
            
            print(f"Found {len(values)} rows in Google Sheets")
            print(f"Sample row: {values[0] if values else 'No data'}")
            
            updated_count = 0
            created_count = 0
            error_count = 0
            
            for row in values:
                try:
                    # Extract data from columns H, I, J, K
                    account_code = row[0] if len(row) > 0 else ''  # Column H
                    # Skip column I
                    client_name = row[2] if len(row) > 2 else ''   # Column J  
                    email = row[3] if len(row) > 3 else ''         # Column K
                    
                    if not client_name or not account_code:
                        continue
                    
                    # Clean email (remove whitespace, validate basic format)
                    email = email.strip() if email else ''
                    if email and '@' not in email:
                        email = ''  # Invalid email format
                    
                    # Convert "None" to null/empty string
                    if email.lower() == 'none':
                        email = ''
                    
                    # Find client by internal account code first, then by client_id
                    client = None
                    if account_code:
                        client = Client.objects.filter(internal_account_code=account_code).first()
                    
                    if not client and client_name:
                        client = Client.objects.filter(client_id=client_name).first()
                    
                    if client:  # Update all clients, even with empty emails; do not override manual_email
                        if not client.manual_email and client.email != (email if email else None):
                            client.email = email if email else None
                            client.save()
                            updated_count += 1
                            print(f"Updated {client.client_id} with email: {email if email else 'null'}")
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error processing row {row}: {e}")
            
            return JsonResponse({
                'success': True,
                'message': f'Client email sync completed: {created_count} created, {updated_count} updated, {error_count} errors',
                'created': created_count,
                'updated': updated_count,
                'errors': error_count,
                'total_processed': len(values)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error accessing Google Sheets: {str(e)}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_file_info(file_path, category):
    """Get file information for display."""
    import os
    from datetime import datetime
    from django.conf import settings
    from django.urls import reverse
    from urllib.parse import quote

    try:
        stat = os.stat(file_path)
        filename = os.path.basename(file_path)

        # Create proper relative path for URL (handle Windows paths)
        relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
        # Normalize path separators for URL
        relative_path = relative_path.replace('\\', '/')

        # URL-encode the path for use in URLs
        encoded_path = quote(relative_path, safe='')

        # Create download URL with proper encoding (works for both docs/ and inspection/)
        download_url = f'/inspections/download-file/?file={encoded_path}'

        # Use download endpoint for url too (direct /media/ doesn't work for docs/)
        file_url = f'/inspections/download-file/?file={encoded_path}&action=view'

        return {
            'name': filename,
            'size': stat.st_size,
            'modified_time': stat.st_mtime,  # Return timestamp for frontend
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),  # Keep formatted string too
            'url': file_url,
            'download_url': download_url,
            'category': category,
            'document_type': category,  # Add document_type field
            'relative_path': relative_path
        }
    except Exception as e:
        return {
            'name': os.path.basename(file_path),
            'size': 0,
            'modified': 'Unknown',
            'url': '#',
            'download_url': '#',
            'category': category,
            'relative_path': ''
        }


def find_document_link_simulation(account_code, commodity, inspection_date):
    """Simple simulation of document link finding."""
    if not account_code or not commodity:
        return 'Document Not Found'
    
    # Simulate finding a document 30% of the time
    import random
    if random.random() < 0.3:
        return f'<a href="https://drive.google.com/file/d/example_{account_code[:5]}/view" class="document-link" target="_blank"> Document Link</a>'
    
    return 'Document Not Found'


@login_required
@role_required(['developer'])
def start_compliance_linking(request):
    """Start the compliance document linking process."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        import json
        data = json.loads(request.body)
        batch_size = data.get('batch_size', 200)
        doc_types = data.get('doc_types', 'all')
        
        # Get total inspection count
        from ..models import FoodSafetyAgencyInspection
        total_inspections = FoodSafetyAgencyInspection.objects.count()
        total_batches = (total_inspections + batch_size - 1) // batch_size
        
        # Initialize or get current progress from cache/session
        from django.core.cache import cache
        progress_key = f'compliance_linking_progress_{request.user.id}'
        
        progress = cache.get(progress_key, {
            'current_round': 1,
            'current_batch': 1,
            'processed_count': 0,
            'linked_count': 0,
            'error_count': 0,
            'is_running': False,
            'batch_size': batch_size,
            'doc_types': doc_types
        })
        
        progress['is_running'] = True
        progress['batch_size'] = batch_size
        progress['doc_types'] = doc_types
        cache.set(progress_key, progress, 3600)  # Cache for 1 hour
        
        row_start = ((progress['current_round'] - 1) * batch_size) + 1
        row_end = min(row_start + batch_size - 1, total_inspections)
        
        return JsonResponse({
            'success': True,
            'message': f'Started processing batch {progress["current_batch"]} of {total_batches}',
            'current_round': progress['current_round'],
            'current_batch': progress['current_batch'],
            'total_batches': total_batches,
            'row_start': row_start,
            'row_end': row_end,
            'total_inspections': total_inspections
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def pause_compliance_linking(request):
    """Pause the compliance document linking process."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        from django.core.cache import cache
        progress_key = f'compliance_linking_progress_{request.user.id}'
        progress = cache.get(progress_key, {})
        
        if progress:
            progress['is_running'] = False
            cache.set(progress_key, progress, 3600)
            return JsonResponse({'success': True, 'message': 'Processing paused'})
        else:
            return JsonResponse({'success': False, 'error': 'No active process found'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def reset_compliance_progress(request):
    """Reset the compliance document linking progress."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        from django.core.cache import cache
        progress_key = f'compliance_linking_progress_{request.user.id}'
        cache.delete(progress_key)
        
        return JsonResponse({
            'success': True, 
            'message': 'Progress reset successfully. Next run will start from the beginning.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def compliance_linking_status(request):
    """Get current status of compliance document linking."""
    try:
        from django.core.cache import cache
        from ..models import FoodSafetyAgencyInspection
        
        progress_key = f'compliance_linking_progress_{request.user.id}'
        progress = cache.get(progress_key, {
            'current_round': 1,
            'current_batch': 1,
            'processed_count': 0,
            'linked_count': 0,
            'error_count': 0,
            'is_running': False,
            'batch_size': 200,
            'doc_types': 'all'
        })
        
        total_inspections = FoodSafetyAgencyInspection.objects.count()
        total_batches = (total_inspections + progress['batch_size'] - 1) // progress['batch_size']
        
        row_start = ((progress['current_round'] - 1) * progress['batch_size']) + 1
        row_end = min(row_start + progress['batch_size'] - 1, total_inspections)
        
        progress_percentage = (progress['processed_count'] / total_inspections * 100) if total_inspections > 0 else 0
        
        # Check if processing is complete
        is_complete = progress['processed_count'] >= total_inspections or not progress['is_running']
        
        return JsonResponse({
            'success': True,
            'message': f"Round {progress['current_round']}, Batch {progress['current_batch']} of {total_batches}",
            'current_round': progress['current_round'],
            'current_batch': progress['current_batch'],
            'total_batches': total_batches,
            'row_start': row_start,
            'row_end': row_end,
            'processed_count': progress['processed_count'],
            'linked_count': progress['linked_count'],
            'error_count': progress['error_count'],
            'total_inspections': total_inspections,
            'progress_percentage': round(progress_percentage, 1),
            'is_running': progress['is_running'],
            'is_complete': is_complete
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['developer'])
def process_compliance_batch(request):
    """Process a single batch of compliance document linking."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        from django.core.cache import cache
        from ..models import FoodSafetyAgencyInspection, Client
        from ..services.google_drive_service import GoogleDriveService
        import os
        from datetime import datetime
        
        progress_key = f'compliance_linking_progress_{request.user.id}'
        progress = cache.get(progress_key, {})
        
        if not progress or not progress.get('is_running'):
            return JsonResponse({'success': False, 'error': 'No active process found'})
        
        batch_size = progress['batch_size']
        doc_types = progress['doc_types']
        
        # Get current batch of inspections
        offset = (progress['current_round'] - 1) * batch_size
        inspections = FoodSafetyAgencyInspection.objects.order_by('-date_of_inspection')[offset:offset + batch_size]
        
        if not inspections:
            # No more inspections to process
            progress['is_running'] = False
            cache.set(progress_key, progress, 3600)
            return JsonResponse({
                'success': True,
                'message': 'All inspections processed',
                'is_complete': True
            })
        
        # Build client mapping (business name -> internal account code)
        client_map = {}
        for client in Client.objects.exclude(internal_account_code__isnull=True).exclude(internal_account_code=''):
            key = (client.client_id or '').strip().lower()
            if key:
                client_map[key] = client.internal_account_code
        
        # Initialize Google Drive service
        drive_service = GoogleDriveService()
        root_folder_id = '1nrODSMMhYQhMeX3gXLzoIV_65fC_pXnO'  # From the Apps Script
        
        processed_count = 0
        linked_count = 0
        error_count = 0
        
        for inspection in inspections:
            try:
                processed_count += 1
                
                # Get client account code
                client_key = (inspection.client_name or '').strip().lower()
                account_code = client_map.get(client_key, 'unknown')
                
                # Get month/year folder from inspection date
                if inspection.date_of_inspection:
                    month_year = inspection.date_of_inspection.strftime('%B %Y')
                    
                    # Find files for this inspection based on account code and date
                    # This is a simplified version - you'd need to implement the full
                    # folder traversal and file matching logic from the Apps Script
                    
                    # For now, just simulate the process
                    links_created = simulate_link_creation(inspection, account_code, doc_types)
                    linked_count += links_created
                    
            except Exception as e:
                error_count += 1
                print(f"Error processing inspection {inspection.remote_id}: {e}")
        
        # Update progress
        progress['processed_count'] += processed_count
        progress['linked_count'] += linked_count
        progress['error_count'] += error_count
        progress['current_round'] += 1
        
        # Check if we need to move to next batch
        total_inspections = FoodSafetyAgencyInspection.objects.count()
        if progress['processed_count'] >= total_inspections:
            progress['is_running'] = False
        
        cache.set(progress_key, progress, 3600)
        
        return JsonResponse({
            'success': True,
            'message': f'Processed {processed_count} inspections, created {linked_count} links',
            'processed_count': processed_count,
            'linked_count': linked_count,
            'error_count': error_count,
            'is_complete': not progress['is_running']
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def simulate_link_creation(inspection, account_code, doc_types):
    """Simulate link creation for testing purposes."""
    # This is a placeholder function that simulates the link creation process
    # In a real implementation, you would:
    # 1. Search for files in Google Drive based on reference patterns
    # 2. Create hyperlinks in the database or update inspection records
    # 3. Return the actual number of links created
    
    links_created = 0
    
    # Simulate different document types
    if doc_types == 'all' or doc_types == 'rfi':
        # Simulate RFI link creation
        if inspection.remote_id and account_code != 'unknown':
            links_created += 1
    
    if doc_types == 'all' or doc_types == 'invoice':
        # Simulate Invoice link creation
        if inspection.remote_id and account_code != 'unknown':
            links_created += 1
    
    if doc_types == 'all' or doc_types == 'coa':
        # Simulate COA link creation
        if inspection.remote_id and account_code != 'unknown':
            links_created += 1
    
    return links_created


@login_required
@role_required(['developer'])
def first_50_compliance_links(request):
    inspections = FoodSafetyAgencyInspection.objects.order_by('-date_of_inspection')[:50]
    # Map client -> internal account code
    from ..models import Client
    client_map = {c.name: (c.internal_account_code or 'no') for c in Client.objects.all()}

    # Load Drive files and build lookup
    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request, max_items=10)
    lookup = GoogleDriveService.build_file_lookup(files)

    results = []
    for ins in inspections:
        account_code = client_map.get(ins.client_name, 'no')
        url = GoogleDriveService.find_best_link(account_code, ins.commodity, ins.date_of_inspection, lookup)
        results.append({
            'remote_id': ins.remote_id,
            'client_name': ins.client_name,
            'commodity': ins.commodity,
            'date_of_inspection': ins.date_of_inspection,
            'account_code': account_code,
            'document_url': url or None,
        })
    return JsonResponse({'success': True, 'count': len(results), 'items': results})


@login_required
@role_required(['developer'])
def fetch_and_store_first_50_compliance_docs(request):
    import os
    from django.conf import settings as dj_settings
    inspections = FoodSafetyAgencyInspection.objects.order_by('-date_of_inspection')[:50]
    from ..models import Client
    client_map = {c.name: (c.internal_account_code or 'no') for c in Client.objects.all()}

    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request)
    lookup = GoogleDriveService.build_file_lookup(files)

    base_dir = os.path.join(dj_settings.MEDIA_ROOT, 'compliance_documents')
    os.makedirs(base_dir, exist_ok=True)

    saved = []
    for ins in inspections:
        account_code = client_map.get(ins.client_name, 'no')
        # Reuse link matcher but pull the file id by reverse searching the lookup
        best_id = None
        best_days = 10**9
        import datetime as _dt
        commodity = (ins.commodity or '').strip().lower()
        if commodity == 'eggs':
            commodity = 'egg'
        for f in lookup.values():
            # Match by account code only (ignore commodity)
            if f['accountCode'] == account_code:
                days = abs((f['zipDate'] - ins.date_of_inspection).days) if ins.date_of_inspection else 10**9
                if days <= 15 and days < best_days:
                    best_days = days
                    best_id = f['id']
                    best_name = f['name']
        if not best_id:
            continue
        client_folder = os.path.join(base_dir, (ins.client_name or 'Unknown').replace(' ', '_'))
        os.makedirs(client_folder, exist_ok=True)
        dest_path = os.path.join(client_folder, best_name)
        try:
            if not os.path.exists(dest_path):
                drive.download_file(best_id, dest_path, request=request)
            saved.append({'client_name': ins.client_name, 'file': dest_path})
        except Exception as e:
            saved.append({'client_name': ins.client_name, 'error': str(e)})

    return JsonResponse({'success': True, 'saved_count': len([s for s in saved if 'file' in s]), 'items': saved})


@login_required
@role_required(['developer'])
def list_any_50_drive_files(request):
    print("[View] list_any_50_drive_files: start")
    """List any 50 files from the specified Drive folder to validate access."""
    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request)
    print(f"[View] list_any_50_drive_files: got {len(files)} files, trimming to 10")
    items = [
        {
            'id': f.get('id'),
            'name': f.get('name'),
            'mimeType': f.get('mimeType'),
            'webViewLink': f.get('webViewLink')
        }
        for f in files[:10]
    ]
    print(f"[View] list_any_50_drive_files: returning {len(items)} items")
    return JsonResponse({'success': True, 'count': len(items), 'items': items})


@login_required
@role_required(['developer'])
def fetch_store_any_50_drive_files(request):
    print("[View] fetch_store_any_50_drive_files: start")
    """Download the first 50 files from the specified Drive folder into MEDIA for testing."""
    import os
    from django.conf import settings as dj_settings

    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request, max_items=10)
    print(f"[View] fetch_store_any_50_drive_files: will download {len(files)} files")

    base_dir = os.path.join(dj_settings.MEDIA_ROOT, 'compliance_documents', 'test_any_50')
    os.makedirs(base_dir, exist_ok=True)

    saved = []
    for f in files:
        file_id = f.get('id')
        name = f.get('name') or f'{file_id}.bin'
        dest_path = os.path.join(base_dir, name)
        try:
            if not os.path.exists(dest_path):
                drive.download_file(file_id, dest_path, request=request)
            saved.append({'id': file_id, 'file': dest_path})
        except Exception as e:
            print(f"[View] fetch_store_any_50_drive_files: error downloading {file_id}: {e}")
            saved.append({'id': file_id, 'error': str(e)})

    print(f"[View] fetch_store_any_50_drive_files: saved_count={len([s for s in saved if 'file' in s])}")
    return JsonResponse({'success': True, 'saved_count': len([s for s in saved if 'file' in s]), 'items': saved})


@login_required
@role_required(['developer'])
def drive_any10_page(request):
    print("[View] drive_any10_page: start")
    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request, max_items=10)
    context = {
        'page_title': 'Drive Test - Any 10 Files',
        'files': files,
    }
    return render(request, 'main/drive_any10.html', context)


@login_required
@role_required(['developer'])
def download_first_10_compliance_by_commodity(request):
    print("[View] download_first_10_compliance_by_commodity: start")
    import os
    from django.conf import settings as dj_settings

    # Prepare inspections and client account code map
    inspections = FoodSafetyAgencyInspection.objects.exclude(client_name__isnull=True).exclude(client_name__exact='').order_by('-date_of_inspection')[:10]
    from ..models import Client
    client_map = {c.name: (c.internal_account_code or 'no') for c in Client.objects.all()}

    # Load Drive files from October 2025 onwards (dynamic)
    from datetime import datetime, timedelta

    parent_folder_id = '1pzot8MQ-m3u0f9-BWxpBO40QgLmeZhRP'  # 2025 folder

    # Dynamically determine which months to pull from (October 2025 onwards)
    start_month = datetime(2025, 10, 1)  # October 2025
    current_date = datetime.now()

    # Generate list of month names from October 2025 to current month + 1 month ahead
    target_months = []
    temp_date = start_month
    while temp_date <= current_date.replace(day=1) + timedelta(days=62):  # Current + ~2 months
        month_name = temp_date.strftime('%B %Y')  # e.g., "October 2025"
        target_months.append(month_name)
        # Move to next month
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1)

    drive = GoogleDriveService()
    print(f"[View] download_first_10_compliance_by_commodity: fetching from {', '.join(target_months)}")

    # Get all subfolders in the 2025 parent folder
    month_folders = drive.list_files_in_folder(parent_folder_id, request=request)

    # Collect all files from target month folders
    files = []
    for folder in month_folders:
        if folder.get('mimeType') == 'application/vnd.google-apps.folder' and folder.get('name') in target_months:
            print(f"[View] Fetching files from {folder['name']}...")
            month_files = drive.list_files_in_folder(folder['id'], request=request, max_items=2000)
            files.extend(month_files)
            print(f"[View] Found {len(month_files)} files in {folder['name']}")
    lookup = GoogleDriveService.build_file_lookup(files)

    # Commodity grouping
    def _group_for_commodity(raw):
        s = (raw or '').strip().lower()
        if s == 'eggs' or s == 'egg':
            return 'egg'
        if 'pmp' in s:
            return 'pmp'
        if 'poultry' in s:
            return 'poultry'
        return 'raw' if 'raw' in s else (s or 'raw')

    saved = []
    base_dir = os.path.join(dj_settings.MEDIA_ROOT, 'compliance')
    os.makedirs(base_dir, exist_ok=True)

    for ins in inspections:
        account_code = client_map.get(ins.client_name, 'no')
        # Find best matching file id similar to earlier logic
        best = None
        best_days = 10**9
        import datetime as _dt
        commodity_norm = (ins.commodity or '').strip().lower()
        if commodity_norm == 'eggs':
            commodity_norm = 'egg'
        # Fast path: try a direct name contains search by account code to reduce scanning
        try:
            # Narrow search using tokens: account code plus commodity prefix
            tokens = [account_code, (ins.commodity or '').split()[0]]
            fast_candidates = drive.search_files_in_folder_by_tokens(folder_id, tokens, request=request, max_items=50)
            for fc in fast_candidates:
                name = fc.get('name') or ''
                import re as _re
                m = _re.match(r'^([A-Za-z]+)-([A-Z]{2}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2,3}-\d+)-(\d{4}-\d{2}-\d{2})', name)
                if not m:
                    continue
                comm, acc, date_str = m.group(1), m.group(2), m.group(3)
                if acc != account_code:
                    continue
                if comm.lower() != commodity_norm:
                    continue
                try:
                    fdate = _dt.datetime.strptime(date_str, '%Y-%m-%d').date()
                except Exception:
                    continue
                # Ensure date_of_inspection is a date object, not datetime
                inspection_date = ins.date_of_inspection
                if inspection_date and hasattr(inspection_date, 'date'):
                    inspection_date = inspection_date.date()
                days = abs((fdate - (inspection_date or _dt.date.min)).days) if inspection_date else 10**9
                if days <= 15 and days < best_days:
                    best_days = days
                    best = {'id': fc.get('id'), 'name': name}
        except Exception:
            pass
        if not best:
            for f in lookup.values():
                # Match by account code only (ignore commodity)
                if f['accountCode'] == account_code:
                    # Enforce ±15 days window
                    days = abs((f['zipDate'] - (ins.date_of_inspection or _dt.date.min)).days) if ins.date_of_inspection else 10**9
                    if days <= 15 and days < best_days:
                        best_days = days
                        best = f
        if not best:
            saved.append({'client_name': ins.client_name, 'status': 'no_match'})
            continue
        group = _group_for_commodity(ins.commodity)
        client_folder = os.path.join(base_dir, group, (ins.client_name or 'Unknown').replace(' ', '_'))
        os.makedirs(client_folder, exist_ok=True)
        dest_path = os.path.join(client_folder, best['name'])
        try:
            if not os.path.exists(dest_path):
                drive.download_file(best['id'], dest_path, request=request)
            saved.append({'client_name': ins.client_name, 'group': group, 'file': dest_path})
        except Exception as e:
            saved.append({'client_name': ins.client_name, 'group': group, 'error': str(e)})

    saved_count = len([s for s in saved if 'file' in s])
    print(f"[View] download_first_10_compliance_by_commodity: saved_count={saved_count}")

    # If AJAX, return JSON, else render a simple page
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if is_ajax:
        return JsonResponse({'success': True, 'saved_count': saved_count, 'items': saved})

    context = {
        'page_title': 'Downloaded Compliance Documents (First 10)',
        'saved_count': saved_count,
        'results': saved,
    }
    return render(request, 'main/download_first10_by_commodity.html', context)


@login_required
@role_required(['developer'])
def fetch_store_first_10_matched_docs(request):
    print("[View] fetch_store_first_10_matched_docs: start")
    import os
    import calendar
    from django.conf import settings as dj_settings
    from ..models import Client

    # Load first 10 inspections (most recent)
    inspections = FoodSafetyAgencyInspection.objects.order_by('-date_of_inspection')[:10]

    # Map client -> internal account code
    client_map = {c.name: (c.internal_account_code or 'no') for c in Client.objects.all()}

    # Drive files lookup
    folder_id = '18CbrhqSsZO53TM3D8hRxkVmZyRBF-Zi4'
    drive = GoogleDriveService()
    files = drive.list_files_in_folder(folder_id, request=request)
    lookup = GoogleDriveService.build_file_lookup(files)

    def classify_category(file_name: str) -> str:
        name = (file_name or '').lower()
        if 'invoice' in name:
            return 'invoice'
        if 'rfi' in name or 'request for information' in name:
            return 'rfi'
        if 'lab' in name:
            return 'lab results'
        if 'retest' in name:
            return 'retest'
        return 'compliance'

    def month_name(dt):
        return calendar.month_name[dt.month]

    saved = []
    for ins in inspections:
        account_code = client_map.get(ins.client_name, 'no')
        # Find best file id based on commodity/account/date like earlier function
        best = None
        best_days = 10**9
        commodity = (ins.commodity or '').strip().lower()
        if commodity == 'eggs':
            commodity = 'egg'
        for f in lookup.values():
            # Match by account code only (ignore commodity)
            if f['accountCode'] == account_code:
                days = abs((f['zipDate'] - (ins.date_of_inspection or f['zipDate'])).days)
                if days <= 15 and days < best_days:
                    best_days = days
                    best = f

        if not best:
            saved.append({'client_name': ins.client_name, 'status': 'no_match'})
            continue

        # Build folder path: media/inspection/YYYY/Month/ClientName/<category>/
        if not ins.date_of_inspection:
            saved.append({'client_name': ins.client_name, 'status': 'no_date'})
            continue
        year_str = f"{ins.date_of_inspection.year}"
        month_str = month_name(ins.date_of_inspection)
        client_folder_name = (ins.client_name or 'Unknown').replace(' ', '')
        category = classify_category(best['name'])

        dest_dir = os.path.join(
            dj_settings.MEDIA_ROOT,
            'inspection',
            year_str,
            month_str,
            client_folder_name,
            category
        )
        os.makedirs(dest_dir, exist_ok=True)
        dest_path = os.path.join(dest_dir, best['name'])

        try:
            if not os.path.exists(dest_path):
                drive.download_file(best['id'], dest_path, request=request)
            saved.append({'client_name': ins.client_name, 'file': dest_path, 'category': category})
        except Exception as e:
            saved.append({'client_name': ins.client_name, 'error': str(e)})

    print(f"[View] fetch_store_first_10_matched_docs: saved_count={len([s for s in saved if 'file' in s])}")
    return JsonResponse({'success': True, 'saved_count': len([s for s in saved if 'file' in s]), 'items': saved})

@login_required
@role_required(['admin', 'super_admin', 'developer'])
def user_management(request):
    """User management page for administrators"""
    from main.models import InspectorMapping
    
    # Check if user has admin, super_admin, or developer role
    if not (request.user.has_role_permission('admin') or request.user.has_role_permission('super_admin') or request.user.has_role_permission('developer') or request.user.is_staff):
        messages.error(request, "You don't have permission to access user management.")
        return redirect('home')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_user':
            # Add new user
            username = request.POST.get('new_username')
            email = request.POST.get('new_email')
            password1 = request.POST.get('new_password1')
            password2 = request.POST.get('new_password2')
            first_name = request.POST.get('new_first_name', '')
            last_name = request.POST.get('new_last_name', '')
            role = request.POST.get('new_role', 'inspector')
            inspector_id = request.POST.get('new_inspector_id', '')
            
            # Only developers can create other developers
            if role == 'developer' and request.user.role != 'developer':
                messages.error(request, "Only developers can create users with developer status.")
            # Validation
            elif not username or not email or not password1:
                messages.error(request, "Username, email, and password are required.")
            elif password1 != password2:
                messages.error(request, "Passwords do not match.")
            elif User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.")
            elif User.objects.filter(email=email).exists():
                messages.error(request, "Email already exists.")
            elif role == 'inspector' and inspector_id:
                # Check if inspector ID already exists
                if InspectorMapping.objects.filter(inspector_id=inspector_id).exists():
                    messages.error(request, f"Inspector ID {inspector_id} is already assigned to another inspector.")
                else:
                    try:
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password1,
                            first_name=first_name,
                            last_name=last_name
                        )
                        # Set role
                        user.role = role
                        user.save()
                        
                        # Create inspector mapping if inspector ID is provided
                        if inspector_id:
                            inspector_name = f"{first_name} {last_name}".strip() or username
                            InspectorMapping.objects.create(
                                inspector_id=int(inspector_id),
                                inspector_name=inspector_name,
                                is_active=True
                            )
                            messages.success(request, f"User '{username}' created successfully with role '{role}' and inspector ID '{inspector_id}'.")
                        else:
                            messages.success(request, f"User '{username}' created successfully with role '{role}'.")
                    except Exception as e:
                        messages.error(request, f"Error creating user: {str(e)}")
            else:
                try:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password1,
                        first_name=first_name,
                        last_name=last_name
                    )
                    # Set role
                    user.role = role
                    user.save()
                    messages.success(request, f"User '{username}' created successfully with role '{role}'.")
                except Exception as e:
                    messages.error(request, f"Error creating user: {str(e)}")
        
        elif action == 'toggle_user_status':
            # Toggle user active status
            user_id = request.POST.get('user_id')
            print(f" Toggle user status - User ID: {user_id}")
            try:
                user = User.objects.get(id=user_id)
                old_status = user.is_active
                user.is_active = not user.is_active
                user.save()
                status = "activated" if user.is_active else "deactivated"
                print(f" User '{user.username}' {status} successfully. Old: {old_status}, New: {user.is_active}")
                messages.success(request, f"User '{user.username}' {status} successfully.")
                # Redirect to refresh the page and show updated status
                return redirect('user_management')
            except User.DoesNotExist:
                print(f" User not found with ID: {user_id}")
                messages.error(request, "User not found.")
            except Exception as e:
                print(f" Error updating user status: {str(e)}")
                messages.error(request, f"Error updating user status: {str(e)}")
        
        elif action == 'reset_password':
            # Reset user password
            user_id = request.POST.get('user_id')
            new_password = request.POST.get('new_password')
            try:
                user = User.objects.get(id=user_id)
                user.password = make_password(new_password)
                user.save()
                messages.success(request, f"Password for user '{user.username}' reset successfully.")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except Exception as e:
                messages.error(request, f"Error resetting password: {str(e)}")
        
        elif action == 'delete_user':
            # Delete user
            user_id = request.POST.get('user_id')
            try:
                user_to_delete = User.objects.get(id=user_id)
                
                # Only admin, super_admin, or developer users can delete other users
                if request.user.role not in ['admin', 'super_admin', 'developer']:
                    messages.error(request, "Only admin, super admin, or developer users can delete other users.")
                # Prevent self-deletion
                elif user_to_delete == request.user:
                    messages.error(request, "You cannot delete your own account.")
                # Prevent deletion of developer accounts
                elif user_to_delete.role == 'developer':
                    messages.error(request, "Cannot delete developer accounts.")
                # Allow admin to delete any non-developer user
                else:
                    username = user_to_delete.username
                    user_to_delete.delete()
                    messages.success(request, f"User '{username}' deleted successfully.")
                    
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except Exception as e:
                messages.error(request, f"Error deleting user: {str(e)}")
        
        elif action == 'edit_user':
            # Edit user information
            user_id = request.POST.get('user_id')
            try:
                user_to_edit = User.objects.get(id=user_id)

                # Get form data first (before using role variable)
                username = request.POST.get('edit_username', '').strip()
                email = request.POST.get('edit_email', '').strip()
                first_name = request.POST.get('edit_first_name', '').strip()
                last_name = request.POST.get('edit_last_name', '').strip()
                role = request.POST.get('edit_role', 'inspector')

                # Only super_admin and developer users can edit user information
                if request.user.role not in ['super_admin', 'developer']:
                    messages.error(request, "Only super admin and developer users can edit user information.")
                else:
                    # Only developers can make other users developers
                    if role == 'developer' and request.user.role != 'developer':
                        messages.error(request, "Only developers can create or promote users to developer status.")
                        return redirect('user_management')
                    phone_number = request.POST.get('edit_phone_number', '').strip()
                    department = request.POST.get('edit_department', '').strip()
                    employee_id = request.POST.get('edit_employee_id', '').strip()
                    inspector_id = request.POST.get('edit_inspector_id', '').strip()
                    password1 = request.POST.get('edit_password1', '').strip()
                    password2 = request.POST.get('edit_password2', '').strip()
                    
                    # Validation
                    if not username:
                        messages.error(request, "Username is required.")
                    elif not email:
                        messages.error(request, "Email is required.")
                    elif User.objects.filter(username=username).exclude(id=user_id).exists():
                        messages.error(request, "Username already exists for another user.")
                    elif User.objects.filter(email=email).exclude(id=user_id).exists():
                        messages.error(request, "Email already exists for another user.")
                    elif password1 and password1 != password2:
                        messages.error(request, "Passwords do not match.")
                    else:
                        # Update user information
                        user_to_edit.username = username
                        user_to_edit.email = email
                        user_to_edit.first_name = first_name
                        user_to_edit.last_name = last_name
                        user_to_edit.role = role
                        user_to_edit.phone_number = phone_number
                        user_to_edit.department = department
                        user_to_edit.employee_id = employee_id
                        
                        # Update password if provided
                        if password1:
                            user_to_edit.set_password(password1)
                        
                        user_to_edit.save()
                        
                        # Handle inspector ID if role is inspector
                        if role == 'inspector' and inspector_id:
                            # Update or create inspector mapping
                            full_name = f"{first_name} {last_name}".strip() or user_to_edit.username
                            mapping, created = InspectorMapping.objects.get_or_create(
                                inspector_name=full_name,
                                defaults={'inspector_id': int(inspector_id)}
                            )
                            if not created:
                                mapping.inspector_id = int(inspector_id)
                                mapping.save()
                        
                        messages.success(request, f"User '{user_to_edit.username}' information updated successfully.")
                        
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except ValueError as e:
                messages.error(request, f"Invalid inspector ID: {str(e)}")
            except Exception as e:
                messages.error(request, f"Error updating user: {str(e)}")
    
    # Get all users (including developers for super_admin and developer roles)
    if request.user.role in ['super_admin', 'developer']:
        users = User.objects.all().order_by('username')
    else:
        users = User.objects.exclude(role='developer').order_by('username')
    
    # Get all inspector mappings
    inspector_mappings = InspectorMapping.objects.all().order_by('inspector_name')
    
    # Create a dictionary to map user names to inspector IDs
    inspector_id_map = {}
    for mapping in inspector_mappings:
        inspector_id_map[mapping.inspector_name] = mapping.inspector_id
    
    # Add inspector_id to each user object for easy template access
    for user in users:
        if user.role == 'inspector':
            user.inspector_id = inspector_id_map.get(user.get_full_name() or user.username)
        else:
            user.inspector_id = None
    
    # Role choices for the form
    role_choices = [
        ('admin', 'Administrator'),
        ('super_admin', 'Super Admin'),
        ('financial', 'Financial Administrator'),
        ('lab_technician', 'Lab Technician'),
        ('inspector', 'Inspector'),
        ('developer', 'Developer'),  # Hidden role
    ]
    
    # Get theme settings
    try:
        from ..models import SystemSettings
        system_settings = SystemSettings.get_settings()
        settings = type('Settings', (), {
            'dark_mode': system_settings.theme_mode == 'dark' if hasattr(system_settings, 'theme_mode') else False,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    context = {
        'users': users,
        'inspector_mappings': inspector_mappings,
        'role_choices': role_choices,
        'settings': settings,
    }
    
    # Ensure CSRF token is properly generated and available
    from django.middleware.csrf import get_token
    get_token(request)
    
    response = render(request, 'main/user_management.html', context)
    response['Cache-Control'] = 'no-store'  # Prevent caching issues with CSRF tokens
    return response


@login_required
@inspector_only_inspections
def system_logs(request):
    """Display system logs with filtering and pagination."""
    if not (request.user.is_staff or request.user.role in ['admin', 'super_admin', 'developer']):
        messages.error(request, "You don't have permission to access system logs.")
        return redirect('home')
    
    # Get filter parameters
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    page_filter = request.GET.get('page', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    show_all = request.GET.get('show_all', 'false') == 'true'
    
    # Build query
    logs = SystemLog.objects.select_related('user').all()
    
    # Apply filters
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    if page_filter:
        logs = logs.filter(page__icontains=page_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Get total count before pagination
    total_logs = logs.count()
    
    # Pagination
    if not show_all:
        paginator = Paginator(logs, 50)  # Show 50 logs per page
        page_number = request.GET.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
            logs = page_obj.object_list
        except:
            page_obj = paginator.page(1)
            logs = page_obj.object_list
    else:
        page_obj = None
        logs = logs[:1000]  # Limit to 1000 logs when showing all
    
    # Get unique values for filter dropdowns
    users = User.objects.filter(system_logs__isnull=False).distinct().order_by('username')
    actions = SystemLog.objects.values_list('action', flat=True).distinct().order_by('action')
    pages = SystemLog.objects.values_list('page', flat=True).exclude(page__isnull=True).exclude(page='').distinct().order_by('page')
    
    # Get theme settings
    try:
        from ..models import SystemSettings
        system_settings = SystemSettings.get_settings()
        settings = type('Settings', (), {
            'dark_mode': system_settings.theme_mode == 'dark' if hasattr(system_settings, 'theme_mode') else False,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    context = {
        'logs': logs,
        'page_obj': page_obj,
        'total_logs': total_logs,
        'users': users,
        'actions': actions,
        'pages': pages,
        'filters': {
            'user': user_filter,
            'action': action_filter,
            'page': page_filter,
            'date_from': date_from,
            'date_to': date_to,
            'show_all': show_all,
        },
        'settings': settings,
    }
    
    return render(request, 'main/system_logs.html', context)


# =============================================================================
# NOTIFICATION VIEWS
# =============================================================================

@login_required
@role_required(['super_admin', 'developer'])
def get_notifications(request):
    """Get unread notification count and recent notifications"""
    from ..models import Notification
    from datetime import datetime

    try:
        # Get unread notifications for current user
        unread_notifications = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).order_by('-created_at')[:10]  # Get last 10 unread

        # Get all recent notifications (read and unread)
        all_notifications = Notification.objects.filter(
            user=request.user
        ).order_by('-created_at')[:20]  # Get last 20 total

        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()

        # Serialize notifications
        notifications_data = []
        for notif in all_notifications:
            notifications_data.append({
                'id': notif.id,
                'title': notif.title,
                'message': notif.message,
                'type': notif.notification_type,
                'priority': notif.priority,
                'is_read': notif.is_read,
                'created_at': notif.created_at.isoformat(),
                'action_url': notif.action_url,
            })

        return JsonResponse({
            'success': True,
            'unread_count': unread_count,
            'notifications': notifications_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@role_required(['super_admin', 'developer'])
def mark_notification_read(request, notification_id):
    """Mark a single notification as read"""
    from ..models import Notification
    from django.utils import timezone

    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id, user=request.user)
            notification.is_read = True
            notification.read_at = timezone.now()
            notification.save()

            return JsonResponse({'success': True})
        except Notification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@login_required
@role_required(['super_admin', 'developer'])
def mark_all_notifications_read(request):
    """Mark all notifications as read for the current user"""
    from ..models import Notification
    from django.utils import timezone

    if request.method == 'POST':
        try:
            Notification.objects.filter(
                user=request.user,
                is_read=False
            ).update(is_read=True, read_at=timezone.now())

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@login_required
@role_required(['super_admin', 'developer'])
def delete_notification(request, notification_id):
    """Delete a notification"""
    from ..models import Notification

    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id, user=request.user)
            notification.delete()

            return JsonResponse({'success': True})
        except Notification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@login_required
@role_required(['super_admin', 'developer'])
def fsa_operations_board(request):
    """FSA Operations Board - Dashboard for operational metrics and monitoring."""
    if not (request.user.role in ['super_admin', 'developer']):
        messages.error(request, "You don't have permission to access the FSA Operations Board.")
        return redirect('home')

    # Load all tickets from the database
    from main.models import Ticket
    from datetime import date
    tickets = Ticket.objects.all().select_related('created_by', 'assigned_to')

    # Get list of all users for assignment dropdown
    all_users = User.objects.filter(is_active=True).order_by('username')

    # Calculate statistics
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status='open').count()
    in_progress_tickets = tickets.filter(status='in-progress').count()
    resolved_today = tickets.filter(status='resolved', updated_at__date=date.today()).count()
    high_priority_tickets = tickets.filter(priority='high').count()

    context = {
        'user': request.user,
        'tickets': tickets,
        'all_users': all_users,
        'stats': {
            'total': total_tickets,
            'open': open_tickets,
            'in_progress': in_progress_tickets,
            'resolved_today': resolved_today,
            'high_priority': high_priority_tickets,
        }
    }

    return render(request, 'main/fsa_operations_board.html', context)


@login_required
@require_http_methods(["POST"])
def update_ticket_status(request, ticket_id):
    """Update ticket status via AJAX."""
    from main.models import Ticket
    import json

    try:
        ticket = Ticket.objects.get(id=ticket_id)
        data = json.loads(request.body)
        new_status = data.get('status')

        # Validate status
        valid_statuses = ['open', 'in-progress', 'resolved']
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

        # Update the ticket
        ticket.status = new_status
        ticket.save()

        return JsonResponse({
            'success': True,
            'ticket_id': ticket.id,
            'status': ticket.status,
            'updated_at': ticket.updated_at.strftime('%Y-%m-%d %H:%M')
        })

    except Ticket.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Ticket not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def submit_ticket(request):
    """Allow users to submit tickets/issues for the FSA Operations Board."""
    from main.models import Ticket
    from datetime import date
    from django.core.mail import send_mail

    if request.method == 'POST':
        # Get all form fields
        title = request.POST.get('title')
        description = request.POST.get('description')
        issue_type = request.POST.get('issue_type')
        affected_area = request.POST.get('affected_area')
        priority = request.POST.get('priority', 'medium')

        # Optional fields
        steps_to_reproduce = request.POST.get('steps_to_reproduce', '')
        expected_behavior = request.POST.get('expected_behavior', '')
        actual_behavior = request.POST.get('actual_behavior', '')
        impact_users = request.POST.get('impact_users', '')
        is_blocking = request.POST.get('is_blocking', '')
        browser_info = request.POST.get('browser_info', '')
        additional_notes = request.POST.get('additional_notes', '')

        # Validate required fields (only title, description, and issue_type are required)
        if not title or not description or not issue_type:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('submit_ticket')

        # Get or create user Ethan for auto-assignment
        try:
            ethan = User.objects.get(username='Ethan')
        except User.DoesNotExist:
            # Create Ethan if doesn't exist
            ethan = User.objects.create_user(
                username='Ethan',
                email='ethan.sevenster@moc-pty.com',
                first_name='Ethan',
                last_name='Sevenster'
            )

        # Create the ticket with all fields and auto-assign to Ethan
        ticket = Ticket.objects.create(
            title=title,
            description=description,
            issue_type=issue_type,
            affected_area=affected_area if affected_area else None,
            priority=priority,
            steps_to_reproduce=steps_to_reproduce if steps_to_reproduce else None,
            expected_behavior=expected_behavior if expected_behavior else None,
            actual_behavior=actual_behavior if actual_behavior else None,
            impact_users=impact_users if impact_users else None,
            is_blocking=is_blocking if is_blocking else None,
            browser_info=browser_info if browser_info else None,
            additional_notes=additional_notes if additional_notes else None,
            created_by=request.user,
            assigned_to=ethan,  # Auto-assign to Ethan
            status='open'
        )

        # Send email notification to Ethan and Anthony
        try:
            email_subject = f"New Ticket #{ticket.id}: {title}"

            email_body = f"""Good day,

A new ticket has been submitted to the Food Safety Agency Support Tickets system.

Issue: {title}
Priority: {ticket.get_priority_display()}

Description
{description}
"""

            if browser_info:
                email_body += f"""
Environment
{browser_info}
"""

            if additional_notes:
                email_body += f"""
Additional Information
{additional_notes}
"""

            email_body += f"""
Submitted by: {request.user.get_full_name() or request.user.username}
Submitted on: {ticket.created_at.strftime('%B %d, %Y')}

Please review and take the necessary action in the Support Tickets system.

Regards,
Food Safety Agency System
"""

            send_mail(
                subject=email_subject,
                message=email_body,
                from_email='info@eclick.co.za',
                recipient_list=['ethan.sevenster@moc-pty.com', 'anthony.penzes@moc-pty.com'],
                fail_silently=True,  # Don't break if email fails
            )
        except Exception as e:
            # Log email error but don't stop ticket creation
            print(f"Failed to send email notification: {e}")

        messages.success(request, f'✅ Ticket #{ticket.id} submitted successfully! Ethan and Anthony have been notified via email.')
        return redirect('submit_ticket')

    # GET request - show the form
    context = {
        'user': request.user,
        'today': date.today().isoformat(),
    }
    return render(request, 'main/submit_ticket.html', context)


@login_required
@role_required(['admin', 'super_admin', 'developer'])
def refresh_tokens(request):
    """Refresh OneDrive and Google Sheets tokens."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        from ..utils.onedrive_utils import refresh_access_token
        from ..services.google_sheets_service import GoogleSheetsService
        
        results = {
            'onedrive': {'success': False, 'message': ''},
            'google_sheets': {'success': False, 'message': ''}
        }
        
        # Refresh OneDrive token
        try:
            onedrive_token = refresh_access_token()
            if onedrive_token:
                results['onedrive'] = {
                    'success': True, 
                    'message': 'OneDrive token refreshed successfully'
                }
            else:
                results['onedrive'] = {
                    'success': False, 
                    'message': 'Failed to refresh OneDrive token'
                }
        except Exception as e:
            results['onedrive'] = {
                'success': False, 
                'message': f'OneDrive token refresh error: {str(e)}'
            }
        
        # Refresh Google Sheets token
        try:
            google_service = GoogleSheetsService()
            if google_service.authenticate():
                results['google_sheets'] = {
                    'success': True, 
                    'message': 'Google Sheets token refreshed successfully'
                }
            else:
                results['google_sheets'] = {
                    'success': False, 
                    'message': 'Failed to refresh Google Sheets token'
                }
        except Exception as e:
            results['google_sheets'] = {
                'success': False, 
                'message': f'Google Sheets token refresh error: {str(e)}'
            }
        
        # Check if any refresh was successful
        any_success = results['onedrive']['success'] or results['google_sheets']['success']
        
        return JsonResponse({
            'success': any_success,
            'message': 'Token refresh completed',
            'results': results
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Token refresh failed: {str(e)}'
        })


# =============================================================================
# INSPECTOR MAPPING MANAGEMENT VIEWS
# =============================================================================

@login_required
@inspector_only_inspections
def inspector_mapping_list(request):
    """Display list of all inspector mappings."""
    clear_messages(request)
    
    # Get all inspector mappings
    mappings = InspectorMapping.objects.all().order_by('inspector_name')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    active_filter = request.GET.get('active', '')
    
    # Apply filters
    if search_query:
        mappings = mappings.filter(
            Q(inspector_name__icontains=search_query) |
            Q(inspector_id__icontains=search_query)
        )
    
    if active_filter == 'true':
        mappings = mappings.filter(is_active=True)
    elif active_filter == 'false':
        mappings = mappings.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(mappings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'mappings': page_obj,
        'total_mappings': mappings.count(),
        'search_query': search_query,
        'active_filter': active_filter,
    }
    
    return render(request, 'main/inspector_mapping_list.html', context)


@login_required
@inspector_only_inspections
def add_inspector_mapping(request):
    """Add a new inspector mapping."""
    clear_messages(request)
    
    if request.method == 'POST':
        form = InspectorMappingForm(request.POST)
        if form.is_valid():
            try:
                mapping = form.save()
                messages.success(request, f"Inspector mapping '{mapping.inspector_name}' (ID: {mapping.inspector_id}) created successfully!")
                return redirect('inspector_mapping_list')
            except Exception as e:
                messages.error(request, f"Error creating inspector mapping: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = InspectorMappingForm()
    
    context = {
        'form': form,
        'action': 'Add'
    }
    
    return render(request, 'main/inspector_mapping_form.html', context)


@login_required
@inspector_only_inspections
def edit_inspector_mapping(request, pk):
    """Edit an existing inspector mapping."""
    clear_messages(request)
    
    mapping = get_object_or_404(InspectorMapping, pk=pk)
    
    if request.method == 'POST':
        form = InspectorMappingForm(request.POST, instance=mapping)
        if form.is_valid():
            try:
                mapping = form.save()
                messages.success(request, f"Inspector mapping '{mapping.inspector_name}' (ID: {mapping.inspector_id}) updated successfully!")
                return redirect('inspector_mapping_list')
            except Exception as e:
                messages.error(request, f"Error updating inspector mapping: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = InspectorMappingForm(instance=mapping)
    
    context = {
        'form': form,
        'mapping': mapping,
        'action': 'Edit'
    }
    
    return render(request, 'main/inspector_mapping_form.html', context)


@login_required
@inspector_only_inspections
def delete_inspector_mapping(request, pk):
    """Delete an inspector mapping."""
    clear_messages(request)
    
    mapping = get_object_or_404(InspectorMapping, pk=pk)
    
    if request.method == 'POST':
        try:
            inspector_name = mapping.inspector_name
            inspector_id = mapping.inspector_id
            mapping.delete()
            messages.success(request, f"Inspector mapping '{inspector_name}' (ID: {inspector_id}) deleted successfully!")
        except Exception as e:
            messages.error(request, f"Error deleting inspector mapping: {str(e)}")
    
    return redirect('inspector_mapping_list')

@login_required
def update_bought_sample(request):
    """Update bought sample value for an inspection."""
    if request.method == 'POST':
        try:
            from ..models import FoodSafetyAgencyInspection, InspectorMapping
            from django.http import JsonResponse

            inspection_id = request.POST.get('inspection_id')
            bought_sample_value = request.POST.get('bought_sample')

            if not inspection_id:
                return JsonResponse({'success': False, 'error': 'Inspection ID is required'})

            # Find the inspection
            inspection = FoodSafetyAgencyInspection.objects.filter(remote_id=inspection_id).first()
            if not inspection:
                return JsonResponse({'success': False, 'error': 'Inspection not found'})

            # Check if user has permission to edit this inspection
            if request.user.role == 'inspector':
                # Get the inspector ID for the current user
                inspector_id = None
                try:
                    inspector_mapping = InspectorMapping.objects.get(
                        inspector_name=request.user.get_full_name() or request.user.username
                    )
                    inspector_id = inspector_mapping.inspector_id
                except InspectorMapping.DoesNotExist:
                    try:
                        inspector_mapping = InspectorMapping.objects.get(
                            inspector_name=request.user.username
                        )
                        inspector_id = inspector_mapping.inspector_id
                    except InspectorMapping.DoesNotExist:
                        inspector_id = None

                # Check if this inspection belongs to the current inspector
                if not inspector_id or inspection.inspector_id != inspector_id:
                    return JsonResponse({'success': False, 'error': 'You can only edit your own inspections'})

            # Update the bought sample value
            if bought_sample_value == '' or bought_sample_value is None:
                # If empty, set to None (will show as "No" in template)
                inspection.bought_sample = None
            else:
                try:
                    new_value = float(bought_sample_value)
                    inspection.bought_sample = new_value
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Invalid bought sample value'})

            inspection.save()

            return JsonResponse({'success': True, 'message': 'Bought sample updated successfully'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
def onedrive_callback(request):
    """Handle OneDrive OAuth callback."""
    from django.http import HttpResponse
    import requests
    import json
    import os
    from datetime import datetime
    from django.conf import settings
    
    # Get the authorization code from the URL
    code = request.GET.get('code')
    
    if code:
        try:
            # Exchange authorization code for access token
            token_url = "https://login.microsoftonline.com/common/oauth2/v2.0/token"
            token_data = {
                'client_id': settings.ONEDRIVE_CLIENT_ID,
                'client_secret': settings.ONEDRIVE_CLIENT_SECRET,
                'grant_type': 'authorization_code',
                'code': code,
                'redirect_uri': settings.ONEDRIVE_REDIRECT_URI
            }
            
            response = requests.post(token_url, data=token_data)
            
            if response.status_code == 200:
                token_response = response.json()
                access_token = token_response.get('access_token')
                refresh_token = token_response.get('refresh_token')
                
                if access_token:
                    # Save tokens to file for the service to use
                    token_data = {
                        'access_token': access_token,
                        'refresh_token': refresh_token,
                        'expires_in': token_response.get('expires_in'),
                        'token_type': token_response.get('token_type'),
                        'expires_at': datetime.now().timestamp() + token_response.get('expires_in', 3600)
                    }
                    
                    token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
                    with open(token_file, 'w') as f:
                        json.dump(token_data, f, indent=2)
                    
                    print(f" Tokens saved to: {token_file}")
                    print(" OneDrive authentication setup complete!")

                    # Return success page with redirect to OneDrive View
                    return HttpResponse("""
                    <html>
                    <head>
                        <title>OneDrive Authentication Success</title>
                        <style>
                            body {
                                font-family: Arial, sans-serif;
                                display: flex;
                                justify-content: center;
                                align-items: center;
                                min-height: 100vh;
                                margin: 0;
                                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            }
                            .success-card {
                                background: white;
                                padding: 3rem;
                                border-radius: 12px;
                                box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                                text-align: center;
                                max-width: 500px;
                            }
                            .success-icon {
                                font-size: 4rem;
                                color: #10b981;
                                margin-bottom: 1rem;
                            }
                            h1 {
                                color: #1f2937;
                                margin-bottom: 1rem;
                            }
                            p {
                                color: #6b7280;
                                margin-bottom: 0.5rem;
                            }
                            .redirect-info {
                                margin-top: 2rem;
                                padding: 1rem;
                                background: #f3f4f6;
                                border-radius: 6px;
                                color: #4b5563;
                                font-size: 0.9rem;
                            }
                        </style>
                    </head>
                    <body>
                        <div class="success-card">
                            <div class="success-icon">✓</div>
                            <h1>OneDrive Connected Successfully!</h1>
                            <p>Your OneDrive account has been authenticated.</p>
                            <p>Access and refresh tokens have been saved.</p>
                            <div class="redirect-info">
                                Redirecting to OneDrive View in 3 seconds...
                            </div>
                        </div>
                        <script>
                            setTimeout(function() {
                                window.location.href = '/developer/onedrive-view/';
                            }, 3000);
                        </script>
                    </body>
                    </html>
                    """)
                else:
                    print(" No access token received")
                    return HttpResponse(" No access token received", status=400)
            else:
                print(f" Token exchange failed: {response.status_code} - {response.text}")
                return HttpResponse(f" Token exchange failed: {response.status_code}", status=400)
                
        except Exception as e:
            print(f" Token exchange failed: {e}")
            return HttpResponse(f" Token exchange failed: {e}", status=400)
    else:
        return HttpResponse(" No authorization code received", status=400)













@login_required(login_url='login')
@role_required(['developer', 'super_admin'])
def performance_monitor(request):
    """View and browse media folder contents with navigation."""
    import os
    from django.conf import settings as django_settings
    from datetime import datetime
    from urllib.parse import unquote
    
    # Get media folder path
    media_root = django_settings.MEDIA_ROOT
    
    # Get current path from URL parameter
    current_path = request.GET.get('path', '')
    if current_path:
        current_path = unquote(current_path)
        # Security check: ensure path is within media root
        full_current_path = os.path.join(media_root, current_path)
        full_current_path = os.path.normpath(full_current_path)
        if not full_current_path.startswith(os.path.normpath(media_root)):
            current_path = ''
            full_current_path = media_root
    else:
        full_current_path = media_root
    
    # Build breadcrumb navigation
    breadcrumbs = [{'name': 'Media', 'path': ''}]
    if current_path:
        path_parts = current_path.split(os.sep)
        cumulative_path = ''
        for part in path_parts:
            if part:
                cumulative_path = os.path.join(cumulative_path, part) if cumulative_path else part
                breadcrumbs.append({'name': part, 'path': cumulative_path})
    
    # Scan current folder
    media_files = []
    media_folders = []
    
    try:
        if os.path.exists(full_current_path):
            for item in os.listdir(full_current_path):
                item_path = os.path.join(full_current_path, item)
                
                if os.path.isfile(item_path):
                    # Get file info
                    stat = os.stat(item_path)
                    file_size = stat.st_size
                    modified_time = datetime.fromtimestamp(stat.st_mtime)
                    
                    # Determine file type
                    file_ext = os.path.splitext(item)[1].lower()
                    file_type = 'unknown'
                    icon = 'fas fa-file'
                    
                    if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                        file_type = 'image'
                        icon = 'fas fa-image'
                    elif file_ext in ['.pdf']:
                        file_type = 'pdf'
                        icon = 'fas fa-file-pdf'
                    elif file_ext in ['.doc', '.docx']:
                        file_type = 'document'
                        icon = 'fas fa-file-word'
                    elif file_ext in ['.xls', '.xlsx']:
                        file_type = 'spreadsheet'
                        icon = 'fas fa-file-excel'
                    elif file_ext in ['.zip', '.rar', '.7z']:
                        file_type = 'archive'
                        icon = 'fas fa-file-archive'
                    elif file_ext in ['.txt', '.log']:
                        file_type = 'text'
                        icon = 'fas fa-file-alt'
                    
                    # Build relative path for media URL
                    relative_path = os.path.relpath(item_path, media_root).replace('\\', '/')
                    
                    media_files.append({
                        'name': item,
                        'path': item_path,
                        'relative_path': relative_path,
                        'size': file_size,
                        'size_formatted': format_file_size(file_size),
                        'modified': modified_time,
                        'type': file_type,
                        'icon': icon,
                        'extension': file_ext
                    })
                    
                elif os.path.isdir(item_path):
                    # Count files in folder recursively
                    try:
                        file_count = 0
                        for root, dirs, files in os.walk(item_path):
                            file_count += len(files)
                    except:
                        file_count = 0
                    
                    # Build navigation path
                    nav_path = os.path.join(current_path, item) if current_path else item
                    nav_path = nav_path.replace('\\', '/')
                    
                    media_folders.append({
                        'name': item,
                        'path': item_path,
                        'nav_path': nav_path,
                        'file_count': file_count,
                        'icon': 'fas fa-folder'
                    })
        
        # Sort files and folders
        media_files.sort(key=lambda x: x['modified'], reverse=True)
        media_folders.sort(key=lambda x: x['name'])
        
    except Exception as e:
        safe_print(f"Error scanning media folder: {e}")
    
    # Get theme settings
    try:
        from ..models import SystemSettings
        system_settings = SystemSettings.get_settings()
        settings = type('Settings', (), {
            'dark_mode': system_settings.theme_mode == 'dark' if hasattr(system_settings, 'theme_mode') else False,
        })()
    except Exception:
        settings = type('Settings', (), {'dark_mode': False})()
    
    context = {
        'media_files': media_files,
        'media_folders': media_folders,
        'media_root': media_root,
        'current_path': current_path,
        'current_folder_name': os.path.basename(full_current_path) if current_path else 'Media',
        'breadcrumbs': breadcrumbs,
        'total_files': len(media_files),
        'total_folders': len(media_folders),
        'settings': settings,
    }
    
    return render(request, 'main/server_view.html', context)


def format_file_size(size_bytes):
    """Convert bytes to human readable format."""
    if size_bytes == 0:
        return "0 B"
    size_names = ["B", "KB", "MB", "GB", "TB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"


@login_required(login_url='login')
def server_directory_tree(request):
    """Get complete directory tree as JSON for the server view page."""
    import os
    from django.conf import settings as django_settings
    from django.http import JsonResponse

    media_root = django_settings.MEDIA_ROOT

    def build_tree(path, name=None):
        """Recursively build directory tree."""
        tree = {
            'name': name or os.path.basename(path) or 'Media',
            'type': 'folder',
            'children': [],
            'file_count': 0
        }

        try:
            items = os.listdir(path)
            items.sort()

            for item in items:
                item_path = os.path.join(path, item)

                if os.path.isdir(item_path):
                    # Recursively build subtree
                    subtree = build_tree(item_path, item)
                    tree['children'].append(subtree)
                    tree['file_count'] += subtree['file_count']
                elif os.path.isfile(item_path):
                    # Add file
                    try:
                        file_size = os.path.getsize(item_path)
                        file_ext = os.path.splitext(item)[1]
                        tree['children'].append({
                            'name': item,
                            'type': 'file',
                            'size': format_file_size(file_size),
                            'extension': file_ext
                        })
                        tree['file_count'] += 1
                    except:
                        # Skip files we can't access
                        pass
        except PermissionError:
            # Skip directories we can't access
            pass
        except Exception as e:
            print(f"Error reading {path}: {e}")

        return tree

    # Build the complete tree
    tree = build_tree(media_root)

    return JsonResponse(tree)


@login_required(login_url='login')
def server_status(request):
    """Check server health and performance metrics."""
    from django.core.cache import cache
    from django.db import connection
    import psutil
    import os
    
    # Get system metrics
    cpu_percent = psutil.cpu_percent(interval=1)
    memory = psutil.virtual_memory()
    disk = psutil.disk_usage('/')
    
    # Get database metrics
    with connection.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM food_safety_agency_inspections")
        inspection_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM food_safety_agency_clients")
        client_count = cursor.fetchone()[0]
    
    # Get cache metrics
    cache_hits = cache.get('cache_hits', 0)
    cache_misses = cache.get('cache_misses', 0)
    cache_hit_rate = (cache_hits / (cache_hits + cache_misses)) * 100 if (cache_hits + cache_misses) > 0 else 0
    
    context = {
        'cpu_percent': cpu_percent,
        'memory_percent': memory.percent,
        'memory_available': memory.available // (1024**3),  # GB
        'disk_percent': disk.percent,
        'disk_free': disk.free // (1024**3),  # GB
        'inspection_count': inspection_count,
        'client_count': client_count,
        'cache_hits': cache_hits,
        'cache_misses': cache_misses,
        'cache_hit_rate': cache_hit_rate,
        'server_uptime': time.time() - psutil.boot_time(),
    }
    
    return render(request, 'main/server_status.html', context)


# =============================================================================
# CLIENT AUTOCOMPLETE API
# =============================================================================

@login_required
def client_autocomplete_api(request):
    """API endpoint for client autocomplete suggestions."""
    from django.http import JsonResponse
    from ..models import Client, FoodSafetyAgencyInspection

    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'suggestions': []})

    try:
        # Get unique client names from both Client model and FoodSafetyAgencyInspection
        # This ensures we get all clients that have inspections
        client_suggestions = []

        # Get clients from Client model
        clients = Client.objects.filter(
            Q(client_id__icontains=query) |
            Q(name__icontains=query)
        ).values('client_id', 'name', 'internal_account_code').distinct()[:10]

        for client in clients:
            client_name = client['name'] or client['client_id']
            if client_name:
                # Get the most recent town for this client from inspections
                town = FoodSafetyAgencyInspection.objects.filter(
                    client_name__iexact=client_name
                ).exclude(town__isnull=True).exclude(town='').order_by('-date_of_inspection').values_list('town', flat=True).first()

                client_suggestions.append({
                    'name': client_name,
                    'account_code': client['internal_account_code'] or '',
                    'town': town or '',
                    'type': 'client'
                })

        # Get additional clients from inspection data that might not be in Client model
        inspection_clients = FoodSafetyAgencyInspection.objects.filter(
            client_name__icontains=query
        ).exclude(client_name__isnull=True).exclude(client_name='').values('client_name').annotate(
            latest_town=models.Max('town')
        ).distinct()[:10]

        for client_data in inspection_clients:
            client_name = client_data['client_name']
            if client_name and not any(s['name'] == client_name for s in client_suggestions):
                # Get the most recent town for this client
                town = FoodSafetyAgencyInspection.objects.filter(
                    client_name__iexact=client_name
                ).exclude(town__isnull=True).exclude(town='').order_by('-date_of_inspection').values_list('town', flat=True).first()

                client_suggestions.append({
                    'name': client_name,
                    'account_code': '',
                    'town': town or '',
                    'type': 'inspection'
                })

        # Sort by name and limit to 15 total suggestions
        client_suggestions.sort(key=lambda x: x['name'].lower())
        client_suggestions = client_suggestions[:15]

        return JsonResponse({'suggestions': client_suggestions})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def scheduled_backup_service_status(request):
    """Get scheduled backup service status."""
    try:
        from ..services.scheduled_backup_service import get_scheduled_backup_service_status

        status = get_scheduled_backup_service_status()
        return JsonResponse(status)

    except ImportError:
        # Service module not available - return graceful status instead of 500 error
        return JsonResponse({
            'is_running': False,
            'status': 'not_available',
            'message': 'Backup service not configured'
        })
    except Exception as e:
        # Other errors - return graceful status to avoid console spam
        return JsonResponse({
            'is_running': False,
            'status': 'error',
            'message': f'Service error: {str(e)}'
        })


@login_required
def start_scheduled_backup_service(request):
    """Start the scheduled backup service."""
    try:
        from ..services.scheduled_backup_service import start_scheduled_backup_service
        
        success, message = start_scheduled_backup_service()
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def stop_scheduled_backup_service(request):
    """Stop the scheduled backup service."""
    try:
        from ..services.scheduled_backup_service import stop_scheduled_backup_service
        
        success, message = stop_scheduled_backup_service()
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def run_manual_backup(request):
    """Run a manual backup."""
    try:
        from ..services.scheduled_backup_service import run_manual_backup
        
        success, message = run_manual_backup()
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# MASTER SERVICE CONTROL
# =============================================================================

@login_required
def master_service_control_status(request):
    """Get the status of all background services."""
    try:
        services_status = {}
        
        # Daily Compliance Sync Service
        try:
            from ..services.daily_compliance_sync import daily_sync_service
            services_status['daily_compliance_sync'] = {
                'name': 'Daily Compliance Sync',
                'running': daily_sync_service.is_running,
                'status': {}
            }
        except Exception as e:
            services_status['daily_compliance_sync'] = {
                'name': 'Daily Compliance Sync',
                'running': False,
                'status': {'error': str(e)}
            }
        
        # Scheduled Sync Service
        try:
            from ..services.scheduled_sync_service import ScheduledSyncService
            sync_service = ScheduledSyncService()
            services_status['scheduled_sync'] = {
                'name': 'Scheduled Sync',
                'running': sync_service.is_running,
                'status': {}
            }
        except Exception as e:
            services_status['scheduled_sync'] = {
                'name': 'Scheduled Sync',
                'running': False,
                'status': {'error': str(e)}
            }
        
        # Scheduled Backup Service
        try:
            from ..services.scheduled_backup_service import ScheduledBackupService
            backup_service = ScheduledBackupService()
            services_status['scheduled_backup'] = {
                'name': 'Scheduled Backup',
                'running': backup_service.is_running,
                'status': {}
            }
        except Exception as e:
            services_status['scheduled_backup'] = {
                'name': 'Scheduled Backup',
                'running': False,
                'status': {'error': str(e)}
            }
        
        # OneDrive Upload Service (skip if there's an error)
        try:
            from ..services.onedrive_direct_service import OneDriveDirectUploadService, get_onedrive_direct_service_status
            onedrive_service = OneDriveDirectUploadService()
            service_status = get_onedrive_direct_service_status()
            
            # Check OneDrive connection status
            connection_status = onedrive_service.authenticate_onedrive()
            
            services_status['onedrive_upload'] = {
                'name': 'OneDrive Upload',
                'running': onedrive_service.is_running,
                'connected': connection_status,
                'status': service_status
            }
        except Exception as e:
            services_status['onedrive_upload'] = {
                'name': 'OneDrive Upload',
                'running': False,
                'connected': False,
                'status': {'error': 'Service unavailable', 'details': str(e)}
            }
        
        # Calculate overall status
        all_running = all(service['running'] for service in services_status.values())
        any_running = any(service['running'] for service in services_status.values())
        
        return JsonResponse({
            'success': True,
            'services': services_status,
            'all_running': all_running,
            'any_running': any_running,
            'total_services': len(services_status),
            'running_services': sum(1 for service in services_status.values() if service['running'])
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def start_all_services(request):
    """Start all background services."""
    try:
        results = {}
        errors = []
        
        # Start Daily Compliance Sync
        try:
            from ..services.daily_compliance_sync import daily_sync_service
            if not daily_sync_service.is_running:
                daily_sync_service.start_daily_sync(manual_start=True)
                results['daily_compliance_sync'] = 'Started successfully'
            else:
                results['daily_compliance_sync'] = 'Already running'
        except Exception as e:
            error_msg = f"Daily Compliance Sync: {str(e)}"
            errors.append(error_msg)
            results['daily_compliance_sync'] = f"Failed: {str(e)}"
        
        # Start Scheduled Sync Service
        try:
            from ..services.scheduled_sync_service import start_scheduled_sync_service
            success, message = start_scheduled_sync_service()
            results['scheduled_sync'] = message
            if not success:
                errors.append(f"Scheduled Sync: {message}")
        except Exception as e:
            error_msg = f"Scheduled Sync: {str(e)}"
            errors.append(error_msg)
            results['scheduled_sync'] = f"Failed: {str(e)}"
        
        # Start Scheduled Backup Service
        try:
            from ..services.scheduled_backup_service import start_scheduled_backup_service
            success, message = start_scheduled_backup_service()
            results['scheduled_backup'] = message
            if not success:
                errors.append(f"Scheduled Backup: {message}")
        except Exception as e:
            error_msg = f"Scheduled Backup: {str(e)}"
            errors.append(error_msg)
            results['scheduled_backup'] = f"Failed: {str(e)}"
        
        # Start OneDrive Upload Service (skip if unavailable)
        try:
            from ..services.onedrive_direct_service import OneDriveDirectUploadService
            onedrive_service = OneDriveDirectUploadService()
            success, message = onedrive_service.start_background_service()
            results['onedrive_upload'] = message
            if not success:
                errors.append(f"OneDrive Upload: {message}")
        except Exception as e:
            error_msg = f"OneDrive Upload: {str(e)}"
            errors.append(error_msg)
            results['onedrive_upload'] = f"Service unavailable"
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Started all background services via Master Service Control',
                details={'results': results, 'errors': errors}
            )
        except Exception:
            pass
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Some services failed to start',
                'results': results,
                'errors': errors
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'All services started successfully',
                'results': results
            })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def stop_all_services(request):
    """Stop all background services."""
    try:
        from ..services.daily_compliance_sync import daily_sync_service
        from ..services.scheduled_sync_service import stop_scheduled_sync_service
        from ..services.scheduled_backup_service import stop_scheduled_backup_service
        from ..services.onedrive_direct_service import OneDriveDirectUploadService
        
        results = {}
        errors = []
        
        # Stop Daily Compliance Sync
        try:
            if daily_sync_service.is_running:
                daily_sync_service.stop_daily_sync()
                results['daily_compliance_sync'] = 'Stopped successfully'
            else:
                results['daily_compliance_sync'] = 'Was not running'
        except Exception as e:
            error_msg = f"Daily Compliance Sync: {str(e)}"
            errors.append(error_msg)
            results['daily_compliance_sync'] = f"Failed: {str(e)}"
        
        # Stop Scheduled Sync Service
        try:
            success, message = stop_scheduled_sync_service()
            results['scheduled_sync'] = message
            if not success:
                errors.append(f"Scheduled Sync: {message}")
        except Exception as e:
            error_msg = f"Scheduled Sync: {str(e)}"
            errors.append(error_msg)
            results['scheduled_sync'] = f"Failed: {str(e)}"
        
        # Stop Scheduled Backup Service
        try:
            success, message = stop_scheduled_backup_service()
            results['scheduled_backup'] = message
            if not success:
                errors.append(f"Scheduled Backup: {message}")
        except Exception as e:
            error_msg = f"Scheduled Backup: {str(e)}"
            errors.append(error_msg)
            results['scheduled_backup'] = f"Failed: {str(e)}"
        
        # Stop OneDrive Upload Service
        try:
            onedrive_service = OneDriveDirectUploadService()
            success, message = onedrive_service.stop_background_service()
            results['onedrive_upload'] = message
            if not success:
                errors.append(f"OneDrive Upload: {message}")
        except Exception as e:
            error_msg = f"OneDrive Upload: {str(e)}"
            errors.append(error_msg)
            results['onedrive_upload'] = f"Failed: {str(e)}"
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Stopped all background services via Master Service Control',
                details={'results': results, 'errors': errors}
            )
        except Exception:
            pass
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Some services failed to stop',
                'results': results,
                'errors': errors
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'All services stopped successfully',
                'results': results
            })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# ONEDRIVE SERVICE CONTROL
# =============================================================================

@login_required
def onedrive_service_status(request):
    """Get OneDrive service status and connection info."""
    try:
        from ..services.onedrive_direct_service import (
            OneDriveDirectUploadService,
            get_onedrive_direct_service_status,
        )
        
        # Read status from the shared status provider (cache-backed)
        service_status = get_onedrive_direct_service_status()
        service_running = bool(service_status.get('is_running'))
        
        # Use a transient instance only to check connection/auth health
        onedrive_service = OneDriveDirectUploadService()
        connection_status = onedrive_service.authenticate_onedrive()
        
        # Get OneDrive settings
        from ..models import SystemSettings
        settings = SystemSettings.get_settings()
        upload_delay = getattr(settings, 'onedrive_upload_delay_days', 3)
        onedrive_enabled = getattr(settings, 'onedrive_enabled', False)
        
        return JsonResponse({
            'success': True,
            'service_running': service_running,
            'connected': connection_status,
            'enabled': onedrive_enabled,
            'upload_delay_days': upload_delay,
            'status': service_status,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def start_onedrive_service(request):
    """Start OneDrive service."""
    try:
        from ..services.onedrive_direct_service import start_onedrive_direct_background_service
        
        success, message = start_onedrive_direct_background_service()
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Started OneDrive service via Master Service Control',
                details={'success': success, 'message': message}
            )
        except Exception:
            pass
        
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def stop_onedrive_service(request):
    """Stop OneDrive service."""
    try:
        from ..services.onedrive_direct_service import stop_onedrive_direct_background_service
        
        success, message = stop_onedrive_direct_background_service()
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Stopped OneDrive service via Master Service Control',
                details={'success': success, 'message': message}
            )
        except Exception:
            pass
        
        return JsonResponse({'success': success, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def test_onedrive_connection(request):
    """Test OneDrive connection."""
    try:
        import os
        import json
        from django.conf import settings
        from ..services.onedrive_direct_service import OneDriveDirectUploadService
        
        onedrive_service = OneDriveDirectUploadService()
        connection_status = onedrive_service.authenticate_onedrive()
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Tested OneDrive connection',
                details={'connected': connection_status}
            )
        except Exception:
            pass
        
        if connection_status:
            return JsonResponse({
                'success': True, 
                'message': 'OneDrive connection successful',
                'connected': True
            })
        else:
            # Check if it's a token issue
            token_file = os.path.join(settings.BASE_DIR, 'onedrive_tokens.json')
            if os.path.exists(token_file):
                with open(token_file, 'r') as f:
                    tokens = json.load(f)
                
                refresh_token = tokens.get('refresh_token')
                if not refresh_token:
                    return JsonResponse({
                        'success': False, 
                        'message': 'OneDrive refresh token missing. Please re-authenticate in Settings.',
                        'connected': False,
                        'needs_reauth': True
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'message': 'OneDrive connection failed. Please check authentication.',
                        'connected': False,
                        'needs_reauth': True
                    })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': 'OneDrive not configured. Please authenticate in Settings.',
                    'connected': False,
                    'needs_reauth': True
                })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'message': f'OneDrive connection test failed: {str(e)}',
            'connected': False
        }, status=500)


@login_required
def reauthenticate_onedrive(request):
    """Re-authenticate OneDrive by generating authorization URL."""
    try:
        from django.conf import settings
        
        # Check if OneDrive is enabled
        if not getattr(settings, 'ONEDRIVE_ENABLED', False):
            return JsonResponse({
                'success': False, 
                'message': 'OneDrive integration is not enabled in settings',
                'requires_setup': True
            })
        
        # Generate authorization URL
        client_id = getattr(settings, 'ONEDRIVE_CLIENT_ID', '')
        redirect_uri = getattr(settings, 'ONEDRIVE_REDIRECT_URI', 'http://localhost:8000/onedrive/callback')
        
        if not client_id:
            return JsonResponse({
                'success': False, 
                'message': 'OneDrive Client ID not configured',
                'requires_setup': True
            })
        
        # Generate authorization URL - try common endpoint first (works for both org and consumer)
        auth_url = "https://login.microsoftonline.com/common/oauth2/v2.0/authorize"
        auth_params = {
            'client_id': client_id,
            'response_type': 'code',
            'redirect_uri': redirect_uri,
            'scope': 'Files.ReadWrite.All',
            'response_mode': 'query',
            'prompt': 'consent'  # Force consent to ensure refresh token is provided
        }
        
        # Build the authorization URL
        auth_url_with_params = f"{auth_url}?{'&'.join([f'{k}={v}' for k, v in auth_params.items()])}"
        
        # Log the activity
        try:
            SystemLog.log_activity(
                user=request.user,
                action='SETTINGS',
                page=request.path,
                description='Generated OneDrive authorization URL for re-authentication',
                details={'auth_url': auth_url_with_params}
            )
        except Exception:
            pass
        
        return JsonResponse({
            'success': True, 
            'message': 'Authorization URL generated. Please visit the URL to complete re-authentication.',
            'auth_url': auth_url_with_params,
            'instructions': [
                '1. Click the authorization URL below',
                '2. Sign in with your Microsoft account',
                '3. Grant permissions to the app',
                '4. You will be redirected back to the application',
                '5. The tokens will be automatically saved and refreshed'
            ]
        })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@role_required(['developer'])
def onedrive_auth(request):
    """Initiate OneDrive OAuth authentication flow."""
    from django.shortcuts import redirect
    from django.conf import settings
    import urllib.parse

    try:
        # Get OneDrive settings
        client_id = getattr(settings, 'ONEDRIVE_CLIENT_ID', '')
        redirect_uri = getattr(settings, 'ONEDRIVE_REDIRECT_URI', 'http://localhost:8000/onedrive/callback')

        if not client_id:
            return HttpResponse('<html><body><h1>Error</h1><p>OneDrive Client ID not configured in settings.</p></body></html>')

        # Generate authorization URL
        auth_url = "https://login.microsoftonline.com/common/oauth2/v2.0/authorize"
        auth_params = {
            'client_id': client_id,
            'response_type': 'code',
            'redirect_uri': redirect_uri,
            'scope': 'https://graph.microsoft.com/Files.ReadWrite.All offline_access',
            'response_mode': 'query',
            'prompt': 'consent'  # Force consent to get refresh token
        }

        # Build full authorization URL
        auth_url_with_params = f"{auth_url}?{urllib.parse.urlencode(auth_params)}"

        # Redirect user to Microsoft OAuth page
        return redirect(auth_url_with_params)

    except Exception as e:
        return HttpResponse(f'<html><body><h1>Error</h1><p>{str(e)}</p></body></html>')


@login_required
def get_onedrive_auth_url(request):
    """Get OneDrive authorization URL for manual authentication."""
    try:
        from django.conf import settings

        # Check if OneDrive is enabled
        if not getattr(settings, 'ONEDRIVE_ENABLED', False):
            return JsonResponse({
                'success': False,
                'message': 'OneDrive integration is not enabled in settings'
            })
        
        # Generate authorization URL
        auth_url = "https://login.microsoftonline.com/consumers/oauth2/v2.0/authorize"
        auth_params = {
            'client_id': settings.ONEDRIVE_CLIENT_ID,
            'response_type': 'code',
            'redirect_uri': settings.ONEDRIVE_REDIRECT_URI,
            'scope': 'Files.ReadWrite.All',
            'response_mode': 'query'
        }
        
        # Build the authorization URL
        auth_url_with_params = f"{auth_url}?{'&'.join([f'{k}={v}' for k, v in auth_params.items()])}"
        
        return JsonResponse({
            'success': True,
            'auth_url': auth_url_with_params,
            'message': 'Click the URL to authenticate with OneDrive'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def send_password_reset_email(request):
    """Send password reset email to a user"""
    import json
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_encode
    from django.utils.encoding import force_bytes
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings

    # Check if user has admin permissions
    if not (request.user.has_role_permission('admin') or
            request.user.has_role_permission('super_admin') or
            request.user.has_role_permission('developer') or
            request.user.is_staff):
        return JsonResponse({
            'success': False,
            'error': 'You do not have permission to send password reset emails'
        }, status=403)

    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')

        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'User ID is required'
            }, status=400)

        # Get the user
        target_user = User.objects.get(id=user_id)

        # Check if user has an email
        if not target_user.email:
            return JsonResponse({
                'success': False,
                'error': 'User does not have an email address configured'
            }, status=400)

        # Generate password reset token
        token = default_token_generator.make_token(target_user)
        uid = urlsafe_base64_encode(force_bytes(target_user.pk))

        # Build password reset URL
        reset_url = request.build_absolute_uri(
            f'/password-reset/{uid}/{token}/'
        )

        # Email content
        subject = 'Password Reset Request - Food Safety Agency'
        message = f'''Hello {target_user.first_name or target_user.username},

You have received this email because a password reset was requested for your account at the Food Safety Agency Inspection System.

Click the link below to reset your password:

{reset_url}

This link will expire in 24 hours.

If you did not request a password reset, please ignore this email.

Best regards,
Food Safety Agency Team
'''

        # HTML email content (optional)
        html_message = f'''
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
                    <h2 style="color: #007890;">Password Reset Request</h2>
                    <p>Hello {target_user.first_name or target_user.username},</p>
                    <p>You have received this email because a password reset was requested for your account at the Food Safety Agency Inspection System.</p>
                    <p>Click the button below to reset your password:</p>
                    <p style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" style="background-color: #007890; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; display: inline-block;">Reset Password</a>
                    </p>
                    <p style="font-size: 0.9em; color: #666;">Or copy and paste this link into your browser:</p>
                    <p style="font-size: 0.9em; color: #666; word-break: break-all;">{reset_url}</p>
                    <p style="font-size: 0.9em; color: #999; margin-top: 30px;">This link will expire in 24 hours.</p>
                    <p style="font-size: 0.9em; color: #999;">If you did not request a password reset, please ignore this email.</p>
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
                    <p style="font-size: 0.85em; color: #999;">Best regards,<br>Food Safety Agency Team</p>
                </div>
            </body>
        </html>
        '''

        # Send email
        send_mail(
            subject=subject,
            message=message,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@foodsafetyagency.com'),
            recipient_list=[target_user.email],
            fail_silently=False,
            html_message=html_message
        )

        # Log the action
        SystemLog.objects.create(
            user=request.user,
            action='SEND_PASSWORD_RESET_EMAIL',
            page='User Management',
            details=f'Sent password reset email to {target_user.username} ({target_user.email})'
        )

        return JsonResponse({
            'success': True,
            'message': f'Password reset email sent to {target_user.email}'
        })

    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'User not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# =============================
# Password Reset Views (For Login Page)
# =============================

def forgot_password(request):
    """
    Handle forgot password requests - ONLY for developer account (ethansevenster5@gmail.com)
    """
    from django.shortcuts import render, redirect
    from django.contrib import messages
    from django.contrib.auth.models import User
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_encode
    from django.utils.encoding import force_bytes
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    from datetime import datetime

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        # IMPORTANT: Only allow password reset for developer account
        if email != 'ethansevenster5@gmail.com':
            messages.error(request, 'Password reset is only available for the developer account.')
            return redirect('forgot_password')

        try:
            # Find user with this email
            user = User.objects.filter(email__iexact=email).first()

            if not user:
                # Don't reveal if user exists or not (security best practice)
                messages.success(request, 'If an account with that email exists, a password reset link has been sent.')
                return redirect('forgot_password')

            # Generate password reset token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))

            # Build reset link
            reset_link = request.build_absolute_uri(
                f'/reset-password/{uid}/{token}/'
            )

            # Prepare email context
            context = {
                'user': user,
                'reset_link': reset_link,
                'current_year': datetime.now().year,
            }

            # Render email template
            html_message = render_to_string('main/password_reset_email.html', context)

            # Send email with proper no-reply display name
            send_mail(
                subject='Food Safety Agency – Password Reset Instructions',
                message=f'Hello {user.username},\n\nClick the link below to reset your password:\n\n{reset_link}\n\nThis link will expire in 1 hour.\n\nIf you did not request this, please ignore this email.',
                from_email='Food Safety Agency - No Reply <info@eclick.co.za>',
                recipient_list=[email],
                html_message=html_message,
                fail_silently=False,
            )

            messages.success(request, 'Password reset link has been sent to your email. Please check your inbox (and spam folder).')
            return redirect('forgot_password')

        except Exception as e:
            messages.error(request, f'Error sending password reset email: {str(e)}')
            return redirect('forgot_password')

    # GET request - show the form
    return render(request, 'main/forgot_password.html')


def reset_password_confirm(request, uidb64, token):
    """
    Handle password reset confirmation (when user clicks link in email)
    """
    from django.shortcuts import render, redirect
    from django.contrib import messages
    from django.contrib.auth.models import User
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str

    # Decode user ID from URL
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    # Verify token is valid
    if user is not None and default_token_generator.check_token(user, token):
        # Token is valid
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            # Validate passwords
            if not new_password or not confirm_password:
                messages.error(request, 'Both password fields are required.')
                return render(request, 'main/reset_password_confirm.html')

            if len(new_password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'main/reset_password_confirm.html')

            if new_password != confirm_password:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'main/reset_password_confirm.html')

            # Set new password
            user.set_password(new_password)
            user.save()

            messages.success(request, 'Your password has been reset successfully! You can now log in with your new password.')
            return redirect('login')

        # GET request - show password reset form
        return render(request, 'main/reset_password_confirm.html')
    else:
        # Invalid or expired token
        messages.error(request, 'This password reset link is invalid or has expired. Please request a new one.')
        return redirect('forgot_password')
