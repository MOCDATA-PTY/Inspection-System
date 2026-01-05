from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils.dateparse import parse_date
from ..models import Shipment, Client
from .utils import apply_filters, clear_messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime
import os
import threading
import time
import shutil
from pathlib import Path
import re
import sys

# Add the project root to Python path to import the config
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

try:
    from eclick_mysql_config import (
        MYSQL_CONFIG, 
        SQLSERVER_CONFIG, 
        SQLSERVER_CONNECTION_STRING,
        FSA_INSPECTION_QUERY
    )
except ImportError:
    # Fallback configurations if import fails
    MYSQL_CONFIG = {
        'host': '77.37.121.135',
        'user': 'admin',
        'password': 'mk7z@Geg123',
        'database': 'E-click-Project-management',
        'port': 3306,
        'charset': 'utf8mb4',
        'autocommit': True,
        'connect_timeout': 60,
        'read_timeout': 60,
        'write_timeout': 60,
        'use_unicode': True,
    }
    
    SQLSERVER_CONFIG = {
        'ENGINE': 'mssql',
        'NAME': 'AFS',
        'USER': 'FSAUser2',
        'PASSWORD': 'password',
        'HOST': '102.67.140.12',
        'PORT': '1053',
        'OPTIONS': {
            'driver': 'ODBC Driver 17 for SQL Server',
            'trusted_connection': 'no',
        },
    }
    
    SQLSERVER_CONNECTION_STRING = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=102.67.140.12,1053;'
        'DATABASE=AFS;'
        'UID=FSAUser2;'
        'PWD=password;'
        'TrustServerCertificate=yes;'
        'Encrypt=yes;'
    )
    
    # Fallback FSA query (comprehensive inspection data) - FROM OCTOBER 2025 ONWARDS
    # IMPORTANT: Removed DISTINCT so that one inspection with multiple products = multiple rows (one per product)
    # IMPORTANT: Now includes InternalAccountNumber to match with Google Sheets client names
    # IMPORTANT: Pulls inspections from 2025-10-01 onwards (no end date limit)
    # IMPORTANT: Changed GPS JOIN to OUTER APPLY to prevent duplicate inspections
    #            - Before: One inspection with 9 GPS records = 9 duplicate inspection rows
    #            - After: One inspection with 9 GPS records = 1 inspection row (using first GPS record)
    FSA_INSPECTION_QUERY = '''
        SELECT 'POULTRY' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName as ProductName FROM AFS.dbo.PoultryLabelInspectionChecklistRecords OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryLabelChecklistInspectionRecordId = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ClientId where AFS.dbo.[PoultryLabelInspectionChecklistRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName IS NOT NULL AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName != ''
        UNION ALL
        SELECT 'EGGS' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForInspection as IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryEggInspectionRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer as ProductName FROM [AFS].[dbo].[PoultryEggInspectionRecords] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryEggInspectionRecordId = [AFS].[dbo].[PoultryEggInspectionRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryEggInspectionRecords].ClientId where AFS.dbo.[PoultryEggInspectionRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer IS NOT NULL AND [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer != ''
        UNION ALL
        SELECT 'RAW' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.RawRMPInspectionLabSampleLinks WHERE RawRMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[RawRMPInspectionRecordTypes].Id as Id, COALESCE(prod.NewClientName, clt.Name) as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.NewProductItemDetails as ProductName FROM [AFS].[dbo].[RawRMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE RawRMPInspectionRecordId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.RawRMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[RawRMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND prod.NewProductItemDetails IS NOT NULL AND prod.NewProductItemDetails != ''
        UNION ALL
        SELECT 'PMP' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.PMPInspectionLabSampleLinks WHERE PMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PMPInspectionRecordTypes].Id as Id, COALESCE(prod.NewClientName, clt.Name) as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.PMPItemDetails as ProductName FROM [AFS].[dbo].[PMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PMPInspectionRecordId = [AFS].[dbo].[PMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.PMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[PMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND prod.PMPItemDetails IS NOT NULL AND prod.PMPItemDetails != ''
    '''

# Mapping of InspectorId to human-readable name (module-level so all views can use it)
INSPECTOR_NAME_MAP = {
    132: 'PAKI OLIFANT', 142: 'MARIANA DU TOIT', 143: 'MOKGADI SELONE', 140: 'AGREEMENT MOSIA',
    68: 'BEN VISAGIE', 144: 'VHAHANGWELE RALULIMI', 97: 'NAKISANI BALOYI', 133: 'CALVIN CLAASSENS',
    118: 'NEO NOE', 131: 'EDITH SELEPE', 124: 'SANDISIWE DLISANI', 85: 'JOEL MHANGWA',
    86: 'ASISIPHO LANDE', 125: 'MARIUS CARSTENS', 145: 'WILSON MAIFO', 82: 'JOE ROSENBLATT',
    95: 'EVANS NKWINIKA', 93: 'VICTOR MATHEBULA', 77: 'DELAREY RIBBENS', 78: 'DEWALD KORF',
    103: 'GERRIT PEKEMA', 76: 'PETRUS POOL', 71: 'PALESA MPANA', 134: 'ARMAND VISAGIE',
    113: 'KATLEGO MOKHUA', 147: 'BRIAN XULU', 146: 'ELIAS THEKISO', 72: 'CHARMAINE NEL',
    148: 'COLLEN DLAMINI', 153: 'JOFRED STEYN', 164: 'HENNIE CHILOANE', 166: 'THATO SEKHOTHO',
    69: 'AFS DUMMY', 177: 'CINGA NGONGO', 91: 'MONTI RAMAANO', 79: 'WENDY CHAKA',
    174: 'LWANDILE MAQINA', 96: 'MCAULEY MUSUNDA', 92: 'THABO MAGWAZA', 149: 'THEMBA SHABANGU',
    175: 'PETRUS POOL', 70: 'THERESA DIOGO', 73: 'RAM RAMBURAN', 74: 'HEIN NEL',
    75: 'SIBONELO ZONDI', 80: 'NIPHO NGOMANE', 81: 'CHELESILE MOYO', 83: 'SIMON SWART',
    84: 'ANDREAS LETABA', 87: 'LOUIS VISAGIE', 88: 'RASSIE SMIT', 89: 'NICOLE BERGH',
    90: 'Spare1 Spare2', 94: 'FRANCOIS PRETORIUS', 98: 'PLACEHOLDER PLACEHOLDER', 99: 'TEST GMAIL',
    102: 'ALI MODIKOE', 106: 'ARMAND VISAGIE', 111: 'DUMMY TESTER', 112: 'TESTER TESTER',
    114: 'DUMMY PLACEHOLDER', 115: 'TESTER PERSON', 116: 'FINANCE TWO', 119: 'TESTER TESTER',
    122: 'Spare Spare', 123: 'LIZELLE BREEDT', 126: 'THAPELO MAPOTSE', 127: 'THAPELO MAPOTSE',
    141: 'IT IT', 150: 'CHRISNA POOL', 154: 'SEUN SEBOLAI', 160: 'LERATO MODIBA',
    163: 'DENNIS CELE', 171: 'MARNUS BADENHORST', 173: 'HLENGIWE GUMEDE', 178: 'NTABISENG MASEKO',
    179: 'KABELO MOKGALAKA', 183: 'MPHO MOTAUNG', 184: 'XOLA MPELUZA', 185: 'PERCY MALEKA',
    186: 'KUTLWANO KUNTWANE', 187: 'SIPHO NDAMASE', 188: 'GLADYS MANGANYE', 193: 'ANTHONY PENZES',
    194: 'PHUMZILE MASOMBUKA', 196: 'NELISA NTOYAPHI',
}

# =============================================================================
# SQL DATABASE VIEWS
# =============================================================================

@login_required
def remote_sqlserver_data_view(request):
    """View to display data from remote SQL Server database (FSA Server)."""
    sqlserver_data = []
    error_message = None
    
    try:
        import pymssql

        # Use pymssql - NO ODBC DRIVERS NEEDED!
        connection = pymssql.connect(
            server='102.67.140.12',
            port=1053,
            user='FSAUser2',
            password='password',
            database='AFS',
            timeout=30
        )
        cursor = connection.cursor()
        
        # Execute the FSA inspection query
        query = FSA_INSPECTION_QUERY
        
        cursor.execute(query)
        
        # Fetch all results and convert to list of dictionaries
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            row_dict = dict(zip(columns, row))
            inspector_id = row_dict.get('InspectorId')
            try:
                # Ensure numeric for mapping
                inspector_id_int = int(inspector_id) if inspector_id is not None else None
            except (TypeError, ValueError):
                inspector_id_int = None
            # Add human-readable inspector name and keep id
            row_dict['Inspector'] = INSPECTOR_NAME_MAP.get(inspector_id_int, 'Unknown')
            sqlserver_data.append(row_dict)
        
        # Close connection
        cursor.close()
        connection.close()
        
    except ImportError as e:
        error_message = f"SQL Server connector not installed. Please install pyodbc. Error: {str(e)}"
    except pyodbc.Error as e:
        error_message = f"SQL Server connection error: {str(e)}"
    except Exception as e:
        error_message = f"Unexpected error: {str(e)}"
        import traceback
        traceback.print_exc()
    
    context = {
        'mysql_data': sqlserver_data,  # Reuse the same template
        'error_message': error_message,
        'data_count': len(sqlserver_data) if sqlserver_data else 0,
        'data_source': 'Remote FSA SQL Server Database - Inspection Records'
    }
    
    return render(request, 'main/remote_data.html', context)


# =============================================================================
# EXPORT VIEWS
# =============================================================================

@login_required(login_url='login')
def export_shipments(request):
    """Export shipment data to different formats (Excel, CSV, PDF)."""
    clear_messages(request)
    
    # Get export format and other parameters
    export_format = request.GET.get('format', 'excel')
    client_id = request.GET.get('client')
    
    # Get shipments with filters if provided - Use optimized queryset
    shipments = Shipment.objects.select_related('client').all()
    shipments = apply_filters(request, shipments)
    
    # Get client name for filename
    client_name = "all_clients"
    if client_id:
        try:
            client = Client.objects.get(pk=client_id)
            client_name = client.name.replace(" ", "_").replace("/", "_")
        except Client.DoesNotExist:
            pass
    
    # Generate timestamp for filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create filename based on client and date
    filename_base = f"claims_{client_name}_{timestamp}"
    

    
    # Export based on selected format
    if export_format == 'excel':
        response = export_to_excel(shipments, filename_base)
        return response
    elif export_format == 'csv':
        response = export_to_csv(shipments, filename_base)
        return response
    elif export_format == 'pdf':
        response = export_to_pdf(shipments, filename_base)
        return response
    else:
        messages.error(request, f"Unsupported export format: {export_format}")
        return redirect('shipment_list')


@login_required(login_url='login')
def export_shipments_excel(request):
    """Legacy function that redirects to the more flexible export_shipments function."""
    return export_shipments(request)


# =============================================================================
# EXPORT HELPER FUNCTIONS - MATCHING TABLE COLUMNS EXACTLY
# =============================================================================

def export_to_excel(shipments, filename_base):
    """Helper function to export data to Excel format - matches table columns exactly."""
    # Create a workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'
    
    # Define headers exactly as shown in the table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, shipment in enumerate(shipments, 2):
        # Use the new client-specific shipment numbers
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields as icons/text
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status badges
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓ Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗ Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~ Partial'
        else:
            settlement_status = '-'
        
        status_display = ''
        if shipment.Status == 'OPEN':
            status_display = '● Open'
        elif shipment.Status == 'CLOSED':
            status_display = '✓ Closed'
        elif shipment.Status == 'PENDING':
            status_display = '⏳ Pending'
        elif shipment.Status == 'REJECTED':
            status_display = '✗ Rejected'
        elif shipment.Status == 'UNDER_REVIEW':
            status_display = '◐ Under Review'
        else:
            status_display = shipment.Status
        
        # Create row data matching table exactly
        row_data = [
            shipment.Claim_No,  # This will now be the new format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Add row to worksheet
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        worksheet.column_dimensions[column_letter].width = adjusted_width
    

    
    # Create response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response


def export_to_csv(shipments, filename_base):
    """Helper function to export data to CSV format - matches table columns exactly."""
    # Create a file-like buffer for response
    response_buffer = io.StringIO()
    
    # Create CSV writer
    response_writer = csv.writer(response_buffer)
    
    # Write header row exactly as shown in table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    response_writer.writerow(headers)
    
    # Write data rows
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "Yes" if shipment.Intent_To_Claim == 'YES' else "No"
        formal_claim = "Yes" if shipment.Formal_Claim_Received == 'YES' else "No"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = 'Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = 'Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = 'Partial'
        else:
            settlement_status = '-'
        
        status_display = shipment.get_Status_display() if shipment.Status else 'Open'
        
        row = [
            shipment.Claim_No,  # New format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Write to response
        response_writer.writerow(row)
    

    
    # Create response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write(response_buffer.getvalue())
    
    return response


def export_to_pdf(shipments, filename_base):
    """Helper function to export data to PDF format - matches table columns exactly."""
    # Create a file-like buffer for the PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object with landscape orientation for all columns
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title=f"Claims Report - {filename_base}",
        topMargin=20,
        bottomMargin=20,
        leftMargin=20,
        rightMargin=20
    )
    
    # Create PDF path for backup
    pdf_backup_path = os.path.join(backup_dir, 'pdf', f"{filename_base}.pdf")
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Add title
    title = Paragraph(f"Claims Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Add spacing
    
    # Define table data exactly matching the web table
    data = [
        ['Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 'Intent', 'Intent Date', 
         'Formal', 'Formal Date', 'Value', 'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 
         'Savings', 'Settlement', 'Exposure', 'Status', 'Closed']
    ]
    
    # Add shipment data
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~'
        else:
            settlement_status = '-'
        
        status_symbol = ''
        if shipment.Status == 'OPEN':
            status_symbol = '●'
        elif shipment.Status == 'CLOSED':
            status_symbol = '✓'
        elif shipment.Status == 'PENDING':
            status_symbol = '⏳'
        elif shipment.Status == 'REJECTED':
            status_symbol = '✗'
        elif shipment.Status == 'UNDER_REVIEW':
            status_symbol = '◐'
        else:
            status_symbol = shipment.Status[:3] if shipment.Status else 'OPN'
        
        # Truncate shipment number for PDF to fit better
        shipment_no_display = shipment.Claim_No
        if len(shipment.Claim_No) > 15:
            shipment_no_display = shipment.Claim_No[:12] + '...'
        
        row = [
            shipment_no_display,  # Truncated for PDF display
            (shipment.Brand or '-')[:8] + '...' if shipment.Brand and len(shipment.Brand) > 8 else (shipment.Brand or '-'),
            (shipment.Claimant or '-')[:10] + '...' if shipment.Claimant and len(shipment.Claimant) > 10 else (shipment.Claimant or '-'),
            client_id,
            client_name[:12] + '...' if len(client_name) > 12 else client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_symbol,
            closed_date
        ]
        data.append(row)
    
    # Create table with smaller font to fit all columns
    table = Table(data, repeatRows=1)
    
    # Add style to table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Header bottom padding
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Data alignment
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Data font
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Smaller data font size
        ('TOPPADDING', (0, 1), (-1, -1), 2),  # Data top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # Data bottom padding
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Grid style
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Box style
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        # Right align amount columns
        ('ALIGN', (9, 1), (12, -1), 'RIGHT'),  # Value, ISCM, Carrier, Insurance
        ('ALIGN', (14, 1), (14, -1), 'RIGHT'),  # Savings
        ('ALIGN', (16, 1), (16, -1), 'RIGHT'),  # Exposure
    ]))
    
    # Add table to elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    

    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    response.write(pdf_data)
    
    return response


# =============================================================================
# IMPORT VIEWS - UPDATED FOR NEW CLIENT-SPECIFIC SHIPMENT IDs
# =============================================================================

@login_required(login_url='login')
def import_shipments(request):
    """Import shipment data from an Excel file, with detailed feedback on the process."""
    clear_messages(request)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return render(request, 'main/import_shipments.html')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            worksheet = wb.active
            skipped_entries, created_entries, error_entries = process_excel_data(worksheet)
            if created_entries == 0 and not skipped_entries and not error_entries:
                messages.info(request, 'No new entries were created. Check if the data is already up to date.')
            else:
                messages.success(request, f'Successfully created {created_entries} entries. Skipped {len(skipped_entries)} duplicate entries.')
                if error_entries:
                    messages.error(request, f'Errors occurred in {len(error_entries)} entries. {", ".join(error_entries)}')
        except Exception as e:
            messages.error(request, f'An error occurred while processing the file: {str(e)}')

        return render(request, 'main/import_shipments.html')
    return render(request, 'main/import_shipments.html')


def process_excel_data(worksheet):
    """Process Excel data and save valid entries to the database."""
    skipped_entries = []
    created_entries = 0
    error_entries = []
    
    # Updated column order for import:
    # 0: Shipment Number (can be blank - will auto-generate), 1: Brand, 2: Claimant, 
    # 3: Intent To Claim, 4: Intent To Claim Date, 5: Formal Claim, 6: Formal Claim Date, 
    # 7: Value, 8: Paid By ISCM, 9: Paid By Carrier, 10: Paid By Insurance, 11: Branch, 
    # 12: Total Savings, 13: Settled or Not Settled, 14: Financial Exposure, 15: Status
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:  # Skip completely empty rows
            continue
            
        # Try to get claim number from column 0, or leave blank for auto-generation
        claim_no = str(row[0]).strip() if row[0] else ""
        
        # Skip if claim number exists and is already in database
        if claim_no and Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue
            
        try:
            # Get Brand (column 1)
            brand = ""
            if len(row) > 1 and row[1]:
                brand = str(row[1]).strip()
            
            # Get Claimant (column 2) - REQUIRED for client identification
            claimant = ""
            if len(row) > 2 and row[2]:
                claimant = str(row[2]).strip()
            
            # Skip rows without claimant as we need it to identify/create client
            if not claimant:
                error_entries.append(f'Row {worksheet.iter_rows().__next__()}: Missing claimant name')
                continue
            
            # Get or create client based on claimant name
            client, created = Client.objects.get_or_create(
                name__iexact=claimant,
                defaults={'name': claimant}
            )
            
            # Handle date conversions for Intent To Claim Date (column 4)
            intend_date = None
            if len(row) > 4 and row[4]:
                if isinstance(row[4], datetime.date):
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    try:
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            intend_date = None
            
            # Handle date conversions for Formal Claim Date (column 6)
            formal_date = None
            if len(row) > 6 and row[6]:
                if isinstance(row[6], datetime.date):
                    formal_date = row[6]
                elif isinstance(row[6], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[6], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            formal_date = datetime.datetime.strptime(row[6], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None
            
            # Handle numeric conversions safely
            # Value (column 7)
            claimed_amount = 0
            if len(row) > 7 and row[7]:
                if isinstance(row[7], (int, float)):
                    claimed_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)
            
            # Paid By ISCM (column 8)
            iscm_amount = 0
            if len(row) > 8 and row[8]:
                if isinstance(row[8], (int, float)):
                    iscm_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        iscm_amount = float(clean_str)
            
            # Paid By Carrier (column 9)
            carrier_amount = 0
            if len(row) > 9 and row[9]:
                if isinstance(row[9], (int, float)):
                    carrier_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)
            
            # Paid By Insurance (column 10)
            insurance_amount = 0
            if len(row) > 10 and row[10]:
                if isinstance(row[10], (int, float)):
                    insurance_amount = float(row[10])
                elif isinstance(row[10], str) and row[10].strip():
                    clean_str = ''.join(c for c in row[10] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)
            
            # Total Savings (column 12)
            total_savings = 0
            if len(row) > 12 and row[12]:
                if isinstance(row[12], (int, float)):
                    total_savings = float(row[12])
                elif isinstance(row[12], str) and row[12].strip():
                    clean_str = ''.join(c for c in row[12] if c.isdigit() or c == '.')
                    if clean_str:
                        total_savings = float(clean_str)
            
            # Financial Exposure (column 14)
            financial_exposure = 0
            if len(row) > 14 and row[14]:
                if isinstance(row[14], (int, float)):
                    financial_exposure = float(row[14])
                elif isinstance(row[14], str) and row[14].strip():
                    clean_str = ''.join(c for c in row[14] if c.isdigit() or c == '.')
                    if clean_str:
                        financial_exposure = float(clean_str)
            
            # Convert Intent To Claim to YES/NO format (column 3)
            intent_claim = "NO"
            if len(row) > 3 and row[3]:
                if isinstance(row[3], bool):
                    intent_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        intent_claim = "YES"
            
            # Convert Formal Claim to YES/NO format (column 5)
            formal_claim = "NO"
            if len(row) > 5 and row[5]:
                if isinstance(row[5], bool):
                    formal_claim = "YES" if row[5] else "NO"
                elif isinstance(row[5], str):
                    if row[5].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"
            
            # Get branch value (column 11)
            branch = ""
            if len(row) > 11 and row[11]:
                branch = str(row[11]).strip()
                # Validate branch code
                if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                    branch = ""  # Set to empty if invalid
            
            # Handle Settlement Status (column 13)
            settlement_status = None
            if len(row) > 13 and row[13]:
                settlement_value = str(row[13]).upper().strip()
                if settlement_value in ["SETTLED", "YES", "Y", "TRUE", "1"]:
                    settlement_status = "SETTLED"
                elif settlement_value in ["NOT SETTLED", "NOT_SETTLED", "NO", "N", "FALSE", "0"]:
                    settlement_status = "NOT_SETTLED"
                elif settlement_value in ["PARTIAL", "PARTIALLY SETTLED", "PARTIAL_SETTLED"]:
                    settlement_status = "PARTIAL"
            
            # Handle Status (column 15)
            status = "OPEN"  # Default status
            if len(row) > 15 and row[15]:
                status_value = str(row[15]).upper().strip()
                valid_statuses = [choice[0] for choice in Shipment.STATUS_CHOICES]
                if status_value in valid_statuses:
                    status = status_value
                elif status_value in ["OPEN", "PENDING", "CLOSED", "REJECTED", "UNDER_REVIEW"]:
                    status = status_value
            
            # Create shipment object - Claim_No will be auto-generated if blank
            shipment = Shipment(
                Claim_No=claim_no,  # Leave blank for auto-generation
                client=client,
                Brand=brand,
                Claimant=claimant,
                Intent_To_Claim=intent_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Received=formal_claim,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Awa=iscm_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Branch=branch,
                Total_Savings=total_savings,
                Settlement_Status=settlement_status,
                Financial_Exposure=financial_exposure,
                Status=status
            )
            shipment.save()  # Auto-generates Claim_No if blank
            created_entries += 1
            
        except Exception as e:
            error_entries.append(f'Row with Claimant {claimant}: {str(e)}')

    return skipped_entries, created_entries, error_entries


@login_required
def get_inspection_fees(request):
    """Get all inspection fees with current and historical rate information"""
    from ..models import InspectionFee

    fees = InspectionFee.objects.all()
    fees_data = []

    for fee in fees:
        # Get the most recent history entry to find current effective date
        latest_history = fee.history.order_by('-effective_date').first()

        fee_dict = {
            'id': fee.id,
            'fee_code': fee.fee_code,
            'fee_name': fee.fee_name,
            'rate': float(fee.rate),
            'description': fee.description,
            'last_updated': fee.last_updated.isoformat() if fee.last_updated else None,
            'effective_date': latest_history.effective_date.isoformat() if latest_history else None,
            'has_history': fee.history.exists(),
            'history_count': fee.history.count()
        }
        fees_data.append(fee_dict)

    return JsonResponse({'fees': fees_data})


@login_required
@require_POST
def update_inspection_fees(request):
    """
    Update inspection fees with versioning support.
    Creates FeeHistory entries instead of directly overwriting rates.
    """
    import json
    from ..models import InspectionFee, FeeHistory
    from datetime import date
    from decimal import Decimal

    try:
        data = json.loads(request.body)
        fees_data = data.get('fees', [])
        effective_date_str = data.get('effective_date')  # Optional: allow setting effective date

        # Parse effective date or default to today
        if effective_date_str:
            try:
                effective_date = date.fromisoformat(effective_date_str)
            except (ValueError, TypeError):
                effective_date = date.today()
        else:
            effective_date = date.today()

        updated_count = 0
        errors = []

        for fee_data in fees_data:
            fee_id = fee_data.get('id')
            new_rate = fee_data.get('rate')

            if fee_id and new_rate is not None:
                try:
                    fee = InspectionFee.objects.get(id=fee_id)
                    new_rate_decimal = Decimal(str(new_rate))

                    # Check if rate actually changed
                    if fee.rate != new_rate_decimal:
                        # Create a new FeeHistory entry
                        FeeHistory.objects.create(
                            fee=fee,
                            rate=new_rate_decimal,
                            effective_date=effective_date,
                            created_by=request.user,
                            notes=fee_data.get('notes', '')  # Optional notes about the change
                        )

                        # Update the current rate on the fee
                        fee.rate = new_rate_decimal
                        fee.updated_by = request.user
                        fee.save()

                        updated_count += 1

                except InspectionFee.DoesNotExist:
                    errors.append(f"Fee with ID {fee_id} not found")
                except Exception as e:
                    errors.append(f"Error updating fee {fee_id}: {str(e)}")

        response_data = {
            'success': True,
            'message': f'Updated {updated_count} fees successfully',
            'updated_count': updated_count,
            'effective_date': effective_date.isoformat()
        }

        if errors:
            response_data['warnings'] = errors

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_inspection_fee_history(request):
    """Get complete history of all fee changes"""
    from ..models import FeeHistory

    # Get all fee history ordered by effective date (most recent first)
    history = FeeHistory.objects.select_related('fee', 'created_by').order_by('-effective_date', '-created_at')

    history_data = []
    for record in history:
        # Get the previous rate by finding the next older history record for the same fee
        previous_history = FeeHistory.objects.filter(
            fee=record.fee,
            effective_date__lt=record.effective_date
        ).order_by('-effective_date').first()

        history_dict = {
            'id': record.id,
            'fee_name': record.fee.fee_name,
            'fee_code': record.fee.fee_code,
            'rate': float(record.rate),
            'previous_rate': float(previous_history.rate) if previous_history else None,
            'effective_date': record.effective_date.isoformat(),
            'created_at': record.created_at.isoformat(),
            'created_by': record.created_by.username if record.created_by else None,
            'notes': record.notes or ''
        }
        history_data.append(history_dict)

    return JsonResponse({'history': history_data})


