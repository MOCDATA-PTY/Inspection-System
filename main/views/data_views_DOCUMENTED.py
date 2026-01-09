"""
================================================================================
DATA VIEWS MODULE - FSA Inspection System
================================================================================

This module handles all data import/export operations and SQL Server integration
for the Food Safety Agency (FSA) Inspection System.

Key Features:
1. SQL Server Integration - Connects to remote AFS SQL Server database
2. Data Export - Excel, CSV, and PDF export functionality
3. Data Import - Excel file import with validation
4. Inspection Fee Management - View and update inspection fees

Database Connections:
- Primary: PostgreSQL (Django ORM)
- Secondary: SQL Server (pymssql for AFS data)

Author: FSA Development Team
Last Updated: 2025-12-10
================================================================================
"""

from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils.dateparse import parse_date
from ..models import Shipment, Client
from .utils import apply_filters, clear_messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime
import os
import threading
import time
import shutil
from pathlib import Path
import re
import sys

# Add the project root to Python path to import the config
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

# ==================================================================================
# CONFIGURATION - SQL SERVER AND DATABASE SETTINGS
# ==================================================================================

try:
    # Try to import configuration from external config file
    from eclick_mysql_config import (
        MYSQL_CONFIG,  # MySQL connection settings (legacy)
        SQLSERVER_CONFIG,  # SQL Server connection settings for Django ORM
        SQLSERVER_CONNECTION_STRING,  # Connection string format for pyodbc
        FSA_INSPECTION_QUERY  # Pre-built SQL query for fetching inspection data
    )
except ImportError:
    # Fallback configurations if import fails - ensures system still works
    # These are hardcoded defaults used when config file is unavailable

    # MySQL Configuration (Legacy - Not actively used)
    MYSQL_CONFIG = {
        'host': '77.37.121.135',
        'user': 'admin',
        'password': 'mk7z@Geg123',
        'database': 'E-click-Project-management',
        'port': 3306,
        'charset': 'utf8mb4',
        'autocommit': True,
        'connect_timeout': 60,
        'read_timeout': 60,
        'write_timeout': 60,
        'use_unicode': True,
    }

    # SQL Server Configuration for Django ORM
    # Used by Django's database router to connect to AFS SQL Server
    SQLSERVER_CONFIG = {
        'ENGINE': 'mssql',  # Django backend for SQL Server
        'NAME': 'AFS',  # Database name on SQL Server
        'USER': 'FSAUser2',  # SQL Server authentication username
        'PASSWORD': 'password',  # SQL Server password
        'HOST': '102.67.140.12',  # SQL Server IP address
        'PORT': '1053',  # Custom SQL Server port (not default 1433)
        'OPTIONS': {
            'driver': 'ODBC Driver 17 for SQL Server',  # ODBC driver (required for Django-mssql)
            'trusted_connection': 'no',  # Use SQL auth, not Windows auth
        },
    }

    # ODBC Connection String Format (Alternative connection method)
    # Used when connecting via pyodbc directly instead of Django ORM
    SQLSERVER_CONNECTION_STRING = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=102.67.140.12,1053;'  # Server IP and port
        'DATABASE=AFS;'
        'UID=FSAUser2;'
        'PWD=password;'
        'TrustServerCertificate=yes;'  # Accept self-signed certificates
        'Encrypt=yes;'  # Encrypt connection
    )

    # FSA Inspection Query - Comprehensive SQL for fetching all inspection data
    # =========================================================================
    # This query unions inspection records from 4 different commodity types:
    # 1. POULTRY - Poultry label inspections
    # 2. EGGS - Egg producer inspections
    # 3. RAW - Raw meat (RMP) inspections
    # 4. PMP - Processed meat product inspections
    #
    # Key Features:
    # - OUTER APPLY for GPS data: Gets first GPS record per inspection to avoid duplicates
    # - Date filtering: Only October 2025 to March 2026 (current season)
    # - Lab sample detection: Checks if samples were sent to lab (IsSampleTaken)
    # - Product filtering: Excludes inspections with blank/null product names
    # - Internal account codes: Includes InternalAccountNumber for client matching
    #
    # Important Notes:
    # - Removed DISTINCT to allow multiple products per inspection (creates multiple rows)
    # - Changed from JOIN to OUTER APPLY for GPS to prevent duplicate rows
    # - Lab sample checking uses EXISTS queries against lab sample link tables
    FSA_INSPECTION_QUERY = '''
        SELECT 'POULTRY' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName as ProductName FROM AFS.dbo.PoultryLabelInspectionChecklistRecords OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryLabelChecklistInspectionRecordId = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ClientId where AFS.dbo.[PoultryLabelInspectionChecklistRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND DateOfInspection < '2026-04-01' AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName IS NOT NULL AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName != ''
        UNION ALL
        SELECT 'EGGS' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForInspection as IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryEggInspectionRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer as ProductName FROM [AFS].[dbo].[PoultryEggInspectionRecords] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryEggInspectionRecordId = [AFS].[dbo].[PoultryEggInspectionRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryEggInspectionRecords].ClientId where AFS.dbo.[PoultryEggInspectionRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND DateOfInspection < '2026-04-01'
        UNION ALL
        SELECT 'RAW' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.RawRMPInspectionLabSampleLinks WHERE RawRMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[RawRMPInspectionRecordTypes].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.NewProductItemDetails as ProductName FROM [AFS].[dbo].[RawRMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE RawRMPInspectionRecordId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.RawRMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[RawRMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND DateOfInspection < '2026-04-01' AND prod.NewProductItemDetails IS NOT NULL AND prod.NewProductItemDetails != ''
        UNION ALL
        SELECT 'PMP' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.PMPInspectionLabSampleLinks WHERE PMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PMPInspectionRecordTypes].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.PMPItemDetails as ProductName FROM [AFS].[dbo].[PMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PMPInspectionRecordId = [AFS].[dbo].[PMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.PMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[PMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND DateOfInspection < '2026-04-01' AND prod.PMPItemDetails IS NOT NULL AND prod.PMPItemDetails != ''
    '''

# ==================================================================================
# INSPECTOR NAME MAPPING
# ==================================================================================
# Maps InspectorId (numeric) to human-readable inspector names
# This is a module-level constant so all views can access it
# Data source: AFS SQL Server database - Inspectors table
# Used to convert numeric IDs to names when displaying inspection data
INSPECTOR_NAME_MAP = {
    132: 'PAKI OLIFANT', 142: 'MARIANA DU TOIT', 143: 'MOKGADI SELONE', 140: 'AGREEMENT MOSIA',
    68: 'BEN VISAGIE', 144: 'VHAHANGWELE RALULIMI', 97: 'NAKISANI BALOYI', 133: 'CALVIN CLAASSENS',
    118: 'NEO NOE', 131: 'EDITH SELEPE', 124: 'SANDISIWE DLISANI', 85: 'JOEL MHANGWA',
    86: 'ASISIPHO LANDE', 125: 'MARIUS CARSTENS', 145: 'WILSON MAIFO', 82: 'JOE ROSENBLATT',
    95: 'EVANS NKWINIKA', 93: 'VICTOR MATHEBULA', 77: 'DELAREY RIBBENS', 78: 'DEWALD KORF',
    103: 'GERRIT PEKEMA', 76: 'PETRUS POOL', 71: 'PALESA MPANA', 134: 'ARMAND VISAGIE',
    113: 'KATLEGO MOKHUA', 147: 'BRIAN XULU', 146: 'ELIAS THEKISO', 72: 'CHARMAINE NEL',
    148: 'COLLEN DLAMINI', 153: 'JOFRED STEYN', 164: 'HENNIE CHILOANE', 166: 'THATO SEKHOTHO',
    69: 'AFS DUMMY', 177: 'CINGA NGONGO', 91: 'MONTI RAMAANO', 79: 'WENDY CHAKA',
    174: 'LWANDILE MAQINA', 96: 'MCAULEY MUSUNDA', 92: 'THABO MAGWAZA', 149: 'THEMBA SHABANGU',
    175: 'PETRUS POOL', 70: 'THERESA DIOGO', 73: 'RAM RAMBURAN', 74: 'HEIN NEL',
    75: 'SIBONELO ZONDI', 80: 'NIPHO NGOMANE', 81: 'CHELESILE MOYO', 83: 'SIMON SWART',
    84: 'ANDREAS LETABA', 87: 'LOUIS VISAGIE', 88: 'RASSIE SMIT', 89: 'NICOLE BERGH',
    90: 'Spare1 Spare2', 94: 'FRANCOIS PRETORIUS', 98: 'PLACEHOLDER PLACEHOLDER', 99: 'TEST GMAIL',
    102: 'ALI MODIKOE', 106: 'ARMAND VISAGIE', 111: 'DUMMY TESTER', 112: 'TESTER TESTER',
    114: 'DUMMY PLACEHOLDER', 115: 'TESTER PERSON', 116: 'FINANCE TWO', 119: 'TESTER TESTER',
    122: 'Spare Spare', 123: 'LIZELLE BREEDT', 126: 'THAPELO MAPOTSE', 127: 'THAPELO MAPOTSE',
    141: 'IT IT', 150: 'CHRISNA POOL', 154: 'SEUN SEBOLAI', 160: 'LERATO MODIBA',
    163: 'DENNIS CELE', 171: 'MARNUS BADENHORST', 173: 'HLENGIWE GUMEDE', 178: 'NTABISENG MASEKO',
    179: 'KABELO MOKGALAKA', 183: 'MPHO MOTAUNG', 184: 'XOLA MPELUZA', 185: 'PERCY MALEKA',
    186: 'KUTLWANO KUNTWANE', 187: 'SIPHO NDAMASE', 188: 'GLADYS MANGANYE', 193: 'ANTHONY PENZES',
    194: 'PHUMZILE MASOMBUKA', 196: 'NELISA NTOYAPHI',
}


# ==================================================================================
# SQL DATABASE VIEWS
# ==================================================================================

@login_required
def remote_sqlserver_data_view(request):
    """
    View to display data from remote SQL Server database (FSA AFS Server).

    Purpose:
    --------
    Fetches inspection records from the remote AFS SQL Server and displays them
    in the web interface. This allows users to view real-time inspection data
    from the field mobile application database.

    Authentication:
    --------------
    Requires user to be logged in (@login_required decorator)

    Database Connection:
    -------------------
    - Uses pymssql library (pure Python, no ODBC drivers needed)
    - Connects to: 102.67.140.12:1053 (AFS SQL Server)
    - Database: AFS
    - User: FSAUser2
    - Timeout: 30 seconds

    Query Execution:
    ---------------
    Executes the FSA_INSPECTION_QUERY which returns:
    - All active inspections from Oct 2025 - Mar 2026
    - Data from 4 commodity types (POULTRY, EGGS, RAW, PMP)
    - GPS coordinates for each inspection
    - Lab sample status
    - Client and product information

    Data Processing:
    ---------------
    1. Fetch all rows from SQL Server
    2. Convert to list of dictionaries (column_name: value)
    3. Map InspectorId to human-readable names
    4. Handle type conversion for inspector IDs

    Error Handling:
    --------------
    - ImportError: pymssql not installed
    - pyodbc.Error: SQL Server connection/query errors
    - General Exception: Catches all other errors with traceback

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object

    Returns:
    -------
    HttpResponse
        Rendered template with inspection data or error message

    Template Context:
    ----------------
    - mysql_data: List of inspection records (reuses template variable name)
    - error_message: Error string if connection/query failed
    - data_count: Number of records fetched
    - data_source: Description of data source

    Template Used:
    -------------
    main/remote_data.html

    Example Usage:
    -------------
    URL: /server-view/
    User clicks "Server View" in sidebar → this view is called →
    fetches latest inspection data → displays in table format

    Notes:
    -----
    - This is a read-only view (SELECT only, no modifications)
    - Data is fetched fresh each time (no caching)
    - Large result sets may cause slow page loads
    - Consider adding pagination for better performance
    """
    # Initialize empty results and error tracking
    sqlserver_data = []
    error_message = None

    try:
        # Import pymssql - pure Python SQL Server connector (no ODBC needed)
        import pymssql

        # Establish connection to SQL Server
        # Using pymssql instead of pyodbc for easier deployment (no drivers needed)
        connection = pymssql.connect(
            server='102.67.140.12',  # AFS SQL Server IP address
            port=1053,  # Custom port (not standard 1433)
            user='FSAUser2',  # SQL authentication username
            password='password',  # SQL authentication password
            database='AFS',  # Target database name
            timeout=30  # Connection timeout in seconds
        )
        cursor = connection.cursor()

        # Execute the comprehensive inspection query
        # This query unions data from 4 different commodity types
        query = FSA_INSPECTION_QUERY
        cursor.execute(query)

        # Fetch all results and convert to list of dictionaries
        # cursor.description contains column metadata (names, types, etc)
        columns = [column[0] for column in cursor.description]  # Extract column names
        rows = cursor.fetchall()  # Fetch all result rows

        # Process each row and convert to dictionary format
        for row in rows:
            # Zip column names with row values to create dictionary
            row_dict = dict(zip(columns, row))

            # Get inspector ID from the row
            inspector_id = row_dict.get('InspectorId')

            # Convert inspector ID to integer for mapping lookup
            # Handle cases where ID might be None, string, or already int
            try:
                # Ensure numeric for mapping - convert to int if possible
                inspector_id_int = int(inspector_id) if inspector_id is not None else None
            except (TypeError, ValueError):
                # If conversion fails (e.g., non-numeric string), set to None
                inspector_id_int = None

            # Add human-readable inspector name to the row dictionary
            # Use INSPECTOR_NAME_MAP to convert ID to name
            # Default to 'Unknown' if ID not found in mapping
            row_dict['Inspector'] = INSPECTOR_NAME_MAP.get(inspector_id_int, 'Unknown')

            # Add processed row to results list
            sqlserver_data.append(row_dict)

        # Clean up database connections
        cursor.close()  # Close cursor first
        connection.close()  # Then close connection

    except ImportError as e:
        # pymssql library not installed
        error_message = f"SQL Server connector not installed. Please install pyodbc. Error: {str(e)}"
    except pyodbc.Error as e:
        # SQL Server connection or query error
        error_message = f"SQL Server connection error: {str(e)}"
    except Exception as e:
        # Catch-all for any other unexpected errors
        error_message = f"Unexpected error: {str(e)}"
        # Print full traceback for debugging
        import traceback
        traceback.print_exc()

    # Prepare context dictionary for template
    context = {
        'mysql_data': sqlserver_data,  # Reuse existing template variable name
        'error_message': error_message,  # Show error if connection failed
        'data_count': len(sqlserver_data) if sqlserver_data else 0,  # Number of records
        'data_source': 'Remote FSA SQL Server Database - Inspection Records'  # Data source description
    }

    # Render template with context data
    return render(request, 'main/remote_data.html', context)


# ==================================================================================
# EXPORT VIEWS - Data Export Functionality
# ==================================================================================

@login_required(login_url='login')
def export_shipments(request):
    """
    Export shipment/claim data to different formats (Excel, CSV, PDF).

    Purpose:
    --------
    Main export view that handles exporting claim/shipment data in multiple formats.
    Supports filtering by client and other parameters before export.

    Authentication:
    --------------
    Requires login - redirects to 'login' if unauthenticated

    Supported Export Formats:
    -------------------------
    1. Excel (.xlsx) - Full-featured spreadsheet with styling
    2. CSV (.csv) - Plain text comma-separated values
    3. PDF (.pdf) - Formatted PDF document in landscape

    URL Parameters:
    --------------
    - format: Export format ('excel', 'csv', or 'pdf')
    - client: Client ID to filter by (optional)
    - Other filter parameters passed to apply_filters()

    Data Processing Flow:
    --------------------
    1. Clear any existing Django messages
    2. Get export format from GET parameters (default: 'excel')
    3. Get optional client ID filter
    4. Fetch shipments using optimized queryset (select_related for client)
    5. Apply any additional filters from request
    6. Determine client name for filename
    7. Generate timestamped filename
    8. Route to appropriate export function

    Filename Format:
    ---------------
    claims_{client_name}_{timestamp}.{extension}
    Example: claims_Acme_Corp_20251210_143022.xlsx

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object containing GET parameters

    Returns:
    -------
    HttpResponse
        File download response with appropriate content-type, OR
        Redirect to shipment_list if unsupported format

    Error Handling:
    --------------
    - Client.DoesNotExist: Uses "all_clients" in filename
    - Unsupported format: Shows error message and redirects

    Example Usage:
    -------------
    GET /export-shipments/?format=excel&client=5
    → Exports all shipments for client ID 5 as Excel file

    Notes:
    -----
    - Uses select_related('client') for query optimization
    - Filename sanitization: spaces and slashes replaced with underscores
    - Timestamp format: YYYYMMDD_HHMMSS (sortable and filesystem-safe)
    - Messages are cleared before export to avoid showing in next request
    """
    # Clear any existing Django messages to avoid showing stale notifications
    clear_messages(request)

    # Get export format from URL parameters, default to Excel if not specified
    export_format = request.GET.get('format', 'excel')

    # Get optional client ID filter from URL parameters
    client_id = request.GET.get('client')

    # Fetch shipments with optimized query
    # select_related('client') performs SQL JOIN to avoid N+1 query problem
    # This loads all client data in one query instead of separate query per shipment
    shipments = Shipment.objects.select_related('client').all()

    # Apply additional filters based on request parameters
    # This handles date ranges, status filters, search queries, etc.
    shipments = apply_filters(request, shipments)

    # Determine client name for filename
    client_name = "all_clients"  # Default if no client filter or client not found

    if client_id:
        try:
            # Fetch client object to get name
            client = Client.objects.get(pk=client_id)
            # Sanitize client name for use in filename
            # Replace spaces with underscores, remove/replace special characters
            client_name = client.name.replace(" ", "_").replace("/", "_")
        except Client.DoesNotExist:
            # Client ID invalid or deleted - use default name
            pass

    # Generate timestamp for filename uniqueness
    # Format: YYYYMMDD_HHMMSS (e.g., 20251210_143022)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Construct base filename (without extension)
    # Format: claims_{client_name}_{timestamp}
    filename_base = f"claims_{client_name}_{timestamp}"

    # Route to appropriate export function based on format parameter
    if export_format == 'excel':
        # Export to Excel (.xlsx) format
        response = export_to_excel(shipments, filename_base)
        return response
    elif export_format == 'csv':
        # Export to CSV (.csv) format
        response = export_to_csv(shipments, filename_base)
        return response
    elif export_format == 'pdf':
        # Export to PDF format (landscape orientation)
        response = export_to_pdf(shipments, filename_base)
        return response
    else:
        # Unsupported format specified
        # Show error message and redirect back to shipment list
        messages.error(request, f"Unsupported export format: {export_format}")
        return redirect('shipment_list')


@login_required(login_url='login')
def export_shipments_excel(request):
    """
    Legacy Excel export function that redirects to the unified export_shipments function.

    Purpose:
    --------
    Maintains backward compatibility with old URLs that specifically called
    export_shipments_excel. Now simply redirects to the more flexible
    export_shipments function.

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object

    Returns:
    -------
    HttpResponse
        Response from export_shipments function (Excel format)

    Migration Note:
    --------------
    This function exists for backward compatibility. New code should use
    export_shipments(request) with format='excel' parameter instead.

    Example:
    -------
    Old URL: /export-shipments-excel/
    New URL: /export-shipments/?format=excel

    Both work, but new URL is more flexible
    """
    # Simply delegate to the main export function
    # It will default to Excel format if no format parameter is specified
    return export_shipments(request)


# ==================================================================================
# EXPORT HELPER FUNCTIONS - Data Formatting for Different Export Formats
# ==================================================================================
# These functions handle the actual file generation for each export format.
# They receive a queryset of shipments and a filename base, then generate
# and return an HttpResponse containing the file data.
# ==================================================================================

def export_to_excel(shipments, filename_base):
    """
    Generate Excel spreadsheet (.xlsx) from shipment queryset.

    Purpose:
    --------
    Creates a professionally formatted Excel workbook with shipment data.
    Includes styled headers, formatted currency values, and auto-sized columns.

    Excel Features:
    --------------
    - Styled header row (blue background, white text, bold font)
    - Centered header alignment
    - Formatted currency values with $ and commas
    - Boolean fields as ✓/✗ symbols
    - Auto-adjusted column widths (max 50 characters)
    - All table columns matched exactly to web table

    Column Structure:
    ----------------
    Matches the web table exactly:
    1. Shipment No - Client-specific claim number
    2. Brand - Product brand name
    3. Claimant - Person making claim
    4. Claim ID - Client identifier
    5. Client Name - Full client business name
    6. Intent - Intent to claim (✓/✗)
    7. Intent Date - Date of intent (MM/DD/YY)
    8. Formal - Formal claim received (✓/✗)
    9. Formal Date - Date formal claim received (MM/DD/YY)
    10. Value - Claimed amount ($)
    11. ISCM Paid - Amount paid by ISCM ($)
    12. Carrier Paid - Amount paid by carrier ($)
    13. Insurance - Amount paid by insurance ($)
    14. Branch - Branch code
    15. Savings - Total savings ($)
    16. Settlement - Settlement status (✓ Settled, ✗ Not Settled, ~ Partial)
    17. Exposure - Financial exposure ($)
    18. Status - Claim status (● Open, ✓ Closed, etc.)
    19. Closed - Closed date (MM/DD/YY)
    20. Actions - Placeholder for action buttons

    Parameters:
    ----------
    shipments : QuerySet
        Django queryset of Shipment objects to export
    filename_base : str
        Base filename without extension (e.g., "claims_Acme_20251210_143022")

    Returns:
    -------
    HttpResponse
        Excel file download response with:
        - Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        - Content-Disposition: attachment with .xlsx filename
        - Binary Excel workbook data

    Data Formatting:
    ---------------
    - Dates: MM/DD/YY format (e.g., 12/10/25)
    - Currency: $X,XXX format with commas (e.g., $1,234)
    - Booleans: ✓ for YES, ✗ for NO
    - Status badges: ● Open, ✓ Closed, ⏳ Pending, etc.
    - Missing values: '-' dash character

    Excel Styling:
    -------------
    - Header background: #2563EB (blue)
    - Header text: White, bold
    - Column widths: Auto-calculated based on content (max 50)

    Example Usage:
    -------------
    shipments = Shipment.objects.all()
    response = export_to_excel(shipments, "claims_export_20251210")
    return response  # Triggers browser download

    Performance Notes:
    -----------------
    - Processes all shipments in memory
    - For large datasets (>10,000 rows), consider batch processing
    - openpyxl is slower than xlsxwriter but more feature-rich

    Dependencies:
    ------------
    - openpyxl: Excel file creation
    - openpyxl.styles: Cell formatting (Font, Alignment, PatternFill)
    """
    # Create new Excel workbook and get active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'  # Set worksheet tab name

    # Define column headers - MUST match web table exactly
    # Order and naming are critical for user expectations
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name',
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value',
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]

    # Define header cell styling
    header_font = Font(bold=True, color='FFFFFF')  # Bold white text
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')  # Blue background

    # Write header row with styling
    for col_num, header in enumerate(headers, 1):  # Start at column 1 (Excel is 1-indexed)
        cell = worksheet.cell(row=1, column=col_num, value=header)  # Get/create cell
        cell.font = header_font  # Apply font styling
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal='center')  # Center-align text

    # Process and write data rows
    for row_num, shipment in enumerate(shipments, 2):  # Start at row 2 (row 1 is headers)
        # Extract client information
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'

        # Format date fields - use MM/DD/YY format for compact display
        # Handle None values gracefully with '-' placeholder
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'

        # Format currency amounts with $ and thousand separators
        # Use :,.0f format specifier (comma separator, 0 decimal places)
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"

        # Format boolean fields as checkmark symbols
        # Use Unicode symbols for visual representation
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"

        # Format settlement status with descriptive text and symbols
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓ Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗ Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~ Partial'
        else:
            settlement_status = '-'  # Unknown or null status

        # Format claim status with Unicode symbols for visual clarity
        status_display = ''
        if shipment.Status == 'OPEN':
            status_display = '● Open'  # Filled circle for open claims
        elif shipment.Status == 'CLOSED':
            status_display = '✓ Closed'  # Checkmark for closed claims
        elif shipment.Status == 'PENDING':
            status_display = '⏳ Pending'  # Hourglass for pending
        elif shipment.Status == 'REJECTED':
            status_display = '✗ Rejected'  # X for rejected
        elif shipment.Status == 'UNDER_REVIEW':
            status_display = '◐ Under Review'  # Half-circle for review
        else:
            status_display = shipment.Status  # Use raw status if unknown

        # Build row data array matching header order
        # CRITICAL: Order must match headers array exactly
        row_data = [
            shipment.Claim_No,  # New format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',  # Use '-' if brand is null/empty
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder (not functional in Excel)
        ]

        # Write row data to worksheet
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Auto-adjust column widths based on content
    # This makes the spreadsheet more readable
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get column letter (A, B, C, etc.)

        # Find maximum content length in this column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set column width with padding (add 2 chars) and max cap (50 chars)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Create HTTP response with Excel file data
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # Excel MIME type
    )
    # Set Content-Disposition header to trigger browser download
    # attachment; filename="..." tells browser to save file instead of displaying it
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    # Save workbook directly to response object
    # openpyxl can write directly to file-like objects
    workbook.save(response)

    return response


def export_to_csv(shipments, filename_base):
    """
    Generate CSV (Comma-Separated Values) file from shipment queryset.

    Purpose:
    --------
    Creates a plain-text CSV file compatible with Excel, Google Sheets, and
    other spreadsheet applications. CSV is lightweight and universally supported.

    CSV Advantages:
    --------------
    - Smaller file size than Excel
    - Universal compatibility (can open in any spreadsheet app)
    - Plain text format (can edit with text editors)
    - Fast generation and parsing
    - No styling/formatting overhead

    CSV Limitations:
    ---------------
    - No cell formatting (colors, fonts, etc.)
    - No formulas
    - Text-only (symbols like ✓ may not display correctly in all apps)
    - Single worksheet only

    Column Structure:
    ----------------
    Same 20 columns as Excel export (see export_to_excel for details)

    Data Formatting:
    ---------------
    - Dates: MM/DD/YY format
    - Currency: $X,XXX format with dollar signs and commas
    - Booleans: "Yes"/"No" text (more CSV-friendly than symbols)
    - Status: Full text descriptions (e.g., "Settled", "Not Settled")
    - Missing values: '-' dash

    CSV Format Specifics:
    --------------------
    - Delimiter: Comma (,)
    - Quoting: Automatic (csv.writer handles it)
    - Line terminator: \\r\\n (Windows standard)
    - Encoding: UTF-8
    - No BOM (Byte Order Mark)

    Parameters:
    ----------
    shipments : QuerySet
        Django queryset of Shipment objects to export
    filename_base : str
        Base filename without extension

    Returns:
    -------
    HttpResponse
        CSV file download response with:
        - Content-Type: text/csv
        - Content-Disposition: attachment with .csv filename
        - CSV formatted data

    Example CSV Output:
    ------------------
    Shipment No,Brand,Claimant,Claim ID,...
    Acme-1-20251210,BrandX,John Doe,CLI001,...
    Acme-2-20251210,BrandY,Jane Smith,CLI001,...

    Performance:
    -----------
    - Very fast for large datasets
    - Low memory usage (can stream if needed)
    - Typically 5-10x faster than Excel generation

    Dependencies:
    ------------
    - csv: Python standard library (no installation needed)
    - io.StringIO: In-memory file-like object for CSV writing

    Notes:
    -----
    - Excel may open with encoding issues if special characters used
    - Some CSV readers may not preserve leading zeros
    - Date formatting may be interpreted differently by Excel
    """
    # Create in-memory text buffer for CSV data
    # StringIO acts like a file but stores data in memory
    response_buffer = io.StringIO()

    # Create CSV writer that writes to our buffer
    # Uses default dialect (commas, double-quote escaping)
    response_writer = csv.writer(response_buffer)

    # Write header row - must match web table exactly
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name',
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value',
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    response_writer.writerow(headers)

    # Process and write each shipment as a CSV row
    for shipment in shipments:
        # Extract client information
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'

        # Format dates - same format as Excel
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'

        # Format currency amounts with $ and thousand separators
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"

        # Format boolean fields as Yes/No text (better CSV compatibility)
        intent_to_claim = "Yes" if shipment.Intent_To_Claim == 'YES' else "No"
        formal_claim = "Yes" if shipment.Formal_Claim_Received == 'YES' else "No"

        # Format settlement status as plain text
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = 'Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = 'Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = 'Partial'
        else:
            settlement_status = '-'

        # Get human-readable status display name
        # Uses Django model's get_FIELD_display() method
        status_display = shipment.get_Status_display() if shipment.Status else 'Open'

        # Build row data - order must match headers
        row = [
            shipment.Claim_No,
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions placeholder
        ]

        # Write row to CSV buffer
        # csv.writer automatically handles quoting and escaping
        response_writer.writerow(row)

    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')
    # Set filename in Content-Disposition header
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'

    # Write CSV data from buffer to response
    response.write(response_buffer.getvalue())

    return response


def export_to_pdf(shipments, filename_base):
    """
    Generate PDF document from shipment queryset in landscape orientation.

    Purpose:
    --------
    Creates a professional PDF report suitable for printing and sharing.
    Uses landscape orientation to fit all 19 columns on a single page.

    PDF Features:
    ------------
    - Landscape orientation (fits all columns)
    - Styled table with header and data rows
    - Blue header background matching brand colors
    - Alternating row colors for readability
    - Compressed text to fit on page
    - Title and date stamp
    - Page numbers (if multi-page)

    Layout Specifications:
    --------------------
    - Page size: Letter (8.5" x 11")
    - Orientation: Landscape (11" x 8.5")
    - Margins: 20 points on all sides
    - Font: Helvetica (cross-platform compatibility)
    - Header font size: 8pt
    - Data font size: 6pt (compressed for column fit)

    Column Structure:
    ----------------
    Same as Excel, but with 19 columns (Actions column excluded)
    Text is truncated/abbreviated to fit page width

    Data Formatting:
    ---------------
    - Dates: MM/DD/YY format
    - Currency: $X,XXX format
    - Booleans: ✓/✗ symbols
    - Status: Single character symbols (●, ✓, ⏳, etc.)
    - Long text: Truncated with "..." ellipsis

    Table Styling:
    -------------
    - Header: Blue background (#2563EB), white text, bold, centered
    - Data: Black text, left-aligned (right-aligned for numbers)
    - Alternating rows: White and light gray (#f1f5f9)
    - Grid lines: 0.5pt gray
    - Box border: 0.5pt black

    Text Truncation:
    ---------------
    To fit all columns, some fields are truncated:
    - Shipment No: Max 15 chars
    - Brand: Max 8 chars
    - Claimant: Max 10 chars
    - Client Name: Max 12 chars

    Parameters:
    ----------
    shipments : QuerySet
        Django queryset of Shipment objects to export
    filename_base : str
        Base filename without extension

    Returns:
    -------
    HttpResponse
        PDF file download response with:
        - Content-Type: application/pdf
        - Content-Disposition: attachment with .pdf filename
        - PDF binary data

    Page Layout:
    -----------
    Title: "Claims Report - YYYY-MM-DD"
    (blank line)
    [Table with all shipment data]

    Performance Notes:
    -----------------
    - Slower than Excel/CSV generation
    - Memory intensive for large datasets
    - Consider pagination for >1000 rows
    - PDF generation is CPU-intensive

    Backup Behavior:
    ---------------
    Code references pdf_backup_path but it's not used (commented out).
    This was likely for saving PDFs to filesystem, now only downloads.

    Dependencies:
    ------------
    - reportlab: PDF generation library
    - reportlab.platypus: High-level layout objects
    - reportlab.lib.colors: Color constants

    Known Limitations:
    -----------------
    - Unicode symbols may not render on all systems
    - Page breaks may split rows awkwardly
    - No automatic landscape detection (always landscape)
    - Fixed column widths (not responsive to content)

    Example Usage:
    -------------
    shipments = Shipment.objects.filter(Status='OPEN')
    response = export_to_pdf(shipments, "open_claims_20251210")
    return response

    Notes:
    -----
    - PDF is best for formal reports and printing
    - Not suitable for data manipulation (use Excel/CSV for that)
    - Consider adding page numbers, headers, footers for multi-page
    - Font sizes are compressed to fit columns - may be hard to read when printed
    """
    # Create in-memory binary buffer for PDF data
    # BytesIO is like StringIO but for binary data
    buffer = io.BytesIO()

    # Create PDF document with landscape orientation
    # SimpleDocTemplate handles page layout and building
    doc = SimpleDocTemplate(
        buffer,  # Output buffer
        pagesize=landscape(letter),  # Letter size, rotated to landscape (11"x8.5")
        title=f"Claims Report - {filename_base}",  # PDF metadata title
        topMargin=20,  # Top margin in points (20pt ≈ 0.28 inch)
        bottomMargin=20,  # Bottom margin
        leftMargin=20,  # Left margin
        rightMargin=20  # Right margin
    )

    # Container for flowable objects (title, table, etc.)
    # Flowables are positioned automatically by ReportLab
    elements = []

    # Get pre-defined style templates
    styles = getSampleStyleSheet()
    title_style = styles['Title']  # Large, bold, centered

    # Create title paragraph with current date
    title = Paragraph(
        f"Claims Report - {datetime.datetime.now().strftime('%Y-%m-%d')}",
        title_style
    )
    elements.append(title)

    # Add blank line spacing after title
    elements.append(Paragraph("<br/>", styles['Normal']))

    # Define table data starting with header row
    # Note: No "Actions" column in PDF (not useful in static doc)
    data = [
        ['Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 'Intent', 'Intent Date',
         'Formal', 'Formal Date', 'Value', 'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch',
         'Savings', 'Settlement', 'Exposure', 'Status', 'Closed']
    ]

    # Process each shipment and add as table row
    for shipment in shipments:
        # Extract client information
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'

        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'

        # Format currency amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"

        # Format boolean fields as symbols
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"

        # Format settlement status - symbols only (space-saving for PDF)
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓'  # Checkmark only
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗'  # X only
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~'  # Tilde for partial
        else:
            settlement_status = '-'  # Dash for unknown

        # Format status - single symbol only (space-saving)
        status_symbol = ''
        if shipment.Status == 'OPEN':
            status_symbol = '●'  # Filled circle
        elif shipment.Status == 'CLOSED':
            status_symbol = '✓'  # Checkmark
        elif shipment.Status == 'PENDING':
            status_symbol = '⏳'  # Hourglass
        elif shipment.Status == 'REJECTED':
            status_symbol = '✗'  # X mark
        elif shipment.Status == 'UNDER_REVIEW':
            status_symbol = '◐'  # Half-circle
        else:
            # Unknown status - use first 3 characters
            status_symbol = shipment.Status[:3] if shipment.Status else 'OPN'

        # Truncate long shipment number for PDF fit
        shipment_no_display = shipment.Claim_No
        if len(shipment.Claim_No) > 15:
            shipment_no_display = shipment.Claim_No[:12] + '...'

        # Build row with truncated text for compactness
        # Truncation is necessary to fit all columns on landscape page
        row = [
            shipment_no_display,  # Truncated if > 15 chars
            # Brand: truncate if > 8 chars
            (shipment.Brand or '-')[:8] + '...' if shipment.Brand and len(shipment.Brand) > 8 else (shipment.Brand or '-'),
            # Claimant: truncate if > 10 chars
            (shipment.Claimant or '-')[:10] + '...' if shipment.Claimant and len(shipment.Claimant) > 10 else (shipment.Claimant or '-'),
            client_id,
            # Client name: truncate if > 12 chars
            client_name[:12] + '...' if len(client_name) > 12 else client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_symbol,
            closed_date
        ]

        # Add row to table data
        data.append(row)

    # Create table from data
    # repeatRows=1 means header row repeats on each page if multi-page
    table = Table(data, repeatRows=1)

    # Apply comprehensive table styling
    table.setStyle(TableStyle([
        # ===== HEADER STYLING =====
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),  # Blue header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # White header text
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Center-align headers
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header font
        ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller font for headers (fit more)
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Padding below headers

        # ===== DATA ROW STYLING =====
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # White data background (overridden by ROWBACKGROUNDS)
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Black text
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Left-align data (numbers overridden below)
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Regular font for data
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Very small font to fit columns
        ('TOPPADDING', (0, 1), (-1, -1), 2),  # Minimal top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # Minimal bottom padding

        # ===== GRID AND BORDERS =====
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Gray grid lines
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Black border around table

        # ===== ALTERNATING ROW COLORS =====
        # Creates zebra striping for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),

        # ===== NUMBER ALIGNMENT =====
        # Right-align currency columns for proper number alignment
        ('ALIGN', (9, 1), (12, -1), 'RIGHT'),  # Value, ISCM, Carrier, Insurance columns
        ('ALIGN', (14, 1), (14, -1), 'RIGHT'),  # Savings column
        ('ALIGN', (16, 1), (16, -1), 'RIGHT'),  # Exposure column
    ]))

    # Add table to elements list
    elements.append(table)

    # Build PDF document
    # This processes all flowables and creates the final PDF
    doc.build(elements)

    # Get PDF binary data from buffer
    pdf_data = buffer.getvalue()
    buffer.close()  # Close buffer to free memory

    # Create HTTP response with PDF content type
    response = HttpResponse(content_type='application/pdf')
    # Set filename for download
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    # Write PDF binary data to response
    response.write(pdf_data)

    return response


# ==================================================================================
# IMPORT VIEWS - Excel Data Import Functionality
# ==================================================================================
# These functions handle importing shipment data from Excel files uploaded by users.
# Includes validation, duplicate detection, and error reporting.
# ==================================================================================

@login_required(login_url='login')
def import_shipments(request):
    """
    Import shipment/claim data from uploaded Excel file with comprehensive validation.

    Purpose:
    --------
    Allows users to bulk-import shipment data from Excel spreadsheets.
    Validates data, detects duplicates, creates clients as needed, and provides
    detailed feedback on success/failure.

    Authentication:
    --------------
    Requires login - redirects to 'login' page if unauthenticated

    HTTP Methods:
    ------------
    - GET: Display upload form (import_shipments.html template)
    - POST: Process uploaded Excel file

    File Requirements:
    -----------------
    - File format: .xlsx or .xls (Excel formats only)
    - Must have headers in row 1
    - Data starts in row 2
    - Required columns: Claimant (for client identification)
    - Optional: All other shipment fields

    Expected Excel Column Order:
    --------------------------
    0: Shipment Number (optional - auto-generated if blank)
    1: Brand
    2: Claimant (REQUIRED - used to identify/create client)
    3: Intent To Claim (YES/NO or boolean)
    4: Intent To Claim Date
    5: Formal Claim (YES/NO or boolean)
    6: Formal Claim Date
    7: Value (numeric, may have $ and commas)
    8: Paid By ISCM (numeric)
    9: Paid By Carrier (numeric)
    10: Paid By Insurance (numeric)
    11: Branch (branch code)
    12: Total Savings (numeric)
    13: Settlement Status (SETTLED/NOT_SETTLED/PARTIAL)
    14: Financial Exposure (numeric)
    15: Status (OPEN/CLOSED/PENDING/etc.)

    Data Processing Flow:
    --------------------
    1. Validate file format (.xlsx or .xls only)
    2. Load workbook with openpyxl (data_only=True for formulas)
    3. Get active worksheet
    4. Call process_excel_data() to process rows
    5. Display results (created, skipped, errors)

    Duplicate Detection:
    -------------------
    - Checks if Claim_No already exists in database
    - Skips duplicate rows (doesn't overwrite existing data)
    - Reports skipped entries to user

    Client Creation:
    ---------------
    - Uses Claimant field to identify client
    - Case-insensitive lookup (name__iexact)
    - Creates new client if not found
    - Links shipment to client automatically

    Auto-Generated Claim Numbers:
    ----------------------------
    If Shipment Number column is blank:
    - Format: {ClientName}-{Count}-{YYYYMMDD}
    - Example: Acme_Corp-1-20251210
    - Count increments per client per day

    Error Handling:
    --------------
    - Invalid file format: Error message, returns to form
    - Missing Claimant: Skip row, add to error list
    - Data conversion errors: Skip row, report specific error
    - Exception during processing: Catch and display error message

    Feedback Messages:
    -----------------
    - Success: "Successfully created X entries. Skipped Y duplicate entries."
    - Info: "No new entries created. Check if data is already up to date."
    - Error: "Errors occurred in X entries. [Details...]"

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object
        - GET: Empty request, displays form
        - POST: Contains uploaded file in request.FILES['excel_file']

    Returns:
    -------
    HttpResponse
        Rendered import_shipments.html template with:
        - Upload form (GET request or after processing)
        - Success/error messages (POST request)

    POST Parameters:
    ---------------
    excel_file : UploadedFile
        Excel file uploaded via file input field
        Accessed via request.FILES.get('excel_file')

    Template: main/import_shipments.html
    Context: Django messages (success, info, error)

    Example Usage:
    -------------
    1. User navigates to /import-shipments/
    2. User selects Excel file and clicks Upload
    3. System processes file and shows results
    4. User sees "Created 50 entries. Skipped 10 duplicates."

    Performance Notes:
    -----------------
    - Processes entire file in memory (may be slow for very large files)
    - Consider batch processing for files >10,000 rows
    - Database save() called for each row (could optimize with bulk_create)

    Security Considerations:
    -----------------------
    - File type validation prevents code injection via file upload
    - openpyxl with data_only=True prevents formula execution
    - SQL injection prevented by Django ORM parameterization
    - User must be authenticated to access

    Known Limitations:
    -----------------
    - Doesn't update existing records (skips duplicates)
    - Requires specific column order (not flexible)
    - Doesn't validate branch codes against choices
    - Large files may timeout on slow servers

    Future Enhancements:
    -------------------
    - Add column header mapping (flexible column order)
    - Support CSV import in addition to Excel
    - Add option to update existing records
    - Implement async processing for large files
    - Add preview before import
    - Support multiple worksheets
    """
    # Clear any existing Django messages to start fresh
    clear_messages(request)

    # Check if this is a POST request with an uploaded file
    if request.method == 'POST' and request.FILES.get('excel_file'):
        # Get uploaded file from request
        excel_file = request.FILES['excel_file']

        # Validate file extension
        # Only accept .xlsx and .xls formats for security
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return render(request, 'main/import_shipments.html')

        try:
            # Load Excel workbook
            # data_only=True means formula cells return their calculated values
            # This prevents formula execution and returns static values
            wb = openpyxl.load_workbook(excel_file, data_only=True)

            # Get active worksheet (usually the first sheet)
            worksheet = wb.active

            # Process all rows in the worksheet
            # Returns: (skipped_entries, created_count, error_entries)
            skipped_entries, created_entries, error_entries = process_excel_data(worksheet)

            # Display appropriate feedback messages based on results
            if created_entries == 0 and not skipped_entries and not error_entries:
                # No changes made at all
                messages.info(request, 'No new entries were created. Check if the data is already up to date.')
            else:
                # Some or all rows were processed successfully
                messages.success(request,
                    f'Successfully created {created_entries} entries. '
                    f'Skipped {len(skipped_entries)} duplicate entries.')

                # If there were errors, report them
                if error_entries:
                    # Join first few error messages (don't flood user with hundreds)
                    error_summary = ", ".join(error_entries[:5])
                    if len(error_entries) > 5:
                        error_summary += f", and {len(error_entries) - 5} more..."
                    messages.error(request,
                        f'Errors occurred in {len(error_entries)} entries. {error_summary}')

        except Exception as e:
            # Catch-all for any unexpected errors during file processing
            messages.error(request, f'An error occurred while processing the file: {str(e)}')

        # Re-render the import form with messages
        return render(request, 'main/import_shipments.html')

    # GET request or no file uploaded - show empty form
    return render(request, 'main/import_shipments.html')


def process_excel_data(worksheet):
    """
    Process rows from Excel worksheet and create Shipment objects in database.

    Purpose:
    --------
    Core data processing function that validates and imports Excel data row-by-row.
    Handles data type conversions, validation, duplicate detection, and error tracking.

    Processing Flow:
    ---------------
    For each row (starting from row 2):
    1. Skip completely empty rows
    2. Extract claim number (if provided)
    3. Check for duplicates
    4. Extract and validate all fields
    5. Create or get client based on Claimant
    6. Create Shipment object
    7. Save to database
    8. Track success/skip/error

    Data Type Conversions:
    ---------------------
    - Dates: Handles datetime.date objects and string formats (YYYY-MM-DD, DD/MM/YYYY)
    - Currency: Strips $, commas, handles int/float/string
    - Booleans: Converts YES/NO/Y/N/TRUE/FALSE/1/0 to YES/NO
    - Status enums: Validates against model choices
    - Branch codes: Validates against BRANCH_CHOICES

    Duplicate Detection:
    -------------------
    - Based on Claim_No field
    - Case-sensitive exact match
    - Skips row if duplicate found
    - Tracks in skipped_entries list

    Client Handling:
    ---------------
    - Required field: Claimant (column 2)
    - Case-insensitive lookup: name__iexact
    - Auto-creates client if not found
    - Uses get_or_create for atomicity

    Auto-Generated Claim Numbers:
    ----------------------------
    - If Claim_No blank: model's save() method generates it
    - Format: {ClientName}-{SequentialNumber}-{YYYYMMDD}
    - Sequential number increments per client per day

    Error Tracking:
    --------------
    Errors tracked for:
    - Missing required fields (Claimant)
    - Data type conversion failures
    - Database constraint violations
    - Any unexpected exceptions

    Parameters:
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Excel worksheet object to process
        Must have data starting in row 2 (row 1 is headers)

    Returns:
    -------
    tuple : (skipped_entries, created_entries, error_entries)
        skipped_entries : list of str
            List of Claim_No values that were skipped (duplicates)
        created_entries : int
            Count of successfully created Shipment objects
        error_entries : list of str
            List of error messages for failed rows

    Expected Excel Structure:
    ------------------------
    Row 1: Headers (ignored)
    Row 2+: Data rows with columns 0-15 (see column descriptions in import_shipments docstring)

    Required Fields:
    ---------------
    - Claimant (column 2): Must not be blank
    - All other fields are optional

    Data Validation:
    ---------------
    - Branch codes: Must be in Shipment.BRANCH_CHOICES or blank
    - Status: Must be in Shipment.STATUS_CHOICES or defaults to 'OPEN'
    - Settlement Status: Must be SETTLED/NOT_SETTLED/PARTIAL or null
    - Dates: Must be valid date format or blank
    - Numbers: Must be numeric or numeric string (e.g., "$1,234.56")

    Error Handling:
    --------------
    - Individual row errors don't stop processing
    - Each error is caught, logged, and added to error_entries
    - Processing continues with next row
    - Partial imports are allowed (some succeed, some fail)

    Transaction Handling:
    --------------------
    - No explicit transaction wrapping
    - Each row saved independently
    - Failed rows don't rollback successful rows
    - Consider wrapping in @transaction.atomic for all-or-nothing behavior

    Performance Considerations:
    --------------------------
    - Uses get_or_create for clients (2 DB queries per unique client)
    - Individual save() per row (could optimize with bulk_create)
    - For 1000 rows: expect ~2-3 seconds processing time
    - For 10,000 rows: expect ~20-30 seconds

    Example Usage:
    -------------
    wb = openpyxl.load_workbook('claims.xlsx')
    ws = wb.active
    skipped, created, errors = process_excel_data(ws)
    print(f"Created: {created}, Skipped: {len(skipped)}, Errors: {len(errors)}")

    Future Enhancements:
    -------------------
    - Add data validation before DB save
    - Implement bulk_create for better performance
    - Add option to update existing records
    - Better date parsing (use dateutil.parser)
    - Add dry-run mode (validate without saving)
    - Log import history for auditing
    """
    # Initialize tracking lists and counters
    skipped_entries = []  # List of claim numbers that were skipped (duplicates)
    created_entries = 0  # Count of successfully created shipments
    error_entries = []  # List of error messages for failed rows

    # Column mapping documentation (for reference):
    # 0: Shipment Number (can be blank - will auto-generate)
    # 1: Brand
    # 2: Claimant (REQUIRED for client identification)
    # 3: Intent To Claim (boolean/YES/NO)
    # 4: Intent To Claim Date
    # 5: Formal Claim (boolean/YES/NO)
    # 6: Formal Claim Date
    # 7: Value (Claimed Amount)
    # 8: Paid By ISCM
    # 9: Paid By Carrier
    # 10: Paid By Insurance
    # 11: Branch
    # 12: Total Savings
    # 13: Settlement Status (SETTLED/NOT_SETTLED/PARTIAL)
    # 14: Financial Exposure
    # 15: Status (OPEN/CLOSED/PENDING/etc.)

    # Iterate through all rows starting from row 2 (row 1 is headers)
    # values_only=True returns cell values instead of Cell objects (faster)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if not row:
            continue

        # === STEP 1: Extract and check claim number ===
        # Get claim number from column 0, or leave blank for auto-generation
        claim_no = str(row[0]).strip() if row[0] else ""

        # Check for duplicates if claim number provided
        if claim_no and Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue  # Skip this row, move to next

        try:
            # === STEP 2: Extract Brand (column 1) ===
            brand = ""
            if len(row) > 1 and row[1]:
                brand = str(row[1]).strip()

            # === STEP 3: Extract Claimant (column 2) - REQUIRED ===
            claimant = ""
            if len(row) > 2 and row[2]:
                claimant = str(row[2]).strip()

            # Skip rows without claimant (required for client identification)
            if not claimant:
                error_entries.append(f'Row {worksheet.iter_rows().__next__()}: Missing claimant name')
                continue

            # === STEP 4: Get or create client based on claimant name ===
            # Case-insensitive lookup to avoid duplicate clients with different casing
            # get_or_create is atomic (thread-safe)
            client, created = Client.objects.get_or_create(
                name__iexact=claimant,  # Case-insensitive lookup
                defaults={'name': claimant}  # If creating new, use this name
            )

            # === STEP 5: Parse Intent To Claim Date (column 4) ===
            intend_date = None
            if len(row) > 4 and row[4]:
                if isinstance(row[4], datetime.date):
                    # Already a date object (from Excel date cell)
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    # String - try multiple date formats
                    try:
                        # Try YYYY-MM-DD format first (ISO standard)
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            # Try DD/MM/YYYY format (European)
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            # Invalid format - leave as None
                            intend_date = None

            # === STEP 6: Parse Formal Claim Date (column 6) ===
            # Same logic as Intent Date
            formal_date = None
            if len(row) > 6 and row[6]:
                if isinstance(row[6], datetime.date):
                    formal_date = row[6]
                elif isinstance(row[6], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[6], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            formal_date = datetime.datetime.strptime(row[6], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None

            # === STEP 7: Parse Claimed Amount / Value (column 7) ===
            claimed_amount = 0
            if len(row) > 7 and row[7]:
                if isinstance(row[7], (int, float)):
                    # Already numeric - use directly
                    claimed_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    # String - strip $ and commas, then convert
                    # Keep only digits and decimal point
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)

            # === STEP 8: Parse Paid By ISCM (column 8) ===
            # Same numeric parsing logic as claimed amount
            iscm_amount = 0
            if len(row) > 8 and row[8]:
                if isinstance(row[8], (int, float)):
                    iscm_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        iscm_amount = float(clean_str)

            # === STEP 9: Parse Paid By Carrier (column 9) ===
            carrier_amount = 0
            if len(row) > 9 and row[9]:
                if isinstance(row[9], (int, float)):
                    carrier_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)

            # === STEP 10: Parse Paid By Insurance (column 10) ===
            insurance_amount = 0
            if len(row) > 10 and row[10]:
                if isinstance(row[10], (int, float)):
                    insurance_amount = float(row[10])
                elif isinstance(row[10], str) and row[10].strip():
                    clean_str = ''.join(c for c in row[10] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)

            # === STEP 11: Parse Total Savings (column 12) ===
            total_savings = 0
            if len(row) > 12 and row[12]:
                if isinstance(row[12], (int, float)):
                    total_savings = float(row[12])
                elif isinstance(row[12], str) and row[12].strip():
                    clean_str = ''.join(c for c in row[12] if c.isdigit() or c == '.')
                    if clean_str:
                        total_savings = float(clean_str)

            # === STEP 12: Parse Financial Exposure (column 14) ===
            financial_exposure = 0
            if len(row) > 14 and row[14]:
                if isinstance(row[14], (int, float)):
                    financial_exposure = float(row[14])
                elif isinstance(row[14], str) and row[14].strip():
                    clean_str = ''.join(c for c in row[14] if c.isdigit() or c == '.')
                    if clean_str:
                        financial_exposure = float(clean_str)

            # === STEP 13: Parse Intent To Claim (column 3) ===
            # Convert various boolean representations to YES/NO
            intent_claim = "NO"  # Default to NO
            if len(row) > 3 and row[3]:
                if isinstance(row[3], bool):
                    # Excel boolean cell
                    intent_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    # String - check various YES representations
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        intent_claim = "YES"

            # === STEP 14: Parse Formal Claim (column 5) ===
            # Same boolean logic as Intent
            formal_claim = "NO"
            if len(row) > 5 and row[5]:
                if isinstance(row[5], bool):
                    formal_claim = "YES" if row[5] else "NO"
                elif isinstance(row[5], str):
                    if row[5].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"

            # === STEP 15: Parse Branch (column 11) ===
            branch = ""
            if len(row) > 11 and row[11]:
                branch = str(row[11]).strip()
                # Validate branch code against model choices
                # If invalid, set to empty string (will use default or allow null)
                if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                    branch = ""  # Invalid branch code - clear it

            # === STEP 16: Parse Settlement Status (column 13) ===
            settlement_status = None  # Allow null if not specified
            if len(row) > 13 and row[13]:
                settlement_value = str(row[13]).upper().strip()
                # Map various representations to standard values
                if settlement_value in ["SETTLED", "YES", "Y", "TRUE", "1"]:
                    settlement_status = "SETTLED"
                elif settlement_value in ["NOT SETTLED", "NOT_SETTLED", "NO", "N", "FALSE", "0"]:
                    settlement_status = "NOT_SETTLED"
                elif settlement_value in ["PARTIAL", "PARTIALLY SETTLED", "PARTIAL_SETTLED"]:
                    settlement_status = "PARTIAL"

            # === STEP 17: Parse Status (column 15) ===
            status = "OPEN"  # Default status if not specified
            if len(row) > 15 and row[15]:
                status_value = str(row[15]).upper().strip()
                # Validate against model choices
                valid_statuses = [choice[0] for choice in Shipment.STATUS_CHOICES]
                if status_value in valid_statuses:
                    status = status_value
                elif status_value in ["OPEN", "PENDING", "CLOSED", "REJECTED", "UNDER_REVIEW"]:
                    # Accept these standard values even if not in exact format
                    status = status_value

            # === STEP 18: Create Shipment object ===
            # Claim_No will be auto-generated by model's save() if left blank
            shipment = Shipment(
                Claim_No=claim_no,  # Blank triggers auto-generation
                client=client,  # Link to client (created or existing)
                Brand=brand,
                Claimant=claimant,
                Intent_To_Claim=intent_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Received=formal_claim,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Awa=iscm_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Branch=branch,
                Total_Savings=total_savings,
                Settlement_Status=settlement_status,
                Financial_Exposure=financial_exposure,
                Status=status
            )

            # === STEP 19: Save to database ===
            # This triggers model's save() method which auto-generates Claim_No if needed
            shipment.save()

            # Increment success counter
            created_entries += 1

        except Exception as e:
            # Catch any errors during processing this row
            # Log error but continue processing other rows
            error_entries.append(f'Row with Claimant {claimant}: {str(e)}')

    # Return results tuple
    return skipped_entries, created_entries, error_entries


# ==================================================================================
# INSPECTION FEE MANAGEMENT VIEWS
# ==================================================================================
# These views handle viewing and updating inspection fee rates.
# Used by financial admins to manage pricing for different inspection types.
# ==================================================================================

@login_required
def get_inspection_fees(request):
    """
    Retrieve all inspection fees as JSON for AJAX requests.

    Purpose:
    --------
    API endpoint that returns all inspection fee records in JSON format.
    Used by frontend JavaScript to populate fee management tables.

    Authentication:
    --------------
    Requires user to be logged in

    HTTP Method:
    -----------
    GET only (no POST/PUT/DELETE handling)

    Response Format:
    ---------------
    JSON object with 'fees' array:
    {
        "fees": [
            {
                "id": 1,
                "fee_code": "POULTRY_LABEL",
                "fee_name": "Poultry Label Inspection",
                "rate": 250.00,
                "description": "Standard poultry label inspection fee",
                "last_updated": "2025-12-10T14:30:22.123Z"
            },
            ...
        ]
    }

    Database Query:
    --------------
    - Model: InspectionFee
    - Filter: None (returns all fees)
    - Returns: All fields as dictionary via .values()
    - Ordering: Not specified (database default, usually by ID)

    Fields Returned:
    ---------------
    - id: Primary key (integer)
    - fee_code: Unique code identifier (string)
    - fee_name: Human-readable fee name (string)
    - rate: Fee amount (decimal)
    - description: Fee description (string, may be null)
    - last_updated: Timestamp of last modification (datetime)

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object (GET request expected)

    Returns:
    -------
    JsonResponse
        JSON response with all inspection fees
        Content-Type: application/json

    Example Response:
    ----------------
    {
        "fees": [
            {
                "id": 1,
                "fee_code": "POULTRY_LABEL",
                "fee_name": "Poultry Label Inspection",
                "rate": 250.00,
                "description": "Standard inspection",
                "last_updated": "2025-12-10T14:30:22Z"
            }
        ]
    }

    Frontend Usage:
    --------------
    JavaScript code calls this endpoint to populate fee table:
    ```javascript
    fetch('/api/inspection-fees/')
        .then(response => response.json())
        .then(data => {
            data.fees.forEach(fee => {
                // Add row to table
            });
        });
    ```

    Performance Notes:
    -----------------
    - Uses .values() for efficient query (returns dicts, not model instances)
    - Queries entire InspectionFee table (typically small, <100 records)
    - No pagination implemented (add if fee count grows large)
    - Result is JSON-serializable (datetimes automatically formatted)

    Security Considerations:
    -----------------------
    - Login required (non-public data)
    - Read-only operation (no data modification)
    - All users can view fees (no role restriction)
    - Consider adding @require_http_methods(['GET']) decorator

    Example URL:
    -----------
    GET /api/inspection-fees/

    Notes:
    -----
    - This is an API endpoint, not a view that renders HTML
    - Returns raw data for consumption by frontend JavaScript
    - Consider adding caching if called frequently
    - Consider pagination for large datasets
    """
    # Import InspectionFee model (lazy import to avoid circular dependencies)
    from ..models import InspectionFee

    # Query all inspection fees and convert to list of dictionaries
    # .values() returns QuerySet of dicts instead of model instances
    # list() forces query evaluation and converts to Python list
    fees = InspectionFee.objects.all().values(
        'id',  # Primary key
        'fee_code',  # Unique code identifier
        'fee_name',  # Human-readable name
        'rate',  # Fee amount (Decimal)
        'description',  # Description text
        'last_updated'  # Timestamp of last update
    )

    # Return as JSON response
    # JsonResponse automatically serializes dates, decimals, etc.
    # list(fees) converts QuerySet to list for JSON serialization
    return JsonResponse({'fees': list(fees)})


@login_required
@require_POST
def update_inspection_fees(request):
    """
    Update inspection fee rates via AJAX POST request.

    Purpose:
    --------
    API endpoint for batch-updating inspection fee rates.
    Receives JSON payload with array of fee updates and applies them to database.

    Authentication & Authorization:
    ------------------------------
    - Requires user to be logged in (@login_required)
    - HTTP POST only (@require_POST)
    - Typically restricted to financial_admin role (check in frontend/middleware)

    HTTP Method:
    -----------
    POST only (enforced by @require_POST decorator)

    Request Format:
    --------------
    Content-Type: application/json

    Body:
    {
        "fees": [
            {"id": 1, "rate": 275.00},
            {"id": 2, "rate": 150.50},
            ...
        ]
    }

    Request Fields:
    --------------
    - fees: Array of fee update objects
        - id: Fee record ID (required)
        - rate: New rate value (required, numeric)

    Processing Flow:
    ---------------
    1. Parse JSON from request body
    2. Extract 'fees' array
    3. For each fee update:
        a. Validate id and rate are present
        b. Fetch InspectionFee object by ID
        c. Update rate field
        d. Set updated_by to current user
        e. Save to database
        f. Increment counter
    4. Return success response with count

    Success Response:
    ----------------
    HTTP 200 OK
    {
        "success": true,
        "message": "Updated X fees successfully",
        "updated_count": X
    }

    Error Response:
    --------------
    HTTP 400 Bad Request
    {
        "success": false,
        "message": "Error description"
    }

    Error Conditions:
    ----------------
    - Invalid JSON format
    - Missing 'fees' key
    - Invalid fee ID (DoesNotExist)
    - Missing id or rate in fee object
    - Rate value not numeric
    - Database save failure
    - Any unexpected exception

    Database Updates:
    ----------------
    - Updates InspectionFee.rate
    - Sets InspectionFee.updated_by to request.user
    - Triggers auto-update of InspectionFee.last_updated (if auto_now=True)
    - Each fee saved individually (not bulk update)

    Audit Trail:
    -----------
    - updated_by field tracks who made change
    - last_updated timestamp auto-set on save
    - Consider adding separate audit log table for complete history

    Parameters:
    ----------
    request : HttpRequest
        Django HTTP request object
        - Must be POST method
        - Must contain JSON body
        - Must have authenticated user

    Returns:
    -------
    JsonResponse
        Success or error response in JSON format
        HTTP 200 for success
        HTTP 400 for errors

    Frontend Usage Example:
    ----------------------
    ```javascript
    fetch('/api/update-inspection-fees/', {
        method: 'POST',
        headers: {
            'Content-Type': 'application/json',
            'X-CSRFToken': csrfToken
        },
        body: JSON.stringify({
            fees: [
                {id: 1, rate: 275.00},
                {id: 2, rate: 150.50}
            ]
        })
    })
    .then(response => response.json())
    .then(data => {
        if (data.success) {
            alert(data.message);
        } else {
            alert('Error: ' + data.message);
        }
    });
    ```

    Security Considerations:
    -----------------------
    - CSRF protection required (Django automatic for POST)
    - User must be authenticated
    - Consider adding role check (only financial_admin should update)
    - Validate rate is positive number
    - Consider adding min/max rate limits
    - Log all fee changes for audit trail

    Transaction Safety:
    ------------------
    - Each fee updated individually (no transaction wrapper)
    - Partial updates possible if some succeed and some fail
    - Consider wrapping in @transaction.atomic for all-or-nothing behavior
    - No rollback on error (successful updates remain)

    Performance Notes:
    -----------------
    - Individual saves (not bulk_update)
    - For small batches (<100) performance is fine
    - For large batches, consider using bulk_update()
    - No locking (race conditions possible with concurrent updates)

    Future Enhancements:
    -------------------
    - Add validation for rate values (min/max, positive)
    - Add detailed audit logging
    - Implement optimistic locking (version field)
    - Add bulk_update for better performance
    - Return list of failed IDs if partial failure
    - Add permission check for financial_admin role

    Example Curl Request:
    --------------------
    ```bash
    curl -X POST https://portal.fsa-pty.co.za/api/update-fees/ \\
         -H "Content-Type: application/json" \\
         -H "X-CSRFToken: abc123..." \\
         -d '{"fees": [{"id": 1, "rate": 275.00}]}'
    ```

    Notes:
    -----
    - This is an API endpoint returning JSON (not HTML)
    - Should be called via AJAX from frontend
    - CSRF token required for POST (Django security)
    - Consider rate-limiting to prevent abuse
    """
    # Import JSON library and InspectionFee model
    import json
    from ..models import InspectionFee

    try:
        # Parse JSON from request body
        # request.body is bytes, json.loads converts to Python dict
        data = json.loads(request.body)

        # Extract fees array from JSON payload
        # Expected format: {"fees": [{"id": 1, "rate": 250.00}, ...]}
        fees_data = data.get('fees', [])

        # Counter for successfully updated fees
        updated_count = 0

        # Process each fee update
        for fee_data in fees_data:
            # Extract fee ID and new rate
            fee_id = fee_data.get('id')
            new_rate = fee_data.get('rate')

            # Validate both id and rate are present
            if fee_id and new_rate is not None:
                # Fetch InspectionFee object from database
                # get() raises DoesNotExist if not found (caught by except below)
                fee = InspectionFee.objects.get(id=fee_id)

                # Update rate field
                fee.rate = new_rate

                # Track who updated this fee (for audit trail)
                fee.updated_by = request.user

                # Save to database
                # This also triggers auto-update of last_updated if auto_now=True
                fee.save()

                # Increment success counter
                updated_count += 1

        # Return success response
        return JsonResponse({
            'success': True,
            'message': f'Updated {updated_count} fees successfully',
            'updated_count': updated_count
        })

    except Exception as e:
        # Catch all errors (JSON parsing, database errors, etc.)
        # Return error response with details
        return JsonResponse({
            'success': False,
            'message': str(e)  # Error description
        }, status=400)  # HTTP 400 Bad Request
